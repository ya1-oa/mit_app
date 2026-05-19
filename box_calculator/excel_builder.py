"""
PPR Box Count Excel report builder.

Generates a formatted .xlsx report matching the PPR standard:
  Room | Small | Medium | Large | Box/Wrapped | Plant/Vase | TV |
  Wardrobe | Mattress | Dish Pack | Glass Pack | Boots & Pans | Total

Styling mirrors the team's existing report format.
"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .ppr_analyzer import PPR_COLUMNS, PPR_COLUMN_LABELS

_HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")   # light blue
_TOTAL_FILL  = PatternFill("solid", fgColor="FFF2CC")   # light yellow
_ERROR_FILL  = PatternFill("solid", fgColor="FFE0E0")   # light red (error rows)

_thin = Side(style="thin")
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _ppr_room_to_dict(ppr_room) -> dict:
    """Convert a BoxCalcPPRRoom ORM instance to a plain counts dict."""
    return {col: getattr(ppr_room, col, 0) or 0 for col in PPR_COLUMNS}


def build_ppr_excel(session) -> bytes:
    """
    Build the PPR box count Excel report for a BoxCalcPPRSession.

    Returns the raw .xlsx bytes ready to stream as a file download.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "PPR Box Count Report"

    # ── Title block ──────────────────────────────────────────────────────────
    ws.merge_cells("A1:M1")
    title_cell = ws["A1"]
    title_cell.value = f"Pre-Packout Report — {session.client.pOwner}"
    title_cell.font = Font(bold=True, size=13)
    title_cell.alignment = Alignment(horizontal="center")

    if session.client.claimNumber:
        ws.merge_cells("A2:M2")
        sub_cell = ws["A2"]
        sub_cell.value = f"Claim #{session.client.claimNumber}"
        sub_cell.alignment = Alignment(horizontal="center")
        sub_cell.font = Font(italic=True, color="555555")

    ws.append([])  # blank row

    # ── Headers ───────────────────────────────────────────────────────────────
    headers = ["Room"] + [PPR_COLUMN_LABELS[c] for c in PPR_COLUMNS] + ["Total"]
    ws.append(headers)

    header_row_idx = ws.max_row
    for cell in ws[header_row_idx]:
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = _BORDER

    # ── Room data rows ────────────────────────────────────────────────────────
    rooms = session.rooms.order_by("order", "room_name").all()
    grand = {col: 0 for col in PPR_COLUMNS}
    room_rows = []

    for ppr_room in rooms:
        counts = _ppr_room_to_dict(ppr_room)
        row_total = sum(counts.values())
        row = [ppr_room.room_name] + [counts[c] or "" for c in PPR_COLUMNS] + [row_total or ""]
        ws.append(row)
        room_rows.append(ws.max_row)

        for col in PPR_COLUMNS:
            grand[col] += counts[col]

        # Style data cells
        row_obj = ws[ws.max_row]
        for cell in row_obj:
            cell.border = _BORDER
            cell.alignment = Alignment(horizontal="center")

        # Dim error/incomplete rows
        if ppr_room.status == "error":
            for cell in row_obj:
                cell.fill = _ERROR_FILL

    # ── Blank separator ───────────────────────────────────────────────────────
    ws.append([])

    # ── Grand totals row ──────────────────────────────────────────────────────
    grand_total = sum(grand.values())
    grand_row = ["GRAND TOTALS"] + [grand[c] or "" for c in PPR_COLUMNS] + [grand_total]
    ws.append(grand_row)

    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
        cell.fill = _TOTAL_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = _BORDER

    # ── Column widths ─────────────────────────────────────────────────────────
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = max(
            (len(str(cell.value)) if cell.value is not None else 0)
            for cell in col_cells
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = max(max_len + 3, 8)

    # Wider room name column
    ws.column_dimensions["A"].width = 22

    ws.freeze_panes = f"A{header_row_idx + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
