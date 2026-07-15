"""
Build a Schedule of Loss Excel file using openpyxl.
Print-optimised: landscape, fit to 1 page wide, page-break per room,
repeating header rows.
"""
from __future__ import annotations
import datetime
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.pagebreak import Break


# ── Column definitions ────────────────────────────────────────────────────────
# Removed: Disposition, Condition, Model #, Serial #, Repl.Source,
#          Purchase Each, Purchase Total.
# Kept: #, Room, Description, Brand, QTY, Retailer, Age Y, Age M,
#       Repl.Value Each, Repl.Value Total.
COLUMNS = [
    # (header_row3,       header_row4,  width)
    ('#',                 '',           4),
    ('Room',              '',          18),
    ('Description',       '',          28),
    ('Brand',             '',          14),
    ('QTY',               '',           5),
    ('Retailer',          '',          16),
    ('Age',               'Y',          5),
    ('',                  'M',          5),
    ('Repl.\nValue',      'Each',      12),
    ('',                  'Total',     12),
]

NUM_COLS    = len(COLUMNS)
LAST_COL    = get_column_letter(NUM_COLS)

COL_RV_EACH  = 9
COL_RV_TOTAL = 10

CURRENCY_COLS = {9, 10}
PCT_COLS      = set()
NUM_COLS_SET  = {5, 7, 8}   # QTY, Age Y, Age M

# Colour palette
CLR_HEADER_BG    = 'FF1F3864'
CLR_HEADER_FG    = 'FFFFFFFF'
CLR_SUBHEADER_BG = 'FF2F5496'
CLR_SUBHEADER_FG = 'FFFFFFFF'
CLR_ROOM_BG      = 'FFD6E4BC'
CLR_TOTAL_BG     = 'FFFFF2CC'
CLR_ALT_ROW      = 'FFF2F2F2'
CLR_SIG_SIGNED   = 'FFD1FAE5'
CLR_SIG_PENDING  = 'FFFAFAFA'
CLR_SIG_LINK     = 'FF2563EB'


# ── Style helpers ─────────────────────────────────────────────────────────────

def _font(bold=False, color='FF000000', size=9):
    return Font(name='Calibri', bold=bold, color=color, size=size)

def _fill(hex_color):
    return PatternFill(fill_type='solid', fgColor=hex_color)

def _border():
    s = Side(style='thin', color='FFB8B8B8')
    return Border(left=s, right=s, top=s, bottom=s)

def _center(wrap=False):
    return Alignment(horizontal='center', vertical='center', wrap_text=wrap)

def _left(wrap=False):
    return Alignment(horizontal='left', vertical='center', wrap_text=wrap)

def _right():
    return Alignment(horizontal='right', vertical='center')

def _currency_fmt():
    return '$#,##0.00'


# ── Page setup ────────────────────────────────────────────────────────────────

def _apply_print_settings(ws):
    """Landscape, fit to 1 page wide, repeat header rows, narrow margins."""
    ws.page_setup.orientation           = 'landscape'
    ws.page_setup.fitToPage             = True
    ws.page_setup.fitToWidth            = 1
    ws.page_setup.fitToHeight           = 0
    ws.page_setup.paperSize             = ws.PAPERSIZE_LETTER
    ws.print_title_rows                 = '1:4'   # rows 1-4 repeat on every page
    ws.page_margins = PageMargins(
        left=0.4, right=0.4, top=0.5, bottom=0.5,
        header=0.3, footer=0.3,
    )


# ── Header builders ───────────────────────────────────────────────────────────

def _build_header_rows(ws, session):
    """Rows 1–2: compact claim metadata banner."""
    # Row 1 — title
    ws.row_dimensions[1].height = 22
    ws.merge_cells(f'A1:{LAST_COL}1')
    c = ws['A1']
    c.value     = 'Schedule of Loss — Detailed Report'
    c.font      = _font(bold=True, color=CLR_HEADER_FG, size=13)
    c.fill      = _fill(CLR_HEADER_BG)
    c.alignment = _center()

    # Row 2 — all claim info on one line
    ws.row_dimensions[2].height = 20
    ws.merge_cells(f'A2:{LAST_COL}2')
    c2 = ws['A2']

    _human_id  = getattr(session.client, 'claimID', '') or ''
    _enc_id    = session.encircle_claim_id or ''
    _name      = session.insured_name or ''
    _claim_num = session.claim_number or ''
    _loss_date = session.loss_date.strftime('%b %d, %Y') if session.loss_date else ''
    _today     = datetime.date.today().strftime('%b %d, %Y')

    parts = []
    if _name:                       parts.append(f'Insured: {_name}')
    if _human_id:                   parts.append(f'Claim ID: {_human_id}')
    if _enc_id:                     parts.append(f'Customer ID: {_enc_id}')
    if _loss_date:                  parts.append(f'Date of Loss: {_loss_date}')
    if _claim_num:                  parts.append(f'Claim #: {_claim_num}')
    parts.append(f'Report Date: {_today}')

    c2.value     = '   •   '.join(parts)
    c2.font      = _font(color=CLR_HEADER_FG, size=8)
    c2.fill      = _fill(CLR_SUBHEADER_BG)
    c2.alignment = _left()


def _build_column_headers(ws, start_row=3):
    """Rows 3–4: two-row column headers."""
    ws.row_dimensions[start_row].height     = 28
    ws.row_dimensions[start_row + 1].height = 14

    col_groups: dict[str, list[int]] = {}
    for col_idx, (h3, _, _) in enumerate(COLUMNS, start=1):
        if h3:
            col_groups.setdefault(h3, []).append(col_idx)

    written: set[str] = set()
    for col_idx, (h3, h4, _) in enumerate(COLUMNS, start=1):
        c3 = ws.cell(row=start_row,     column=col_idx)
        c4 = ws.cell(row=start_row + 1, column=col_idx)

        for c in (c3, c4):
            c.border = _border()

        c3.fill      = _fill(CLR_HEADER_BG)
        c3.font      = _font(bold=True, color=CLR_HEADER_FG, size=8)
        c3.alignment = _center(wrap=True)

        c4.fill      = _fill(CLR_SUBHEADER_BG)
        c4.font      = _font(bold=True, color=CLR_HEADER_FG, size=8)
        c4.alignment = _center()

        if h3 and h3 not in written:
            siblings = col_groups.get(h3, [col_idx])
            if len(siblings) > 1:
                start_c = get_column_letter(siblings[0])
                end_c   = get_column_letter(siblings[-1])
                try:
                    ws.merge_cells(f'{start_c}{start_row}:{end_c}{start_row}')
                except Exception:
                    pass
            ws.cell(row=start_row, column=col_idx).value = h3
            written.add(h3)

        if h4:
            c4.value = h4
        elif h3:
            try:
                col_l = get_column_letter(col_idx)
                ws.merge_cells(f'{col_l}{start_row}:{col_l}{start_row+1}')
            except Exception:
                pass


# ── Per-room writers ──────────────────────────────────────────────────────────

def _write_room_header(ws, row: int, room) -> None:
    ws.row_dimensions[row].height = 18
    ws.merge_cells(f'A{row}:{LAST_COL}{row}')
    c = ws[f'A{row}']
    c.value     = f"{room.room_number}  —  {room.room_name}"
    c.font      = _font(bold=True, size=10)
    c.fill      = _fill(CLR_ROOM_BG)
    c.alignment = _left()
    c.border    = _border()


def _write_item_row(ws, row: int, item, item_num: int, room_name: str = '',
                    alt: bool = False) -> None:
    ws.row_dimensions[row].height = 14
    fill = _fill(CLR_ALT_ROW) if alt else None

    rv_each  = float(item.replacement_value_each or 0)
    qty      = item.qty or 1
    rv_total = rv_each * qty

    row_data = [
        item_num,
        room_name,
        item.description,
        item.brand or '',
        item.qty,
        item.retailer or '',
        item.age_years if item.age_years is not None else '',
        item.age_months if item.age_months is not None else '',
        rv_each,
        rv_total,
    ]

    for col_idx, value in enumerate(row_data, start=1):
        c = ws.cell(row=row, column=col_idx, value=value)
        c.border    = _border()
        c.font      = _font(size=9)
        c.alignment = _left(wrap=(col_idx == 3))
        if fill:
            c.fill = fill
        if col_idx in CURRENCY_COLS and isinstance(value, (int, float)):
            c.number_format = _currency_fmt()
            c.alignment     = _right()
        elif col_idx in PCT_COLS and isinstance(value, (int, float)):
            c.number_format = '0.0"%"'
            c.alignment     = _center()
        elif col_idx in NUM_COLS_SET:
            c.alignment = _center()


def _write_room_total_row(ws, row: int, room) -> None:
    ws.row_dimensions[row].height = 15
    label_end = get_column_letter(COL_RV_EACH - 1)
    ws.merge_cells(f'A{row}:{label_end}{row}')
    ws[f'A{row}'].value     = f"Room Total — {room.room_name}"
    ws[f'A{row}'].font      = _font(bold=True, size=9)
    ws[f'A{row}'].fill      = _fill(CLR_TOTAL_BG)
    ws[f'A{row}'].alignment = _right()

    for col_idx in range(1, NUM_COLS + 1):
        c = ws.cell(row=row, column=col_idx)
        c.fill   = _fill(CLR_TOTAL_BG)
        c.border = _border()

    items = list(room.items.filter(structural=False))
    if items:
        rv_sum = sum(float(i.replacement_value_each or 0) * (i.qty or 1) for i in items)
        c = ws.cell(row=row, column=COL_RV_TOTAL, value=rv_sum)
        c.number_format = _currency_fmt()
        c.font          = _font(bold=True, size=9)
        c.fill          = _fill(CLR_TOTAL_BG)
        c.alignment     = _right()
        c.border        = _border()


def _write_grand_total(ws, row: int, session) -> None:
    ws.row_dimensions[row].height = 18
    label_end = get_column_letter(COL_RV_EACH - 1)
    ws.merge_cells(f'A{row}:{label_end}{row}')
    ws[f'A{row}'].value     = 'GRAND TOTAL'
    ws[f'A{row}'].font      = _font(bold=True, color=CLR_HEADER_FG, size=10)
    ws[f'A{row}'].fill      = _fill(CLR_HEADER_BG)
    ws[f'A{row}'].alignment = _right()

    for col_idx in range(1, NUM_COLS + 1):
        c = ws.cell(row=row, column=col_idx)
        c.fill   = _fill(CLR_HEADER_BG)
        c.border = _border()

    all_items = [
        item
        for room in session.rooms.prefetch_related('items').all()
        for item in room.items.filter(structural=False)
    ]
    rv_grand = sum(float(i.replacement_value_each or 0) * (i.qty or 1) for i in all_items)
    c = ws.cell(row=row, column=COL_RV_TOTAL, value=rv_grand)
    c.number_format = _currency_fmt()
    c.font          = _font(bold=True, color=CLR_HEADER_FG, size=10)
    c.fill          = _fill(CLR_HEADER_BG)
    c.alignment     = _right()
    c.border        = _border()


def _write_room_signature_row(ws, row: int, room, share_url: str | None = None) -> None:
    ws.row_dimensions[row].height = 16
    ws.merge_cells(f'A{row}:{LAST_COL}{row}')
    c = ws[f'A{row}']

    for col_idx in range(1, NUM_COLS + 1):
        ws.cell(row=row, column=col_idx).border = _border()

    if room.signature_name and room.signed_at:
        signed_str = room.signed_at.strftime('%B %d, %Y at %I:%M %p')
        c.value     = f"✔  Signed by: {room.signature_name}   |   {signed_str}"
        c.font      = Font(name='Calibri', bold=True, color='FF065F46', size=9)
        c.fill      = _fill(CLR_SIG_SIGNED)
    else:
        if share_url:
            c.value     = f"⬜  Awaiting client signature — Share link: {share_url}"
            c.hyperlink = share_url
            c.font      = Font(name='Calibri', color=CLR_SIG_LINK, size=9, underline='single')
        else:
            c.value = "⬜  Awaiting client signature"
            c.font  = Font(name='Calibri', color='FF64748B', size=9, italic=True)
        c.fill = _fill(CLR_SIG_PENDING)
    c.alignment = _left()


# ── Main entry point ──────────────────────────────────────────────────────────

def build_excel(session, share_url: str | None = None) -> bytes:
    """
    Generate the Schedule of Loss Excel file for a CPSReportSession.
    Returns raw bytes of the .xlsx file.

    Print layout: landscape, fit-to-1-page-wide, one room per printed page.
    Requires session to have been loaded with select_related('client').
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Schedule of Loss'

    _apply_print_settings(ws)
    ws.freeze_panes = 'A5'

    for col_idx, (_, _, width) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Header (rows 1-2) + column headers (rows 3-4)
    _build_header_rows(ws, session)
    _build_column_headers(ws, start_row=3)

    current_row     = 5
    global_item_num = 1
    rooms = list(
        session.rooms
        .prefetch_related('items')
        .order_by('order', 'room_number')
        .all()
    )

    for room_idx, room in enumerate(rooms):
        items = list(room.items.filter(structural=False).order_by('order'))
        if not items:
            continue

        _write_room_header(ws, current_row, room)
        current_row += 1

        room_label = f"{room.room_number} {room.room_name}".strip()
        for i, item in enumerate(items):
            _write_item_row(ws, current_row, item, global_item_num,
                            room_name=room_label, alt=(i % 2 == 1))
            global_item_num += 1
            current_row += 1

        _write_room_total_row(ws, current_row, room)
        current_row += 1

        _write_room_signature_row(ws, current_row, room, share_url=share_url)
        current_row += 1

        if room_idx < len(rooms) - 1:
            pb = Break(id=current_row - 1)
            ws.row_breaks.append(pb)
            current_row += 1

    _write_grand_total(ws, current_row, session)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
