"""
Build per-room summary views for a CPSReportSession.

A "summary" condenses each room to a single row:
  Room | # Items | RCV Total

Outputs: HTML context dict, PDF bytes (ReportLab), Excel bytes (openpyxl).
"""
from __future__ import annotations

import datetime
from io import BytesIO


# ── Data helpers ──────────────────────────────────────────────────────────────

def compute_summary(session) -> dict:
    """Return a summary dict ready for template context or export builders."""
    rows = []
    grand_items = 0
    grand_rcv = 0.0

    for room in session.rooms.prefetch_related('items').order_by('order', 'room_number'):
        items = list(room.items.all())
        rcv = sum(float(i.replacement_value_each or 0) * (i.qty or 1) for i in items)

        grand_items += len(items)
        grand_rcv   += rcv

        rows.append({
            'room_number':    room.room_number,
            'room_name':      room.room_name,
            'item_count':     len(items),
            'rcv_total':      rcv,
            'status':         room.status,
            'ai_confidence':  room.ai_confidence,
        })

    return {
        'session':      session,
        'rows':         rows,
        'grand_items':  grand_items,
        'grand_rcv':    grand_rcv,
        'generated_at': datetime.datetime.now(),
    }


# ── PDF builder ───────────────────────────────────────────────────────────────

def build_summary_pdf(session) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        BaseDocTemplate, Frame, PageTemplate,
        Paragraph, Spacer, Table, TableStyle,
    )

    C_HEADER_BG = colors.HexColor('#1e40af')
    C_HEADER_FG = colors.white
    C_ROOM_BG   = colors.HexColor('#059669')
    C_TOTAL_BG  = colors.HexColor('#1e3a5f')
    C_ALT       = colors.HexColor('#f0fdf4')
    C_TEXT      = colors.HexColor('#0f172a')
    C_MUTED     = colors.HexColor('#64748b')
    C_RULE      = colors.HexColor('#e2e8f0')

    def _fmt(v):
        try:
            return f"${float(v):,.2f}"
        except (TypeError, ValueError):
            return "$0.00"

    def _hf(canvas, doc):
        canvas.saveState()
        w, h = letter
        canvas.setStrokeColor(C_RULE)
        canvas.setLineWidth(0.5)
        canvas.line(0.6 * inch, h - 0.45 * inch, w - 0.6 * inch, h - 0.45 * inch)
        canvas.setFont('Helvetica', 7)
        canvas.setFillColor(C_MUTED)
        canvas.drawString(0.6 * inch, 0.35 * inch, 'CPS Summary Report — Confidential')
        canvas.drawRightString(w - 0.6 * inch, 0.35 * inch, f'Page {doc.page}')
        canvas.restoreState()

    data = compute_summary(session)
    buf  = BytesIO()
    styles = getSampleStyleSheet()

    h1   = ParagraphStyle('H1', fontSize=20, textColor=C_HEADER_FG, fontName='Helvetica-Bold', leading=24)
    h2   = ParagraphStyle('H2', fontSize=11, textColor=C_HEADER_FG, fontName='Helvetica-Bold', leading=14)
    body = ParagraphStyle('Body', fontSize=9, textColor=C_TEXT, leading=12)
    muted= ParagraphStyle('Muted', fontSize=8, textColor=C_MUTED, leading=10)

    doc = BaseDocTemplate(
        buf, pagesize=letter,
        leftMargin=0.6 * inch, rightMargin=0.6 * inch,
        topMargin=0.7 * inch, bottomMargin=0.6 * inch,
    )
    pw = letter[0] - 1.2 * inch

    frame = Frame(doc.leftMargin, doc.bottomMargin, pw, letter[1] - 1.3 * inch, id='main')
    doc.addPageTemplates([PageTemplate(id='all', frames=[frame], onPage=_hf)])

    story = []

    # Cover block
    cover_data = [[Paragraph('CPS Summary Report', h1)]]
    cover_tbl = Table(cover_data, colWidths=[pw])
    cover_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), C_HEADER_BG),
        ('TOPPADDING',    (0, 0), (-1, -1), 18),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 18),
        ('LEFTPADDING',   (0, 0), (-1, -1), 14),
    ]))
    story.append(cover_tbl)
    story.append(Spacer(1, 6))

    meta_lines = [
        f"Insured: {session.insured_name or session.client.pOwner or '—'}",
        f"Claim #: {session.claim_number or session.encircle_claim_id or '—'}",
        f"Generated: {data['generated_at'].strftime('%B %d, %Y %I:%M %p')}",
    ]
    for line in meta_lines:
        story.append(Paragraph(line, muted))
    story.append(Spacer(1, 14))

    # Table header
    col_widths = [pw * 0.07, pw * 0.33, pw * 0.12, pw * 0.48]
    hdr = ['#', 'Room', 'Items', 'RCV Total']
    table_data = [hdr]

    for row in data['rows']:
        table_data.append([
            row['room_number'],
            row['room_name'],
            str(row['item_count']),
            _fmt(row['rcv_total']),
        ])

    table_data.append([
        '', 'GRAND TOTAL',
        str(data['grand_items']),
        _fmt(data['grand_rcv']),
    ])

    tbl = Table(table_data, colWidths=col_widths, repeatRows=1)

    ts = [
        ('BACKGROUND',    (0, 0), (-1, 0), C_ROOM_BG),
        ('TEXTCOLOR',     (0, 0), (-1, 0), colors.white),
        ('FONTNAME',      (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0), (-1, -1), 8),
        ('ALIGN',         (2, 0), (-1, -1), 'RIGHT'),
        ('ALIGN',         (0, 0), (1, -1), 'LEFT'),
        ('TOPPADDING',    (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING',   (0, 0), (-1, -1), 6),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 6),
        ('GRID',          (0, 0), (-1, -1), 0.3, C_RULE),
        # Grand total row
        ('BACKGROUND',    (0, -1), (-1, -1), C_TOTAL_BG),
        ('TEXTCOLOR',     (0, -1), (-1, -1), colors.white),
        ('FONTNAME',      (0, -1), (-1, -1), 'Helvetica-Bold'),
    ]
    # Alternating row tints
    for i, _ in enumerate(data['rows'], start=1):
        if i % 2 == 0:
            ts.append(('BACKGROUND', (0, i), (-1, i), C_ALT))

    tbl.setStyle(TableStyle(ts))
    story.append(tbl)

    doc.build(story)
    return buf.getvalue()


# ── Excel builder ─────────────────────────────────────────────────────────────

def build_summary_excel(session) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side,
    )
    from openpyxl.utils import get_column_letter

    def _fmt(v):
        try:
            return round(float(v), 2)
        except (TypeError, ValueError):
            return 0.0

    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    BLUE_FILL   = PatternFill('solid', fgColor='1E40AF')
    GREEN_FILL  = PatternFill('solid', fgColor='059669')
    YELLOW_FILL = PatternFill('solid', fgColor='FFF2CC')
    NAVY_FILL   = PatternFill('solid', fgColor='1E3A5F')
    ALT_FILL    = PatternFill('solid', fgColor='F0FDF4')

    WHITE_BOLD  = Font(bold=True, color='FFFFFF', size=10)
    WHITE_FONT  = Font(color='FFFFFF', size=9)
    BOLD        = Font(bold=True, size=9)
    NORMAL      = Font(size=9)

    data = compute_summary(session)

    wb = Workbook()
    ws = wb.active
    ws.title = 'PPR Summary'
    ws.freeze_panes = 'A5'

    # Title
    ws.merge_cells('A1:D1')
    c = ws['A1']
    c.value = 'PPR Schedule of Loss — Summary'
    c.font  = Font(bold=True, color='FFFFFF', size=14)
    c.fill  = BLUE_FILL
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # Meta row
    ws.merge_cells('A2:D2')
    insured = session.insured_name or (session.client.pOwner if session.client else '—')
    claim   = session.claim_number or session.encircle_claim_id or '—'
    ws['A2'].value = f"Insured: {insured}   |   Claim #: {claim}"
    ws['A2'].font  = NORMAL
    ws['A2'].fill  = PatternFill('solid', fgColor='EFF6FF')
    ws['A2'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A3:D3')
    ws['A3'].value = f"Generated: {data['generated_at'].strftime('%B %d, %Y %I:%M %p')}"
    ws['A3'].font  = Font(size=8, color='64748B')
    ws['A3'].alignment = Alignment(horizontal='center')

    ws.row_dimensions[4].height = 4  # spacer

    # Header row (row 5)
    headers = ['Room #', 'Room Name', '# Items', 'RCV Total']
    for col, hdr in enumerate(headers, 1):
        c = ws.cell(row=5, column=col, value=hdr)
        c.font      = WHITE_BOLD
        c.fill      = GREEN_FILL
        c.alignment = Alignment(horizontal='center' if col > 2 else 'left', vertical='center')
        c.border    = border
    ws.row_dimensions[5].height = 18

    # Data rows
    for i, row in enumerate(data['rows'], start=6):
        fill = ALT_FILL if i % 2 == 0 else None
        vals = [
            row['room_number'],
            row['room_name'],
            row['item_count'],
            _fmt(row['rcv_total']),
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.font   = NORMAL
            c.border = border
            c.alignment = Alignment(
                horizontal='right' if col > 2 else 'left',
                vertical='center',
            )
            if fill:
                c.fill = fill
            if col >= 4:
                c.number_format = '"$"#,##0.00'

    # Grand total row
    gt_row = 6 + len(data['rows'])
    gt_vals = ['', 'GRAND TOTAL', data['grand_items'], _fmt(data['grand_rcv'])]
    for col, val in enumerate(gt_vals, 1):
        c = ws.cell(row=gt_row, column=col, value=val)
        c.font   = WHITE_BOLD
        c.fill   = NAVY_FILL
        c.border = border
        c.alignment = Alignment(horizontal='right' if col > 2 else 'left', vertical='center')
        if col >= 4:
            c.number_format = '"$"#,##0.00'
    ws.row_dimensions[gt_row].height = 18

    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 18

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
