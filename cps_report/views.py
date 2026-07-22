"""
PPR Schedule of Loss Report views.
"""
import json
import logging
import re

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, render
from django.utils import timezone
from django.views.decorators.http import require_POST

from docsAppR.models import Client

from .models import CPSReportSession, CPSReportRoom, CPSReportItem

logger = logging.getLogger(__name__)


def _ppr_resolve_url(key):
    """Resolve a storage key to a fresh presigned URL; pass http(s) URLs through unchanged."""
    if not key:
        return key
    if key.startswith(('http://', 'https://')):
        return key
    try:
        from django.core.files.storage import default_storage
        return default_storage.url(key)
    except Exception:
        return key


@login_required
def cps_home(request):
    """Landing page — select a claim to run a CPS report."""
    # Group all sessions by claim so the user can browse and compare runs
    all_sessions = (
        CPSReportSession.objects
        .select_related('client')
        .prefetch_related('rooms')
        .order_by('insured_name', '-created_at')
    )

    # Build grouped structure: {encircle_claim_id: {label, sessions[]}}
    from collections import OrderedDict
    grouped = OrderedDict()
    for s in all_sessions:
        key = s.encircle_claim_id
        if key not in grouped:
            grouped[key] = {
                'label': s.insured_name or s.claim_number or s.encircle_claim_id,
                'claim_number': s.claim_number,
                'sessions': [],
            }
        grouped[key]['sessions'].append(s)

    return render(request, 'cps_report/home.html', {
        'grouped_sessions': grouped,
        # keep for any legacy template references
        'recent_sessions': list(all_sessions.order_by('-updated_at')[:10]),
    })


@login_required
def api_search_clients(request):
    """
    AJAX claim search for the CPS report selector.
    Queries Encircle directly (cached 15 min) so all claims are visible,
    not just the ones imported into the local DB.
    """
    from django.core.cache import cache
    from docsAppR.encircle_client import EncircleAPIClient, EncircleDataProcessor

    q = request.GET.get('q', '').strip().lower()
    try:
        page = max(1, int(request.GET.get('page', 1)))
    except (ValueError, TypeError):
        page = 1

    per_page = 25

    # Fetch all Encircle claims (cached 15 min)
    CACHE_KEY = 'cps_encircle_claims_list'
    claims = cache.get(CACHE_KEY)
    if claims is None:
        try:
            api       = EncircleAPIClient()
            processor = EncircleDataProcessor()
            raw       = api.get_all_claims()
            processed = processor.process_claims_list(raw)
            claims = [
                {
                    'id':                  str(c.get('id') or ''),
                    'pOwner':              c.get('policyholder_name') or '',
                    'pAddress':            c.get('full_address') or '',
                    'claimNumber':         c.get('policy_number') or '',
                    'encircle_claim_id':   str(c.get('id') or ''),
                    'insurance_company':   c.get('insurance_company_name') or '',
                }
                for c in processed
            ]
            claims.sort(key=lambda c: c['pOwner'].lower())
            cache.set(CACHE_KEY, claims, 900)
        except Exception as e:
            logger.error(f"api_search_clients: Encircle fetch failed: {e}", exc_info=True)
            claims = []

    # Search
    if q:
        claims = [
            c for c in claims
            if q in c['pOwner'].lower()
            or q in c['pAddress'].lower()
            or q in c['claimNumber'].lower()
            or q in c['encircle_claim_id'].lower()
        ]

    total      = len(claims)
    offset     = (page - 1) * per_page
    page_items = claims[offset:offset + per_page]

    return JsonResponse({
        'clients':     page_items,
        'total':       total,
        'page':        page,
        'per_page':    per_page,
        'total_pages': max(1, (total + per_page - 1) // per_page),
    })


@login_required
def session_view(request, session_id):
    """View / manage an existing CPS report session."""
    session = get_object_or_404(CPSReportSession.objects.select_related('client'), id=session_id)
    rooms = list(session.rooms.prefetch_related('items').order_by('order', 'room_number'))
    share_url = request.build_absolute_uri(f'/cps-report/sign/{session.share_token}/')
    other_sessions = (
        CPSReportSession.objects
        .filter(encircle_claim_id=session.encircle_claim_id)
        .exclude(id=session.id)
        .prefetch_related('rooms')
        .order_by('-created_at')
    )
    from django.core.files.storage import default_storage
    photo_pdf_ready = default_storage.exists(f'cps_photo_pdfs/{session.id}.pdf')

    # Resolve storage keys to fresh presigned URLs for template display.
    # item_photo_pairs: {item.id: [[storage_key, presigned_url], ...]}
    # item_url_map:     {item.id: {storage_key: presigned_url}} — for JS data-url-map attr
    item_photo_pairs = {}
    item_url_map = {}
    for room in rooms:
        for item in room.items.all():
            if item.source_image_urls:
                pairs = [[k, _ppr_resolve_url(k)] for k in item.source_image_urls]
                item_photo_pairs[item.id] = pairs
                item_url_map[item.id] = {k: url for k, url in pairs}

    return render(request, 'cps_report/session.html', {
        'session': session,
        'rooms': rooms,
        'share_url': share_url,
        'other_sessions': other_sessions,
        'photo_pdf_ready': photo_pdf_ready,
        'item_photo_pairs': item_photo_pairs,
        'item_url_map': item_url_map,
    })


# ---------------------------------------------------------------------------
# API endpoints
# ---------------------------------------------------------------------------

@login_required
def api_fetch_rooms(request):
    """
    GET — fetch Encircle rooms for a claim grouped by series (100s/300s/400s/BU).
    Called by the room-type selector modal before analysis starts.
    """
    encircle_claim_id = request.GET.get('encircle_claim_id', '').strip()
    if not encircle_claim_id:
        return JsonResponse({'error': 'encircle_claim_id required'}, status=400)
    try:
        from docsAppR.encircle_client import EncircleAPIClient
        api = EncircleAPIClient()
        structures = api.get_claim_structures(encircle_claim_id)
        if not structures or not structures.get('list'):
            return JsonResponse({'error': 'No structures found for this claim'}, status=404)
        structure_id = structures['list'][0]['id']
        all_rooms    = api.get_all_structure_rooms(encircle_claim_id, structure_id)

        return JsonResponse({'success': True, 'groups': _group_encircle_rooms(all_rooms)})
    except Exception as e:
        logger.error(f"api_fetch_rooms error: {e}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)


def _group_encircle_rooms(all_rooms):
    """
    Classify Encircle rooms by numeric series prefix into 100s / 300s / 400s / bu.
    BU only matches rooms NOT already classified by their numeric prefix, so a
    room named "401 BU PICS" lands in 400s (not both).
    """
    groups = {'400s': [], '300s': [], '100s': [], 'bu': []}
    for room in all_rooms:
        label = (room.get('label') or room.get('name') or '').strip()
        rid   = room.get('id', '')
        m = re.match(r'^(\d+)', label)
        classified = False
        if m:
            num  = int(m.group(1))
            info = {'id': rid, 'label': label, 'number': m.group(1)}
            if 100 <= num < 200:
                groups['100s'].append(info)
                classified = True
            elif 300 <= num < 400:
                groups['300s'].append(info)
                classified = True
            elif 400 <= num < 500:
                groups['400s'].append(info)
                classified = True
        if not classified and re.search(r'\bbu\b', label, re.IGNORECASE):
            groups['bu'].append({'id': rid, 'label': label, 'number': ''})
    return groups


@login_required
@require_POST
def api_start_session(request):
    """
    Create a new PPR session.  Accepts room_sources (default ["400s","100s","bu"])
    which controls which Encircle room series are included and in what order.
    """
    try:
        data = json.loads(request.body)
        encircle_claim_id = str(data.get('encircle_claim_id') or data.get('client_id') or '').strip()
        if not encircle_claim_id:
            return JsonResponse({'error': 'encircle_claim_id required'}, status=400)

        pricing_mode = str(data.get('pricing_mode') or 'normal').strip()
        if pricing_mode not in ('normal', 'premium'):
            pricing_mode = 'normal'

        room_sources = list(data.get('room_sources') or ['400s', '100s', 'bu'])

        # Find or stub a Client record for the FK
        client = Client.unscoped.filter(encircle_claim_id=encircle_claim_id).first()
        if not client:
            from docsAppR.encircle_client import EncircleAPIClient as _API, EncircleDataProcessor as _P
            _api = _API()
            try:
                details   = _api.get_claim_details(encircle_claim_id)
                processed = _P.process_claim_details(details)
                client = Client.unscoped.create(
                    encircle_claim_id = encircle_claim_id,
                    pOwner            = processed.get('policyholder_name') or '',
                    pAddress          = processed.get('full_address') or '',
                    claimNumber       = processed.get('policy_number') or '',
                )
            except Exception:
                client = Client.unscoped.create(encircle_claim_id=encircle_claim_id)

        # Fetch and group all rooms from Encircle
        from docsAppR.encircle_client import EncircleAPIClient
        api = EncircleAPIClient()

        structures = api.get_claim_structures(encircle_claim_id)
        if not structures or not structures.get('list'):
            return JsonResponse({'error': 'No structures found for this claim'}, status=404)

        structure_id = structures['list'][0]['id']
        all_rooms    = api.get_all_structure_rooms(encircle_claim_id, structure_id)
        grouped      = _group_encircle_rooms(all_rooms)

        has_400s = '400s' in room_sources and grouped['400s']
        has_300s = '300s' in room_sources and grouped['300s']
        has_100s = '100s' in room_sources and grouped['100s']
        has_bu   = 'bu'   in room_sources and grouped['bu']

        if not has_400s and not has_300s:
            available = [k for k in ('400s', '300s') if grouped[k]]
            return JsonResponse({
                'error': 'No primary rooms (400s/300s) found or selected. '
                         f'Available: {available or "none"}',
            }, status=404)

        # Build paired primary rooms — when both 400s and 300s are selected,
        # pair them by numeric suffix (e.g. "401" ↔ "301") so the AI sees
        # both photo sets together for the same physical room.
        paired_rooms = []  # [(primary_info, secondary_info|None), ...]
        if has_400s and has_300s:
            from collections import defaultdict
            by_suffix = defaultdict(dict)
            for r in grouped['400s']:
                by_suffix[r['number'][1:]][r['number'][0]] = r
            for r in grouped['300s']:
                by_suffix[r['number'][1:]][r['number'][0]] = r
            for suffix in sorted(by_suffix.keys()):
                sm = by_suffix[suffix]
                primary   = sm.get('4') or sm.get('3')
                secondary = sm.get('3') if ('4' in sm and '3' in sm) else None
                paired_rooms.append((primary, secondary))
        elif has_400s:
            paired_rooms = [(r, None) for r in sorted(grouped['400s'], key=lambda x: x['number'])]
        else:
            paired_rooms = [(r, None) for r in sorted(grouped['300s'], key=lambda x: x['number'])]

        # Create session
        session = CPSReportSession.objects.create(
            client=client,
            encircle_claim_id=encircle_claim_id,
            claim_number=client.claimNumber or '',
            insured_name=client.pOwner or '',
            encircle_structure_id=structure_id,
            pricing_mode=pricing_mode,
            room_sources=room_sources,
            status='pending',
        )

        def _clean_name(label):
            return re.sub(r'^\d+\s*[\-–—·\.]*\s*', '', label).strip() or label

        # Primary rooms (order 0+)
        for order, (primary, secondary) in enumerate(paired_rooms):
            CPSReportRoom.objects.create(
                session=session,
                room_name=_clean_name(primary['label']),
                room_number=primary['number'],
                encircle_room_id=primary['id'],
                encircle_room_label=primary['label'],
                encircle_room_id_secondary=secondary['id'] if secondary else '',
                encircle_room_label_secondary=secondary['label'] if secondary else '',
                room_source=CPSReportRoom.ROOM_SOURCE_PRIMARY,
                order=order,
                status='pending',
            )

        # Overview rooms — 100s (order 100+)
        if has_100s:
            for ov_idx, r in enumerate(sorted(grouped['100s'], key=lambda x: x['number'])):
                CPSReportRoom.objects.create(
                    session=session,
                    room_name=_clean_name(r['label']),
                    room_number=r['number'],
                    encircle_room_id=r['id'],
                    encircle_room_label=r['label'],
                    room_source=CPSReportRoom.ROOM_SOURCE_OVERVIEW,
                    order=100 + ov_idx,
                    status='pending',
                )

        # BU rooms (order 200+)
        if has_bu:
            for bu_idx, r in enumerate(grouped['bu']):
                CPSReportRoom.objects.create(
                    session=session,
                    room_name=_clean_name(r['label']),
                    room_number=r.get('number', ''),
                    encircle_room_id=r['id'],
                    encircle_room_label=r['label'],
                    room_source=CPSReportRoom.ROOM_SOURCE_BU,
                    order=200 + bu_idx,
                    status='pending',
                )

        rooms_out = [
            {
                'id': r.id,
                'room_name': r.room_name,
                'room_number': r.room_number,
                'encircle_room_id': r.encircle_room_id,
                'room_source': r.room_source,
                'status': r.status,
            }
            for r in session.rooms.order_by('order').all()
        ]

        # Launch Celery task
        from .tasks import process_cps_session_task
        task = process_cps_session_task.delay(session.id)
        session.celery_task_id = task.id
        session.status = 'processing'
        session.save(update_fields=['celery_task_id', 'status'])

        return JsonResponse({
            'success': True,
            'session_id': session.id,
            'rooms': rooms_out,
            'total_rooms': len(rooms_out),
            'progress_url': f'/cps-report/session/{session.id}/progress/',
        })

    except Exception as e:
        logger.error(f"api_start_session error: {e}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_POST
def api_process_room(request):
    """
    Run AI analysis on a single room.
    Downloads Encircle images and calls Claude to generate item list.
    """
    try:
        data = json.loads(request.body)
        room_id = data.get('room_id')
        if not room_id:
            return JsonResponse({'error': 'room_id required'}, status=400)

        room = get_object_or_404(CPSReportRoom, id=room_id)
        session = room.session

        room.status = 'processing'
        room.save(update_fields=['status'])

        from .ai_analyzer import analyze_room_for_ppr, fetch_all_claim_media
        all_claim_media = fetch_all_claim_media(session.encircle_claim_id)
        result = analyze_room_for_ppr(
            room_name=f"{room.room_number} {room.room_name}",
            room_number=room.room_number,
            prefetched_media=all_claim_media,
            pricing_mode=session.pricing_mode or 'normal',
        )

        room.images_used = result.get('images_used', 0)
        room.ai_confidence = result.get('confidence', '')
        room.ai_notes = result.get('room_summary', '')
        room.status = 'complete' if result.get('success') else 'error'
        room.save(update_fields=['images_used', 'ai_confidence', 'ai_notes', 'status'])

        return JsonResponse({
            'success': True,
            'room_id': room.id,
            'items': result.get('items', []),
            'confidence': result.get('confidence', ''),
            'room_summary': result.get('room_summary', ''),
            'images_used': result.get('images_used', 0),
            'error': result.get('error'),
        })

    except Exception as e:
        logger.error(f"api_process_room error: {e}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_POST
def api_save_room_items(request):
    """
    Persist the (possibly edited) item list for a room.
    Replaces existing items for that room.
    """
    try:
        data = json.loads(request.body)
        room_id = data.get('room_id')
        items_data = data.get('items', [])
        if not room_id:
            return JsonResponse({'error': 'room_id required'}, status=400)

        room = get_object_or_404(CPSReportRoom, id=room_id)

        room.items.all().delete()
        for order, item_dict in enumerate(items_data):
            CPSReportItem.objects.create(
                room=room,
                order=order,
                description=str(item_dict.get('description', ''))[:500],
                brand=str(item_dict.get('brand', ''))[:200],
                disposition=str(item_dict.get('disposition', 'Replacement'))[:100],
                condition=str(item_dict.get('condition', ''))[:50],
                qty=max(1, int(item_dict.get('qty', 1) or 1)),
                model_number=str(item_dict.get('model_number', ''))[:200],
                serial_number=str(item_dict.get('serial_number', ''))[:200],
                retailer=str(item_dict.get('retailer', ''))[:200],
                replacement_source=str(item_dict.get('replacement_source', 'Retail'))[:200],
                purchase_price_each=float(item_dict.get('purchase_price_each', 0) or 0),
                age_years=max(0, min(5, int(item_dict.get('age_years', 0) or 0))),
                age_months=max(0, min(11, int(item_dict.get('age_months', 0) or 0))),
                replacement_value_each=float(item_dict.get('replacement_value_each', 0) or 0),
                notes=str(item_dict.get('notes', ''))[:500],
                ai_suggested=bool(item_dict.get('ai_suggested', True)),
                structural=bool(item_dict.get('structural', False)),
                source_image_urls=list(item_dict.get('source_image_urls', []) or []),
            )

        room.status = 'complete'
        room.save(update_fields=['status'])

        # Update session status
        session = room.session
        all_statuses = set(session.rooms.values_list('status', flat=True))
        if all_statuses == {'complete'}:
            session.status = 'complete'
            session.save(update_fields=['status'])
            _auto_generate_summary(session)

        # Invalidate cached photo PDF so the next download reflects edits.
        # For large sessions that can't generate on-demand, also queue a rebuild.
        from django.core.files.storage import default_storage
        _pdf_path = f'cps_photo_pdfs/{session.id}.pdf'
        if default_storage.exists(_pdf_path):
            default_storage.delete(_pdf_path)
            room_count = session.rooms.count()
            if room_count > 15:
                from .tasks import regenerate_photo_pdf_task
                regenerate_photo_pdf_task.delay(session.id)

        return JsonResponse({
            'success': True,
            'room_id': room.id,
            'items_saved': room.items.count(),
        })

    except Exception as e:
        logger.error(f"api_save_room_items error: {e}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def session_progress(request, session_id):
    """Live progress page — polls status API while Celery task runs."""
    session = get_object_or_404(CPSReportSession, id=session_id)
    rooms = session.rooms.order_by('room_number').all()
    return render(request, 'cps_report/progress.html', {
        'session': session,
        'rooms': rooms,
    })


@login_required
def api_session_status(request, session_id):
    """Return current session + room statuses with item counts and RCV totals."""
    session = get_object_or_404(CPSReportSession, id=session_id)
    rooms = session.rooms.prefetch_related('items').order_by('room_number').all()

    total_items = 0
    total_images = 0
    total_rcv = 0.0

    room_data = []
    for r in rooms:
        items = list(r.items.filter(structural=False))
        room_rcv = sum(
            float(i.replacement_value_each or 0) * (i.qty or 1)
            for i in items
        )
        total_items += len(items)
        total_images += r.images_used or 0
        total_rcv += room_rcv
        room_data.append({
            'id': r.id,
            'room_name': r.room_name,
            'room_number': r.room_number,
            'status': r.status,
            'images_used': r.images_used,
            'ai_confidence': r.ai_confidence,
            'ai_notes': r.ai_notes,
            'item_count': len(items),
            'rcv_total': round(room_rcv, 2),
        })

    return JsonResponse({
        'session_id': session.id,
        'status': session.status,
        'total_rooms': len(room_data),
        'total_items': total_items,
        'total_images': total_images,
        'total_rcv': round(total_rcv, 2),
        'rooms': room_data,
    })


def api_session_logs(request, session_id):
    """Return live log lines written by the Celery task for this session."""
    from django.core.cache import cache
    key = f'ppr:live_logs:{session_id}'
    logs = cache.get(key) or []
    try:
        after = int(request.GET.get('after', 0))
    except (TypeError, ValueError):
        after = 0
    return JsonResponse({'logs': logs[after:], 'total': len(logs)})


@login_required
def api_room_items(request, room_id):
    """Return saved items for a room."""
    room = get_object_or_404(CPSReportRoom, id=room_id)
    return JsonResponse({
        'room_id': room.id,
        'room_name': room.room_name,
        'items': [i.to_dict() for i in room.items.order_by('order').all()],
    })


@login_required
@require_POST
def api_reassign_photo(request):
    """
    Move a photo URL from one CPSReportItem to another.
    POST JSON: { from_item_id, to_item_id, photo_url }
    """
    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'error': 'invalid JSON'}, status=400)

    from_id  = data.get('from_item_id')
    to_id    = data.get('to_item_id')
    url      = (data.get('photo_url') or '').strip()

    if not url or not from_id or not to_id:
        return JsonResponse({'error': 'from_item_id, to_item_id, photo_url required'}, status=400)
    if from_id == to_id:
        return JsonResponse({'error': 'source and destination are the same item'}, status=400)

    from_item = get_object_or_404(CPSReportItem, id=from_id)
    to_item   = get_object_or_404(CPSReportItem, id=to_id)

    if from_item.room.session_id != to_item.room.session_id:
        return JsonResponse({'error': 'items must belong to the same session'}, status=403)

    from_urls = list(from_item.source_image_urls or [])
    to_urls   = list(to_item.source_image_urls or [])

    if url not in from_urls:
        return JsonResponse({'error': 'photo_url not found on source item'}, status=400)

    from_urls.remove(url)
    if url not in to_urls:
        to_urls.append(url)

    from_item.source_image_urls = from_urls
    from_item.save(update_fields=['source_image_urls'])
    to_item.source_image_urls = to_urls
    to_item.save(update_fields=['source_image_urls'])

    _invalidate_photo_pdf_cache(from_item.room.session_id)

    all_keys = list(dict.fromkeys(from_urls + to_urls))
    return JsonResponse({
        'success': True,
        'from_source_image_urls': from_urls,
        'to_source_image_urls': to_urls,
        'url_map': {k: _ppr_resolve_url(k) for k in all_keys},
    })


def _invalidate_photo_pdf_cache(session_id):
    """Delete the pre-built Photo Evidence PDF so a photo edit forces a rebuild
    the next time the Photo PDF button is used. Without this, edits would not
    appear in the cached PDF served from storage."""
    try:
        from django.core.files.storage import default_storage
        path = f'cps_photo_pdfs/{session_id}.pdf'
        if default_storage.exists(path):
            default_storage.delete(path)
    except Exception as exc:
        logger.warning(f"Could not invalidate photo PDF cache for session {session_id}: {exc}")


@login_required
@require_POST
def api_delete_photo(request):
    """
    Remove a photo URL from a CPSReportItem entirely (human review layer).
    The photo will no longer appear in the Photo Evidence report.
    POST JSON: { item_id, photo_url }
    """
    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'error': 'invalid JSON'}, status=400)

    item_id = data.get('item_id')
    url     = (data.get('photo_url') or '').strip()

    if not item_id or not url:
        return JsonResponse({'error': 'item_id and photo_url required'}, status=400)

    item = get_object_or_404(CPSReportItem, id=item_id)
    urls = list(item.source_image_urls or [])

    if url not in urls:
        return JsonResponse({'error': 'photo_url not found on item'}, status=400)

    urls.remove(url)
    item.source_image_urls = urls
    item.save(update_fields=['source_image_urls'])

    _invalidate_photo_pdf_cache(item.room.session_id)

    return JsonResponse({'success': True, 'source_image_urls': urls})


@login_required
def export_pdf(request, session_id):
    """Generate and return the Schedule of Loss PDF file."""
    session = get_object_or_404(CPSReportSession.objects.select_related('client'), id=session_id)
    try:
        from .pdf_builder import build_pdf
        pdf_bytes = build_pdf(session)
        filename = f"PPR_Schedule_Of_Loss_{session.claim_number or session.encircle_claim_id}_{session.updated_at:%Y%m%d}.pdf"
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    except Exception as e:
        logger.error(f"export_pdf error: {e}", exc_info=True)
        return HttpResponse(f"Error generating PDF: {e}", status=500)


@login_required
def export_photo_pdf(request, session_id):
    """Serve the Photo Evidence PDF — uses the pre-generated copy from the PPR
    Celery task if available.  On-demand generation is allowed only for small
    claims (≤ 15 rooms); large claims must use the Celery-pre-generated file to
    avoid OOM-killing the web worker (502 Bad Gateway)."""
    session = get_object_or_404(CPSReportSession.objects.select_related('client'), id=session_id)
    try:
        from django.core.files.storage import default_storage
        from .photo_pdf_builder import build_photo_pdf

        _pdf_path = f'cps_photo_pdfs/{session_id}.pdf'

        # ?rebuild=1 — flush the cached copy and queue a fresh build
        if request.GET.get('rebuild'):
            if default_storage.exists(_pdf_path):
                default_storage.delete(_pdf_path)
            from .tasks import regenerate_photo_pdf_task
            regenerate_photo_pdf_task.delay(session_id)
            return HttpResponse(
                'Photo PDF rebuild queued. Wait a few minutes, then download again '
                '(the Photo PDF button will show "building…" until it is ready).',
                content_type='text/plain',
            )

        if default_storage.exists(_pdf_path):
            with default_storage.open(_pdf_path, 'rb') as _f:
                pdf_bytes = _f.read()
        else:
            room_count = session.rooms.count()
            if room_count > 15:
                return HttpResponse(
                    f"The photo PDF for this claim ({room_count} rooms) is too large to generate "
                    f"on demand. It is generated automatically after the PPR analysis completes — "
                    f"please run or re-run the PPR report and then download the photo PDF.",
                    status=503,
                    content_type='text/plain',
                )
            pdf_bytes = build_photo_pdf(session)

        filename = f"NON_SALVAGEABLE_PPR_Photo_Evidence_Report_{session.claim_number or session.encircle_claim_id}_{session.updated_at:%Y%m%d}.pdf"
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    except Exception as e:
        logger.error(f"export_photo_pdf error: {e}", exc_info=True)
        return HttpResponse(f"Error generating Photo PDF: {e}", status=500)


@login_required
def regenerate_photo_pdf(request, session_id):
    """Queue a Celery task to rebuild the photo PDF for an existing session.
    Returns JSON so it can be called from the session page via fetch()."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    session = get_object_or_404(CPSReportSession, id=session_id)
    from .tasks import regenerate_photo_pdf_task
    regenerate_photo_pdf_task.delay(session_id)
    return JsonResponse({'queued': True, 'session_id': session_id})


@login_required
def photo_pdf_status_api(request, session_id):
    """Lightweight check — does the pre-built photo PDF file exist in storage?"""
    get_object_or_404(CPSReportSession, id=session_id)
    from django.core.files.storage import default_storage
    ready = default_storage.exists(f'cps_photo_pdfs/{session_id}.pdf')
    return JsonResponse({'ready': ready, 'session_id': session_id})


@login_required
def api_debug_claim_media(request, claim_id):
    """
    Diagnostic: fetch ALL media for a claim and break it down by room.
    Shows total items, unique filenames, duplicate versions, and source types
    so we can understand exactly what the Encircle API is returning.

    GET /cps-report/api/debug/media/<claim_id>/
    """
    from collections import defaultdict
    from docsAppR.encircle_client import EncircleAPIClient

    try:
        api = EncircleAPIClient()
        all_media = []
        after_cursor = None
        while True:
            params = {'limit': 100}
            if after_cursor:
                params['after'] = after_cursor
            resp = api._make_request(f"property_claims/{claim_id}/media", params=params)
            if not resp or 'list' not in resp:
                break
            all_media.extend(resp['list'])
            after_cursor = resp.get('cursor', {}).get('after')
            if not after_cursor:
                break

        # --- global stats ---
        total = len(all_media)
        content_type_counts = defaultdict(int)
        source_type_counts  = defaultdict(int)
        for m in all_media:
            content_type_counts[(m.get('content_type') or 'unknown').split(';')[0].strip()] += 1
            source_type_counts[(m.get('source') or {}).get('type') or 'unknown'] += 1

        # --- per-room breakdown (keyed by source.primary_id) ---
        by_room = defaultdict(list)
        for m in all_media:
            src = (m.get('source') or {})
            key = f"{src.get('type','?')}:{src.get('primary_id','?')}"
            by_room[key].append(m)

        room_rows = []
        for key, items in sorted(by_room.items(), key=lambda x: -len(x[1])):
            filenames  = [m.get('filename') or '' for m in items]
            unique_fns = list(dict.fromkeys(filenames))  # preserves order, dedupes
            cts        = list({(m.get('content_type') or 'unknown').split(';')[0].strip() for m in items})
            room_rows.append({
                'source_key':       key,
                'total_media':      len(items),
                'unique_filenames': len(unique_fns),
                'duplicate_extra':  len(items) - len(unique_fns),
                'content_types':    cts,
                'sample_filenames': unique_fns[:5],
            })

        # Raw sample: first 2 items of each source type so we can see all fields
        raw_samples = {}
        for m in all_media:
            src_type = (m.get('source') or {}).get('type') or 'unknown'
            if src_type not in raw_samples:
                raw_samples[src_type] = m
            if len(raw_samples) >= 5:
                break

        return JsonResponse({
            'claim_id':            claim_id,
            'total_media_items':   total,
            'content_type_counts': dict(content_type_counts),
            'source_type_counts':  dict(source_type_counts),
            'unique_source_keys':  len(by_room),
            'raw_sample_by_type':  raw_samples,
            'rooms': room_rows,
        })

    except Exception as e:
        logger.error(f"api_debug_claim_media error: {e}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)


def sign_session(request, token):
    """Public (no login) page where the client signs off on each room."""
    from django.db.models import Prefetch
    session = get_object_or_404(CPSReportSession, share_token=token)
    rooms = session.rooms.prefetch_related(
        Prefetch('items',
                 queryset=CPSReportItem.objects.filter(structural=False).order_by('order'),
                 to_attr='billable_items')
    ).order_by('order', 'room_number')
    return render(request, 'cps_report/sign.html', {
        'session': session,
        'rooms': rooms,
        'token': str(token),
    })


def sign_room_direct(request, token):
    """Public (no login) page where the client signs a single room via its own token."""
    from django.db.models import Prefetch
    from .models import CPSReportRoom
    room = get_object_or_404(
        CPSReportRoom.objects.prefetch_related(
            Prefetch('items',
                     queryset=CPSReportItem.objects.filter(structural=False).order_by('order'),
                     to_attr='billable_items')
        ),
        share_token=token,
    )
    return render(request, 'cps_report/sign_room.html', {
        'session': room.session,
        'room': room,
        'token': str(token),
    })


@require_POST
def api_sign_room_direct(request, token):
    """Public POST — sign a single room using the room's own share token."""
    from .models import CPSReportRoom
    from .tasks import send_cps_room_signing_notification
    
    room = get_object_or_404(CPSReportRoom, share_token=token)
    try:
        if room.signature_name:
            return JsonResponse({'error': 'Room already signed'}, status=400)
        data = json.loads(request.body)
        name = (data.get('name') or '').strip()
        if not name:
            return JsonResponse({'error': 'name is required'}, status=400)

        x_forwarded = request.META.get('HTTP_X_FORWARDED_FOR')
        ip = x_forwarded.split(',')[0].strip() if x_forwarded else request.META.get('REMOTE_ADDR')

        room.signature_name = name
        room.signed_at = timezone.now()
        room.signer_ip = ip
        room.save(update_fields=['signature_name', 'signed_at', 'signer_ip'])

        # Trigger email notification asynchronously
        client_email = None
        try:
            client = room.session.client
            client_email = getattr(client, 'email', None) or getattr(client, 'pEmail', None)
        except Exception:
            pass
        
        send_cps_room_signing_notification.delay(room.id, client_email)

        return JsonResponse({
            'success': True,
            'room_id': room.id,
            'signed_at': room.signed_at.strftime('%B %d, %Y at %I:%M %p'),
        })
    except Exception as e:
        logger.error(f"api_sign_room_direct error: {e}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def get_room_share_link(request, session_id, room_id):
    """Return the per-room public share URL (shows only that room to the client)."""
    session = get_object_or_404(CPSReportSession, id=session_id)
    room = get_object_or_404(CPSReportRoom, id=room_id, session=session)
    url = request.build_absolute_uri(f'/cps-report/sign/room/{room.share_token}/')
    return JsonResponse({'url': url})


@login_required
@require_POST
def api_cancel_session(request, session_id):
    """Cancel a stuck processing session — revoke Celery task and mark as error."""
    session = get_object_or_404(CPSReportSession, id=session_id)
    if session.celery_task_id:
        try:
            from celery.app.control import Control
            from django.conf import settings
            import celery as _celery
            app = _celery.current_app
            app.control.revoke(session.celery_task_id, terminate=True, signal='SIGTERM')
        except Exception as e:
            logger.warning(f"Could not revoke Celery task {session.celery_task_id}: {e}")
    session.status = 'error'
    session.save(update_fields=['status'])
    session.rooms.filter(status__in=['processing', 'pending']).update(status='error')
    return JsonResponse({'success': True})


@login_required
@require_POST
def api_rerun_session(request, session_id):
    """Reset a session and re-fire the Celery task to reprocess all rooms."""
    session = get_object_or_404(CPSReportSession, id=session_id)
    # Allow caller to override pricing mode for this re-run
    try:
        body = json.loads(request.body or b'{}')
        new_mode = str(body.get('pricing_mode') or session.pricing_mode or 'normal').strip()
        if new_mode not in ('normal', 'premium'):
            new_mode = 'normal'
    except Exception:
        new_mode = session.pricing_mode or 'normal'
    if new_mode != session.pricing_mode:
        session.pricing_mode = new_mode
        session.save(update_fields=['pricing_mode'])
    # Reset rooms — clear old items so the task starts fresh
    for room in session.rooms.all():
        room.items.all().delete()
        room.status = 'pending'
        room.ai_confidence = ''
        room.ai_notes = ''
        room.images_used = 0
        room.save(update_fields=['status', 'ai_confidence', 'ai_notes', 'images_used'])
    session.status = 'pending'
    session.save(update_fields=['status'])
    # Re-fire the task
    from .tasks import process_cps_session_task
    task = process_cps_session_task.delay(session.id)
    session.celery_task_id = task.id
    session.status = 'processing'
    session.save(update_fields=['celery_task_id', 'status'])
    return JsonResponse({'success': True, 'redirect': f'/cps-report/session/{session.id}/progress/'})


@require_POST
def api_sign_room(request, token):
    """Public POST endpoint — save a typed-name signature for one room."""
    from .tasks import send_cps_room_signing_notification
    
    session = get_object_or_404(CPSReportSession, share_token=token)
    try:
        data = json.loads(request.body)
        room_id = data.get('room_id')
        name = (data.get('name') or '').strip()
        if not room_id or not name:
            return JsonResponse({'error': 'room_id and name are required'}, status=400)

        room = get_object_or_404(CPSReportRoom, id=room_id, session=session)
        if room.signature_name:
            return JsonResponse({'error': 'Room already signed'}, status=400)

        x_forwarded = request.META.get('HTTP_X_FORWARDED_FOR')
        ip = x_forwarded.split(',')[0].strip() if x_forwarded else request.META.get('REMOTE_ADDR')

        room.signature_name = name
        room.signed_at = timezone.now()
        room.signer_ip = ip
        room.save(update_fields=['signature_name', 'signed_at', 'signer_ip'])

        # Trigger email notification asynchronously
        client_email = None
        try:
            client = room.session.client
            client_email = getattr(client, 'email', None) or getattr(client, 'pEmail', None)
        except Exception:
            pass
        
        send_cps_room_signing_notification.delay(room.id, client_email)

        return JsonResponse({
            'success': True,
            'room_id': room.id,
            'signed_at': room.signed_at.strftime('%B %d, %Y at %I:%M %p'),
        })
    except Exception as e:
        logger.error(f"api_sign_room error: {e}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def get_share_link(request, session_id):
    """Return the public share URL for a session."""
    session = get_object_or_404(CPSReportSession, id=session_id)
    url = request.build_absolute_uri(f'/cps-report/sign/{session.share_token}/')
    return JsonResponse({'url': url})


@login_required
@require_POST
def api_clear_signatures(request, session_id):
    """Clear all room signatures for a session so the client can re-sign."""
    session = get_object_or_404(CPSReportSession, id=session_id)
    updated = session.rooms.update(signature_name='', signed_at=None, signer_ip=None)
    return JsonResponse({'success': True, 'rooms_cleared': updated})


@login_required
@require_POST
def api_clear_room_signature(request, session_id, room_id):
    """Clear the signature for a single room (e.g. remove a test signature)."""
    session = get_object_or_404(CPSReportSession, id=session_id)
    room = get_object_or_404(CPSReportRoom, id=room_id, session=session)
    room.signature_name = ''
    room.signed_at = None
    room.signer_ip = None
    room.save(update_fields=['signature_name', 'signed_at', 'signer_ip'])
    return JsonResponse({'success': True, 'room_id': room.id})


@login_required
def export_excel(request, session_id):
    """Generate and return the Schedule of Loss Excel file."""
    session = get_object_or_404(CPSReportSession.objects.select_related('client'), id=session_id)
    try:
        from .excel_builder import build_excel
        share_url = request.build_absolute_uri(f'/cps-report/sign/{session.share_token}/')
        xlsx_bytes = build_excel(session, share_url=share_url)
        filename = f"PPR_Schedule_Of_Loss_{session.claim_number or session.encircle_claim_id}_{session.updated_at:%Y%m%d}.xlsx"

        # Best-effort: save to claim folder + notify via email
        _cps_save_and_notify(session, xlsx_bytes, filename)

        response = HttpResponse(
            xlsx_bytes,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    except Exception as e:
        logger.error(f"export_excel error: {e}", exc_info=True)
        return HttpResponse(f"Error generating Excel: {e}", status=500)


@login_required
@require_POST
def api_import_excel(request):
    """
    Import a Schedule of Loss Excel file (any version) as a new session.
    Detects column positions from the header row so old format files work
    regardless of when they were generated.
    """
    import io
    import re as _re
    from openpyxl import load_workbook

    try:
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            return JsonResponse({'error': 'No file uploaded'}, status=400)
        if not excel_file.name.lower().endswith(('.xlsx', '.xls')):
            return JsonResponse({'error': 'File must be an Excel file (.xlsx)'}, status=400)

        wb = load_workbook(io.BytesIO(excel_file.read()), data_only=True)
        ws = wb.active

        # ── Scan all rows for claim metadata (flexible — works across formats) ─
        encircle_claim_id = ''
        loss_type = ''
        insured_name = ''
        claim_number = ''

        for scan_row in range(1, min(15, ws.max_row + 1)):
            for col in range(1, ws.max_column + 1):
                raw = str(ws.cell(row=scan_row, column=col).value or '')
                if not raw:
                    continue
                m = _re.search(r'Claim Id[:\s]+(\S+)', raw, _re.I)
                if m and not encircle_claim_id:
                    encircle_claim_id = m.group(1).strip()
                m = _re.search(r'(?:Type of Loss|Loss)[:\s]+(.+?)(?:\s*\||$)', raw, _re.I)
                if m and not loss_type:
                    loss_type = m.group(1).strip()
                m = _re.search(r'Insured[:\s]+(.+?)(?:\s*\||$)', raw, _re.I)
                if m and not insured_name:
                    insured_name = m.group(1).strip()
                m = _re.search(r'Claim\s*#[:\s]+(.+?)(?:\s*\||$)', raw, _re.I)
                if m and not claim_number:
                    claim_number = m.group(1).strip()

        # ── Detect column positions from header row ───────────────────────────
        # Scan for the row containing "Description" — that anchors all columns.
        # Works for any format (old or new) regardless of how many prefix columns exist.
        desc_col = None
        data_start_row = 7   # fallback

        for scan_row in range(1, min(20, ws.max_row + 1)):
            for col in range(1, ws.max_column + 1):
                v = ws.cell(row=scan_row, column=col).value
                if v is not None and str(v).strip().lower() == 'description':
                    desc_col = col
                    data_start_row = scan_row + 2   # header is 2 rows (row5 + row6)
                    break
            if desc_col is not None:
                break

        # Column map relative to description. Offsets match excel_builder.py layout:
        # #, [Room,] Description, Brand, Disposition, Condition, QTY, Model#, Serial#,
        # Retailer, Repl.Source, Purchase(Each, Total), Age(Y, M), Repl.Value(Each, Total)
        D = desc_col if desc_col is not None else 5   # 5 = pre-Room-column legacy default

        # ── Find or create Client ─────────────────────────────────────────────
        # Use unscoped so pre-tenant-migration clients (tenant_id=NULL) are found.
        client = None
        if encircle_claim_id:
            client = Client.unscoped.filter(encircle_claim_id=encircle_claim_id).first()
        if not client:
            client = Client.unscoped.create(
                encircle_claim_id=encircle_claim_id or 'imported',
                pOwner=insured_name,
                claimNumber=claim_number,
            )

        # ── Create a new CPSReportSession ─────────────────────────────────────
        session = CPSReportSession.objects.create(
            client=client,
            encircle_claim_id=encircle_claim_id or client.encircle_claim_id,
            claim_number=claim_number or client.claimNumber,
            insured_name=insured_name or client.pOwner,
            loss_type=loss_type,
            status='complete',
            notes='Imported from Excel file',
        )

        # ── Parse rooms and items ─────────────────────────────────────────────
        # Row types (matched in priority order):
        #   "GRAND TOTAL"           → stop
        #   starts with "Room Total" → room subtotal row — skip
        #   contains "Signed by" / "Awaiting client signature" → sig row — skip
        #   matches ^\d+\s*[—–-]+  → room header row
        #   isinstance int/float    → item data row
        current_room = None
        room_order = 0
        item_order = 0
        rooms_created = 0
        items_created = 0

        def _gs(r, col):
            v = ws.cell(row=r, column=col).value
            return str(v).strip() if v is not None else ''

        def _gf(r, col):
            v = ws.cell(row=r, column=col).value
            try:
                return float(v) if v is not None else 0.0
            except (TypeError, ValueError):
                return 0.0

        def _gi(r, col):
            v = ws.cell(row=r, column=col).value
            try:
                return int(float(v)) if v is not None else 0
            except (TypeError, ValueError):
                return 0

        for row_idx in range(data_start_row, ws.max_row + 1):
            val_a = ws.cell(row=row_idx, column=1).value
            if val_a is None:
                continue
            val_str = str(val_a).strip()
            if not val_str:
                continue

            if val_str == 'GRAND TOTAL':
                break

            if val_str.startswith('Room Total') or 'Signed by' in val_str or 'Awaiting client signature' in val_str:
                continue

            # Room header: "401  —  Living Room" or "401  -  Living Room   (AI confidence: high)"
            if _re.match(r'^\d+\s*[—–\-]+\s*', val_str):
                m = _re.match(r'^(\d+)\s*[—–\-]+\s*(.+?)(?:\s+\(AI confidence:.*)?$', val_str)
                if m:
                    room_num = m.group(1).strip()
                    room_name = m.group(2).strip()
                else:
                    parts = _re.split(r'\s*[—–\-]+\s*', val_str, 1)
                    room_num = parts[0].strip()
                    room_name = parts[1].strip() if len(parts) > 1 else val_str

                current_room = CPSReportRoom.objects.create(
                    session=session,
                    room_name=room_name,
                    room_number=room_num,
                    order=room_order,
                    status='complete',
                )
                room_order += 1
                rooms_created += 1
                item_order = 0
                continue

            # Item row: column A holds the global item counter (integer)
            if isinstance(val_a, (int, float)) and current_room is not None:
                try:
                    if float(val_a) != int(float(val_a)):
                        continue  # not a whole-number item counter
                except (TypeError, ValueError):
                    continue

                CPSReportItem.objects.create(
                    room=current_room,
                    order=item_order,
                    description=_gs(row_idx, D),
                    brand=_gs(row_idx, D + 1),
                    disposition=_gs(row_idx, D + 2) or 'Replacement',
                    condition=_gs(row_idx, D + 3),
                    qty=max(1, _gi(row_idx, D + 4)),
                    model_number=_gs(row_idx, D + 5),
                    serial_number=_gs(row_idx, D + 6),
                    retailer=_gs(row_idx, D + 7),
                    replacement_source=_gs(row_idx, D + 8),
                    purchase_price_each=_gf(row_idx, D + 9),
                    age_years=min(200, max(0, _gi(row_idx, D + 11))),
                    age_months=min(11, max(0, _gi(row_idx, D + 12))),
                    replacement_value_each=_gf(row_idx, D + 13),
                    ai_suggested=False,
                )
                item_order += 1
                items_created += 1

        logger.info(
            f"Excel import: session {session.id} — {rooms_created} rooms, {items_created} items "
            f"(claim {encircle_claim_id or 'unknown'})"
        )
        return JsonResponse({
            'success': True,
            'session_id': session.id,
            'rooms_created': rooms_created,
            'items_created': items_created,
            'redirect_url': f'/cps-report/session/{session.id}/',
        })

    except Exception as e:
        logger.error(f"api_import_excel error: {e}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def pricing_audit_view(request, session_id):
    """
    Pricing audit / difference report for a PPR session.

    Displays per-room and total-level breakdown of:
      - Baseline RCV  (from a companion normal-mode session, or back-estimated)
      - Expected Premium RCV  (baseline × PREMIUM_EXPECTED_LIFT)
      - Actual AI-Generated Premium RCV
      - Delta / Variance  (actual − expected)

    The companion normal-mode session is selected as the most recently
    completed normal-pricing run for the same Encircle claim ID.
    When no normal session exists, baseline is back-estimated by dividing
    each room's premium total by PREMIUM_EXPECTED_LIFT.
    """
    from .ai_analyzer import (
        PREMIUM_SOFT_THRESHOLD,
        PREMIUM_LOG_SCALE_FACTOR,
        PREMIUM_HARD_CEILING,
        PREMIUM_EXPECTED_LIFT,
        CATEGORY_BASELINES,
    )

    session = get_object_or_404(CPSReportSession, id=session_id)

    # ── Find companion normal session ─────────────────────────────────────────
    normal_session = (
        CPSReportSession.objects
        .filter(
            encircle_claim_id=session.encircle_claim_id,
            pricing_mode='normal',
            status='complete',
        )
        .order_by('-updated_at')
        .first()
    )

    # ── Build room-level index for normal session (room_number → items) ───────
    normal_room_rcv: dict[str, float] = {}
    if normal_session:
        for room in normal_session.rooms.prefetch_related('items').all():
            normal_room_rcv[room.room_number] = float(sum(
                (float(i.replacement_value_each or 0) * (i.qty or 1))
                for i in room.items.all()
            ))

    # ── Build per-room audit rows ─────────────────────────────────────────────
    room_rows = []
    total_baseline  = 0.0
    total_expected  = 0.0
    total_actual    = 0.0
    cap_hit_count   = 0
    total_items     = 0

    for room in session.rooms.prefetch_related('items').order_by('order', 'room_number'):
        items = list(room.items.all())
        total_items += len(items)

        actual_rcv = float(sum(
            (float(i.replacement_value_each or 0) * (i.qty or 1))
            for i in items
        ))

        # Baseline: use companion normal session if available, else back-estimate
        if room.room_number in normal_room_rcv:
            baseline_rcv = normal_room_rcv[room.room_number]
            baseline_source = 'normal_session'
        elif actual_rcv > 0:
            baseline_rcv = round(actual_rcv / PREMIUM_EXPECTED_LIFT, 2)
            baseline_source = 'estimated'
        else:
            baseline_rcv = 0.0
            baseline_source = 'estimated'

        expected_rcv = round(baseline_rcv * PREMIUM_EXPECTED_LIFT, 2)
        delta        = round(actual_rcv - expected_rcv, 2)
        delta_pct    = round((delta / expected_rcv * 100) if expected_rcv else 0, 1)

        # Count items that hit the cap (annotated in notes)
        capped_items = sum(1 for i in items if 'cap-applied' in (i.notes or ''))
        cap_hit_count += capped_items

        # Health signal per room
        if abs(delta_pct) <= 10:
            health = 'ok'
        elif abs(delta_pct) <= 25:
            health = 'warn'
        else:
            health = 'over' if delta > 0 else 'under'

        room_rows.append({
            'room_number':    room.room_number,
            'room_name':      room.room_name,
            'baseline_rcv':   baseline_rcv,
            'expected_rcv':   expected_rcv,
            'actual_rcv':     actual_rcv,
            'delta':          delta,
            'delta_pct':      delta_pct,
            'capped_items':   capped_items,
            'item_count':     len(items),
            'health':         health,
            'baseline_source': baseline_source,
        })

        total_baseline += baseline_rcv
        total_expected += expected_rcv
        total_actual   += actual_rcv

    total_delta     = round(total_actual - total_expected, 2)
    total_delta_pct = round((total_delta / total_expected * 100) if total_expected else 0, 1)

    if abs(total_delta_pct) <= 10:
        overall_health = 'ok'
    elif abs(total_delta_pct) <= 25:
        overall_health = 'warn'
    else:
        overall_health = 'over' if total_delta > 0 else 'under'

    context = {
        'session':           session,
        'normal_session':    normal_session,
        'room_rows':         room_rows,
        # Totals
        'total_baseline':    round(total_baseline, 2),
        'total_expected':    round(total_expected, 2),
        'total_actual':      round(total_actual, 2),
        'total_delta':       total_delta,
        'total_delta_pct':   total_delta_pct,
        'overall_health':    overall_health,
        # Calibration metadata
        'cap_hit_count':     cap_hit_count,
        'total_items':       total_items,
        'soft_threshold':    PREMIUM_SOFT_THRESHOLD,
        'log_scale_factor':  PREMIUM_LOG_SCALE_FACTOR,
        'hard_ceiling':      PREMIUM_HARD_CEILING,
        'expected_lift':     PREMIUM_EXPECTED_LIFT,
        'expected_lift_pct': round((PREMIUM_EXPECTED_LIFT - 1) * 100, 1),
    }
    return render(request, 'cps_report/pricing_audit.html', context)


def _auto_generate_summary(session) -> None:
    """Best-effort: pre-build summary exports so on-demand page loads instantly."""
    try:
        from .summary_builder import build_summary_excel, build_summary_pdf
        import os
        from django.conf import settings as _cfg

        client = session.client
        if not client:
            return
        folder_path = getattr(client, 'get_server_folder_path', lambda: None)()
        if not folder_path:
            return

        summary_dir = os.path.join(folder_path, '92-CPS', 'summaries')
        os.makedirs(summary_dir, exist_ok=True)
        base = f"CPS_Summary_{session.claim_number or session.encircle_claim_id}_{session.updated_at:%Y%m%d}"

        xlsx = build_summary_excel(session)
        with open(os.path.join(summary_dir, f'{base}.xlsx'), 'wb') as f:
            f.write(xlsx)

        pdf = build_summary_pdf(session)
        with open(os.path.join(summary_dir, f'{base}.pdf'), 'wb') as f:
            f.write(pdf)

        logger.info(f"Auto-generated summary for session {session.id}")
    except Exception as e:
        logger.warning(f"Auto-generate summary failed (session {session.id}): {e}")


@login_required
def session_summary(request, session_id):
    """Render the per-room summary page for a CPS session."""
    session = get_object_or_404(CPSReportSession, id=session_id)
    from .summary_builder import compute_summary
    summary = compute_summary(session)
    return render(request, 'cps_report/summary.html', {
        'session': session,
        **summary,
    })


@login_required
def export_summary_pdf(request, session_id):
    """Stream the summary as a PDF file."""
    session = get_object_or_404(CPSReportSession, id=session_id)
    try:
        from .summary_builder import build_summary_pdf
        pdf_bytes = build_summary_pdf(session)
        filename = f"CPS_Summary_{session.claim_number or session.encircle_claim_id}_{session.updated_at:%Y%m%d}.pdf"
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    except Exception as e:
        logger.error(f"export_summary_pdf error: {e}", exc_info=True)
        return HttpResponse(f"Error generating summary PDF: {e}", status=500)


@login_required
def export_summary_excel(request, session_id):
    """Stream the summary as an Excel file."""
    session = get_object_or_404(CPSReportSession, id=session_id)
    try:
        from .summary_builder import build_summary_excel
        xlsx_bytes = build_summary_excel(session)
        filename = f"CPS_Summary_{session.claim_number or session.encircle_claim_id}_{session.updated_at:%Y%m%d}.xlsx"
        response = HttpResponse(
            xlsx_bytes,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    except Exception as e:
        logger.error(f"export_summary_excel error: {e}", exc_info=True)
        return HttpResponse(f"Error generating summary Excel: {e}", status=500)


def _cps_save_and_notify(session, xlsx_bytes: bytes, filename: str) -> None:
    """
    Save the CPS Excel to the claim's 92-CPS folder and email notification
    list. Both steps are best-effort — failures are logged but never raised.
    """
    import os, json as _json
    from django.conf import settings as _cfg

    # ── 1. Save to disk ──────────────────────────────────────────────────────
    try:
        client = session.client
        if client:
            folder_path = client.get_server_folder_path()
            if folder_path:
                cps_dir = os.path.join(folder_path, '92-CPS')
                os.makedirs(cps_dir, exist_ok=True)
                save_path = os.path.join(cps_dir, filename)
                with open(save_path, 'wb') as fh:
                    fh.write(xlsx_bytes)
                logger.info(f"CPS Excel saved: {save_path}")
    except Exception as e:
        logger.warning(f"CPS Excel save failed: {e}")

    # ── 2. Email notification ─────────────────────────────────────────────────
    try:
        settings_path = os.path.join(_cfg.MEDIA_ROOT, 'config', 'excel_hub_settings.json')
        recipients = []
        if os.path.exists(settings_path):
            with open(settings_path) as f:
                recipients = [e.strip() for e in _json.load(f).get('emails', []) if e.strip()]

        if not recipients:
            return

        from django.core.mail import EmailMessage
        subject = (
            f"New CPS Schedule of Loss — "
            f"{session.insured_name or session.claim_number or session.encircle_claim_id}"
        )
        body = (
            f"A new CPS Schedule of Loss has been generated.\n\n"
            f"  Insured : {session.insured_name or '—'}\n"
            f"  Claim # : {session.claim_number or '—'}\n"
            f"  Date    : {session.updated_at.strftime('%B %d, %Y')}\n\n"
            f"The Excel file is attached.\n\nSent automatically by Claimet App"
        )
        msg = EmailMessage(
            subject=subject, body=body,
            from_email=_cfg.DEFAULT_FROM_EMAIL, to=recipients,
        )
        msg.attach(
            filename, xlsx_bytes,
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        msg.send()
        logger.info(f"CPS Excel emailed to {recipients}")
    except Exception as e:
        logger.warning(f"CPS Excel email failed: {e}")


# ── PPR Box Count ─────────────────────────────────────────────────────────────

@login_required
def ppr_box_count_pdf(request, session_id):
    """Stream the NON SALVAGEABLE / PPR Box Count PDF.

    Same data as the client's CPS box count (BoxCalcCPSSession) — identical
    layout, rebranded title. The CPS box count itself stays untouched at
    /box-calculator/.
    """
    from box_calculator.models import BoxCalcCPSSession
    from box_calculator.ppr_pdf_builder import build_ppr_pdf

    session = get_object_or_404(CPSReportSession, id=session_id)
    cps_session = (
        BoxCalcCPSSession.unscoped
        .filter(client=session.client)
        .order_by('-updated_at')
        .first()
    )
    if cps_session is None:
        return HttpResponse(
            'No box count data exists for this client yet.\n'
            'Run the CPS box calculator for this client first '
            '(Box Calculator app -> CPS), then download this report again.',
            status=404, content_type='text/plain',
        )

    pdf_bytes = build_ppr_pdf(cps_session)

    client    = session.client
    safe_name = (client.pOwner or 'Unknown').replace(' ', '_').replace('/', '-')
    claim_num = (client.claimNumber or '').replace(' ', '_').replace('/', '-')
    suffix    = f"_{claim_num}" if claim_num else ''
    filename  = f"NON_SALVAGEABLE_PPR_Box_Count{suffix}_{safe_name}.pdf"

    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
