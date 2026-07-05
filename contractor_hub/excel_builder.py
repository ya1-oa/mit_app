"""
Contractor Hub — GC Estimate Excel Builder
============================================
Generates a professional Xactimate-format Excel workbook using openpyxl.

Public API:
    generate_gc_estimate_excel(estimate) -> io.BytesIO
"""

import io
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers as xl_numbers,
)
from openpyxl.utils import get_column_letter

# ── Colour constants (ARGB — no leading #) ───────────────────────────────────
C_PURPLE       = 'FF7C3AED'
C_PURPLE_LIGHT = 'FFF3E8FF'
C_DARK         = 'FF0F172A'
C_MID          = 'FF475569'
C_LIGHT        = 'FF94A3B8'
C_WHITE        = 'FFFFFFFF'
C_ROW_ALT      = 'FFF8FAFC'
C_BORDER       = 'FFE2E8F0'
C_INNER        = 'FFF1F5F9'
C_GREEN_BG     = 'FFD1FAE5'
C_GREEN_FG     = 'FF065F46'
C_AMBER_BG     = 'FFFEF3C7'
C_AMBER_FG     = 'FF92400E'

FMT_MONEY  = '"$"#,##0.00'
FMT_NUMBER = '#,##0.00'

# ── Helper factories ─────────────────────────────────────────────────────────

def _fill(argb):
    return PatternFill(fill_type='solid', fgColor=argb)

def _font(bold=False, size=10, color=C_DARK, italic=False, name='Calibri'):
    return Font(name=name, size=size, bold=bold, italic=italic, color=color)

def _align(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border(color=C_BORDER, style='thin'):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _thin_bottom(color=C_BORDER):
    s = Side(style='thin', color=color)
    return Border(bottom=s)

def _apply(ws, row, col, value=None, font=None, fill=None, align=None,
           border=None, number_format=None):
    cell = ws.cell(row=row, column=col)
    if value is not None:
        cell.value = value
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if align:
        cell.alignment = align
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format
    return cell


def generate_gc_estimate_excel(estimate) -> io.BytesIO:
    """
    Build a complete GC Estimate Excel workbook and return as BytesIO.

    Args:
        estimate: GCEstimate instance (sections + line_items prefetched)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'GC Estimate'

    # ── Sheet setup ───────────────────────────────────────────────────────────
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation   = 'landscape'
    ws.page_setup.fitToWidth    = 1
    ws.print_title_rows         = '1:1'

    # Column widths:  A     B      C           D     E     F        G        H    I
    # Corresponds to: CAT  SEL   Description  Qty  Unit  Remove   Replace  Tax  Total
    col_widths = [7, 11, 45, 8, 6, 12, 12, 5, 13]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 1  # current row cursor

    # ── Title banner ──────────────────────────────────────────────────────────
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    title_cell = ws.cell(row=r, column=1, value='GC ESTIMATE')
    title_cell.font      = _font(bold=True, size=16, color=C_WHITE)
    title_cell.fill      = _fill(C_PURPLE)
    title_cell.alignment = _align('center')
    ws.row_dimensions[r].height = 32
    r += 1

    # ── Claim info ────────────────────────────────────────────────────────────
    client = estimate.client
    gc     = estimate.gc_contractor

    def _info_row(label, value, start_col=1, end_col=4, label_col_span=None):
        nonlocal r
        # Label cell
        lc = ws.cell(row=r, column=start_col, value=label)
        lc.font      = _font(bold=True, size=8, color=C_LIGHT)
        lc.fill      = _fill(C_ROW_ALT)
        lc.alignment = _align('left')
        # Value cell
        vc = ws.cell(row=r, column=start_col + 1, value=value)
        vc.font      = _font(size=9)
        vc.fill      = _fill(C_WHITE)
        vc.alignment = _align('left')
        ws.merge_cells(start_row=r, start_column=start_col + 1, end_row=r, end_column=end_col)

    def _info_pair(l1, v1, l2, v2):
        nonlocal r
        c = ws.cell(row=r, column=1, value=l1)
        c.font = _font(bold=True, size=8, color=C_LIGHT); c.fill = _fill(C_ROW_ALT); c.alignment = _align()
        c = ws.cell(row=r, column=2, value=v1)
        c.font = _font(size=9); c.alignment = _align()
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        c = ws.cell(row=r, column=5, value=l2)
        c.font = _font(bold=True, size=8, color=C_LIGHT); c.fill = _fill(C_ROW_ALT); c.alignment = _align()
        c = ws.cell(row=r, column=6, value=v2)
        c.font = _font(size=9); c.alignment = _align()
        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=9)
        ws.row_dimensions[r].height = 16
        r += 1

    addr = (client.pAddress or '') + (f', {client.pCityStateZip}' if client.pCityStateZip else '')
    _info_pair('INSURED',    client.pOwner or '—',          'ADDRESS',   addr or '—')
    _info_pair('CLAIM #',    client.claimNumber or '—',      'CAUSE',     client.causeOfLoss or '—')
    _info_pair('GC',         gc.name,                         'ESTIMATOR', estimate.estimator.name if estimate.estimator else '—')
    est_date = estimate.date_entered.strftime('%b %d, %Y') if estimate.date_entered else '—'
    _info_pair('PRICE LIST', estimate.price_list or '—',    'DATE',      est_date)
    _info_pair('EST. NUMBER',estimate.estimate_number or '—','TYPE',      estimate.type_of_estimate or '—')
    _info_pair('STATUS',     estimate.get_status_display(),  'O&P',       f'Overhead {estimate.overhead_pct}% / Profit {estimate.profit_pct}%')

    r += 1  # blank separator row

    # ── Column headers (persistent, repeated each section) ───────────────────
    HEADERS = ['CAT', 'SEL', 'Description', 'Qty', 'Unit', 'Remove', 'Replace', 'Tax', 'Total']

    def _section_header(section):
        nonlocal r
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
        c = ws.cell(row=r, column=1,
                    value=f'{section.order}. {section.section_label}')
        c.font      = _font(bold=True, size=10, color=C_WHITE)
        c.fill      = _fill(C_PURPLE)
        c.alignment = _align('left')
        ws.row_dimensions[r].height = 20
        r += 1

        # Sub-header showing subcontractor
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
        sub_name = (section.subcontractor.name if section.subcontractor
                    else f'GC Direct — {estimate.gc_contractor.name}')
        c = ws.cell(row=r, column=1, value=sub_name)
        c.font      = _font(size=8, color=C_MID, italic=True)
        c.fill      = _fill(C_PURPLE_LIGHT)
        c.alignment = _align('left')
        ws.row_dimensions[r].height = 14
        r += 1

        # Column header row
        for col, hdr in enumerate(HEADERS, start=1):
            c = ws.cell(row=r, column=col, value=hdr)
            c.font      = _font(bold=True, size=8, color=C_MID)
            c.fill      = _fill(C_ROW_ALT)
            c.alignment = _align('right' if col >= 4 else 'left')
            c.border    = _border(C_BORDER)
        ws.row_dimensions[r].height = 16
        r += 1

    def _add_line(li, row_idx):
        nonlocal r
        alt = (row_idx % 2 == 1)
        bg  = _fill(C_ROW_ALT if alt else C_WHITE)

        vals = [
            li.cat, li.sel + (' [*]' if li.is_bid_item else ''),
            li.description, float(li.quantity), li.unit,
            float(li.remove_rate), float(li.replace_rate),
            'T' if li.taxable else '', float(li.line_total),
        ]
        fmts = [None, None, None, FMT_NUMBER, None, FMT_MONEY, FMT_MONEY, None, FMT_MONEY]
        aligns = ['left', 'left', 'left', 'right', 'center', 'right', 'right', 'center', 'right']
        bolds  = [True, True, False, False, False, False, False, False, True]

        for col, (v, fmt, aln, bold) in enumerate(zip(vals, fmts, aligns, bolds), start=1):
            c = ws.cell(row=r, column=col, value=v)
            c.fill      = bg
            c.font      = _font(bold=bold, size=8, color=C_PURPLE if col <= 2 else C_DARK)
            c.alignment = _align(aln, wrap=(col == 3))
            c.border    = _thin_bottom()
            if fmt:
                c.number_format = fmt
        ws.row_dimensions[r].height = 14
        r += 1

    def _section_subtotal(section):
        nonlocal r
        for col in range(1, 8):
            c = ws.cell(row=r, column=col)
            c.fill = _fill(C_PURPLE_LIGHT)
        # Label spans cols 1-8
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        lc = ws.cell(row=r, column=1, value='Section Subtotal')
        lc.font      = _font(bold=True, size=9, color=C_MID)
        lc.fill      = _fill(C_PURPLE_LIGHT)
        lc.alignment = _align('right')
        # Total in col 9
        tc = ws.cell(row=r, column=9, value=float(section.section_subtotal))
        tc.font         = _font(bold=True, size=9, color=C_PURPLE)
        tc.fill         = _fill(C_PURPLE_LIGHT)
        tc.alignment    = _align('right')
        tc.number_format= FMT_MONEY
        tc.border       = _border(C_PURPLE, 'medium')
        ws.row_dimensions[r].height = 16
        r += 2  # blank after section

    # ── Write all sections ────────────────────────────────────────────────────
    sections = estimate.sections.prefetch_related('line_items', 'subcontractor').order_by('order')

    for section in sections:
        _section_header(section)
        line_items = list(section.line_items.order_by('order'))
        if not line_items:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
            c = ws.cell(row=r, column=1, value='No line items.')
            c.font = _font(size=8, color=C_LIGHT, italic=True)
            c.alignment = _align('center')
            ws.row_dimensions[r].height = 14
            r += 1
        else:
            for i, li in enumerate(line_items):
                _add_line(li, i)
        _section_subtotal(section)

    # ── Totals block ─────────────────────────────────────────────────────────
    r += 1  # extra space before totals
    totals = [
        ('Line Item Total',                  float(estimate.line_item_total), False),
        (f'Overhead ({estimate.overhead_pct}%)', float(estimate.overhead_amount), False),
        (f'Profit ({estimate.profit_pct}%)',   float(estimate.profit_amount),   False),
        (f'Tax ({estimate.tax_rate}%)',         float(estimate.tax_amount),      False),
        ('GRAND TOTAL',                       float(estimate.grand_total),      True),
    ]

    for label, val, is_grand in totals:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        lc = ws.cell(row=r, column=1, value=label)
        lc.font      = _font(bold=True, size=11 if is_grand else 9,
                             color=C_PURPLE if is_grand else C_MID)
        lc.fill      = _fill(C_PURPLE_LIGHT if is_grand else C_WHITE)
        lc.alignment = _align('right')

        vc = ws.cell(row=r, column=9, value=val)
        vc.font         = _font(bold=True, size=12 if is_grand else 9,
                                color=C_PURPLE if is_grand else C_DARK)
        vc.fill         = _fill(C_PURPLE_LIGHT if is_grand else C_WHITE)
        vc.alignment    = _align('right')
        vc.number_format= FMT_MONEY
        if is_grand:
            vc.border = _border(C_PURPLE, 'thick')
        else:
            vc.border = _thin_bottom()
        ws.row_dimensions[r].height = 20 if is_grand else 15
        r += 1

    # ── Notes sheet (if any) ─────────────────────────────────────────────────
    if estimate.notes:
        wn = wb.create_sheet('Notes')
        wn['A1'] = 'Notes'
        wn['A1'].font = _font(bold=True, size=12, color=C_PURPLE)
        wn['A2'] = estimate.notes
        wn['A2'].alignment = _align('left', wrap=True)
        wn.column_dimensions['A'].width = 80

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
