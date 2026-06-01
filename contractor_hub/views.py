"""
Contractor Bid Hub — Views
"""
import csv
import io
import json

from django.contrib.auth.decorators import login_required
from django.db.models import Max
from django.http import JsonResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.utils import timezone

from docsAppR.models import Client

from .models import (
    Contractor, GCEstimate, GCSection, GCLineItem,
    RateItem, PriceListVersion, SectionType, SECTION_ORDER, SUBCONTRACTED_SECTIONS,
    BoxCountReport, LineItemTemplate, BoxType,
)

# ── Column aliases (same as management command) ─────────────────────────────
_COLUMN_ALIASES = {
    'cat':          ['cat', 'category', 'cat.', 'category code', 'trade'],
    'sel':          ['sel', 'selector', 'sel.', 'item code', 'activity code', 'code', 'item'],
    'description':  ['activity', 'description', 'act', 'desc', 'line item', 'item description'],
    'unit':         ['unit', 'unit of measure', 'uom', 'calc', 'unit/calc'],
    'remove_rate':  ['remove', 'rem', 'remove rate', 'rem rate', 'remove price'],
    'replace_rate': ['replace', 'rep', 'replace rate', 'rep rate', 'unit price', 'price', 'replace price'],
    'taxable':      ['tax', 'taxable', 'tax?', 'tax flag', 'tx'],
}

def _detect_col(headers):
    """Map actual CSV headers → our field names. Returns dict field→index."""
    h_lower = [h.strip().lower() for h in headers]
    mapping = {}
    for field, aliases in _COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in h_lower:
                mapping[field] = h_lower.index(alias)
                break
    return mapping

def _parse_csv_file(file_obj):
    """
    Parse an uploaded CSV/TSV and return (headers, rows_as_dicts, error_str).
    rows_as_dicts: list of {cat, sel, description, unit, remove_rate, replace_rate, taxable}
    """
    raw = file_obj.read()
    for enc in ('utf-8-sig', 'utf-8', 'latin-1'):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        return None, None, 'Cannot decode file — try saving as UTF-8.'

    # Detect delimiter
    sample = text[:2048]
    delimiter = '\t' if sample.count('\t') > sample.count(',') else ','

    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows = list(reader)
    if not rows:
        return None, None, 'File is empty.'

    headers = rows[0]
    col = _detect_col(headers)
    required = {'cat', 'sel', 'description', 'replace_rate'}
    missing = required - set(col.keys())
    if missing:
        return headers, None, f'Missing required columns: {", ".join(missing)}. Found: {", ".join(headers)}'

    parsed = []
    for i, row in enumerate(rows[1:], start=2):
        if not row or not any(row):
            continue
        def g(field, default=''):
            idx = col.get(field)
            return row[idx].strip() if idx is not None and idx < len(row) else default

        cat = g('cat').upper()
        sel = g('sel').upper()
        if not cat or not sel:
            continue

        try:
            remove_rate  = float(g('remove_rate',  '0').replace('$', '').replace(',', '') or '0')
            replace_rate = float(g('replace_rate', '0').replace('$', '').replace(',', '') or '0')
        except ValueError:
            continue

        taxable_raw = g('taxable', 'Y').upper()
        taxable = taxable_raw not in ('N', 'NO', 'FALSE', '0', 'F')

        parsed.append({
            'cat': cat, 'sel': sel,
            'description': g('description'),
            'unit': g('unit', 'EA').upper() or 'EA',
            'remove_rate': remove_rate,
            'replace_rate': replace_rate,
            'taxable': taxable,
        })

    return headers, parsed, None


def _parse_xactimate_excel(file_obj):
    """
    Parse a .xlsx/.xlsm Xactimate export.
    Tries the 'PRICES XACT IN' tab first, then any tab with price-like columns.
    Returns (headers, rows_as_dicts, error_str) — same contract as _parse_csv_file.
    """
    try:
        import openpyxl
    except ImportError:
        return None, None, 'openpyxl is required for Excel imports (pip install openpyxl).'

    try:
        raw = file_obj.read()
        wb  = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as e:
        return None, None, f'Cannot open Excel file: {e}'

    # Find the right sheet
    preferred = ['prices xact in', 'prices xact', 'xact in', 'price list', 'xactimate prices', 'rates']
    sheet = None
    for name in wb.sheetnames:
        if name.strip().lower() in preferred:
            sheet = wb[name]
            break
    if sheet is None:
        sheet = wb.worksheets[0]  # fall back to first sheet

    rows_raw = list(sheet.iter_rows(values_only=True))
    wb.close()
    if not rows_raw:
        return None, None, 'Selected sheet is empty.'

    # Find header row (first row that looks like column headers)
    header_idx = 0
    for i, row in enumerate(rows_raw[:10]):
        cells = [str(c or '').strip().lower() for c in row]
        if any(alias in cells for aliases in _COLUMN_ALIASES.values() for alias in aliases):
            header_idx = i
            break

    raw_headers = [str(c or '').strip() for c in rows_raw[header_idx]]
    col = _detect_col(raw_headers)
    required = {'cat', 'sel', 'description', 'replace_rate'}
    missing = required - set(col.keys())
    if missing:
        return raw_headers, None, (
            f'Missing required columns: {", ".join(missing)}. '
            f'Found: {", ".join(h for h in raw_headers if h)}'
        )

    parsed = []
    for row in rows_raw[header_idx + 1:]:
        if not row or all(c is None or str(c).strip() == '' for c in row):
            continue

        def g(field, default=''):
            idx = col.get(field)
            if idx is None or idx >= len(row):
                return default
            val = row[idx]
            return str(val).strip() if val is not None else default

        cat = g('cat').upper()
        sel = g('sel').upper()
        if not cat or not sel:
            continue

        try:
            remove_rate  = float(str(g('remove_rate',  '0')).replace('$', '').replace(',', '') or '0')
            replace_rate = float(str(g('replace_rate', '0')).replace('$', '').replace(',', '') or '0')
        except ValueError:
            continue

        taxable_raw = g('taxable', 'Y').upper()
        taxable = taxable_raw not in ('N', 'NO', 'FALSE', '0', 'F')

        parsed.append({
            'cat': cat, 'sel': sel,
            'description': g('description'),
            'unit': g('unit', 'EA').upper() or 'EA',
            'remove_rate': remove_rate,
            'replace_rate': replace_rate,
            'taxable': taxable,
        })

    if not parsed:
        return raw_headers, None, 'No valid price rows found in the sheet.'
    return raw_headers, parsed, None


# Box type keyword matcher for file imports
_BOX_KEYWORDS = {
    'small_boxes':     ['small'],
    'medium_boxes':    ['medium'],
    'large_boxes':     ['large'],
    'xl_items':        ['xl', 'unboxed', 'x-large'],
    'mirror_boxes':    ['mirror', 'picture'],
    'lamp_boxes':      ['lamp', 'plant', 'vase'],
    'tv_boxes':        ['tv', 'television'],
    'wardrobe_boxes':  ['wardrobe'],
    'mattress_boxes':  ['mattress'],
    'dishpack_boxes':  ['dish pack', 'dishpack', 'dish'],
    'glasspack_boxes': ['glass pack', 'glasspack', 'glass'],
    'pots_boxes':      ['pots', 'pans'],
}

def _match_box_field(text):
    """Return BCR field name for a row label, or None."""
    t = text.strip().lower()
    for field, keywords in _BOX_KEYWORDS.items():
        for kw in keywords:
            if kw in t:
                return field
    return None


def _parse_box_count_file(file_obj):
    """
    Parse a box count report from CSV or Excel (xlsm/xlsx).
    Supports:
      - Columnar: rows like "Small Boxes | 52"
      - Row-header: first row = box types, second row = counts
    Returns (counts_dict, source_name, error_str).
    counts_dict keys = BCR field names (small_boxes, medium_boxes, etc.)
    """
    name     = getattr(file_obj, 'name', '')
    ext      = name.rsplit('.', 1)[-1].lower() if '.' in name else ''
    counts   = {}

    if ext in ('xlsx', 'xlsm', 'xls'):
        try:
            import openpyxl
        except ImportError:
            return None, name, 'openpyxl is required for Excel imports.'

        try:
            raw = file_obj.read()
            wb  = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        except Exception as e:
            return None, name, f'Cannot open Excel file: {e}'

        # Find 'Box Summary' sheet (any variant)
        box_sheet = None
        for sname in wb.sheetnames:
            if 'box' in sname.lower() and ('summary' in sname.lower() or 'count' in sname.lower()):
                box_sheet = wb[sname]
                break
        if box_sheet is None:
            # Try any sheet that has box-type keywords
            for sname in wb.sheetnames:
                rows_check = list(wb[sname].iter_rows(max_row=20, values_only=True))
                flat = ' '.join(str(c or '') for row in rows_check for c in row).lower()
                if 'small' in flat and ('medium' in flat or 'large' in flat):
                    box_sheet = wb[sname]
                    break
        if box_sheet is None:
            box_sheet = wb.worksheets[0]

        rows_raw = list(box_sheet.iter_rows(values_only=True))
        wb.close()

        # Strategy 1: two-column layout (label | count)
        for row in rows_raw:
            cells = [str(c or '').strip() for c in row]
            non_empty = [c for c in cells if c]
            if len(non_empty) >= 2:
                label = non_empty[0]
                field = _match_box_field(label)
                if field:
                    try:
                        counts[field] = int(float(non_empty[1]))
                    except (ValueError, TypeError):
                        pass

        # Strategy 2: row-header layout (if strategy 1 found nothing)
        if not counts and rows_raw:
            header_row = [str(c or '').strip() for c in rows_raw[0]]
            if len(rows_raw) > 1:
                data_row = [str(c or '').strip() for c in rows_raw[1]]
                for i, label in enumerate(header_row):
                    field = _match_box_field(label)
                    if field and i < len(data_row):
                        try:
                            counts[field] = int(float(data_row[i]))
                        except (ValueError, TypeError):
                            pass

        # Strategy 3: per-room grid with a Totals row (CPS Box Summary format).
        # Row 5 = column headers (Small Box, Medium Box, …), rows 6-N = per room,
        # last data row first cell = "Total" with column sums.
        if not counts:
            # Find the header row — the first row that has ≥2 box-type keywords
            hdr_idx, hdr_cells = None, None
            for i, row in enumerate(rows_raw):
                cells = [str(c or '').strip() for c in row]
                if sum(1 for c in cells if _match_box_field(c)) >= 2:
                    hdr_idx, hdr_cells = i, cells
                    break
            if hdr_idx is not None:
                # Find the Total row — first non-None cell is "total"
                for row in rows_raw[hdr_idx + 1:]:
                    first = str(row[0] or '').strip().lower() if row else ''
                    if first == 'total':
                        for col_idx, header_cell in enumerate(hdr_cells):
                            field = _match_box_field(header_cell)
                            if field and col_idx < len(row) and row[col_idx] is not None:
                                try:
                                    counts[field] = int(float(str(row[col_idx])))
                                except (ValueError, TypeError):
                                    pass
                        break

    else:
        # CSV path
        raw = file_obj.read()
        for enc in ('utf-8-sig', 'utf-8', 'latin-1'):
            try:
                text = raw.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        else:
            return None, name, 'Cannot decode file.'

        sample = text[:1024]
        delim  = '\t' if sample.count('\t') > sample.count(',') else ','
        reader = csv.reader(io.StringIO(text), delimiter=delim)
        rows   = [r for r in reader if any(c.strip() for c in r)]

        # Two-column layout
        for row in rows:
            if len(row) >= 2:
                field = _match_box_field(row[0])
                if field:
                    try:
                        counts[field] = int(float(row[1].replace(',', '').strip()))
                    except (ValueError, TypeError):
                        pass

        # Row-header layout fallback
        if not counts and len(rows) >= 2:
            headers = rows[0]
            data    = rows[1]
            for i, label in enumerate(headers):
                field = _match_box_field(label)
                if field and i < len(data):
                    try:
                        counts[field] = int(float(data[i].replace(',', '').strip()))
                    except (ValueError, TypeError):
                        pass

    if not counts:
        return None, name, 'No box count data found. Check the file format.'
    return counts, name, None


# ---------------------------------------------------------------------------
# Dashboard
# ---------------------------------------------------------------------------

@login_required
def dashboard(request):
    clients     = Client.objects.prefetch_related('box_count_report').order_by('pOwner')
    subs        = Contractor.objects.filter(is_active=True).exclude(role='gc').order_by('name')
    latest_pl   = PriceListVersion.objects.order_by('-imported_at').first()
    total_rates = RateItem.objects.count()
    clients_with_bcr = sum(1 for c in clients if hasattr(c, 'box_count_report'))
    return render(request, 'contractor_hub/dashboard.html', {
        'clients':          clients,
        'subs':             subs,
        'latest_pl':        latest_pl,
        'total_rates':      total_rates,
        'total_clients':    clients.count(),
        'clients_with_bcr': clients_with_bcr,
    })


# ---------------------------------------------------------------------------
# Estimates
# ---------------------------------------------------------------------------

@login_required
def estimate_create(request):
    from docsAppR.models import Client
    clients  = Client.objects.order_by('pOwner')
    gcs      = Contractor.objects.filter(is_active=True).order_by('name')
    estimators = Contractor.objects.filter(is_active=True).order_by('name')

    if request.method == 'POST':
        client_id    = request.POST.get('client')
        gc_id        = request.POST.get('gc_contractor')
        estimator_id = request.POST.get('estimator') or None
        est_number   = request.POST.get('estimate_number', '')
        price_list   = request.POST.get('price_list', '')
        type_of_est  = request.POST.get('type_of_estimate', 'Fire')
        date_entered = request.POST.get('date_entered') or None
        tax_rate     = request.POST.get('tax_rate', '8.25')
        notes        = request.POST.get('notes', '')

        try:
            client = Client.objects.get(pk=client_id)
            gc     = Contractor.objects.get(pk=gc_id)

            estimate = GCEstimate.objects.create(
                client=client,
                gc_contractor=gc,
                estimator_id=estimator_id,
                estimate_number=est_number,
                price_list=price_list,
                type_of_estimate=type_of_est,
                date_entered=date_entered,
                tax_rate=tax_rate,
                notes=notes,
                created_by=request.user,
            )

            # Auto-create all 8 fixed sections
            _create_default_sections(estimate)

            messages.success(request, f'Estimate {estimate.estimate_number or estimate.id} created.')
            return redirect('contractor_hub:estimate_detail', pk=estimate.pk)

        except Exception as e:
            messages.error(request, f'Error creating estimate: {e}')

    return render(request, 'contractor_hub/estimate_form.html', {
        'clients': clients,
        'gcs': gcs,
        'estimators': estimators,
        'action': 'Create',
        'type_choices': ['Fire', 'Water', 'Wind', 'Mold', 'Other'],
    })


@login_required
def estimate_detail(request, pk):
    estimate = get_object_or_404(
        GCEstimate.objects.select_related('client', 'gc_contractor', 'estimator'),
        pk=pk,
    )
    sections = estimate.sections.prefetch_related(
        'line_items', 'subcontractor'
    ).order_by('order')
    rate_items = RateItem.objects.order_by('cat', 'sel')

    return render(request, 'contractor_hub/estimate_detail.html', {
        'estimate': estimate,
        'sections': sections,
        'rate_items': rate_items,
    })


@login_required
def estimate_edit(request, pk):
    estimate   = get_object_or_404(GCEstimate, pk=pk)
    gcs        = Contractor.objects.filter(is_active=True).order_by('name')
    estimators = Contractor.objects.filter(is_active=True).order_by('name')

    if request.method == 'POST':
        estimate.estimate_number  = request.POST.get('estimate_number', estimate.estimate_number)
        estimate.price_list       = request.POST.get('price_list', estimate.price_list)
        estimate.type_of_estimate = request.POST.get('type_of_estimate', estimate.type_of_estimate)
        estimate.date_entered     = request.POST.get('date_entered') or estimate.date_entered
        estimate.tax_rate         = request.POST.get('tax_rate', estimate.tax_rate)
        estimate.status           = request.POST.get('status', estimate.status)
        estimate.notes            = request.POST.get('notes', estimate.notes)
        gc_id = request.POST.get('gc_contractor')
        if gc_id:
            estimate.gc_contractor_id = gc_id
        estimate.save()
        messages.success(request, 'Estimate updated.')
        return redirect('contractor_hub:estimate_detail', pk=estimate.pk)

    from .models import EstimateStatus
    return render(request, 'contractor_hub/estimate_form.html', {
        'estimate': estimate,
        'gcs': gcs,
        'estimators': estimators,
        'action': 'Edit',
        'type_choices': ['Fire', 'Water', 'Wind', 'Mold', 'Other'],
        'status_choices': EstimateStatus.choices,
    })


# ---------------------------------------------------------------------------
# Section detail (manage line items for one section)
# ---------------------------------------------------------------------------

@login_required
def section_detail(request, pk, section_pk):
    estimate = get_object_or_404(GCEstimate, pk=pk)
    section  = get_object_or_404(GCSection, pk=section_pk, estimate=estimate)
    line_items = section.line_items.order_by('order')
    rate_items = RateItem.objects.filter(section_hint=section.section_type).order_by('cat', 'sel')
    subs       = Contractor.objects.filter(is_active=True).exclude(role='gc')

    if request.method == 'POST':
        # Update section subcontractor / bid status
        sub_id = request.POST.get('subcontractor') or None
        section.subcontractor_id = sub_id
        section.bid_status = request.POST.get('bid_status', section.bid_status)
        section.notes = request.POST.get('notes', section.notes)
        section.save()
        messages.success(request, f'{section.section_label} updated.')
        return redirect('contractor_hub:section_detail', pk=pk, section_pk=section_pk)

    return render(request, 'contractor_hub/section_detail.html', {
        'estimate': estimate,
        'section': section,
        'line_items': line_items,
        'rate_items': rate_items,
        'subs': subs,
    })


@login_required
def section_import_cps(request, pk, section_pk):
    """Auto-populate packing/transport/cleaning sections from the CPS session."""
    estimate = get_object_or_404(GCEstimate, pk=pk)
    section  = get_object_or_404(GCSection, pk=section_pk, estimate=estimate)

    if request.method == 'POST':
        from box_calculator.models import BoxCalcCPSSession
        try:
            cps = BoxCalcCPSSession.objects.filter(
                client=estimate.client
            ).latest('updated_at')
            counts = cps.grand_counts

            lines_created = _auto_populate_from_cps(section, counts)
            messages.success(request, f'{lines_created} line items imported from CPS session.')
        except BoxCalcCPSSession.DoesNotExist:
            messages.error(request, 'No CPS session found for this client.')

    return redirect('contractor_hub:section_detail', pk=pk, section_pk=section_pk)


# ---------------------------------------------------------------------------
# Price List Import (upload → preview → confirm)
# ---------------------------------------------------------------------------

@login_required
def price_list_import(request):
    """
    Step 1 (GET / POST file):  parse CSV, store preview in session, show preview page.
    Step 2 (POST confirm):     read from session, save to DB.
    """
    step = request.GET.get('step', '1')

    # ── Step 2: Confirm ──────────────────────────────────────────────────────
    if request.method == 'POST' and request.POST.get('action') == 'confirm':
        preview_json = request.session.get('pl_preview')
        if not preview_json:
            messages.error(request, 'Session expired. Please re-upload the file.')
            return redirect('contractor_hub:price_list_import')

        meta = preview_json['meta']
        rows = preview_json['rows']

        version = PriceListVersion.objects.create(
            code=meta['code'],
            market=meta.get('market', ''),
            effective_date=meta.get('effective_date') or None,
            source_file=meta.get('filename', ''),
            imported_by=request.user,
        )

        created = updated = skipped = 0
        for r in rows:
            obj, was_created = RateItem.objects.get_or_create(
                cat=r['cat'], sel=r['sel'],
                defaults={
                    'description': r['description'],
                    'unit': r['unit'],
                    'remove_rate': r['remove_rate'],
                    'replace_rate': r['replace_rate'],
                    'taxable': r['taxable'],
                    'price_list_version': version,
                },
            )
            if was_created:
                created += 1
            else:
                # Track previous rates
                obj.previous_remove_rate  = obj.remove_rate
                obj.previous_replace_rate = obj.replace_rate
                obj.remove_rate   = r['remove_rate']
                obj.replace_rate  = r['replace_rate']
                obj.description   = r['description']
                obj.unit          = r['unit']
                obj.taxable       = r['taxable']
                obj.price_list_version = version
                obj.last_updated_at = timezone.now()
                obj.save()
                if obj.remove_rate != obj.previous_remove_rate or \
                   obj.replace_rate != obj.previous_replace_rate:
                    updated += 1
                else:
                    skipped += 1

        version.total_items   = len(rows)
        version.items_created = created
        version.items_updated = updated
        version.items_skipped = skipped
        version.save()

        del request.session['pl_preview']
        messages.success(
            request,
            f'Import complete — {created} new, {updated} updated, {skipped} unchanged.'
        )
        return redirect('contractor_hub:price_list_history')

    # ── Step 1: Upload & Parse ───────────────────────────────────────────────
    if request.method == 'POST' and 'csv_file' in request.FILES:
        uploaded = request.FILES['csv_file']
        code     = request.POST.get('code', '').strip().upper()
        market   = request.POST.get('market', '').strip()
        eff_date = request.POST.get('effective_date', '').strip() or None

        if not code:
            messages.error(request, 'Price list code is required (e.g., OHCL8X_MAR26).')
            return render(request, 'contractor_hub/price_list_import.html', {})

        if PriceListVersion.objects.filter(code=code).exists():
            messages.error(request, f'A price list with code "{code}" already exists. '
                                    f'Choose a different code or delete the existing one.')
            return render(request, 'contractor_hub/price_list_import.html', {
                'form_data': {'code': code, 'market': market, 'effective_date': eff_date},
            })

        # Choose parser based on file extension
        fname = uploaded.name.lower()
        if fname.endswith('.xlsx') or fname.endswith('.xlsm') or fname.endswith('.xls'):
            headers, rows, error = _parse_xactimate_excel(uploaded)
        else:
            headers, rows, error = _parse_csv_file(uploaded)
        if error:
            messages.error(request, error)
            return render(request, 'contractor_hub/price_list_import.html', {})

        # Build change comparison for preview
        preview_rows = []
        for r in rows:
            existing = RateItem.objects.filter(cat=r['cat'], sel=r['sel']).first()
            if existing:
                r['row_exists']   = True
                r['old_remove']   = float(existing.remove_rate)
                r['old_replace']  = float(existing.replace_rate)
                r['row_changed']  = (
                    abs(float(existing.remove_rate)  - r['remove_rate'])  > 0.001 or
                    abs(float(existing.replace_rate) - r['replace_rate']) > 0.001
                )
            else:
                r['row_exists']  = False
                r['row_changed'] = True

            preview_rows.append(r)

        new_count     = sum(1 for r in preview_rows if not r['row_exists'])
        update_count  = sum(1 for r in preview_rows if r['row_exists'] and r['row_changed'])
        skip_count    = sum(1 for r in preview_rows if r['row_exists'] and not r['row_changed'])

        request.session['pl_preview'] = {
            'meta': {'code': code, 'market': market, 'effective_date': eff_date,
                     'filename': uploaded.name},
            'rows': rows,  # clean rows (no _* preview keys) for DB save
        }

        return render(request, 'contractor_hub/price_list_import.html', {
            'step': 2,
            'preview_rows': preview_rows,
            'meta': {'code': code, 'market': market, 'effective_date': eff_date,
                     'filename': uploaded.name, 'total': len(rows)},
            'new_count':    new_count,
            'update_count': update_count,
            'skip_count':   skip_count,
        })

    return render(request, 'contractor_hub/price_list_import.html', {'step': 1})


@login_required
def price_list_history(request):
    versions = PriceListVersion.objects.select_related('imported_by').order_by('-imported_at')
    return render(request, 'contractor_hub/price_list_history.html', {
        'versions': versions,
    })


# ---------------------------------------------------------------------------
# PDF / Excel generation (stubs — Phase 5/6)
# ---------------------------------------------------------------------------

@login_required
def estimate_pdf(request, pk):
    """Phase 5 — Generate and serve a GC Estimate PDF."""
    from django.http import HttpResponse
    from .pdf_builder import generate_gc_estimate_pdf

    estimate = get_object_or_404(
        GCEstimate.objects.select_related('client', 'gc_contractor', 'estimator'),
        pk=pk,
    )
    # Prefetch sections + line items for efficiency
    estimate.sections.prefetch_related('line_items', 'subcontractor')

    try:
        buf = generate_gc_estimate_pdf(estimate)
        safe_name = (
            (estimate.estimate_number or f'EST-{str(estimate.id)[:8]}')
            .replace(' ', '_').replace('/', '-')
        )
        response = HttpResponse(buf.read(), content_type='application/pdf')
        response['Content-Disposition'] = (
            f'attachment; filename="{safe_name}.pdf"'
        )
        return response
    except Exception as e:
        import logging
        logging.getLogger(__name__).error(f'PDF generation failed for {pk}: {e}', exc_info=True)
        messages.error(request, f'PDF generation failed: {e}')
        return redirect('contractor_hub:estimate_detail', pk=pk)


@login_required
def estimate_excel(request, pk):
    """Phase 6 — Generate and serve a GC Estimate Excel workbook."""
    from django.http import HttpResponse
    from .excel_builder import generate_gc_estimate_excel

    estimate = get_object_or_404(
        GCEstimate.objects.select_related('client', 'gc_contractor', 'estimator'),
        pk=pk,
    )
    estimate.sections.prefetch_related('line_items', 'subcontractor')

    try:
        buf = generate_gc_estimate_excel(estimate)
        safe_name = (
            (estimate.estimate_number or f'EST-{str(estimate.id)[:8]}')
            .replace(' ', '_').replace('/', '-')
        )
        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = (
            f'attachment; filename="{safe_name}.xlsx"'
        )
        return response
    except Exception as e:
        import logging
        logging.getLogger(__name__).error(f'Excel generation failed for {pk}: {e}', exc_info=True)
        messages.error(request, f'Excel generation failed: {e}')
        return redirect('contractor_hub:estimate_detail', pk=pk)


@login_required
def section_invoice_pdf(request, pk, section_pk):
    """Subcontractor invoice PDF for a single section."""
    from django.http import HttpResponse
    from .pdf_builder import generate_subcontractor_invoice_pdf

    estimate = get_object_or_404(
        GCEstimate.objects.select_related('client', 'gc_contractor', 'estimator'),
        pk=pk,
    )
    section = get_object_or_404(
        GCSection.objects.prefetch_related('line_items').select_related('subcontractor'),
        pk=section_pk, estimate=estimate,
    )

    try:
        buf = generate_subcontractor_invoice_pdf(estimate, section)
        safe_est  = (estimate.estimate_number or f'EST-{str(estimate.id)[:8]}').replace(' ', '_').replace('/', '-')
        safe_sect = section.section_label.replace(' ', '_').replace('/', '-')
        fname     = f'{safe_est}_{safe_sect}_SUB_INVOICE.pdf'
        response  = HttpResponse(buf.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{fname}"'
        return response
    except Exception as e:
        import logging
        logging.getLogger(__name__).error(f'Sub invoice PDF failed for section {section_pk}: {e}', exc_info=True)
        messages.error(request, f'Sub invoice PDF generation failed: {e}')
        return redirect('contractor_hub:section_detail', pk=pk, section_pk=section_pk)


# ---------------------------------------------------------------------------
# Quick Sub Invoice — 3-input generator (claim + sub + work type → PDF)
# ---------------------------------------------------------------------------

# Which section types can be generated via the quick invoice tool
QUICK_INVOICE_SECTIONS = [
    SectionType.PACKING,
    SectionType.CLEANING,
    SectionType.TRANSPORT,
    SectionType.ADMIN,
    SectionType.DEMO,
]


@login_required
def quick_sub_invoice(request):
    """
    GET  — show the 3-input form (client, subcontractor, work type).
    POST — build line items from LineItemTemplates × BoxCountReport, generate PDF.
    """
    from django.http import HttpResponse
    from .pdf_builder import generate_quick_sub_invoice_pdf
    from decimal import Decimal

    clients         = Client.objects.prefetch_related('box_count_report').order_by('pOwner')
    subs            = Contractor.objects.filter(is_active=True).exclude(role='gc').order_by('name')
    section_choices = [(st.value, st.label) for st in QUICK_INVOICE_SECTIONS]

    if request.method == 'POST':
        client_pk    = request.POST.get('client')
        sub_pk       = request.POST.get('subcontractor')
        section_type = request.POST.get('section_type')

        client = get_object_or_404(Client, pk=client_pk)
        sub    = get_object_or_404(Contractor, pk=sub_pk)

        # Validate section type
        valid_types = [st.value for st in QUICK_INVOICE_SECTIONS]
        if section_type not in valid_types:
            messages.error(request, f'Invalid work type: {section_type}')
            return redirect('contractor_hub:quick_sub_invoice')

        # Box count report — now linked to client directly
        try:
            bcr = client.box_count_report
        except BoxCountReport.DoesNotExist:
            messages.error(request,
                'No Box Count Report for this client. Enter box counts first.')
            return redirect('contractor_hub:box_count_report', pk=client.pk)

        box_counts = bcr.as_dict()

        # Templates for this section type
        templates = LineItemTemplate.objects.filter(
            section_type=section_type
        ).order_by('group_code', 'order')

        if not templates.exists():
            messages.error(request,
                f'No line item templates for "{section_type}". '
                f'Run: python manage.py seed_line_item_templates')
            return redirect('contractor_hub:quick_sub_invoice')

        # Live rate lookup
        rate_lookup = {(r.cat, r.sel): r for r in RateItem.objects.all()}

        line_items_data = []
        for tmpl in templates:
            qty = tmpl.compute_qty(box_counts)
            if qty == Decimal('0.00') and tmpl.box_type != BoxType.FIXED:
                continue

            rate         = rate_lookup.get((tmpl.cat, tmpl.sel))
            remove_rate  = rate.remove_rate  if rate else Decimal('0.00')
            replace_rate = rate.replace_rate if rate else Decimal('0.00')

            line_items_data.append({
                'cat':          tmpl.cat,
                'sel':          tmpl.sel,
                'description':  tmpl.description,
                'qty':          qty,
                'unit':         tmpl.unit,
                'remove_rate':  remove_rate,
                'replace_rate': replace_rate,
                'taxable':      tmpl.taxable,
                'notes':        tmpl.notes,
                'order':        tmpl.order,
                'calc_formula': f'LL×{tmpl.qty_factor}',
            })

        if not line_items_data:
            messages.error(request,
                'All box counts are zero — nothing to generate. Check the Box Count Report.')
            return redirect('contractor_hub:box_count_report', pk=client.pk)

        # Build estimate-like object for the PDF builder
        try:
            estimate_obj = GCEstimate.objects.select_related(
                'gc_contractor', 'estimator'
            ).get(client=client)
        except GCEstimate.DoesNotExist:
            class _ClientEstimate:
                def __init__(self, c):
                    self.estimate_number  = c.claimNumber or f'CLM-{c.pk}'
                    self.id               = c.pk
                    self.client           = c
                    self.gc_contractor    = None
                    self.estimator        = None
                    self.overhead_pct     = Decimal('10.00')
                    self.profit_pct       = Decimal('10.00')
                    self.tax_rate         = Decimal('8.25')
                    self.type_of_estimate = c.causeOfLoss or 'Contents'
                    self.date_entered     = None
                    self.price_list       = ''
            estimate_obj = _ClientEstimate(client)

        try:
            buf         = generate_quick_sub_invoice_pdf(
                estimate_obj, sub, section_type, line_items_data
            )
            safe_client = (client.pOwner or f'CLM-{client.pk}').replace(' ', '_')[:20]
            safe_sub    = sub.name.replace(' ', '_')[:20]
            safe_sect   = section_type.upper()
            fname       = f'{safe_client}_{safe_sub}_{safe_sect}_INVOICE.pdf'
            response    = HttpResponse(buf.read(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{fname}"'
            return response
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f'Quick sub invoice failed: {e}', exc_info=True)
            messages.error(request, f'PDF generation failed: {e}')
            return redirect('contractor_hub:quick_sub_invoice')

    return render(request, 'contractor_hub/quick_sub_invoice.html', {
        'clients':         clients,
        'subs':            subs,
        'section_choices': section_choices,
    })


@login_required
def box_count_report(request, pk):
    """
    GET/POST — create or update the BoxCountReport for a client (pk=client.pk).
    Supports:
      - Manual entry via form fields
      - File import (CSV or Excel with Box Summary tab)
    """
    client = get_object_or_404(Client, pk=pk)
    bcr, _ = BoxCountReport.objects.get_or_create(client=client)

    BCR_FIELDS = [
        'small_boxes', 'medium_boxes', 'large_boxes', 'xl_items',
        'mirror_boxes', 'lamp_boxes', 'tv_boxes', 'wardrobe_boxes',
        'mattress_boxes', 'dishpack_boxes', 'glasspack_boxes', 'pots_boxes',
    ]

    if request.method == 'POST':
        # ── File import path ────────────────────────────────────────────────
        if 'box_file' in request.FILES:
            uploaded = request.FILES['box_file']
            counts, source, error = _parse_box_count_file(uploaded)
            if error:
                messages.error(request, f'Import failed: {error}')
            else:
                for field in BCR_FIELDS:
                    if field in counts:
                        setattr(bcr, field, counts[field])
                bcr.source_file = source
                bcr.notes       = request.POST.get('notes', bcr.notes)
                bcr.save()
                found = ', '.join(f'{k}={v}' for k, v in counts.items())
                messages.success(request,
                    f'Imported from {source} — {bcr.total_boxes} total boxes. ({found})')
                return redirect('contractor_hub:box_count_report', pk=client.pk)

        # ── Manual entry path ───────────────────────────────────────────────
        for f in BCR_FIELDS:
            val = request.POST.get(f, '0')
            try:
                setattr(bcr, f, max(0, int(val)))
            except (ValueError, TypeError):
                setattr(bcr, f, 0)
        bcr.notes = request.POST.get('notes', '')
        bcr.save()
        messages.success(request, f'Box counts saved — {bcr.total_boxes} total boxes.')
        next_url = request.POST.get('next')
        if next_url:
            return redirect(next_url)
        return redirect('contractor_hub:quick_sub_invoice')

    return render(request, 'contractor_hub/box_count_report.html', {
        'client': client,
        'bcr':    bcr,
    })


# ---------------------------------------------------------------------------
# Contractor Registry
# ---------------------------------------------------------------------------

@login_required
def contractor_list(request):
    contractors = Contractor.objects.order_by('role', 'name')
    return render(request, 'contractor_hub/contractor_list.html', {
        'contractors': contractors,
    })


@login_required
def contractor_create(request):
    if request.method == 'POST':
        Contractor.objects.create(
            name=request.POST['name'],
            ein=request.POST.get('ein', ''),
            role=request.POST.get('role', 'other'),
            address=request.POST.get('address', ''),
            city=request.POST.get('city', ''),
            state=request.POST.get('state', ''),
            zip_code=request.POST.get('zip_code', ''),
            phone=request.POST.get('phone', ''),
            phone2=request.POST.get('phone2', ''),
            email=request.POST.get('email', ''),
            contact_person=request.POST.get('contact_person', ''),
            certification=request.POST.get('certification', ''),
            notes=request.POST.get('notes', ''),
            is_active=bool(request.POST.get('is_active')),
        )
        messages.success(request, 'Contractor added.')
        return redirect('contractor_hub:contractor_list')
    return render(request, 'contractor_hub/contractor_form.html', {
        'contractor': None,
        'action': 'Add',
    })


@login_required
def contractor_edit(request, pk):
    contractor = get_object_or_404(Contractor, pk=pk)
    if request.method == 'POST':
        for field in ['name', 'ein', 'role', 'address', 'city', 'state', 'zip_code',
                      'phone', 'phone2', 'email', 'contact_person', 'certification', 'notes']:
            if field in request.POST:
                setattr(contractor, field, request.POST[field])
        contractor.is_active = bool(request.POST.get('is_active'))
        contractor.save()
        messages.success(request, f'{contractor.name} updated.')
        return redirect('contractor_hub:contractor_list')
    return render(request, 'contractor_hub/contractor_form.html', {
        'contractor': contractor,
        'action': 'Edit',
    })


# ---------------------------------------------------------------------------
# JSON API
# ---------------------------------------------------------------------------

@login_required
def api_stats(request):
    return JsonResponse({
        'rate_count':       RateItem.objects.count(),
        'contractor_count': Contractor.objects.filter(is_active=True).count(),
        'estimate_count':   GCEstimate.objects.count(),
    })


@login_required
def api_estimate_totals(request, pk):
    estimate = get_object_or_404(GCEstimate, pk=pk)
    sections_data = [
        {
            'id': s.pk,
            'label': s.section_label,
            'subtotal': float(s.section_subtotal),
        }
        for s in estimate.sections.prefetch_related('line_items').order_by('order')
    ]
    return JsonResponse({
        'line_item_total': float(estimate.line_item_total),
        'overhead':        float(estimate.overhead_amount),
        'profit':          float(estimate.profit_amount),
        'tax':             float(estimate.tax_amount),
        'grand_total':     float(estimate.grand_total),
        'sections':        sections_data,
    })


@login_required
def api_lineitem_add(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    section_pk = request.POST.get('section_pk')
    section = get_object_or_404(GCSection, pk=section_pk)

    # Get rate from library if sel provided
    cat  = request.POST.get('cat', '')
    sel  = request.POST.get('sel', '')
    rate = None
    if cat and sel:
        rate = RateItem.objects.filter(cat=cat, sel=sel).first()

    qty  = request.POST.get('quantity', '0')
    last_order = section.line_items.aggregate(m=Max('order'))['m'] or 0

    li = GCLineItem.objects.create(
        section=section,
        rate_item=rate,
        cat=cat,
        sel=sel,
        description=request.POST.get('description', rate.description if rate else ''),
        quantity=qty,
        unit=request.POST.get('unit', rate.unit if rate else 'EA'),
        remove_rate=request.POST.get('remove_rate', rate.remove_rate if rate else '0'),
        replace_rate=request.POST.get('replace_rate', rate.replace_rate if rate else '0'),
        taxable=request.POST.get('taxable', 'true').lower() == 'true',
        is_bid_item=rate.is_bid_item if rate else False,
        calc_formula=request.POST.get('calc_formula', ''),
        notes=request.POST.get('notes', ''),
        order=last_order + 10,
    )
    return JsonResponse({'id': li.pk, 'line_total': float(li.line_total)})


@login_required
def api_lineitem_update(request, pk):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    li = get_object_or_404(GCLineItem, pk=pk)
    for field in ['quantity', 'remove_rate', 'replace_rate', 'description',
                  'calc_formula', 'notes', 'taxable']:
        if field in request.POST:
            val = request.POST[field]
            if field == 'taxable':
                val = val.lower() == 'true'
            setattr(li, field, val)
    li.save()
    return JsonResponse({'id': li.pk, 'line_total': float(li.line_total)})


@login_required
def api_lineitem_delete(request, pk):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    li = get_object_or_404(GCLineItem, pk=pk)
    li.delete()
    return JsonResponse({'deleted': pk})


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _create_default_sections(estimate):
    """Create all 8 fixed sections for a new estimate, pre-assigned to standard subs."""
    # Default sub assignments by section type (can be overridden per estimate)
    DEFAULT_SUBS = {}
    try:
        DEFAULT_SUBS = {
            SectionType.ADMIN:    Contractor.objects.get(ein='92-0685-963'),  # Adens Perfection
            SectionType.PACKING:  Contractor.objects.get(ein='83-2260563'),   # All Phase
            SectionType.STORAGE:  Contractor.objects.get(ein='83-131-5114'),  # BAL
            SectionType.CLEANING: Contractor.objects.get(ein='92-1783835'),   # Ian His Hands
            SectionType.DEMO:     Contractor.objects.get(ein='84-460-8968'),  # CAL
        }
    except Contractor.DoesNotExist:
        pass  # Subs not yet seeded — will be assigned manually

    for section_type, order in SECTION_ORDER.items():
        GCSection.objects.create(
            estimate=estimate,
            section_type=section_type,
            order=order,
            subcontractor=DEFAULT_SUBS.get(section_type),
        )


def _auto_populate_from_cps(section, counts):
    """
    Populate line items in a section from CPS box counts.
    Returns number of lines created.
    """
    # Mapping: CPS column name → packing line item config
    PACKING_MAP = {
        'small':       {'eval_sel': 'BXMSE', 'eval_rate': '14.15', 'labor_mult': '0.2'},
        'medium':      {'eval_sel': 'BXMME', 'eval_rate': '17.36', 'labor_mult': '0.3'},
        'large':       {'box_sel': 'BX>',  'box_rate': '5.28',  'eval_sel': 'BXMLE', 'eval_rate': '22.09', 'labor_mult': '0.5'},
        'box_wrapped': {'eval_sel': 'BXMXE', 'eval_rate': '29.87', 'cover_sel': 'CVCH', 'labor_mult': '1.0'},  # XL
        'plant_vase':  {'box_sel': 'BXLMP', 'box_rate': '8.91',  'eval_sel': 'BXMME', 'eval_rate': '17.36', 'labor_mult': '0.3'},
        'wardrobe':    {'box_sel': 'BXWDR>', 'box_rate': '27.89', 'eval_sel': 'BXMLE', 'eval_rate': '22.09', 'labor_mult': '0.5'},
        'dish_pack':   {'box_sel': 'BXDISH', 'box_rate': '9.98',  'eval_sel': 'BXDPE', 'eval_rate': '21.02', 'labor_mult': '0.3'},
        'glass_pack':  {'box_sel': 'BXGL',  'box_rate': '19.34', 'eval_sel': 'BXGLE', 'eval_rate': '28.11', 'labor_mult': '0.3'},
        'boots_pans':  {'box_sel': 'BX',    'box_rate': '3.91',  'eval_sel': 'BXPPE', 'eval_rate': '9.48',  'labor_mult': '0.3'},
    }

    CLEANING_MAP = {
        'small':       {'sel': 'BXMS',  'rate': '54.06'},
        'medium':      {'sel': 'BXMM',  'rate': '64.83'},
        'large':       {'sel': 'BXML',  'rate': '81.05'},
        'box_wrapped': {'sel': 'BXMX',  'rate': '107.82'},  # XL
        'plant_vase':  {'sel': 'FLF-',  'rate': '26.02',  'cat': 'CLM'},
        'wardrobe':    {'sel': 'BXML',  'rate': '81.05'},
        'dish_pack':   {'sel': 'BXDP-', 'rate': '33.95'},
        'glass_pack':  {'sel': 'BXGL-', 'rate': '28.76'},
        'boots_pans':  {'sel': 'BXPP',  'rate': '29.64'},
    }

    # Transport multipliers by box type (applied to room LF perimeter)
    TRANSPORT_MULT = {
        'small': '0.2*2', 'medium': '0.3*2', 'large': '1.0',
        'box_wrapped': '1.0', 'plant_vase': '0.3*2',
        'wardrobe': '0.5*2', 'dish_pack': '0.3*2',
        'glass_pack': '0.3*2', 'boots_pans': '0.3*2',
    }

    created = 0
    last_order = section.line_items.aggregate(m=Max('order'))['m'] or 0

    if section.section_type == SectionType.PACKING:
        for box_type, count in counts.items():
            if count == 0 or box_type not in PACKING_MAP:
                continue
            cfg = PACKING_MAP[box_type]
            last_order += 10
            # Evaluate line
            GCLineItem.objects.create(
                section=section,
                cat='CPS', sel=cfg['eval_sel'],
                description=f'Evaluate pack & inventory - {box_type.replace("_", " ")} box',
                quantity=count,
                unit='EA',
                replace_rate=cfg['eval_rate'],
                calc_formula='CPS',
                auto_calculated=True,
                order=last_order,
            )
            created += 1

    elif section.section_type == SectionType.CLEANING:
        for box_type, count in counts.items():
            if count == 0 or box_type not in CLEANING_MAP:
                continue
            cfg = CLEANING_MAP[box_type]
            last_order += 10
            GCLineItem.objects.create(
                section=section,
                cat=cfg.get('cat', 'CGN'), sel=cfg['sel'],
                description=f'Clean - {box_type.replace("_", " ")} box',
                quantity=count,
                unit='EA',
                replace_rate=cfg['rate'],
                calc_formula='CPS',
                auto_calculated=True,
                order=last_order,
            )
            created += 1

    elif section.section_type == SectionType.TRANSPORT:
        for box_type, count in counts.items():
            if count == 0:
                continue
            mult = TRANSPORT_MULT.get(box_type, '1.0')
            last_order += 10
            GCLineItem.objects.create(
                section=section,
                cat='CON', sel='LAB',
                description=f'Content Manipulation - {box_type.replace("_", " ")} box transport',
                quantity=count,
                unit='HR',
                replace_rate='54.90',
                calc_formula=f'LL*{mult}',
                auto_calculated=True,
                order=last_order,
            )
            created += 1

    return created
