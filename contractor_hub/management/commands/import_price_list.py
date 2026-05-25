"""
Management command: import_price_list

Imports an Xactimate price list export (CSV or Excel) into the RateItem table.

Xactimate export formats supported:
  - Xactimate x1 CSV export   (modern, comma-delimited)
  - Xactimate Classic CSV     (older, may be tab-delimited)
  - Excel (.xlsx) export
  - Any CSV with headers that map to CAT / SEL / description / unit / remove / replace

Usage:
    # Dry run — preview changes without saving
    python manage.py import_price_list ohcl8x_jun26.csv --dry-run

    # Full import
    python manage.py import_price_list ohcl8x_jun26.csv --price-list OHCL8X_JUN26

    # Import Excel file
    python manage.py import_price_list pricelist.xlsx --price-list OHCL8X_JUN26

    # Generate a sample template you can fill in
    python manage.py import_price_list --generate-sample

    # Show all imported price list versions
    python manage.py import_price_list --list-versions

    # Only update existing items, don't create new ones
    python manage.py import_price_list ohcl8x_jun26.csv --no-create

    # Only create new items, don't update existing rates
    python manage.py import_price_list ohcl8x_jun26.csv --no-update
"""

import csv
import io
import os
import sys
from datetime import date
from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand, CommandError
from django.utils import timezone

from contractor_hub.models import PriceListVersion, RateItem


# ---------------------------------------------------------------------------
# Column header aliases
# Xactimate has changed their export headers across versions.
# We map every known variation to our internal name.
# ---------------------------------------------------------------------------

COLUMN_ALIASES = {
    'cat': [
        'cat', 'category', 'cat.', 'category code', 'catg',
        'line item category', 'item category',
    ],
    'sel': [
        'sel', 'selector', 'sel.', 'item code', 'line item',
        'item selector', 'code', 'xactimate code',
    ],
    'description': [
        'activity', 'description', 'act', 'act.', 'line item description',
        'item description', 'desc', 'activity description', 'name',
    ],
    'unit': [
        'unit', 'unit of measure', 'uom', 'calc', 'measure',
        'unit type', 'calculation type',
    ],
    'remove_rate': [
        'remove', 'rem', 'rem.', 'remove rate', 'removal',
        'demo rate', 'labor remove', 'remove $', 'remove($)',
    ],
    'replace_rate': [
        'replace', 'rep', 'rep.', 'replace rate', 'replacement',
        'install rate', 'labor replace', 'replace $', 'replace($)',
        'unit price', 'price',
    ],
    'taxable': [
        'tax', 'taxable', 'tax?', 'tax flag', 'taxable (y/n)',
        'tax (y/n)', 'is taxable',
    ],
    'op_flag': [
        'o&p', 'op', 'o and p', 'overhead and profit', 'op flag',
        'o&p (y/n)',
    ],
}

# Values that mean "yes/taxable" in the tax column
TAX_TRUE_VALUES = {'y', 'yes', 'true', '1', 'x', 'taxable', 't'}

# Valid unit codes — map common variations to our standard codes
UNIT_MAP = {
    'ea':   'EA', 'each':  'EA', 'ea.':  'EA',
    'hr':   'HR', 'hour':  'HR', 'hrs':  'HR', 'hr.':   'HR',
    'lf':   'LF', 'lin ft':'LF', 'linear ft': 'LF', 'linear foot': 'LF',
    'sf':   'SF', 'sq ft': 'SF', 'sq. ft.': 'SF', 'square foot': 'SF',
    'cf':   'CF', 'cu ft': 'CF', 'cubic ft': 'CF', 'cubic foot': 'CF',
    'mo':   'MO', 'month': 'MO', 'mo.':  'MO',
    'ls':   'LS', 'lump sum': 'LS', 'lump': 'LS',
    'sy':   'SY', 'sq yd': 'SY',   # not in our choices but handle gracefully
}


class Command(BaseCommand):
    help = 'Import an Xactimate price list CSV/Excel into the RateItem table'

    def add_arguments(self, parser):
        parser.add_argument(
            'file', nargs='?', default=None,
            help='Path to the CSV or Excel price list file',
        )
        parser.add_argument(
            '--price-list', '-p', default=None,
            help='Price list code, e.g. OHCL8X_JUN26 (auto-detected from filename if omitted)',
        )
        parser.add_argument(
            '--market', '-m', default='',
            help='Market description, e.g. "Ohio - Cleveland"',
        )
        parser.add_argument(
            '--effective-date', '-d', default=None,
            help='Effective date YYYY-MM-DD (optional)',
        )
        parser.add_argument(
            '--dry-run', action='store_true',
            help='Preview all changes without saving anything',
        )
        parser.add_argument(
            '--no-create', action='store_true',
            help='Skip creating new line items — only update existing ones',
        )
        parser.add_argument(
            '--no-update', action='store_true',
            help='Skip updating existing rates — only add new line items',
        )
        parser.add_argument(
            '--generate-sample', action='store_true',
            help='Write a sample CSV template to xactimate_pricelist_sample.csv',
        )
        parser.add_argument(
            '--list-versions', action='store_true',
            help='Show all imported price list versions and exit',
        )
        parser.add_argument(
            '--show-changes', action='store_true',
            help='Print every rate that changed (old → new) during import',
        )

    # ── Entry point ──────────────────────────────────────────────────────────

    def handle(self, *args, **options):
        if options['list_versions']:
            return self._list_versions()

        if options['generate_sample']:
            return self._generate_sample()

        file_path = options['file']
        if not file_path:
            raise CommandError(
                'Provide a file path, or use --generate-sample / --list-versions'
            )
        if not os.path.exists(file_path):
            raise CommandError(f'File not found: {file_path}')

        # Auto-detect price list code from filename if not provided
        price_list_code = options['price_list']
        if not price_list_code:
            basename = os.path.splitext(os.path.basename(file_path))[0].upper()
            price_list_code = basename.replace(' ', '_').replace('-', '_')
            self.stdout.write(
                self.style.WARNING(
                    f'--price-list not set, using filename: {price_list_code}'
                )
            )

        dry_run = options['dry_run']
        if dry_run:
            self.stdout.write(self.style.WARNING('DRY RUN — no changes will be saved'))

        # Parse the file
        rows = self._parse_file(file_path)
        if not rows:
            raise CommandError('No data rows found in file.')

        self.stdout.write(f'Parsed {len(rows)} rows from {os.path.basename(file_path)}')

        # Run import
        self._run_import(
            rows=rows,
            price_list_code=price_list_code,
            market=options['market'],
            effective_date=options['effective_date'],
            source_file=os.path.basename(file_path),
            dry_run=dry_run,
            allow_create=not options['no_create'],
            allow_update=not options['no_update'],
            show_changes=options['show_changes'],
        )

    # ── File parsers ─────────────────────────────────────────────────────────

    def _parse_file(self, file_path):
        """Detect file type and parse into list of dicts."""
        ext = os.path.splitext(file_path)[1].lower()

        if ext in ('.xlsx', '.xls'):
            return self._parse_excel(file_path)
        else:
            return self._parse_csv(file_path)

    def _parse_csv(self, file_path):
        """
        Parse CSV — handles:
        - UTF-8 with or without BOM
        - UTF-16 (some Xactimate exports)
        - Comma or tab delimited
        - Quoted fields
        """
        # Try multiple encodings
        for encoding in ('utf-8-sig', 'utf-8', 'utf-16', 'latin-1'):
            try:
                with open(file_path, 'r', encoding=encoding, newline='') as f:
                    sample = f.read(4096)
                    f.seek(0)

                    # Detect delimiter
                    delimiter = '\t' if sample.count('\t') > sample.count(',') else ','

                    reader = csv.DictReader(f, delimiter=delimiter)
                    rows = list(reader)
                    if rows:
                        self.stdout.write(f'  Encoding: {encoding}, Delimiter: {"TAB" if delimiter == chr(9) else "COMMA"}')
                        return self._normalize_rows(rows)
            except (UnicodeDecodeError, UnicodeError):
                continue

        raise CommandError(f'Could not decode {file_path} — try saving as UTF-8 CSV from Excel')

    def _parse_excel(self, file_path):
        """Parse .xlsx file using openpyxl."""
        try:
            import openpyxl
        except ImportError:
            raise CommandError('openpyxl is required for Excel files: pip install openpyxl')

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        # Find the sheet with the most rows (usually the price list sheet)
        target_sheet = max(wb.worksheets, key=lambda ws: ws.max_row)
        self.stdout.write(f'  Using sheet: "{target_sheet.title}"')

        rows_raw = list(target_sheet.iter_rows(values_only=True))
        if not rows_raw:
            return []

        # First non-empty row is the header
        headers = None
        data_start = 0
        for i, row in enumerate(rows_raw):
            non_empty = [c for c in row if c is not None and str(c).strip()]
            if len(non_empty) >= 3:
                headers = [str(c).strip() if c is not None else '' for c in row]
                data_start = i + 1
                break

        if not headers:
            raise CommandError('Could not find header row in Excel file')

        dict_rows = []
        for row in rows_raw[data_start:]:
            if all(c is None or str(c).strip() == '' for c in row):
                continue  # Skip blank rows
            d = {headers[i]: (str(row[i]).strip() if row[i] is not None else '')
                 for i in range(min(len(headers), len(row)))}
            dict_rows.append(d)

        return self._normalize_rows(dict_rows)

    def _normalize_rows(self, dict_rows):
        """
        Map raw column headers to our internal field names.
        Returns list of normalized dicts with keys:
            cat, sel, description, unit, remove_rate, replace_rate, taxable
        """
        if not dict_rows:
            return []

        # Build header → field mapping from first row's keys
        raw_headers = list(dict_rows[0].keys())
        col_map = {}  # field_name → raw_header

        for field, aliases in COLUMN_ALIASES.items():
            for raw in raw_headers:
                if raw.lower().strip() in aliases:
                    col_map[field] = raw
                    break

        # Check required columns
        required = ['cat', 'sel', 'description']
        missing = [f for f in required if f not in col_map]
        if missing:
            self.stdout.write(
                self.style.ERROR(
                    f'Could not find required columns: {missing}\n'
                    f'Headers found: {raw_headers}\n'
                    f'Expected one of: {[COLUMN_ALIASES[f] for f in missing]}'
                )
            )
            raise CommandError(
                'Column mapping failed. Run --generate-sample to see expected format.'
            )

        self.stdout.write('  Column mapping:')
        for field, raw in col_map.items():
            self.stdout.write(f'    {field:20s} ← "{raw}"')

        normalized = []
        for row in dict_rows:
            def get(field, default=''):
                key = col_map.get(field)
                return row.get(key, default).strip() if key else default

            cat = get('cat').upper()
            sel = get('sel').upper()

            # Skip blank / header-repeat rows
            if not cat or not sel or cat == 'CAT':
                continue

            normalized.append({
                'cat':          cat,
                'sel':          sel,
                'description':  get('description'),
                'unit':         self._normalize_unit(get('unit', 'EA')),
                'remove_rate':  self._parse_decimal(get('remove_rate', '0')),
                'replace_rate': self._parse_decimal(get('replace_rate', '0')),
                'taxable':      get('taxable', 'Y').lower() in TAX_TRUE_VALUES,
            })

        return normalized

    # ── Core import logic ────────────────────────────────────────────────────

    def _run_import(self, rows, price_list_code, market, effective_date,
                    source_file, dry_run, allow_create, allow_update, show_changes):

        created = updated = skipped = unchanged = errors = 0
        changes = []  # list of (cat, sel, old_remove, new_remove, old_replace, new_replace)

        # Parse effective_date
        eff_date = None
        if effective_date:
            try:
                eff_date = date.fromisoformat(effective_date)
            except ValueError:
                self.stdout.write(
                    self.style.WARNING(f'Invalid date "{effective_date}" — ignored')
                )

        now = timezone.now()

        for row in rows:
            cat   = row['cat']
            sel   = row['sel']
            desc  = row['description']
            unit  = row['unit']
            rem   = row['remove_rate']
            rep   = row['replace_rate']
            taxbl = row['taxable']

            # Skip rows with zero rates and no description
            if rem == Decimal('0') and rep == Decimal('0') and not desc:
                skipped += 1
                continue

            try:
                existing = RateItem.objects.filter(cat=cat, sel=sel).first()

                if existing:
                    # Check if anything actually changed
                    rate_changed = (
                        existing.remove_rate  != rem or
                        existing.replace_rate != rep
                    )
                    desc_changed = existing.description != desc and desc

                    if not rate_changed and not desc_changed:
                        unchanged += 1
                        continue

                    if not allow_update:
                        skipped += 1
                        continue

                    if show_changes and rate_changed:
                        changes.append((
                            cat, sel,
                            float(existing.remove_rate), float(rem),
                            float(existing.replace_rate), float(rep),
                        ))

                    if not dry_run:
                        # Snapshot previous rates before overwriting
                        if rate_changed:
                            existing.previous_remove_rate  = existing.remove_rate
                            existing.previous_replace_rate = existing.replace_rate
                        if desc_changed:
                            existing.description = desc
                        existing.remove_rate   = rem
                        existing.replace_rate  = rep
                        existing.taxable       = taxbl
                        existing.unit          = unit or existing.unit
                        existing.last_updated_at = now
                        existing.save(update_fields=[
                            'remove_rate', 'replace_rate', 'taxable', 'unit',
                            'description', 'previous_remove_rate',
                            'previous_replace_rate', 'last_updated_at',
                        ])
                    updated += 1

                else:
                    if not allow_create:
                        skipped += 1
                        continue

                    if not dry_run:
                        RateItem.objects.create(
                            cat=cat, sel=sel,
                            description=desc,
                            unit=unit,
                            remove_rate=rem,
                            replace_rate=rep,
                            taxable=taxbl,
                            last_updated_at=now,
                        )
                    created += 1

            except Exception as e:
                self.stdout.write(
                    self.style.ERROR(f'  ERROR on {cat} {sel}: {e}')
                )
                errors += 1

        total = created + updated + skipped + unchanged

        # Print change log
        if show_changes and changes:
            self.stdout.write('\n' + self.style.MIGRATE_HEADING('Rate Changes:'))
            self.stdout.write(
                f'  {"CAT":<6} {"SEL":<12} {"OLD REM":>9} {"NEW REM":>9}  '
                f'{"OLD REP":>9} {"NEW REP":>9}  ΔREPLACE'
            )
            self.stdout.write('  ' + '─' * 75)
            for cat, sel, old_rem, new_rem, old_rep, new_rep in changes:
                delta = new_rep - old_rep
                sign  = '+' if delta >= 0 else ''
                self.stdout.write(
                    f'  {cat:<6} {sel:<12} {old_rem:>9.2f} {new_rem:>9.2f}  '
                    f'{old_rep:>9.2f} {new_rep:>9.2f}  {sign}{delta:.2f}'
                )

        # Summary
        self.stdout.write('')
        self.stdout.write(self.style.MIGRATE_HEADING(
            f'{"[DRY RUN] " if dry_run else ""}Import Summary: {price_list_code}'
        ))
        self.stdout.write(f'  Total rows processed : {total}')
        self.stdout.write(self.style.SUCCESS(f'  Created              : {created}'))
        if updated:
            self.stdout.write(self.style.WARNING(f'  Updated (rate change): {updated}'))
        self.stdout.write(f'  Unchanged            : {unchanged}')
        if skipped:
            self.stdout.write(f'  Skipped              : {skipped}')
        if errors:
            self.stdout.write(self.style.ERROR(f'  Errors               : {errors}'))

        # Save version record
        if not dry_run:
            # Set price_list_version FK on all updated/created items
            version, _ = PriceListVersion.objects.update_or_create(
                code=price_list_code,
                defaults={
                    'market':         market,
                    'effective_date': eff_date,
                    'source_file':    source_file,
                    'total_items':    total,
                    'items_created':  created,
                    'items_updated':  updated,
                    'items_skipped':  skipped,
                    'notes':          f'{errors} errors during import' if errors else '',
                },
            )
            # Stamp all recently-updated items with this version
            RateItem.objects.filter(last_updated_at=now).update(
                price_list_version=version
            )
            self.stdout.write(
                self.style.SUCCESS(f'\nPrice list "{price_list_code}" saved. ID={version.pk}')
            )
        else:
            self.stdout.write(
                self.style.WARNING('\nDRY RUN complete — run without --dry-run to apply changes.')
            )

    # ── Utilities ─────────────────────────────────────────────────────────────

    def _normalize_unit(self, raw):
        """Map raw unit string to our 2-letter standard code."""
        if not raw:
            return 'EA'
        key = raw.strip().lower()
        return UNIT_MAP.get(key, raw.upper()[:5])

    def _parse_decimal(self, raw):
        """Parse a rate string like '$26.00', '26.00', '26', '' → Decimal."""
        if not raw:
            return Decimal('0.00')
        cleaned = raw.replace('$', '').replace(',', '').strip()
        if not cleaned or cleaned == '-':
            return Decimal('0.00')
        try:
            return Decimal(cleaned).quantize(Decimal('0.01'))
        except InvalidOperation:
            return Decimal('0.00')

    # ── Utility commands ──────────────────────────────────────────────────────

    def _list_versions(self):
        versions = PriceListVersion.objects.order_by('-imported_at')
        if not versions.exists():
            self.stdout.write('No price list versions imported yet.')
            return

        self.stdout.write(self.style.MIGRATE_HEADING(
            f'\n{"Code":<20} {"Market":<25} {"Effective":<12} '
            f'{"Items":>6} {"Created":>8} {"Updated":>8} {"Imported At"}'
        ))
        self.stdout.write('─' * 95)
        for v in versions:
            self.stdout.write(
                f'{v.code:<20} {v.market:<25} '
                f'{str(v.effective_date or ""):<12} '
                f'{v.total_items:>6} {v.items_created:>8} {v.items_updated:>8}  '
                f'{v.imported_at.strftime("%Y-%m-%d %H:%M")}'
            )

    def _generate_sample(self):
        """Write a sample CSV template showing the expected format."""
        out_path = 'xactimate_pricelist_sample.csv'
        sample_rows = [
            # Header
            ['Cat', 'Sel', 'Activity', 'Unit', 'Remove', 'Replace', 'Tax'],
            # Sample data — using known rates from OHCL8X_MAR26
            ['WTR', 'DRY',     'Air mover (per 24 hour period) - No monitoring',           'EA', '0.00',   '26.00',  'Y'],
            ['WTR', 'DUCTLF',  'Ducting - lay-flat',                                       'LF', '0.00',   '0.38',   'Y'],
            ['WTR', 'EQ',      'Equipment setup, take down, and monitoring (hourly charge)','HR', '0.00',   '69.50',  'Y'],
            ['WTR', 'EQD',     'Equipment decontamination charge - per piece of equipment', 'EA', '0.00',   '41.28',  'Y'],
            ['WTR', 'FHEPA',   'Add for HEPA filter (for negative air exhaust fan)',        'EA', '0.00',   '219.79', 'Y'],
            ['HMR', 'LABH',    'Hazardous Waste/Mold Cleaning Technician - per hour',       'HR', '0.00',   '73.49',  'Y'],
            ['HMR', 'HBAG',    'Plastic bag - used for hazardous waste cleanup - Medium',   'EA', '0.00',   '1.93',   'Y'],
            ['HMR', 'HEPAVAC', 'HEPA Vacuuming - hourly charge',                           'HR', '0.00',   '76.78',  'Y'],
            ['HMR', 'PPE+',    'Add for personal protective equipment - Heavy duty',        'EA', '0.00',   '50.05',  'Y'],
            ['HMR', 'PROTX',   'Protect - Cover with plastic',                             'SF', '0.00',   '0.31',   'Y'],
            ['HMR', 'CLNJST',  'Clean floor or roof joist system',                         'SF', '0.00',   '1.69',   'Y'],
            ['HMR', 'HEPAVAS', 'HEPA Vacuuming - Detailed - (PER SF)',                     'SF', '0.00',   '0.93',   'Y'],
            ['CON', 'LAB',     'Content Manipulation charge - per hour',                   'HR', '0.00',   '54.90',  'Y'],
            ['CGN', 'BXMS',    'Clean misc items - per Sml box',                           'EA', '0.00',   '54.06',  'Y'],
            ['CGN', 'BXMM',    'Clean misc items - per Med box',                           'EA', '0.00',   '64.83',  'Y'],
            ['CGN', 'BXML',    'Clean misc items - per Lg box',                            'EA', '0.00',   '81.05',  'Y'],
            ['CGN', 'BXMX',    'Clean misc items - per Xlg box',                           'EA', '0.00',   '107.82', 'Y'],
            ['CGN', 'BXDP-',   'Clean dishpack - per box - Light clean',                  'EA', '0.00',   '33.95',  'Y'],
            ['CGN', 'BXGL-',   'Clean glasspack - per box - Light clean',                 'EA', '0.00',   '28.76',  'Y'],
            ['CGN', 'BXPP',    'Clean pots and pans - per medium box',                    'EA', '0.00',   '29.64',  'Y'],
            ['CGN', 'DODROZ',  'Deodorization chamber - Ozone treatment',                 'CF', '0.00',   '0.11',   'Y'],
            ['CLM', 'FLF-',    'Clean floor lamp - flat finish (no shade) - Light clean', 'EA', '0.00',   '26.02',  'Y'],
            ['CPS', 'BXMSE',   'Evaluate pack & inventory misc items - per Sml box',      'EA', '0.00',   '14.15',  'Y'],
            ['CPS', 'BXMME',   'Evaluate pack & inventory misc items - per Med box',      'EA', '0.00',   '17.36',  'Y'],
            ['CPS', 'BXMLE',   'Evaluate pack & inventory misc items - per Lg box',       'EA', '0.00',   '22.09',  'Y'],
            ['CPS', 'BXMXE',   'Evaluate pack & inventory misc items - per Xlg box',      'EA', '0.00',   '29.87',  'Y'],
            ['CPS', 'BXMIR',   'Provide mirror/picture box, packing paper & tape',        'EA', '0.00',   '10.29',  'Y'],
            ['CPS', 'BXLMP',   'Provide lamp box set, packing paper & tape',              'EA', '0.00',   '8.91',   'Y'],
            ['CPS', 'BXWDR>',  'Provide wardrobe box & tape - large size',                'EA', '0.00',   '27.89',  'Y'],
            ['CPS', 'BXDISH',  'Provide dishpack box, packing paper & tape',              'EA', '0.00',   '9.98',   'Y'],
            ['CPS', 'BXDPE',   'Evaluate pack & inventory dishpack - per box',            'EA', '0.00',   '21.02',  'Y'],
            ['CPS', 'BXGL',    'Provide glasspack box, packing paper & tape',             'EA', '0.00',   '19.34',  'Y'],
            ['CPS', 'BXGLE',   'Evaluate pack & inventory glasspack - per box',           'EA', '0.00',   '28.11',  'Y'],
            ['CPS', 'BX',      'Provide box, packing paper & tape - medium size',         'EA', '0.00',   '3.91',   'Y'],
            ['CPS', 'BX>',     'Provide box, packing paper & tape - large size',          'EA', '0.00',   '5.28',   'Y'],
            ['CPS', 'BXPPE',   'Evaluate pack & inventory pots and pans - per Med box',   'EA', '0.00',   '9.48',   'Y'],
            ['CPS', 'CVCH',    'Provide plastic chair cover & packing tape',              'EA', '0.00',   '5.31',   'Y'],
            ['CPS', 'CVCH>',   'Provide plastic couch/sofa cover & packing tape',         'EA', '0.00',   '8.57',   'Y'],
            ['CPS', 'PAD+',    'Provide furniture heavyweight blanket/pad',               'EA', '0.00',   '18.26',  'Y'],
            ['CPS', 'LABS',    'Contents Evaluation and/or Supervisor/Admin - per hour',  'HR', '0.00',   '72.28',  'Y'],
            ['CPS', 'LAB',     'Inventory, Packing, Boxing, and Moving charge - per hour','HR', '0.00',   '56.29',  'Y'],
            ['CLN', 'AV',      'Clean the floor',                                         'SF', '0.00',   '0.50',   'Y'],
            ['DMO', 'PU',      'Haul debris - per pickup truck load - including dump fees','EA', '199.44', '0.00',   'Y'],
            ['DMO', 'LAB',     'General Demolition - per hour',                           'HR', '62.58',  '0.00',   'Y'],
        ]

        with open(out_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerows(sample_rows)

        self.stdout.write(self.style.SUCCESS(
            f'Sample written to: {os.path.abspath(out_path)}\n\n'
            'This file uses the same format as an Xactimate price list export.\n'
            'To import it:\n'
            f'  python manage.py import_price_list {out_path} '
            f'--price-list OHCL8X_MAR26 --dry-run'
        ))
