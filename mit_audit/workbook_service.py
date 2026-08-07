"""
mit_audit/workbook_service.py

Three responsibilities:
  1. find_and_copy_client_workbook(audit)  — locate the client's existing 82-MIT
                                             workbook, copy it to a per-audit
                                             working directory, return the path.
  2. write_dimensions()       — fill room L/W/H into the jobinfo(2) sheet
  3. recalc_and_read_equipment() — trigger LibreOffice UNO recalc, then read
                                    the TOTAL-EQPT tab

Actual workbook structure (from 82-MIT-3DAY.xlsm):
  ┌─────────────────┬────────────────────────────────────────┐
  │ jobinfo(2) sheet │                                        │
  │  Row 52          │ Header: col B=ROOM ID, C=name, E=L,   │
  │                  │          F=W, G=H                      │
  │  Rows 53–102     │ Room data (C=name, E=L, F=W, G=H)     │
  ├─────────────────┴────────────────────────────────────────┤
  │ MIT-EQPT sheet   │ Pulls from jobinfo(2)!C53:G77          │
  │                  │ Calculates per-room equipment via IICRC│
  ├─────────────────┴────────────────────────────────────────┤
  │ TOTAL-EQPT sheet │ Aggregates MIT-EQPT; col C = qty       │
  │  Row 5  DRY      │ Air Movers                             │
  │  Row 10 DHM      │ Dehumidifiers                          │
  │  Row 14 AFD      │ Air Filtration Device                  │
  │  Row 18 BARRZ    │ Zippers / Containment                  │
  │  Row 20 BARRP    │ Tension Poles                          │
  │  Row 23 CCDU     │ Ceiling Cavity Drying Unit             │
  │  Row 26 WCDU     │ Wall Cavity Drying Unit                │
  └──────────────────┴───────────────────────────────────────┘

Why LibreOffice for recalc?
  openpyxl cannot evaluate Excel formulas.  The TOTAL-EQPT totals are
  formula-driven (MIT-EQPT references back to jobinfo(2)).  LibreOffice UNO
  opens the workbook, recalculates every formula in memory, then saves —
  leaving a file openpyxl can re-open with data_only=True to read the
  computed values.

Fallback chain (mirrors docsAppR/tasks.py):
  1. UNO listener on port 2002         (preferred — zero corruption)
  2. LibreOffice subprocess            (slower but also zero corruption)
  3. raw openpyxl data_only read       (formulas show as None — only works
                                        if the workbook already has cached values)
"""
import logging
import os
import shutil
import subprocess
import tempfile
from pathlib import Path

from django.conf import settings

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Dimension cell map — reflects the ACTUAL 82-MIT-3DAY.xlsm layout.
# Update via Admin → MIT Day 3 Config → dimension_cell_map if the template
# ever changes.
# ---------------------------------------------------------------------------
DEFAULT_DIMENSION_MAP = {
    'job_info_sheet':  'jobinfo(2)',
    'room_start_row':  53,       # row 53 = Room/Area 1 (row 52 is the header)
    'room_name_col':   'C',      # column C: room label  (e.g. "Kitchen")
    'length_col':      'E',      # column E: length (ft)
    'width_col':       'F',      # column F: width (ft)
    'height_col':      'G',      # column G: height (ft)
    'max_rows':        50,       # rows 53–102 → 50 rooms maximum
}

# Known row positions in TOTAL-EQPT (qty in col C).
# Used when no explicit equipment_cell_map is configured.
TOTAL_EQPT_ROWS = [
    {'row': 5,  'name': 'Air Movers (DRY)',              'xact_code': 'DRY'},
    {'row': 10, 'name': 'Dehumidifiers (DHM)',           'xact_code': 'DHM'},
    {'row': 14, 'name': 'AFD Air Filtration Device',     'xact_code': 'AFD'},
    {'row': 18, 'name': 'Zippers / Containment (BARRZ)', 'xact_code': 'BARRZ'},
    {'row': 20, 'name': 'Tension Poles (BARRP)',         'xact_code': 'BARRP'},
    {'row': 23, 'name': 'CCDU Ceiling Cavity Drying',   'xact_code': 'CCDU'},
    {'row': 26, 'name': 'WCDU Wall Cavity Drying',      'xact_code': 'WCDU'},
]

# Equipment type keywords for auto-categorisation when reading TOTAL-EQPT.
_CATEGORY_KEYWORDS = {
    'dehumidifier':  ['dehumid', 'dhm', 'lgr', 'desiccant'],
    'air_cleaner':   ['air clean', 'afd', 'scrubber', 'hepa',
                      'filtration', 'negative air', 'nafan', 'hydroxyl'],
    'zipper_wall':   ['zipper', 'barrz', 'containment', 'poly barrier'],
    'tension_poles': ['tension pole', 'barrp'],
    'double_zipper': ['double zipper', 'dbl zipper'],
    'blower':        ['blower', 'air mover', 'dry', 'axial',
                      'centrifugal', 'fan'],
    'wall_cavity':   ['wcdu', 'wall cavity', 'injectidry', 'wall dry'],
    'ceiling_cavity':['ccdu', 'ceiling cavity'],
    'floor_drying':  ['floor mat', 'floor dry', 'drying mat', 'extraction mat'],
    'hydroxyl':      ['hydroxyl', 'dodhy', 'odor counteract'],
    'heater':        ['heater', 'heat'],
}

# Items in these categories require a dedicated "stabilization" photo showing
# the equipment connected and actively running.
_STABILIZATION_TYPES = {
    'dehumidifier', 'air_cleaner', 'zipper_wall', 'double_zipper',
    'ceiling_cavity', 'wall_cavity', 'hydroxyl',
}


def _categorise(name: str) -> tuple[str, bool]:
    """Return (category, requires_stabilization_photo) for a display name."""
    lower = name.lower()
    for cat, keywords in _CATEGORY_KEYWORDS.items():
        if any(kw in lower for kw in keywords):
            return cat, cat in _STABILIZATION_TYPES
    return 'other', False


# ---------------------------------------------------------------------------
# Step 1-helper: Find and copy the client's existing MIT workbook
# ---------------------------------------------------------------------------

def find_and_copy_client_workbook(audit) -> str:
    """
    Locate the client's existing 82-MIT workbook (generated by the main
    document-generation pipeline and stored in the client's server folder),
    copy it to a per-audit working directory, and return the copy's absolute path.

    Lookup order:
      1. ClaimFile records with file_type='82-MIT' for this client (most recent first)
      2. Filesystem glob of the client's Templates folder for 82-MIT*.xlsm / *.xlsx

    Raises FileNotFoundError if no workbook can be found.
    """
    from docsAppR.models import ClaimFile

    client = audit.client

    # --- Try ClaimFile DB records first ---
    claim_file = (
        ClaimFile.objects
        .filter(client=client, file_type='82-MIT')
        .order_by('-created_at')
        .first()
    )
    if claim_file:
        full_path = claim_file.get_full_path()
        if os.path.exists(full_path):
            logger.info('[MIT] Found 82-MIT workbook via ClaimFile: %s', full_path)
            return _copy_to_audit_dir(full_path, audit.pk)
        else:
            logger.warning('[MIT] ClaimFile record points to missing file: %s', full_path)

    # --- Fallback: scan the client's Templates folder ---
    try:
        templates_folder = client.get_templates_folder()
    except Exception:
        templates_folder = None

    if templates_folder and os.path.isdir(templates_folder):
        import glob as glob_mod
        patterns = [
            os.path.join(templates_folder, '82-MIT*.xlsm'),
            os.path.join(templates_folder, '82-MIT*.xlsx'),
            os.path.join(templates_folder, '*MIT*3*DAY*.xlsm'),
            os.path.join(templates_folder, '*MIT*3*DAY*.xlsx'),
        ]
        for pattern in patterns:
            matches = sorted(glob_mod.glob(pattern))
            if matches:
                src = matches[-1]  # most recently created alphabetically
                logger.info('[MIT] Found 82-MIT workbook via filesystem scan: %s', src)
                return _copy_to_audit_dir(src, audit.pk)

    raise FileNotFoundError(
        f'No 82-MIT workbook found for client "{client.pOwner}" '
        f'(id={client.pk}). Generate the MIT documents first via the '
        f'main document pipeline, then re-run the audit.'
    )


def _copy_to_audit_dir(src_path: str, audit_id: int) -> str:
    """Copy src_path → MEDIA_ROOT/mit_audits/{audit_id}/ and return the new path."""
    src = Path(src_path)
    dest_dir = Path(settings.MEDIA_ROOT) / 'mit_audits' / str(audit_id)
    dest_dir.mkdir(parents=True, exist_ok=True)
    dest = dest_dir / f'MIT_Day3_Job_{audit_id}{src.suffix}'
    shutil.copy2(src, dest)
    logger.info('[MIT] Copied client workbook → %s', dest)
    return str(dest)


# ---------------------------------------------------------------------------
# Step 2: Write room dimensions into jobinfo(2) sheet
# ---------------------------------------------------------------------------

def write_dimensions(workbook_path: str, room_dimensions) -> list[dict]:
    """
    Write approved room dimensions into the jobinfo(2) sheet.

    The sheet has pre-labelled rows " Room/Area 1 " … " Room/Area N ".
    We match Encircle room names to existing rows case-insensitively;
    if no match is found we write to the next free name cell.

    Args:
        workbook_path:   Absolute path to the job workbook copy.
        room_dimensions: QuerySet or list of MITRoomDimension instances.

    Returns:
        List of dicts: [{ 'room_name', 'row', 'written': True/False }, ...]
    """
    import openpyxl
    from mit_audit.models import MITDay3Config
    cfg_obj  = MITDay3Config.get()
    cell_map = cfg_obj.dimension_cell_map or DEFAULT_DIMENSION_MAP

    sheet_name = cell_map.get('job_info_sheet', DEFAULT_DIMENSION_MAP['job_info_sheet'])
    start_row  = int(cell_map.get('room_start_row', DEFAULT_DIMENSION_MAP['room_start_row']))
    name_col   = cell_map.get('room_name_col',  DEFAULT_DIMENSION_MAP['room_name_col'])
    length_col = cell_map.get('length_col',     DEFAULT_DIMENSION_MAP['length_col'])
    width_col  = cell_map.get('width_col',      DEFAULT_DIMENSION_MAP['width_col'])
    height_col = cell_map.get('height_col',     DEFAULT_DIMENSION_MAP['height_col'])
    max_rows   = int(cell_map.get('max_rows',   DEFAULT_DIMENSION_MAP['max_rows']))

    approved = [d for d in room_dimensions if d.approved]
    if not approved:
        logger.warning('[MIT] No approved dimensions to write for %s', workbook_path)
        return []

    # .xlsm files must be opened with keep_vba=True or macros are stripped
    is_xlsm = workbook_path.lower().endswith('.xlsm')
    wb = openpyxl.load_workbook(workbook_path, keep_vba=is_xlsm)

    if sheet_name not in wb.sheetnames:
        ws = wb.active
        logger.warning('[MIT] Sheet "%s" not found — writing to active sheet "%s"',
                       sheet_name, ws.title)
    else:
        ws = wb[sheet_name]

    # Build map: normalised room label already in the sheet → row number
    # (The sheet has " Room/Area 1 " in col B and the room name in col C.)
    existing_name_rows: dict[str, int] = {}
    first_free_row: int | None = None
    for r in range(start_row, start_row + max_rows):
        cell_val = ws[f'{name_col}{r}'].value
        if cell_val:
            existing_name_rows[str(cell_val).strip().lower()] = r
        elif first_free_row is None:
            first_free_row = r

    if first_free_row is None:
        first_free_row = start_row + max_rows  # overflow safety

    results = []
    written_rows: set[int] = set()

    for dim in approved:
        name_key   = dim.room_name.strip().lower()
        target_row = existing_name_rows.get(name_key)

        if target_row is None:
            # No matching room label — write the name into the next free row
            target_row = first_free_row + len(written_rows)
            ws[f'{name_col}{target_row}'] = dim.room_name.strip()
            logger.debug('[MIT] New room "%s" → row %d', dim.room_name, target_row)

        ws[f'{length_col}{target_row}'] = float(dim.length) if dim.length else None
        ws[f'{width_col}{target_row}']  = float(dim.width)  if dim.width  else None
        ws[f'{height_col}{target_row}'] = float(dim.height) if dim.height else None

        dim.workbook_row = target_row
        dim.save(update_fields=['workbook_row'])
        written_rows.add(target_row)
        results.append({'room_name': dim.room_name, 'row': target_row, 'written': True})
        logger.debug('[MIT] Wrote %s → row %d (L=%s W=%s H=%s)',
                     dim.room_name, target_row, dim.length, dim.width, dim.height)

    wb.save(workbook_path)
    logger.info('[MIT] Saved workbook with %d room dimensions → %s',
                len(results), workbook_path)
    return results


# ---------------------------------------------------------------------------
# Step 3a: Recalculate via LibreOffice UNO (preferred)
# ---------------------------------------------------------------------------

def recalculate_via_uno(workbook_path: str) -> bool:
    """
    Open the workbook in the persistent LibreOffice UNO listener (port 2002),
    trigger a full recalculation, then save back to the same path.
    Returns True on success, False if UNO is unavailable (caller should fallback).
    """
    try:
        from docsAppR.lo_uno_service import _import_uno, _connect, _make_prop
        uno, PropertyValue, NoConnectException = _import_uno()
        desktop = _connect(uno, NoConnectException)

        file_url = uno.systemPathToFileUrl(workbook_path)
        props = [
            _make_prop('Hidden', True),
            _make_prop('MacroExecutionMode', 4),  # ALWAYS_EXECUTE_NO_WARN
        ]
        doc = desktop.loadComponentFromURL(file_url, '_blank', 0, tuple(props))

        # Force recalculation of all formula cells
        sheet_enum = doc.Sheets.createEnumeration()
        while sheet_enum.hasMoreElements():
            sheet = sheet_enum.nextElement()
            cell_enum = sheet.createEnumeration()
            while cell_enum.hasMoreElements():
                cell_range = cell_enum.nextElement()
                cell_inner = cell_range.createEnumeration()
                while cell_inner.hasMoreElements():
                    cell = cell_inner.nextElement()
                    try:
                        if hasattr(cell, 'FormulaResultType') and cell.FormulaResultType is not None:
                            pass  # accessing the attribute forces evaluation
                    except Exception:
                        pass

        doc.store()
        doc.close(True)
        logger.info('[MIT] LibreOffice UNO recalc + save complete: %s', workbook_path)
        return True
    except ImportError:
        logger.info('[MIT] UNO not importable — skipping UNO recalc')
        return False
    except Exception as exc:
        logger.warning('[MIT] UNO recalc failed (%s), will try subprocess fallback', exc)
        return False


# ---------------------------------------------------------------------------
# Step 3b: Recalculate via LibreOffice subprocess (fallback)
# ---------------------------------------------------------------------------

def recalculate_via_subprocess(workbook_path: str) -> bool:
    """
    Fallback: call LibreOffice headless as a subprocess to convert and save.
    Converts xlsx/xlsm → ods → xlsx (round-trip forces formula evaluation).
    Slower than UNO but does not require python3-uno.
    """
    lo_candidates = [
        '/usr/bin/libreoffice',
        '/usr/bin/soffice',
        '/usr/local/bin/libreoffice',
        '/opt/libreoffice/program/soffice',
    ]
    lo_bin = next((c for c in lo_candidates if os.path.exists(c)), None)
    if not lo_bin:
        logger.warning('[MIT] LibreOffice binary not found — skipping subprocess recalc')
        return False

    src = Path(workbook_path)
    with tempfile.TemporaryDirectory() as tmp:
        try:
            # Convert to ods (forces recalc) then back to xlsx
            subprocess.run(
                [lo_bin, '--headless', '--convert-to', 'ods',
                 '--outdir', tmp, str(src)],
                timeout=120, check=True, capture_output=True
            )
            ods_path = Path(tmp) / (src.stem + '.ods')
            subprocess.run(
                [lo_bin, '--headless', '--convert-to', 'xlsx',
                 '--outdir', tmp, str(ods_path)],
                timeout=120, check=True, capture_output=True
            )
            out_xlsx = Path(tmp) / (src.stem + '.xlsx')
            if out_xlsx.exists():
                shutil.copy2(out_xlsx, workbook_path)
                logger.info('[MIT] Subprocess recalc complete: %s', workbook_path)
                return True
        except subprocess.TimeoutExpired:
            logger.error('[MIT] LibreOffice subprocess timed out')
        except subprocess.CalledProcessError as exc:
            logger.error('[MIT] LibreOffice subprocess error: %s', exc)
    return False


# ---------------------------------------------------------------------------
# Step 4: Read TOTAL-EQPT tab → equipment list
# ---------------------------------------------------------------------------

def read_total_equipment(workbook_path: str) -> list[dict]:
    """
    Read the TOTAL-EQPT tab and return one dict per item where qty > 0.

    Two modes:
      • Explicit cell map (from MITDay3Config.equipment_cell_map) — most reliable
      • Known-rows mode  (TOTAL_EQPT_ROWS constant above) — uses the standard
        82-MIT template layout; reads col B for name, col C for qty at each row.

    Returns:
        [
          {
            'display_name':               'Air Movers (DRY)',
            'equipment_type':             'blower',
            'category':                   'blower',
            'required_quantity':          6,
            'source_sheet':               'TOTAL-EQPT',
            'workbook_row':               5,
            'workbook_cell':              'C5',
            'requires_stabilization_photo': False,
          },
          ...
        ]

    Falls back gracefully: if formula cells read as None (LibreOffice recalc
    did not run), those rows are skipped.  The caller must ensure recalc ran.
    """
    import openpyxl
    from mit_audit.models import MITDay3Config

    cfg_obj    = MITDay3Config.get()
    sheet_name = cfg_obj.total_equipment_sheet or 'TOTAL-EQPT'
    cell_map   = cfg_obj.equipment_cell_map  # list of dicts, may be empty/None

    is_xlsm = workbook_path.lower().endswith('.xlsm')
    wb = openpyxl.load_workbook(workbook_path, data_only=True, keep_vba=is_xlsm)

    if sheet_name not in wb.sheetnames:
        logger.warning('[MIT] Sheet "%s" not found; sheets: %s', sheet_name, wb.sheetnames)
        # Try TOTAL-EQPT case-insensitively
        for s in wb.sheetnames:
            if 'total' in s.lower() and 'eqpt' in s.lower():
                sheet_name = s
                logger.info('[MIT] Using sheet "%s" as TOTAL-EQPT', s)
                break
        else:
            return []

    ws      = wb[sheet_name]
    results = []

    if cell_map:
        # Explicit admin-configured map — most reliable
        for entry in cell_map:
            row      = int(entry['row'])
            name_col = entry.get('name_col', 'B')
            qty_col  = entry.get('qty_col',  'C')
            name_val = ws[f'{name_col}{row}'].value
            qty_val  = ws[f'{qty_col}{row}'].value
            if not name_val or qty_val is None:
                continue
            try:
                qty = int(float(qty_val))
            except (ValueError, TypeError):
                continue
            if qty <= 0:
                continue
            eq_type = entry.get('equipment_type', name_val)
            cat, stab = _categorise(str(name_val))
            results.append({
                'display_name':                str(name_val).strip(),
                'equipment_type':              eq_type,
                'category':                    cat,
                'required_quantity':           qty,
                'source_sheet':                sheet_name,
                'workbook_row':                row,
                'workbook_cell':               f'{qty_col}{row}',
                'requires_stabilization_photo': stab,
            })

    else:
        # Known-rows mode: use the standard 82-MIT template row positions.
        # col B = equipment name, col C = total quantity for that type.
        for entry in TOTAL_EQPT_ROWS:
            row      = entry['row']
            name_val = ws[f'B{row}'].value
            qty_val  = ws[f'C{row}'].value

            # Fall back to the constant name if the cell is empty (merged cells etc.)
            display = (str(name_val).strip() if name_val else entry['name'])
            if qty_val is None:
                logger.debug('[MIT] TOTAL-EQPT row %d "%s" → None (no recalc?)', row, display)
                continue
            try:
                qty = int(float(qty_val))
            except (ValueError, TypeError):
                continue
            if qty <= 0:
                continue
            cat, stab = _categorise(display)
            results.append({
                'display_name':                display,
                'equipment_type':              entry['xact_code'].lower(),
                'category':                    cat,
                'required_quantity':           qty,
                'source_sheet':                sheet_name,
                'workbook_row':                row,
                'workbook_cell':               f'C{row}',
                'requires_stabilization_photo': stab,
            })

    logger.info('[MIT] Read %d equipment items (qty > 0) from %s', len(results), workbook_path)
    return results
