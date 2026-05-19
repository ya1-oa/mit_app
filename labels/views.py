"""
Labels app views — box labels and wall labels.
"""
import io
import logging
import math
import os
import platform
import re
import subprocess
import tempfile
import time
import traceback

from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.core.files.base import ContentFile
from django.core.mail import EmailMessage
from django.http import Http404, HttpResponse, HttpResponseNotAllowed, JsonResponse
from django.shortcuts import get_object_or_404, render
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt

from allauth.account.decorators import login_required

from openpyxl import load_workbook

from docsAppR.models import Client, File

logger = logging.getLogger(__name__)


GEORGIA_TEAM_EMAILS = [
    "galaxielsaga@gmail.com",
    "wsbjoe9@gmail.com",
    "natebrownlee6@gmail.com",
    "ashleyabernathy3001@gmail.com",
    "tonyjonesteam365@gmail.com",
    "wejones729@yahoo.com",
    "owen7768@att.net",
    "jonesmlc5907@gmail.com",
    "bryanworking4joe@gmail.com",
    "calcompany.cle@gmail.com",
    "lavernebrownlee107@gmail.com",
    "tdeonte17@gmail.com",
    "robbybrowniii@gmail.com",
    "mcgheemarcellus@gmail.com",
    "ihsaankhatim@gmail.com",
]

OHIO_TEAM_EMAILS = [
    "galaxielsaga@gmail.com",
    "wsbjoe9@gmail.com",
    "wejones729@yahoo.com",
    "owen7768@att.net",
    "ihsaankhatim@gmail.com",
]


# ── Helper functions ──────────────────────────────────────────────────────────

def safe_filename(name, max_length=120):
    """Create filesystem-safe filename"""
    return re.sub(r'[<>:"/\\|?*]', '_', name)[:max_length]


def get_room_index_from_name(room_name):
    """Extract the room index from room name"""
    match = re.search(r'(\d+)', room_name)
    if match:
        return int(match.group(1))
    return None


def calculate_print_area(num_labels):
    """Calculate the print area based on number of labels requested"""
    if num_labels <= 0:
        return "A1:B4"  # Default to at least one label area

    labels_per_block = 2
    blocks_needed = math.ceil(num_labels / labels_per_block)
    end_row = blocks_needed * 4

    return f"A1:B{end_row}"


def create_excel_from_template(template_path, output_path, sheet_name, room_index, claim_id, client):
    """Creates an Excel file from template with client data populated"""
    try:
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb = load_workbook(template_path, data_only=False)  # Keep formulas as formulas

        if sheet_name not in wb.sheetnames:
            logger.error(f"Sheet '{sheet_name}' not found in template. Available sheets: {wb.sheetnames}")
            return False

        # Populate jobinfo sheet if exists
        jobinfo_sheet = None
        for sheet in wb.sheetnames:
            if 'jobinfo(2)' in sheet.lower():
                jobinfo_sheet = wb[sheet]
                break

        if jobinfo_sheet:
            try:
                field_mapping = {
                    'pOwner': 1, 'pAddress': 2, 'pCityStateZip': 3, 'cEmail': 4,
                    'cPhone': 5, 'coOwner2': 6, 'cPhone2': 7, 'cAddress2': 8,
                    'cCityStateZip2': 9, 'cEmail2': 10, 'causeOfLoss': 11,
                    'dateOfLoss': 12, 'rebuildType1': 13, 'rebuildType2': 14,
                    'rebuildType3': 15, 'demo': 16, 'mitigation': 17,
                    'otherStructures': 18, 'replacement': 19, 'CPSCLNCONCGN': 20,
                    'yearBuilt': 21, 'contractDate': 22, 'lossOfUse': 23,
                    'breathingIssue': 24, 'hazardMaterialRemediation': 25,
                    'insuranceCo_Name': 26, 'insAddressOvernightMail': 27,
                    'insCityStateZip': 28, 'insuranceCoPhone': 29, 'insWebsite': 30,
                    'insMailingAddress': 31, 'insMailCityStateZip': 32,
                    'claimNumber': 33, 'policyNumber': 34, 'emailInsCo': 35,
                    'deskAdjusterDA': 36, 'DAPhone': 37, 'DAPhExt': 38, 'DAEmail': 39,
                    'fieldAdjusterName': 40, 'phoneFieldAdj': 41, 'fieldAdjEmail': 42,
                    'adjContents': 43, 'adjCpsPhone': 44, 'adjCpsEmail': 45,
                    'emsAdj': 46, 'emsAdjPhone': 47, 'emsTmpEmail': 48,
                    'attLossDraftDept': 49, 'newCustomerID': 50,
                    'mortgageCo': 77, 'mortgageAccountCo': 78,
                    'mortgageContactPerson': 79, 'mortgagePhoneContact': 80,
                    'mortgagePhoneExtContact': 81, 'mortgageAttnLossDraftDept': 82,
                    'mortgageOverNightMail': 83, 'mortgageCityStZipOVN': 84,
                    'mortgageEmail': 85, 'mortgageWebsite': 86, 'mortgageCoFax': 87,
                    'mortgageMailingAddress': 88,
                    'mortgageInitialOfferPhase1ContractAmount': 89,
                    'drawRequest': 90, 'coName': 91, 'coWebsite': 92,
                    'coEmailstatus': 93, 'coAddress': 94, 'coCityState': 95,
                    'coAddress2': 96, 'coCityState2': 97, 'coCityState3': 98,
                    'coLogo1': 99, 'coLogo2': 100, 'coLogo3': 101,
                    'coRepPH': 102, 'coREPEmail': 103, 'coPhone2': 104,
                    'TinW9': 105, 'fedExAccount': 106, 'claimReportDate': 107,
                    'insuranceCustomerServiceRep': 108, 'timeOfClaimReport': 109,
                    'phoneExt': 110, 'tarpExtTMPOk': 111, 'IntTMPOk': 112,
                    'DRYPLACUTOUTMOLDSPRAYOK': 113, 'lossOfUseALE': 114,
                    'tenantLesee': 115, 'propertyAddressStreet': 116,
                    'propertyCityStateZip': 117, 'customerEmail': 118,
                    'cstOwnerPhoneNumber': 119,
                    'startDate': 131, 'endDate': 132, 'lessor': 133,
                    'bedrooms': 138, 'termsAmount': 139,
                }

                datetime_fields = ['dateOfLoss', 'contractDate', 'claimReportDate', 'startDate', 'endDate']

                for field_name, row_num in field_mapping.items():
                    cell_ref = f'C{row_num}'
                    try:
                        value = getattr(client, field_name, None)
                        if field_name in datetime_fields and value is not None:
                            if value.tzinfo is not None:
                                value = value.replace(tzinfo=None)
                            jobinfo_sheet[cell_ref] = value
                        elif isinstance(value, bool):
                            jobinfo_sheet[cell_ref] = 'Yes' if value else 'No'
                        else:
                            jobinfo_sheet[cell_ref] = str(value) if value not in [None, ''] else 'TBD'
                    except Exception as e:
                        logger.warning(f"Error setting {field_name} in {cell_ref}: {str(e)}")
                        jobinfo_sheet[cell_ref] = 'ERROR'

                logger.info("Successfully populated client data in jobinfo sheet")

            except Exception as e:
                logger.warning(f"Could not populate client data: {str(e)}", exc_info=True)

        wb.template = False
        wb.save(output_path)
        wb.close()
        return True

    except Exception as e:
        logger.error(f"Failed to create Excel from template: {str(e)}", exc_info=True)
        return False


def create_room_label_pdf_thermal(pdf_path, room_name, claim_name, num_labels,
                                  start_box_number=1):
    """
    Box labels — 4×3 inch thermal.
    Two-column layout: Col A (75%) = room name + claim name, Col B (25%) = BOX # + number.
    """
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import inch as INCH
    from reportlab.lib import colors
    from reportlab.pdfbase.pdfmetrics import stringWidth

    W = 4 * INCH
    H = 3 * INCH
    MARGIN = 0.15 * INCH

    col_a_w = W * 0.75
    col_b_x = col_a_w
    col_b_w = W * 0.25
    b_cx    = col_b_x + col_b_w / 2

    def fit_text(text, max_width, x, y, max_fs, min_fs=7,
                 font="Helvetica-Bold", centered=True):
        fs = max_fs
        while fs >= min_fs and stringWidth(text, font, fs) > max_width:
            fs -= 1
        c.setFont(font, fs)
        if centered:
            c.drawCentredString(x, y, text)
        else:
            c.drawString(x, y, text)

    c = canvas.Canvas(pdf_path, pagesize=(W, H))

    for i in range(num_labels):
        box_num = start_box_number + i

        fit_text(room_name.upper(),
                 max_width=col_a_w - MARGIN * 2,
                 x=col_a_w / 2, y=H * 0.60,
                 max_fs=36, font="Helvetica-Bold")

        fit_text(claim_name,
                 max_width=col_a_w - MARGIN * 2,
                 x=col_a_w / 2, y=H * 0.33,
                 max_fs=15, font="Helvetica")

        c.setStrokeColor(colors.black)
        c.setLineWidth(0.8)
        c.line(col_b_x, MARGIN, col_b_x, H - MARGIN)

        c.setFillColor(colors.black)
        c.setFont("Helvetica-Bold", 9)
        c.drawCentredString(b_cx, H * 0.76, "BOX #")

        c.setLineWidth(0.5)
        c.line(col_b_x + MARGIN * 0.3, H * 0.69,
               W - MARGIN * 0.3, H * 0.69)

        fit_text(str(box_num),
                 max_width=col_b_w - MARGIN,
                 x=b_cx, y=H * 0.24,
                 max_fs=52, min_fs=16, font="Helvetica-Bold")

        if i < num_labels - 1:
            c.showPage()

    c.save()
    logger.info(f"Box labels saved: {pdf_path} ({num_labels} labels, "
                f"#{start_box_number}–{start_box_number + num_labels - 1})")


def create_wall_label_pdf(pdf_path, room_name, claim_name, num_labels,
                          wtm=None, seq=1):
    """
    Wall orientation labels — 4×6 inch thermal.
    """
    import math as _math
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import inch as INCH
    from reportlab.lib.colors import HexColor, black
    from reportlab.pdfbase.pdfmetrics import stringWidth

    if wtm is None:
        wtm = {}

    W    = 4 * INCH
    H    = 6 * INCH
    BLUE = HexColor('#1a3a8a')
    RED  = HexColor('#dc2626')
    GRID = HexColor('#aac0e8')

    LABEL_PAIRS = [
        ((100, "OVERVIEW"),       (500, "DEMO")),
        ((200, "SOURCE of LOSS"), (600, "WTR MIT")),
        ((300, "CPS"),            (700, "HMR")),
        ((400, "PPR"),            (None, "")),
    ]
    ROW_H   = 0.265 * INCH
    COL_MID = W / 2

    def fit_text(text, max_w, x, y, max_fs, min_fs=7,
                 font="Helvetica-Bold", centered=True):
        fs = max_fs
        while fs >= min_fs and stringWidth(text, font, fs) > max_w:
            fs -= 1
        c.setFont(font, fs)
        if centered:
            c.drawCentredString(x, y, text)
        else:
            c.drawString(x, y, text)

    def dotted_hline(x1, x2, y, width=0.6):
        c.setStrokeColor(GRID); c.setLineWidth(width); c.setDash(3, 3)
        c.line(x1, y, x2, y)
        c.setDash()

    def dotted_vline(x, y1, y2, width=0.5):
        c.setStrokeColor(GRID); c.setLineWidth(width); c.setDash(2, 3)
        c.line(x, y1, x, y2)
        c.setDash()

    c = canvas.Canvas(pdf_path, pagesize=(W, H))

    for i in range(num_labels):
        cx = W / 2

        c.setFillColor(black)
        fit_text(claim_name, W - 0.4*INCH,
                 0.18*INCH, H - 0.26*INCH,
                 max_fs=9, font="Helvetica", centered=False)

        c.setFillColor(black)
        fit_text(room_name, W - 0.3*INCH,
                 cx, H - 0.65*INCH,
                 max_fs=26, font="Helvetica-Bold")

        dotted_hline(0.2*INCH, W - 0.2*INCH, H - 0.85*INCH)

        diag_top = H - 1.05*INCH
        diag_bot = H - 3.85*INCH
        diag_mid = (diag_top + diag_bot) / 2

        col_w = W / 3
        c0_cx = col_w * 0.5
        c1_cx = W / 2
        c2_cx = col_w * 2.5
        R     = 0.32 * INCH

        def _up_arrow(ax, base_y, tip_y):
            aw = 0.18*INCH; sw = 0.06*INCH
            hh = min(0.15*INCH, (tip_y - base_y) * 0.42)
            p = c.beginPath()
            p.moveTo(ax,       tip_y)
            p.lineTo(ax+aw/2,  tip_y-hh); p.lineTo(ax+sw/2, tip_y-hh)
            p.lineTo(ax+sw/2,  base_y);   p.lineTo(ax-sw/2, base_y)
            p.lineTo(ax-sw/2,  tip_y-hh); p.lineTo(ax-aw/2, tip_y-hh)
            p.close()
            c.setFillColor(BLUE); c.setStrokeColor(HexColor('#1A4472'))
            c.setLineWidth(0.4); c.drawPath(p, fill=1, stroke=1)

        def _c_arrow(ax, ay, Rr, open_right):
            gap_half = 45; extent = 360 - gap_half * 2
            arc_s = gap_half if open_right else 180 + gap_half
            arc_e = arc_s + extent
            ah_ang = _math.radians(arc_e if open_right else arc_s)
            tx = -_math.sin(ah_ang) if open_right else  _math.sin(ah_ang)
            ty =  _math.cos(ah_ang) if open_right else -_math.cos(ah_ang)
            c.saveState()
            c.setStrokeColor(BLUE); c.setLineWidth(1.1); c.setLineCap(0)
            p = c.beginPath()
            p.arc(ax-Rr, ay-Rr, ax+Rr, ay+Rr, arc_s, extent)
            c.drawPath(p, fill=0, stroke=1)
            c.restoreState()
            hx = ax + Rr*_math.cos(ah_ang); hy = ay + Rr*_math.sin(ah_ang)
            hs = Rr*0.18; pw = Rr*0.12; px = -ty; py = tx
            p2 = c.beginPath()
            p2.moveTo(hx+tx*hs,          hy+ty*hs)
            p2.lineTo(hx-tx*hs+px*pw,    hy-ty*hs+py*pw)
            p2.lineTo(hx-tx*hs-px*pw,    hy-ty*hs-py*pw)
            p2.close()
            c.setFillColor(BLUE); c.setStrokeColor(BLUE)
            c.setLineWidth(0.3); c.drawPath(p2, fill=1, stroke=1)

        c.setFillColor(black)
        c.setFont("Helvetica-Bold", 12)
        c.drawCentredString(c1_cx, diag_mid + 0.82*INCH, "W=2")
        _up_arrow(c1_cx, diag_mid + 0.22*INCH, diag_mid + 0.68*INCH)

        _c_arrow(c0_cx, diag_mid, R, open_right=True)
        c.setFillColor(black); c.setFont("Helvetica-Bold", 11)
        c.drawCentredString(c0_cx, diag_mid - 0.05*INCH, "W=1")

        bw = 0.78*INCH; bh = 0.40*INCH
        c.setStrokeColor(black); c.setLineWidth(1.1)
        c.rect(c1_cx-bw/2, diag_mid-bh/2, bw, bh)
        c.setFillColor(black); c.setFont("Helvetica-Bold", 8)
        c.drawCentredString(c1_cx, diag_mid - 0.05*INCH, "CENTER")

        _c_arrow(c2_cx, diag_mid, R, open_right=False)
        c.setFillColor(black); c.setFont("Helvetica-Bold", 11)
        c.drawCentredString(c2_cx, diag_mid - 0.05*INCH, "W=3")

        c.setFillColor(black)
        c.setFont("Helvetica-Bold", 12)
        c.drawCentredString(c1_cx, diag_mid - 0.83*INCH, "W=4")

        data_top = diag_bot - 0.10*INCH
        dotted_hline(0.1*INCH, W - 0.1*INCH, data_top)

        for row_i, (left_pair, right_pair) in enumerate(LABEL_PAIRS):
            row_y_top = data_top - row_i * ROW_H
            row_y_bot = row_y_top - ROW_H
            text_y    = row_y_bot + ROW_H * 0.30

            dotted_hline(0.1*INCH, W - 0.1*INCH, row_y_bot, width=0.5)
            dotted_vline(COL_MID, row_y_top, row_y_bot)

            for col_i, (wt_id, wt_desc) in enumerate((left_pair, right_pair)):
                if wt_id is None:
                    continue
                value_type = wtm.get(wt_id, '')
                code_str   = str(wt_id + seq)
                cell_x     = (0.12*INCH if col_i == 0 else COL_MID + 0.08*INCH)

                from reportlab.pdfbase.pdfmetrics import stringWidth as sw
                c.setFillColor(black); c.setFont("Helvetica-Bold", 7)
                c.drawString(cell_x, text_y, code_str)
                x_after_code = cell_x + sw(code_str, "Helvetica-Bold", 7) + 3

                c.setFillColor(black); c.setFont("Helvetica", 7)
                c.drawString(x_after_code, text_y, wt_desc)
                x_after_desc = x_after_code + sw(wt_desc, "Helvetica", 7) + 3

                if value_type == 'TRAVEL':
                    c.setFillColor(black); c.setFont("Helvetica-Bold", 7)
                    c.drawString(x_after_desc, text_y, "TRAVEL")
                elif value_type in ('LOS', 'DAMAGED'):
                    c.setFillColor(RED); c.setFont("Helvetica-Bold", 7)
                    c.drawString(x_after_desc, text_y, value_type)

        footer_sep_y = data_top - len(LABEL_PAIRS) * ROW_H - 0.08*INCH
        dotted_hline(0.1*INCH, W - 0.1*INCH, footer_sep_y)
        footer_y = footer_sep_y - 0.18*INCH
        c.setFillColor(black); c.setFont("Helvetica", 7)
        c.drawRightString(W - 0.15*INCH, footer_y, claim_name[:40])

        if i < num_labels - 1:
            c.showPage()

    c.save()
    logger.info(f"Wall labels saved: {pdf_path} ({num_labels} labels)")


def convert_excel_to_pdf_with_pages(excel_path, pdf_path, sheet_name, room_name, p_owner, num_labels):
    """Convert Excel to PDF with proper print area, eliminating formula errors"""
    try:
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            temp_xlsx = tmp_file.name

        try:
            wb = load_workbook(excel_path)
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found in {excel_path}")

            ws = wb[sheet_name]

            for sheet in wb.sheetnames:
                if sheet != sheet_name:
                    wb.remove(wb[sheet])

            labels_per_page = 4
            total_labels = math.ceil(num_labels / labels_per_page) * labels_per_page
            total_rows = total_labels * 2

            room_name_upper = room_name.upper()
            for row_num in range(1, total_rows + 1):
                if row_num % 2 == 1:
                    ws.cell(row=row_num, column=1, value=room_name_upper)
                else:
                    ws.cell(row=row_num, column=1, value=p_owner)

            print_area = calculate_print_area(num_labels)
            if ':' in print_area:
                ws.print_area = print_area
                logger.info(f"Set print area for {sheet_name}: {print_area}")
            else:
                logger.warning(f"Invalid print area format: {print_area}")

            wb.template = False
            wb.security = None
            wb.save(temp_xlsx)
            wb.close()

            temp_dir = os.path.dirname(temp_xlsx)

            if platform.system() == "Windows":
                libreoffice_path = 'C:\\Program Files\\LibreOffice\\program\\soffice.exe'
            else:
                libreoffice_path = '/usr/bin/libreoffice'

            cmd = [
                libreoffice_path,
                '--headless',
                '--convert-to', 'pdf',
                '--outdir', temp_dir,
                temp_xlsx
            ]

            result = subprocess.run(
                cmd,
                check=True,
                timeout=120,
                capture_output=True,
                text=True
            )
            logger.debug(f"LibreOffice output: {result.stdout}")

            generated_pdf = os.path.splitext(temp_xlsx)[0] + '.pdf'
            max_wait = 30
            waited = 0

            while waited < max_wait:
                if os.path.exists(generated_pdf):
                    try:
                        with open(generated_pdf, 'rb') as f:
                            pdf_data = f.read()
                        with open(pdf_path, 'wb') as f:
                            f.write(pdf_data)
                        logger.info(f"PDF successfully generated at {pdf_path}")
                        return True
                    except PermissionError:
                        time.sleep(0.5)
                        waited += 0.5
                else:
                    time.sleep(0.5)
                    waited += 0.5

            raise RuntimeError(f"PDF generation timed out after {max_wait} seconds")

        finally:
            cleanup_files = [temp_xlsx]
            temp_pdf = os.path.splitext(temp_xlsx)[0] + '.pdf'
            if os.path.exists(temp_pdf):
                cleanup_files.append(temp_pdf)
            for file_path in cleanup_files:
                try:
                    os.unlink(file_path)
                except Exception as e:
                    logger.warning(f"Could not delete temp file {file_path}: {e}")

    except subprocess.CalledProcessError as e:
        error_msg = f"LibreOffice conversion failed: {e.stderr}"
        logger.error(error_msg)
        raise RuntimeError(error_msg)
    except Exception as e:
        logger.error(f"PDF conversion error: {str(e)}", exc_info=True)
        raise


# ── Public views ──────────────────────────────────────────────────────────────

@login_required
def generate_combined_labels(request, claim_id):
    """
    Return a single combined PDF containing wall labels + box labels for every
    room in the given claim.  Used by the 'Download All Labels' button on the
    labels page.
    """
    from docsAppR.tasks import _create_combined_wall_labels_pdf, _create_combined_box_labels_pdf

    try:
        client = Client.objects.get(pOwner=claim_id)
    except Client.DoesNotExist:
        raise Http404

    rooms = client.rooms.all().order_by('sequence')
    if not rooms.exists():
        return HttpResponse("No rooms configured for this claim.", status=400)

    # Build wall-labels PDF
    wall_buf = io.BytesIO()
    _create_combined_wall_labels_pdf(wall_buf, client, rooms)
    wall_buf.seek(0)

    # Build box-labels PDF
    box_buf = io.BytesIO()
    _create_combined_box_labels_pdf(box_buf, client, rooms)
    box_buf.seek(0)

    # Merge into one PDF
    try:
        from pypdf import PdfWriter, PdfReader
        writer = PdfWriter()
        for buf in (wall_buf, box_buf):
            reader = PdfReader(buf)
            for page in reader.pages:
                writer.add_page(page)
        merged_buf = io.BytesIO()
        writer.write(merged_buf)
        merged_buf.seek(0)
        combined_bytes = merged_buf.read()
    except ImportError:
        combined_bytes = wall_buf.read()

    safe_name = "".join(
        c for c in (client.pOwner or 'Claim') if c.isalnum() or c in (' ', '-', '_')
    ).strip().replace(' ', '_')

    response = HttpResponse(combined_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{safe_name}_All_Labels.pdf"'
    return response


def generate_wall_labels_download(request, claim_id):
    """Return a wall-labels-only PDF for every room in the given claim."""
    from docsAppR.tasks import _create_combined_wall_labels_pdf

    try:
        client = Client.objects.get(pOwner=claim_id)
    except Client.DoesNotExist:
        raise Http404

    rooms = client.rooms.all().order_by('sequence')
    if not rooms.exists():
        return HttpResponse("No rooms configured for this claim.", status=400)

    buf = io.BytesIO()
    _create_combined_wall_labels_pdf(buf, client, rooms)
    buf.seek(0)

    safe_name = "".join(
        c for c in (client.pOwner or 'Claim') if c.isalnum() or c in (' ', '-', '_')
    ).strip().replace(' ', '_')

    response = HttpResponse(buf.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{safe_name}_Wall_Labels.pdf"'
    return response


@login_required
def email_labels_to_group(request):
    """Email combined wall + box labels PDFs to selected teams and/or custom addresses."""
    if request.method != 'POST':
        return HttpResponseNotAllowed(['POST'])

    from docsAppR.tasks import _create_combined_wall_labels_pdf, _create_combined_box_labels_pdf

    claim_id = request.POST.get('claim', '').strip()
    if not claim_id:
        return JsonResponse({'status': 'error', 'message': 'Missing claim ID'}, status=400)

    try:
        client = get_object_or_404(Client, pOwner=claim_id)
        rooms = client.rooms.all().order_by('sequence')

        if not rooms.exists():
            return JsonResponse({'status': 'error', 'message': 'No rooms found for this claim'}, status=400)

        # Build recipient list from selected teams + custom addresses
        recipients = set()
        if request.POST.get('georgia'):
            recipients.update(GEORGIA_TEAM_EMAILS)
        if request.POST.get('ohio'):
            recipients.update(OHIO_TEAM_EMAILS)
        custom_raw = request.POST.get('custom_emails', '').strip()
        if custom_raw:
            for addr in custom_raw.split(','):
                addr = addr.strip()
                if addr:
                    recipients.add(addr)

        recipients = list(recipients)
        if not recipients:
            return JsonResponse(
                {'status': 'error', 'message': 'Please select a team or enter at least one email address.'},
                status=400
            )

        # Generate PDFs
        wall_buffer = io.BytesIO()
        _create_combined_wall_labels_pdf(wall_buffer, client, rooms)
        wall_buffer.seek(0)

        box_buffer = io.BytesIO()
        _create_combined_box_labels_pdf(box_buffer, client, rooms)
        box_buffer.seek(0)

        claim_name = client.pOwner or 'Unknown'
        safe_claim_name = "".join(
            c for c in claim_name if c.isalnum() or c in (' ', '-', '_')
        ).strip()
        sender_name = request.user.get_full_name() or request.user.email

        teams_sent = []
        if request.POST.get('georgia'):
            teams_sent.append('Georgia Team')
        if request.POST.get('ohio'):
            teams_sent.append('Ohio Team')
        if custom_raw:
            teams_sent.append('custom addresses')
        teams_label = ', '.join(teams_sent) if teams_sent else 'custom addresses'

        subject = f'[LABELS] {claim_name} – Wall & Box Labels'
        body = f"""<html>
<body style="font-family: Arial, sans-serif; color: #333;">
    <h2 style="color: #1e88e5;">Labels – {claim_name}</h2>
    <div style="background: #f5f5f5; padding: 15px; border-radius: 8px; margin: 15px 0;">
        <p><strong>Claim:</strong> {claim_name}</p>
        <p><strong>Rooms:</strong> {rooms.count()}</p>
        <p><strong>Sent to:</strong> {teams_label}</p>
        <p><strong>Sent by:</strong> {sender_name}</p>
    </div>
    <h3>Attached Files:</h3>
    <ul>
        <li><strong>Wall Labels PDF</strong> – 4×6" thermal labels with wall orientation diagram</li>
        <li><strong>Box Labels PDF</strong> – 4×3" thermal labels with room name and box numbers</li>
    </ul>
    <p style="color: #888; font-size: 12px; margin-top: 20px;">
        Sent from the Claims Management System.
    </p>
</body>
</html>"""

        email = EmailMessage(
            subject=subject,
            body=body,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=recipients,
        )
        email.content_subtype = 'html'
        email.attach(f'{safe_claim_name}_Wall_Labels.pdf', wall_buffer.read(), 'application/pdf')
        email.attach(f'{safe_claim_name}_Box_Labels.pdf', box_buffer.read(), 'application/pdf')
        email.send()

        logger.info(
            f"Labels emailed for claim '{claim_id}' to {teams_label} "
            f"({len(recipients)} recipients) by {request.user.email}"
        )
        return JsonResponse({
            'status': 'success',
            'message': f'Labels sent to {len(recipients)} recipient(s) ({teams_label})',
            'recipients_count': len(recipients),
        })

    except Exception as e:
        logger.error(f"Failed to email labels for claim '{claim_id}': {str(e)}", exc_info=True)
        return JsonResponse({'status': 'error', 'message': 'Failed to send email. Please try again.'}, status=500)


@login_required
def labels(request):
    logger.info(f"Labels function called - method: {request.method}")
    logger.info(f"User: {request.user}, Authenticated: {request.user.is_authenticated}")

    # GET request handling - show the form
    if request.method == 'GET':
        try:
            claims = Client.objects.all()
            selected_claim_id = request.GET.get('claim')
            rooms = []

            if selected_claim_id:
                try:
                    client = get_object_or_404(Client, pOwner=selected_claim_id)

                    for room in client.rooms.all().order_by('sequence'):
                        rooms.append({
                            'id': str(room.id),
                            'name': room.room_name,
                            'sequence': room.sequence
                        })

                    logger.info(f"Found {len(rooms)} rooms for claim {selected_claim_id}")

                except Client.DoesNotExist:
                    rooms = []
                    logger.error(f"Client not found for pOwner: {selected_claim_id}")
                except Exception as e:
                    rooms = []
                    logger.error(f"Unexpected error loading rooms for claim {selected_claim_id}: {str(e)}")
                    logger.debug(f"Traceback: {traceback.format_exc()}")

            from django.contrib.auth.models import Group
            user_groups = Group.objects.all().order_by('name')
            context = {
                'claims': claims,
                'rooms': rooms,
                'selected_claim_id': selected_claim_id,
                'user_groups': user_groups,
            }
            return render(request, 'account/labels.html', context)

        except Exception as e:
            logger.error(f"Error in GET request: {str(e)}", exc_info=True)
            return JsonResponse({'status': 'error', 'message': 'Error loading page'}, status=500)

    # POST request handling - generate PDFs
    elif request.method == 'POST':
        try:
            logger.info("=== STARTING LABEL GENERATION ===")

            room_labels = {}
            claim_id = request.POST.get('claim', '').strip()
            logger.info(f"Claim ID from POST: '{claim_id}'")

            if not claim_id:
                logger.error("Missing claim ID in POST data")
                return JsonResponse({'status': 'error', 'message': 'Missing claim ID'}, status=400)

            # Parse room labels from POST data
            logger.info("Parsing room labels from POST data:")
            for key, value in request.POST.items():
                if key.startswith('room_labels['):
                    try:
                        room_name = key[len('room_labels['):-1]
                        count = int(value)
                        if count > 0:
                            room_labels[room_name] = count
                            logger.info(f"  - {room_name}: {count} labels")
                    except ValueError:
                        logger.warning(f"Invalid value for room label {key}: {value}")
                        continue

            logger.info(f"Total room labels parsed: {len(room_labels)}")

            if not room_labels:
                logger.info("No room labels requested")
                return JsonResponse({'status': 'success', 'message': 'No labels requested', 'pdfs': []})

            # Get client data
            logger.info(f"Looking up client with pOwner: '{claim_id}'")
            try:
                client = Client.objects.get(pOwner=claim_id)
                logger.info(f"Client found: {client.pOwner}")
            except Client.DoesNotExist:
                logger.error(f"Client not found for pOwner: {claim_id}")
                return JsonResponse({'status': 'error', 'message': 'Client not found'}, status=404)
            except Exception as e:
                logger.error(f"Error getting client: {str(e)}")
                return JsonResponse({'status': 'error', 'message': 'Error retrieving client data'}, status=500)

            # Create room index mapping from Room model
            room_indices = {}
            logger.info("Creating room index mapping:")
            for room in client.rooms.all().order_by('sequence'):
                room_indices[room.room_name] = room.sequence + 1
                logger.info(f"  - Room {room.sequence + 1}: '{room.room_name}'")

            logger.info(f"Room indices mapping created with {len(room_indices)} entries")

            # Check if template exists
            template_path = os.path.join(settings.BASE_DIR, 'docsAppR', 'templates', 'excel', 'room_labels_template.xlsx')
            logger.info(f"Template path: {template_path}")

            if not os.path.exists(template_path):
                logger.error(f"Template file not found at: {template_path}")
                return JsonResponse({'status': 'error', 'message': 'Template file not found'}, status=500)

            # Start PDF generation
            logger.info("Starting PDF generation in temporary directory")
            with tempfile.TemporaryDirectory() as temp_dir:
                pdfs_info = []
                logger.info(f"Temporary directory created: {temp_dir}")

                for room_name, num_labels in room_labels.items():
                    try:
                        logger.info(f"--- Processing room: '{room_name}', labels: {num_labels} ---")

                        room_index = room_indices.get(room_name.strip())
                        logger.info(f"Room index from mapping: {room_index}")

                        if not room_index:
                            logger.info("Trying to get room index from name using helper function")
                            try:
                                room_index = get_room_index_from_name(room_name)
                                logger.info(f"Room index from helper function: {room_index}")
                            except Exception as e:
                                logger.error(f"Error in get_room_index_from_name: {str(e)}")
                                room_index = None

                            if not room_index:
                                logger.warning(f"No index found for room: '{room_name}'. Skipping.")
                                continue

                        safe_claim = safe_filename(claim_id)
                        safe_room = safe_filename(room_name)
                        excel_filename = f"labels_{safe_claim}_{safe_room}.xlsx"
                        pdf_filename = f"labels_{safe_claim}_{safe_room}.pdf"
                        temp_excel_path = os.path.join(temp_dir, excel_filename)
                        temp_pdf_path = os.path.join(temp_dir, pdf_filename)
                        sheet_name = f"RM ({room_index})"

                        logger.info(f"File details:")
                        logger.info(f"  - Excel: {temp_excel_path}")
                        logger.info(f"  - PDF: {temp_pdf_path}")
                        logger.info(f"  - Sheet: {sheet_name}")

                        # Generate thermal printer PDF directly
                        logger.info("Generating thermal printer PDF...")
                        try:
                            create_room_label_pdf_thermal(
                                pdf_path=temp_pdf_path,
                                room_name=room_name,
                                claim_name=client.pOwner,
                                num_labels=num_labels
                            )
                            logger.info("PDF generation successful")
                        except Exception as e:
                            logger.error(f"PDF generation failed for {room_name}: {str(e)}")
                            continue

                        if os.path.exists(temp_pdf_path):
                            logger.info(f"PDF file exists at: {temp_pdf_path}")
                            try:
                                with open(temp_pdf_path, 'rb') as pdf_file:
                                    pdf_content = pdf_file.read()

                                logger.info(f"PDF content size: {len(pdf_content)} bytes")

                                from docsAppR.claim_folder_utils import copy_file_to_claim_folder

                                client_folder_name = f"{client.pOwner}@{client.pAddress}" if client.pOwner and client.pAddress else f"Client_{client.id}"
                                safe_folder_name = re.sub(r'[<>:"/\\|?*]', '_', client_folder_name)
                                destination_folder = f"Templates {safe_folder_name}"

                                try:
                                    server_pdf_path = copy_file_to_claim_folder(
                                        client=client,
                                        source_file_path=temp_pdf_path,
                                        destination_folder_type=destination_folder,
                                        new_filename=pdf_filename
                                    )
                                    logger.info(f"Saved PDF to server: {server_pdf_path}")
                                except Exception as copy_err:
                                    logger.warning(f"Could not copy to server folder: {str(copy_err)}")
                                    server_pdf_path = None

                                pdf_obj = File(filename=pdf_filename, size=len(pdf_content))
                                pdf_obj.file.save(pdf_filename, ContentFile(pdf_content))

                                pdfs_info.append({
                                    'room_name': room_name,
                                    'pdf_url': pdf_obj.file.url,
                                    'server_path': server_pdf_path,
                                    'num_labels': num_labels,
                                    'print_area': calculate_print_area(num_labels)
                                })

                                logger.info(f"Successfully generated PDF for {room_name} with {num_labels} labels")
                            except Exception as e:
                                logger.error(f"Error saving PDF file: {str(e)}")
                                continue
                        else:
                            logger.error(f"PDF file not found after conversion: {temp_pdf_path}")

                    except Exception as e:
                        logger.error(f"Error processing room {room_name}: {str(e)}", exc_info=True)
                        continue

                logger.info(f"=== PDF GENERATION COMPLETED ===")
                logger.info(f"Generated {len(pdfs_info)} PDFs out of {len(room_labels)} requested")

                combined_url = reverse('generate_combined_labels', args=[claim_id])
                if pdfs_info:
                    return JsonResponse({
                        'status': 'success',
                        'message': f'Generated {len(pdfs_info)} PDF(s) successfully',
                        'pdfs': pdfs_info,
                        'combined_pdf_url': combined_url,
                    })
                else:
                    return JsonResponse({
                        'status': 'success',
                        'message': 'No valid labels generated',
                        'pdfs': [],
                        'combined_pdf_url': combined_url,
                    })

        except Exception as e:
            logger.error(f"=== LABEL GENERATION FAILED ===")
            logger.error(f"Error: {str(e)}", exc_info=True)
            return JsonResponse({
                'status': 'error',
                'message': 'Label generation failed. Please try again.'
            }, status=500)

    # Handle other HTTP methods
    else:
        return HttpResponseNotAllowed(['GET', 'POST'])


@login_required
def wall_labels(request):
    """
    Generate wall orientation labels for thermal printer (4×6 inch labels)
    Shows room name and wall orientation diagram
    """
    logger.info(f"Wall labels function called - method: {request.method}")

    if request.method == 'GET':
        try:
            claims = Client.objects.all()
            selected_claim_id = request.GET.get('claim')
            rooms = []

            if selected_claim_id:
                try:
                    client = get_object_or_404(Client, pOwner=selected_claim_id)
                    for room in client.rooms.all().order_by('sequence'):
                        rooms.append({
                            'id': str(room.id),
                            'name': room.room_name,
                            'sequence': room.sequence
                        })
                    logger.info(f"Found {len(rooms)} rooms for claim {selected_claim_id}")
                except Client.DoesNotExist:
                    rooms = []
                    logger.error(f"Client not found for pOwner: {selected_claim_id}")
                except Exception as e:
                    rooms = []
                    logger.error(f"Unexpected error loading rooms for claim {selected_claim_id}: {str(e)}")

            context = {
                'claims': claims,
                'rooms': rooms,
                'selected_claim_id': selected_claim_id,
                'label_type': 'wall',
            }
            return render(request, 'account/wall_labels.html', context)
        except Exception as e:
            logger.error(f"Error in GET request: {str(e)}", exc_info=True)
            return JsonResponse({'status': 'error', 'message': 'Error loading page'}, status=500)

    elif request.method == 'POST':
        try:
            logger.info("=== STARTING WALL LABEL GENERATION ===")

            room_labels = {}
            claim_id = request.POST.get('claim', '').strip()
            logger.info(f"Claim ID from POST: '{claim_id}'")

            if not claim_id:
                logger.error("Missing claim ID in POST data")
                return JsonResponse({'status': 'error', 'message': 'Missing claim ID'}, status=400)

            for key, value in request.POST.items():
                if key.startswith('room_labels['):
                    try:
                        room_name = key[len('room_labels['):-1]
                        count = int(value)
                        if count > 0:
                            room_labels[room_name] = count
                            logger.info(f"  - {room_name}: {count} wall labels")
                    except ValueError:
                        continue

            if not room_labels:
                return JsonResponse({'status': 'success', 'message': 'No labels requested', 'pdfs': []})

            try:
                client = Client.objects.get(pOwner=claim_id)
                logger.info(f"Client found: {client.pOwner}")
            except Client.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'Client not found'}, status=404)

            from docsAppR.models import Room
            from docsAppR.tasks import _create_combined_wall_labels_pdf
            import io as _io

            rooms_qs = (
                Room.objects
                .filter(client=client)
                .prefetch_related('work_type_values__work_type')
                .order_by('sequence')
            )

            expanded_rooms = []
            for room in rooms_qs:
                count = room_labels.get(room.room_name, 0)
                for _ in range(count):
                    expanded_rooms.append(room)

            if not expanded_rooms:
                return JsonResponse({'status': 'success', 'message': 'No labels requested', 'pdfs': []})

            safe_claim = re.sub(r'[^A-Za-z0-9 _-]', '_', claim_id)
            pdf_filename = f"{safe_claim}_Wall_Labels.pdf"

            buf = _io.BytesIO()
            _create_combined_wall_labels_pdf(buf, client, expanded_rooms)
            buf.seek(0)
            pdf_bytes = buf.read()

            try:
                from docsAppR.claim_folder_utils import copy_file_to_claim_folder
                with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
                    tmp.write(pdf_bytes)
                    tmp_path = tmp.name
                copy_file_to_claim_folder(
                    client=client,
                    source_file_path=tmp_path,
                    destination_folder_type='Labels',
                    new_filename=pdf_filename,
                )
                os.unlink(tmp_path)
                logger.info(f"Combined wall labels saved to Labels/{pdf_filename}")
            except Exception as save_exc:
                logger.warning(f"Could not save combined labels to Labels folder: {save_exc}")

            pdf_obj = File(filename=pdf_filename, size=len(pdf_bytes))
            pdf_obj.file.save(pdf_filename, ContentFile(pdf_bytes))

            logger.info(f"=== WALL LABEL GENERATION COMPLETED — {len(expanded_rooms)} pages ===")

            return JsonResponse({
                'status': 'success',
                'message': f'Generated {len(expanded_rooms)}-page wall labels PDF',
                'pdf_url': pdf_obj.file.url,
                'pdf_filename': pdf_filename,
            })

        except Exception as e:
            logger.error(f"=== WALL LABEL GENERATION FAILED ===")
            logger.error(f"Error: {str(e)}", exc_info=True)
            return JsonResponse({
                'status': 'error',
                'message': 'Wall label generation failed. Please try again.'
            }, status=500)
    else:
        return HttpResponseNotAllowed(['GET', 'POST'])
