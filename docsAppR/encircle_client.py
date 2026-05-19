"""
Encircle API client classes and claim comparison utilities.
Extracted from docsAppR/views.py to improve manageability.

Classes:
    EncircleExcelExporter  - Export Encircle data to Excel workbooks
    EncircleAPIClient      - REST API calls against the Encircle v1/v2 API
    EncircleDataProcessor  - Parse and structure raw API responses
    EncircleMediaDownloader - Download media to disk (legacy, file-system based)
    ZipMediaDownloader     - Download media into an in-memory ZIP file

Functions:
    normalize_text, extract_tokens, extract_location_code,
    calculate_match_score, compare_claims,
    _is_valid_claim, _is_valid_folder, find_duplicates
"""

import datetime as dt
import json
import logging
import math
import os
import re
from collections import defaultdict
from io import BytesIO

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# EncircleExcelExporter
# ---------------------------------------------------------------------------

class EncircleExcelExporter:
    """
    Class to export Encircle claims data to an Excel workbook with multiple worksheets
    """

    def __init__(self):
        self.wb = Workbook()
        self.default_sheet = self.wb.active
        self.default_sheet.title = "Claims Summary"
        self.styles = self._define_styles()

    def _define_styles(self):
        """Define reusable cell styles"""
        return {
            'header': {
                'font': Font(bold=True, color="FFFFFF"),
                'fill': PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid"),
                'alignment': Alignment(horizontal="center", vertical="center", wrap_text=True),
                'border': Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))
            },
            'data': {
                'alignment': Alignment(vertical="center", wrap_text=True),
                'border': Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))
            }
        }

    def _apply_style(self, cell, style_name):
        """Apply a predefined style to a cell"""
        if style_name in self.styles:
            style = self.styles[style_name]
            if 'font' in style:
                cell.font = style['font']
            if 'fill' in style:
                cell.fill = style['fill']
            if 'alignment' in style:
                cell.alignment = style['alignment']
            if 'border' in style:
                cell.border = style['border']

    def _auto_adjust_columns(self, sheet):
        """Auto-adjust column widths based on content"""
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value is not None:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except Exception:
                    pass
            adjusted_width = min(max(max_length + 2, 10), 50) * 1.2
            sheet.column_dimensions[column_letter].width = adjusted_width

    def _safe_cell_value(self, value):
        """Safely convert value to Excel-compatible format"""
        if value is None:
            return 'N/A'
        elif isinstance(value, (list, tuple)):
            return ", ".join(str(item) for item in value)
        elif isinstance(value, dict):
            return str(value)
        else:
            return str(value)

    def export_claims(self, claims_data):
        """Main method to export all claims data to Excel"""
        try:
            if not isinstance(claims_data, dict):
                raise ValueError("claims_data must be a dictionary")

            claims = claims_data.get('claims', [])
            if claims:
                self._create_claims_sheet(claims)

            if 'claim_details' in claims_data and claims_data['claim_details']:
                self._create_details_sheet(claims_data['claim_details'])

            if 'rooms' in claims_data and claims_data['rooms']:
                self._create_rooms_sheet(claims_data['rooms'])

            if 'floor_plan' in claims_data and claims_data['floor_plan']:
                self._create_floor_plan_sheet(claims_data['floor_plan'])

            if len(self.wb.worksheets) > 1:
                claims_sheet = self.wb["Claims Summary"]
                if claims_sheet.max_row == 1:
                    self.wb.remove(claims_sheet)

            timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"encircle_claims_export_{timestamp}.xlsx"

            output = BytesIO()
            self.wb.save(output)
            output.seek(0)

            return output, filename

        except Exception as e:
            raise Exception(f"Error generating Excel file: {str(e)}")

    def _create_claims_sheet(self, claims):
        """Create a vertical claims worksheet matching the exact template format."""
        sheet = self.wb["Claims Summary"]

        template_rows = [
            "Insured Name", "Address", "City, State", "Email", "Phone",
            "TBD", "TBD", "TBD", "TBD", "TBD",
            "Cause of Loss", "Date of Loss", "Type of Loss",
            "TBD", "TBD",
            "Coverage A Dwelling", "Coverage B Other Structures",
            "Coverage C Personal Property", "Coverage D Loss of Use",
            "Coverage E Liability", "Coverage F Medical",
            "Insurance Company", "Policy Number", "Claim Number",
            "Adjuster Email", "Adjuster Name", "Adjuster Phone",
            "Insurance Company Phone", "Insurance Company Fax", "Claims Email",
        ] + [f"TBD" for _ in range(170)]

        for row_num, header in enumerate(template_rows, start=1):
            sheet.cell(row=row_num, column=1, value=header)

        for col_num, claim in enumerate(claims, start=2):
            if not isinstance(claim, dict):
                continue
            data_mapping = {
                1: claim.get('policyholder_name', 'TBD'),
                2: claim.get('full_address', 'TBD'),
                3: f"{claim.get('city', '')}, {claim.get('state', '')}".strip(', '),
                4: claim.get('email', 'TBD'),
                5: claim.get('phone', 'TBD'),
                11: claim.get('cause_of_loss', 'TBD'),
                12: claim.get('date_of_loss', 'TBD'),
                22: claim.get('insurance_company_name', 'TBD'),
                23: claim.get('policy_number', 'TBD'),
                24: claim.get('claim_number', 'TBD'),
                25: claim.get('adjuster_email', 'TBD'),
                26: claim.get('adjuster_name', 'TBD'),
                27: claim.get('adjuster_phone', 'TBD'),
            }
            for row_num, value in data_mapping.items():
                sheet.cell(row=row_num, column=col_num, value=self._safe_cell_value(value))

        self._auto_adjust_columns(sheet)

    def _create_details_sheet(self, claim_details):
        """Create a worksheet for detailed claim information"""
        if not isinstance(claim_details, dict):
            return

        sheet = self.wb.create_sheet("Claim Details")

        sections = {
            "Basic Information": [
                ("Claim ID", claim_details.get('id')),
                ("Brand ID", claim_details.get('brand_id')),
                ("Organization ID", claim_details.get('organization_id')),
                ("Date Created", claim_details.get('date_claim_created')),
                ("Permalink URL", claim_details.get('permalink_url'))
            ],
            "Property Information": [
                ("Full Address", claim_details.get('full_address')),
                ("Policyholder", claim_details.get('policyholder_name')),
                ("Phone", claim_details.get('policyholder_phone_number')),
                ("Email", claim_details.get('policyholder_email_address'))
            ],
            "Insurance Information": [
                ("Insurance Company", claim_details.get('insurance_company_name')),
                ("Policy Number", claim_details.get('policy_number')),
                ("Insurer ID", claim_details.get('insurer_identifier')),
                ("Broker/Agent", claim_details.get('broker_or_agent_name'))
            ],
            "Loss Information": [
                ("Date of Loss", claim_details.get('date_of_loss')),
                ("Type of Loss", claim_details.get('type_of_loss')),
                ("Loss Details", claim_details.get('loss_details')),
                ("CAT Code", claim_details.get('cat_code'))
            ],
            "Adjuster Information": [
                ("Adjuster Name", claim_details.get('adjuster_name')),
                ("Project Manager", claim_details.get('project_manager_name')),
                ("Contractor ID", claim_details.get('contractor_identifier')),
                ("Assignment ID", claim_details.get('assignment_identifier'))
            ],
            "Financial Information": [
                ("Repair Estimate", claim_details.get('repair_estimate')),
                ("Contents Estimate", claim_details.get('contents_estimate')),
                ("Emergency Estimate", claim_details.get('emergency_estimate')),
                ("Sales Tax", claim_details.get('sales_tax')),
                ("Default Depreciation", claim_details.get('default_depreciation')),
                ("Max Depreciation", claim_details.get('max_depreciation'))
            ]
        }

        row_num = 1
        for section_name, items in sections.items():
            if sheet.max_column >= 2:
                sheet.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
            cell = sheet.cell(row=row_num, column=1, value=section_name)
            cell.font = Font(bold=True, size=14)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            row_num += 1

            for label, value in items:
                label_cell = sheet.cell(row=row_num, column=1, value=label)
                label_cell.font = Font(bold=True)
                sheet.cell(row=row_num, column=2, value=self._safe_cell_value(value))
                row_num += 1

            row_num += 1

        self._auto_adjust_columns(sheet)

    def _create_rooms_sheet(self, rooms):
        """Create a worksheet for room data"""
        if not isinstance(rooms, list) or not rooms:
            return

        sheet = self.wb.create_sheet("Rooms")
        headers = [
            "Claim ID", "Claim Name", "Room Name", "Type", "Floor",
            "Ceiling Height", "Length", "Width",
            "Area", "Perimeter", "Damage Present", "Contents Affected"
        ]

        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            self._apply_style(cell, 'header')

        for row_num, room in enumerate(rooms, 2):
            if not isinstance(room, dict):
                continue
            row_data = [
                room.get('claim_id'), room.get('claim_name'),
                room.get('name'), room.get('room_type'), room.get('floor_name'),
                room.get('ceiling_height'), room.get('length'), room.get('width'),
                room.get('area'), room.get('perimeter'),
                "Yes" if room.get('damage_present') else "No",
                "Yes" if room.get('contents_affected') else "No"
            ]
            for col_num, value in enumerate(row_data, 1):
                cell = sheet.cell(row=row_num, column=col_num, value=self._safe_cell_value(value))
                self._apply_style(cell, 'data')

        self._auto_adjust_columns(sheet)

    def _create_floor_plan_sheet(self, floor_plan):
        """Create a worksheet for floor plan dimensional data"""
        if not isinstance(floor_plan, dict) or not floor_plan:
            return

        sheet = self.wb.create_sheet("Floor Plans")
        headers = [
            "Floor", "Room Name", "Primary Length",
            "Primary Width", "Bounding Width", "Bounding Height",
            "Area", "Ceiling Height"
        ]

        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            self._apply_style(cell, 'header')

        row_num = 2
        for floor_name, rooms in floor_plan.items():
            if not isinstance(rooms, dict):
                continue
            for room_name, dimensions in rooms.items():
                if not isinstance(dimensions, dict):
                    continue
                primary_dims = dimensions.get('primary_dimensions', {})
                bounding_box = dimensions.get('bounding_box', {})
                row_data = [
                    floor_name.replace('Floor_', 'Floor '),
                    room_name,
                    primary_dims.get('length') if isinstance(primary_dims, dict) else None,
                    primary_dims.get('width') if isinstance(primary_dims, dict) else None,
                    bounding_box.get('width') if isinstance(bounding_box, dict) else None,
                    bounding_box.get('height') if isinstance(bounding_box, dict) else None,
                    dimensions.get('area'),
                    dimensions.get('ceiling_height')
                ]
                for col_num, value in enumerate(row_data, 1):
                    cell = sheet.cell(row=row_num, column=col_num, value=self._safe_cell_value(value))
                    self._apply_style(cell, 'data')
                row_num += 1

        self._auto_adjust_columns(sheet)


# ---------------------------------------------------------------------------
# EncircleAPIClient
# ---------------------------------------------------------------------------

class EncircleAPIClient:
    """
    Centralized API client for Encircle API operations.
    Following Single Responsibility Principle.
    """

    def __init__(self):
        self.api_key = "367382d2-0b2d-4b01-9d06-8f18fd492f5e"
        self.base_url = "https://api.encircleapp.com/v1"
        self.v2_base_url = "https://api.encircleapp.com/v2"
        self.headers = {"Authorization": f"Bearer {self.api_key}"}

    def _make_request(self, endpoint, params=None):
        """Generic GET to v1 API."""
        try:
            url = f"{self.base_url}/{endpoint}"
            response = requests.get(url, headers=self.headers, params=params)
            response.raise_for_status()
            return response.json()
        except requests.exceptions.RequestException as e:
            logger.error(f"API request failed for {endpoint}: {str(e)}")
            raise Exception(f"API request failed: {str(e)}")

    def _make_v2_request(self, endpoint, params=None):
        """Generic GET to v2 API."""
        try:
            url = f"{self.v2_base_url}/{endpoint}"
            response = requests.get(url, headers=self.headers, params=params)
            if response.status_code != 200:
                logger.error(f"API request failed for {endpoint}: {response.status_code} - {response.text}")
                raise Exception(f"API request failed: {response.status_code} - {response.text}")
            return response.json()
        except requests.exceptions.RequestException as e:
            logger.error(f"API request failed for {endpoint}: {str(e)}")
            raise Exception(f"API request failed: {str(e)}")

    def _make_post_request(self, endpoint, payload):
        """Generic POST to v1 API."""
        try:
            url = f"{self.base_url}/{endpoint}"
            headers = {**self.headers, "Content-Type": "application/json"}
            response = requests.post(url, headers=headers, json=payload)
            if not response.ok:
                try:
                    body = response.json()
                except Exception:
                    body = response.text
                logger.error(f"POST request failed for {endpoint}: {response.status_code} — {body}")
                logger.error(f"POST payload was: {payload}")
                response.raise_for_status()
            return response.json()
        except requests.exceptions.RequestException as e:
            logger.error(f"POST request failed for {endpoint}: {str(e)}")
            raise Exception(f"POST request failed: {str(e)}")

    def _make_patch_request(self, endpoint, payload):
        """Generic PATCH to v1 API."""
        try:
            url = f"{self.base_url}/{endpoint}"
            headers = {**self.headers, "Content-Type": "application/json"}
            response = requests.patch(url, headers=headers, json=payload)
            response.raise_for_status()
            return response.json()
        except requests.exceptions.RequestException as e:
            logger.error(f"PATCH request failed for {endpoint}: {str(e)}")
            raise Exception(f"PATCH request failed: {str(e)}")

    def get_all_claims(self, limit=100, order='newest'):
        """Fetch all property claims using cursor-based pagination."""
        all_claims = {'list': []}
        after_cursor = None
        page_count = 0

        while True:
            page_count += 1
            params = {'limit': min(limit, 100), 'order': order}
            if after_cursor:
                params['after'] = after_cursor

            try:
                response = self._make_request("property_claims", params=params)
                claims = response.get('list', [])

                if not claims:
                    break

                all_claims['list'].extend(claims)

                # Safe cursor extraction — same pattern as get_all_structure_rooms
                after_cursor = (response.get('cursor') or {}).get('after')
                if not after_cursor:
                    break

            except Exception as e:
                logger.error(f"Error fetching claims page {page_count}: {e}")
                break

        return all_claims

    def get_claim_details(self, claim_id):
        return self._make_request(f"property_claims/{claim_id}")

    def get_claim_structures(self, claim_id):
        return self._make_request(f"property_claims/{claim_id}/structures")

    def get_claim_rooms(self, claim_id, structure_id):
        return self._make_request(f"property_claims/{claim_id}/structures/{structure_id}/rooms")

    def get_all_structure_rooms(self, claim_id, structure_id):
        """Fetch ALL rooms in a structure using cursor-based pagination."""
        all_rooms = []
        after_cursor = None
        while True:
            params = {'limit': 100}
            if after_cursor:
                params['after'] = after_cursor
            resp = self._make_request(
                f"property_claims/{claim_id}/structures/{structure_id}/rooms",
                params=params,
            )
            items = resp.get('list', []) if isinstance(resp, dict) else []
            all_rooms.extend(items)
            after_cursor = (resp.get('cursor') or {}).get('after') if isinstance(resp, dict) else None
            if not after_cursor:
                break
        return all_rooms

    def get_claim_floor_plan(self, claim_id):
        return self._make_v2_request(f"property_claims/{claim_id}/floor_plan_dimensions")

    def get_account_ids(self):
        """Fetch organization_id, brand_id, and contractor_identifier from the first existing claim."""
        result = {'organization_id': None, 'brand_id': None, 'contractor_identifier': None}
        try:
            resp = self._make_request("property_claims", params={"limit": 1})
            claims = resp.get("list", []) if isinstance(resp, dict) else []
            if claims:
                c = claims[0]
                result['organization_id'] = str(c['organization_id']) if c.get('organization_id') else None
                result['brand_id'] = str(c['brand_id']) if c.get('brand_id') else None
                result['contractor_identifier'] = str(c['contractor_identifier']) if c.get('contractor_identifier') else None
        except Exception as exc:
            logger.warning(f"get_account_ids: failed — {exc}", exc_info=True)
        return result

    def create_claim(self, claim_data):
        """Create a new property claim in Encircle."""
        payload = dict(claim_data)
        needs_ids = (
            not payload.get('organization_id')
            or not payload.get('brand_id')
            or not payload.get('contractor_identifier')
        )
        if needs_ids:
            ids = self.get_account_ids()
            if ids['organization_id'] and not payload.get('organization_id'):
                payload['organization_id'] = ids['organization_id']
            if ids['brand_id'] and not payload.get('brand_id'):
                payload['brand_id'] = ids['brand_id']
            if ids['contractor_identifier'] and not payload.get('contractor_identifier'):
                payload['contractor_identifier'] = ids['contractor_identifier']
        if not payload.get('type_of_loss'):
            payload['type_of_loss'] = 'Other'
        return self._make_post_request("property_claims", payload)

    def get_or_create_default_structure(self, encircle_claim_id):
        """Return the first structure on a claim, creating one if none exist."""
        structures_resp = self.get_claim_structures(encircle_claim_id)
        structures = structures_resp.get("list", structures_resp) if isinstance(structures_resp, dict) else structures_resp
        if structures:
            return structures[0]
        return self._make_post_request(
            f"property_claims/{encircle_claim_id}/structures",
            {"name": "Main Building"}
        )

    def create_room(self, encircle_claim_id, structure_id, room_payload):
        """Create a room inside a structure."""
        return self._make_post_request(
            f"property_claims/{encircle_claim_id}/structures/{structure_id}/rooms",
            room_payload
        )

    def get_room_media(self, claim_id, structure_id, room_id):
        """Fetch all media items attached to a specific room (paginated)."""
        all_media = []
        after_cursor = None
        while True:
            params = {'limit': 100}
            if after_cursor:
                params['after'] = after_cursor
            resp = self._make_request(
                f"property_claims/{claim_id}/structures/{structure_id}/rooms/{room_id}/media",
                params=params,
            )
            items = resp.get('list', []) if isinstance(resp, dict) else []
            all_media.extend(items)
            after_cursor = (resp.get('cursor') or {}).get('after') if isinstance(resp, dict) else None
            if not after_cursor:
                break
        return all_media

    def delete_room(self, claim_id, structure_id, room_id):
        """Delete a room from a structure. Returns the HTTP status code."""
        url = f"{self.base_url}/property_claims/{claim_id}/structures/{structure_id}/rooms/{room_id}"
        resp = requests.delete(url, headers=self.headers)
        resp.raise_for_status()
        return resp.status_code

    def reassign_media(self, claim_id, media_id, new_room_id):
        """Move a media item to a different room."""
        return self._make_patch_request(
            f"property_claims/{claim_id}/media/{media_id}",
            {"source": {"type": "room", "primary_id": str(new_room_id)}},
        )

    def get_all_claim_media(self, claim_id):
        """Fetch ALL media items for a claim using claim-level pagination."""
        all_media = []
        after_cursor = None
        while True:
            params = {'limit': 100}
            if after_cursor:
                params['after'] = after_cursor
            resp = self._make_request(f"property_claims/{claim_id}/media", params=params)
            items = resp.get('list', []) if isinstance(resp, dict) else []
            all_media.extend(items)
            after_cursor = (resp.get('cursor') or {}).get('after') if isinstance(resp, dict) else None
            if not after_cursor:
                break
        return all_media

    def upload_media_to_room(self, dest_claim_id, dest_room_id, file_bytes, filename, content_type):
        """Upload a photo (bytes) into a specific room on a destination claim."""
        url = f"{self.base_url}/property_claims/{dest_claim_id}/media"
        auth_headers = {"Authorization": self.headers["Authorization"]}
        files = {'file': (filename, file_bytes, content_type)}
        data = {
            'source[type]': 'room',
            'source[primary_id]': str(dest_room_id),
        }
        try:
            resp = requests.post(url, headers=auth_headers, files=files, data=data)
            if not resp.ok:
                try:
                    body = resp.json()
                except Exception:
                    body = resp.text
                logger.error(f"upload_media_to_room failed: {resp.status_code} — {body}")
                resp.raise_for_status()
            return resp.json()
        except requests.exceptions.RequestException as e:
            logger.error(f"upload_media_to_room request error: {e}")
            raise Exception(f"Media upload failed: {e}")


# ---------------------------------------------------------------------------
# EncircleDataProcessor
# ---------------------------------------------------------------------------

class EncircleDataProcessor:
    """Data processing utilities for Encircle API responses."""

    @staticmethod
    def _get_polygon_dimensions(coordinates):
        """Calculate width and length from polygon coordinates (bounding box)."""
        points = coordinates[0]
        x_coords = [point[0] for point in points]
        y_coords = [point[1] for point in points]
        width = (max(x_coords) - min(x_coords)) / 30.48
        height = (max(y_coords) - min(y_coords)) / 30.48
        return width, height

    @staticmethod
    def _get_edge_lengths(coordinates):
        """Calculate actual edge lengths of the polygon."""
        points = coordinates[0]
        edge_lengths = []
        for i in range(len(points) - 1):
            x1, y1 = points[i]
            x2, y2 = points[i + 1]
            length = math.sqrt((x2 - x1) ** 2 + (y2 - y1) ** 2) / 30.48
            edge_lengths.append(length)
        return edge_lengths

    @staticmethod
    def process_floor_plan_data(raw_floor_plan_data):
        """Process floor plan data to calculate room dimensions."""
        if not raw_floor_plan_data or 'list' not in raw_floor_plan_data:
            return {}

        room_dimensions = {}

        for floor_idx, floor in enumerate(raw_floor_plan_data['list'][0]['floors']):
            floor_rooms = {}
            for feature in floor['features']:
                room_name = feature['properties']['name']
                coordinates = feature['geometry']['coordinates']
                width, height = EncircleDataProcessor._get_polygon_dimensions(coordinates)
                edge_lengths = EncircleDataProcessor._get_edge_lengths(coordinates)

                floor_rooms[room_name] = {
                    'bounding_box': {'width': round(width, 2), 'height': round(height, 2)},
                    'ceiling_height': feature['properties'].get('ceiling_height', 'N/A'),
                    'area': round(width * height, 2)
                }

            room_dimensions[f'Floor_{floor_idx + 1}'] = floor_rooms

        return room_dimensions

    @staticmethod
    def get_room_dimensions_simple(room_feature):
        """Get width and length for a single room feature."""
        coordinates = room_feature['geometry']['coordinates'][0]
        x_coords = [point[0] for point in coordinates]
        y_coords = [point[1] for point in coordinates]
        width = (max(x_coords) - min(x_coords)) / 12
        height = (max(y_coords) - min(y_coords)) / 12
        return {'width': round(width, 2), 'length': round(height, 2), 'area': round(width * height, 2)}

    @staticmethod
    def process_claims_list(raw_claims_data):
        """Process and structure the claims list data."""
        if not raw_claims_data or 'list' not in raw_claims_data:
            return []

        processed_claims = []
        for claim in raw_claims_data['list']:
            processed_claims.append({
                'id': claim.get('id'),
                'brand_id': claim.get('brand_id'),
                'organization_id': claim.get('organization_id'),
                'permalink_url': claim.get('permalink_url'),
                'type_of_loss': claim.get('type_of_loss'),
                'full_address': claim.get('full_address'),
                'date_claim_created': claim.get('date_claim_created'),
                'date_of_loss': claim.get('date_of_loss'),
                'policyholder_name': claim.get('policyholder_name'),
                'insurance_company_name': claim.get('insurance_company_name'),
                'policy_number': claim.get('policy_number'),
                'adjuster_name': claim.get('adjuster_name'),
                'project_manager_name': claim.get('project_manager_name'),
                'total_rooms': 0,
                'room_types': []
            })

        return processed_claims

    @staticmethod
    def process_claim_rooms(raw_rooms_data):
        """Process room data for a specific claim."""
        if not raw_rooms_data or 'list' not in raw_rooms_data:
            return []

        processed_rooms = []
        room_types = set()

        for room in raw_rooms_data['list']:
            room_data = {
                'id': room.get('id'),
                'name': room.get('name', 'Unnamed Room'),
                'room_type': room.get('room_type', 'Unknown'),
                'floor_name': room.get('floor_name', 'Unknown Floor'),
                'ceiling_height': room.get('ceiling_height', 0),
                'length': room.get('length', 0),
                'width': room.get('width', 0),
                'area': room.get('area', 0),
                'perimeter': room.get('perimeter', 0),
                'damage_present': room.get('damage_present', False),
                'contents_affected': room.get('contents_affected', False)
            }
            processed_rooms.append(room_data)
            room_types.add(room_data['room_type'])

        return processed_rooms, list(room_types)

    @staticmethod
    def process_claim_details(raw_claim_data):
        """Process detailed claim information."""
        if not raw_claim_data:
            return {}

        return {
            'id': raw_claim_data.get('id'),
            'brand_id': raw_claim_data.get('brand_id'),
            'organization_id': raw_claim_data.get('organization_id'),
            'permalink_url': raw_claim_data.get('permalink_url'),
            'type_of_loss': raw_claim_data.get('type_of_loss'),
            'locale': raw_claim_data.get('locale'),
            'adjuster_name': raw_claim_data.get('adjuster_name'),
            'assignment_identifier': raw_claim_data.get('assignment_identifier'),
            'cat_code': raw_claim_data.get('cat_code'),
            'contents_estimate': raw_claim_data.get('contents_estimate'),
            'contractor_identifier': raw_claim_data.get('contractor_identifier'),
            'date_claim_created': raw_claim_data.get('date_claim_created'),
            'date_of_loss': raw_claim_data.get('date_of_loss'),
            'default_depreciation': raw_claim_data.get('default_depreciation'),
            'emergency_estimate': raw_claim_data.get('emergency_estimate'),
            'full_address': raw_claim_data.get('full_address'),
            'insurer_identifier': raw_claim_data.get('insurer_identifier'),
            'loss_details': raw_claim_data.get('loss_details'),
            'max_depreciation': raw_claim_data.get('max_depreciation'),
            'policyholder_email_address': raw_claim_data.get('policyholder_email_address'),
            'policyholder_name': raw_claim_data.get('policyholder_name'),
            'policyholder_phone_number': raw_claim_data.get('policyholder_phone_number'),
            'project_manager_name': raw_claim_data.get('project_manager_name'),
            'repair_estimate': raw_claim_data.get('repair_estimate'),
            'sales_tax': raw_claim_data.get('sales_tax'),
            'broker_or_agent_name': raw_claim_data.get('broker_or_agent_name'),
            'insurance_company_name': raw_claim_data.get('insurance_company_name'),
            'policy_number': raw_claim_data.get('policy_number')
        }


# ---------------------------------------------------------------------------
# EncircleMediaDownloader  (downloads to disk)
# ---------------------------------------------------------------------------

class EncircleMediaDownloader:
    """
    Downloads media files from Encircle API and organizes them by room labels.
    Files are numbered sequentially and saved in room-specific folders.
    """

    def __init__(self, api_client, target_rooms=None):
        self.api_client = api_client
        self.target_rooms = target_rooms or []
        self.downloaded_files = 0
        self.failed_downloads = 0
        self.base_dir = "encircle_media_downloads"
        os.makedirs(self.base_dir, exist_ok=True)

    def download_claim_media(self, property_claim_id):
        """Main method to download all media for a specific claim."""
        try:
            media_list = self._get_media_list(property_claim_id)
            if not media_list:
                return
            for idx, media_item in enumerate(media_list, 1):
                if self._should_download(media_item):
                    self._process_media_item(media_item, idx, len(media_list))
        except Exception as e:
            logging.error(f"Error downloading media for claim {property_claim_id}: {str(e)}")
            raise

    def _should_download(self, media_item):
        if not self.target_rooms:
            return True
        labels = media_item.get('labels', [])
        return any(room.lower() in [label.lower() for label in labels] for room in self.target_rooms)

    def _get_media_list(self, property_claim_id):
        all_media = []
        after_cursor = None
        while True:
            params = {'limit': 100}
            if after_cursor:
                params['after'] = after_cursor
            response = self.api_client._make_request(f"property_claims/{property_claim_id}/media", params=params)
            if not response or 'list' not in response:
                break
            all_media.extend(response['list'])
            after_cursor = response.get('cursor', {}).get('after')
            if not after_cursor:
                break
        return all_media

    def _process_media_item(self, media_item, current_index, total_items):
        try:
            folder_name = self._get_folder_name(media_item)
            file_extension = self._get_file_extension(media_item['content_type'])
            seq_num = str(current_index).zfill(len(str(total_items)))
            clean_filename = self._sanitize_filename(f"{seq_num}_{media_item['filename']}")
            if not clean_filename.lower().endswith(file_extension.lower()):
                clean_filename += file_extension
            folder_path = os.path.join(self.base_dir, folder_name)
            os.makedirs(folder_path, exist_ok=True)
            file_path = os.path.join(folder_path, clean_filename)
            self._download_file(media_item['download_uri'], file_path)
            self.downloaded_files += 1
            self._create_metadata_file(media_item, file_path)
        except Exception as e:
            self.failed_downloads += 1
            logging.error(f"Failed to process media item: {str(e)}")

    def _get_folder_name(self, media_item):
        labels = media_item.get('labels', [])
        if not labels:
            return "unlabeled_media"
        valid_labels = [label.strip() for label in labels if label.strip()]
        if not valid_labels:
            return "unlabeled_media"
        return os.path.join(*[self._sanitize_folder_name(label) for label in valid_labels[:3]])

    def _sanitize_filename(self, filename):
        for char in '<>:"/\\|?*':
            filename = filename.replace(char, '_')
        return filename

    def _sanitize_folder_name(self, foldername):
        for char in '<>:"/\\|?*':
            foldername = foldername.replace(char, '_')
        return foldername[:50]

    def _get_file_extension(self, content_type):
        return {
            'image/jpeg': '.jpg',
            'image/png': '.png',
            'application/pdf': '.pdf',
            'video/mp4': '.mp4',
            'application/vnd.openxmlformats-officedocument.wordprocessingml.document': '.docx',
        }.get(content_type, '.bin')

    def _download_file(self, url, file_path):
        response = requests.get(url, stream=True)
        response.raise_for_status()
        with open(file_path, 'wb') as f:
            for chunk in response.iter_content(chunk_size=8192):
                f.write(chunk)

    def _create_metadata_file(self, media_item, file_path):
        metadata = {
            'original_filename': media_item['filename'],
            'content_type': media_item['content_type'],
            'source_type': media_item['source']['type'],
            'source_id': media_item['source']['primary_id'],
            'creator': media_item['creator']['actor_identifier'],
            'created_date': media_item['primary_server_created'],
            'labels': media_item.get('labels', []),
            'download_time': dt.datetime.now().isoformat()
        }
        with open(f"{file_path}.meta.json", 'w') as f:
            json.dump(metadata, f, indent=2)


# ---------------------------------------------------------------------------
# ZipMediaDownloader  (downloads into an in-memory ZIP)
# ---------------------------------------------------------------------------

class ZipMediaDownloader:
    """
    Downloads media files from Encircle API and organizes them in a zip file by room labels.
    """

    IMAGE_CONTENT_TYPES = {
        'image/jpeg', 'image/jpg', 'image/png', 'image/gif',
        'image/webp', 'image/heic', 'image/heif', 'image/tiff',
    }

    def __init__(self, api_client, target_rooms=None, zip_file=None,
                 organize_by_rooms=True, images_only=False):
        self.api_client = api_client
        self.target_rooms = target_rooms or []
        self.downloaded_files = 0
        self.failed_downloads = 0
        self.zip_file = zip_file
        self.organize_by_rooms = organize_by_rooms
        self.images_only = images_only

    def download_claim_media(self, property_claim_id):
        """Main method to download all media for a specific claim into the zip file."""
        try:
            media_list = self._get_media_list(property_claim_id)
            if not media_list:
                return
            for idx, media_item in enumerate(media_list, 1):
                if self._should_download(media_item):
                    self._process_media_item(media_item, idx, len(media_list))
        except Exception as e:
            logging.error(f"Error downloading media for claim {property_claim_id}: {str(e)}")
            raise

    def _should_download(self, media_item):
        if self.images_only:
            ct = (media_item.get('content_type') or '').lower().split(';')[0].strip()
            if not ct.startswith('image/'):
                return False
        if not self.target_rooms:
            return True
        labels = media_item.get('labels', [])
        return any(room.lower() in [label.lower() for label in labels] for room in self.target_rooms)

    def _get_media_list(self, property_claim_id):
        all_media = []
        after_cursor = None
        while True:
            params = {'limit': 100}
            if after_cursor:
                params['after'] = after_cursor
            response = self.api_client._make_request(f"property_claims/{property_claim_id}/media", params=params)
            if not response or 'list' not in response:
                break
            all_media.extend(response['list'])
            after_cursor = response.get('cursor', {}).get('after')
            if not after_cursor:
                break
        return all_media

    def _process_media_item(self, media_item, current_index, total_items):
        try:
            folder_name = self._get_folder_name(media_item)
            file_extension = self._get_file_extension(media_item['content_type'])
            seq_num = str(current_index).zfill(len(str(total_items)))
            clean_filename = self._sanitize_filename(f"{seq_num}_{media_item['filename']}")
            if not clean_filename.lower().endswith(file_extension.lower()):
                clean_filename += file_extension
            zip_path = os.path.join(folder_name, clean_filename)
            file_content = self._download_file_content(media_item['download_uri'])
            self.zip_file.writestr(zip_path, file_content)
            self.downloaded_files += 1
            self._add_metadata_file(media_item, zip_path)
        except Exception as e:
            self.failed_downloads += 1
            logging.error(f"Failed to process media item: {str(e)}")

    def _get_folder_name(self, media_item):
        if not self.organize_by_rooms:
            return "images"
        labels = media_item.get('labels', [])
        if not labels:
            return "unlabeled_media"
        valid_labels = [label.strip() for label in labels if label.strip()]
        if not valid_labels:
            return "unlabeled_media"
        return os.path.join(*[self._sanitize_folder_name(label) for label in valid_labels[:3]])

    def _sanitize_filename(self, filename):
        for char in '<>:"/\\|?*':
            filename = filename.replace(char, '_')
        return filename

    def _sanitize_folder_name(self, foldername):
        for char in '<>:"/\\|?*':
            foldername = foldername.replace(char, '_')
        return foldername[:50]

    def _get_file_extension(self, content_type):
        return {
            'image/jpeg': '.jpg',
            'image/png': '.png',
            'application/pdf': '.pdf',
            'video/mp4': '.mp4',
            'application/vnd.openxmlformats-officedocument.wordprocessingml.document': '.docx',
        }.get(content_type, '.bin')

    def _download_file_content(self, url):
        response = requests.get(url, stream=True)
        response.raise_for_status()
        return response.content

    def _add_metadata_file(self, media_item, file_path):
        metadata = {
            'original_filename': media_item['filename'],
            'content_type': media_item['content_type'],
            'source_type': media_item['source']['type'],
            'source_id': media_item['source']['primary_id'],
            'creator': media_item['creator']['actor_identifier'],
            'created_date': media_item['primary_server_created'],
            'labels': media_item.get('labels', []),
            'download_time': dt.datetime.now().isoformat()
        }
        self.zip_file.writestr(f"{file_path}.meta.json", json.dumps(metadata, indent=2))


# ---------------------------------------------------------------------------
# Claim comparison utilities
# ---------------------------------------------------------------------------

_TEST_EXCLUDE_PATTERNS = [
    'HOW2', 'TEST', 'TEMPLATE', 'SAMPLE', 'ROOMLISTS', 'READINGS',
    'TMPL', 'CHECKLIST', 'TRAILER', 'WAREHOUSE', 'DEFAULT', 'TEMP',
    'PLACEHOLDER', 'EXAMPLE', 'DEMO', 'XXXX', 'AAA', '===', 'BACKEND', 'TUTORIAL'
]


def normalize_text(text):
    """Basic normalization: uppercase, strip extra spaces."""
    if not text:
        return ""
    return ' '.join(text.upper().split())


def extract_tokens(text):
    """Extract meaningful tokens from text. Returns a set of normalized words."""
    if not text:
        return set()
    parts = text.split('@')
    main_part = parts[0] if parts else text
    address_part = parts[1] if len(parts) > 1 else ""
    tokens = re.findall(r'\b[A-Z0-9]{2,}\b', main_part.upper())
    if address_part:
        tokens.extend(re.findall(r'\b[A-Z0-9]{2,}\b', address_part.upper()))
    noise_words = {'LLC', 'INC', 'THE', 'AND', 'FOR', 'CLAIM', 'EST', 'FIRE', 'WATER', 'STORM'}
    return {t for t in tokens if t not in noise_words and len(t) >= 2}


def extract_location_code(text):
    """Extract location codes like GA22, OH24, GA22A."""
    if not text:
        return None
    match = re.search(r'\b([A-Z]{2,3}\d{2,3}[A-Z0-9\-]*)\b', text.upper())
    return match.group(1) if match else None


def calculate_match_score(encircle_contractor, folder_name):
    """Simple fuzzy matching between contractor ID and folder name. Returns score 0-1."""
    if not encircle_contractor or not folder_name:
        return 0.0

    contractor_norm = normalize_text(encircle_contractor)
    folder_norm = normalize_text(folder_name)

    score = 0.0

    if contractor_norm in folder_norm or folder_norm in contractor_norm:
        score += 0.4

    contractor_location = extract_location_code(contractor_norm)
    folder_location = extract_location_code(folder_norm)

    if contractor_location and folder_location:
        if contractor_location == folder_location:
            score += 0.3
        elif contractor_location[:4] == folder_location[:4]:
            score += 0.2

    contractor_tokens = extract_tokens(contractor_norm)
    folder_tokens = extract_tokens(folder_norm)

    if contractor_tokens and folder_tokens:
        common_tokens = contractor_tokens.intersection(folder_tokens)
        token_score = len(common_tokens) / max(len(contractor_tokens), len(folder_tokens))
        score += token_score * 0.3

    return min(score, 1.0)


def _is_valid_claim(claim):
    """Filter out test/placeholder claims from Encircle."""
    if not claim.get('policyholder_name') and not claim.get('contractor_identifier'):
        return False
    policyholder = (claim.get('policyholder_name') or '').upper()
    contractor = (claim.get('contractor_identifier') or '').upper()
    return not any(p in policyholder or p in contractor for p in _TEST_EXCLUDE_PATTERNS)


def _is_valid_folder(folder_claim):
    """Filter out test/placeholder folders from OneDrive."""
    folder_name = (folder_claim.get('folder_name') or '').upper()
    exclude_patterns = _TEST_EXCLUDE_PATTERNS + [
        'CLOSED CLAIMS', 'PROOF OF LOSS', 'DRAWINGS', 'APPRAISALS', 'FOLDER', 'TEXT'
    ]
    if any(p in folder_name for p in exclude_patterns):
        return False
    clean_alpha = re.sub(r'[^A-Z]', '', folder_name)
    return len(clean_alpha) >= 3


def find_duplicates(encircle_claims, onedrive_claims):
    """Find duplicate claims in both systems."""
    duplicates = {'encircle_duplicates': [], 'onedrive_duplicates': []}

    contractor_count = defaultdict(list)
    for claim in encircle_claims:
        contractor_id = claim.get('contractor_identifier', '')
        if contractor_id and contractor_id.strip():
            contractor_count[contractor_id].append(claim)

    for contractor_id, claims in contractor_count.items():
        if len(claims) > 1:
            duplicates['encircle_duplicates'].append({
                'contractor_id': contractor_id,
                'count': len(claims),
                'claims': claims
            })

    folder_count = defaultdict(list)
    for claim in onedrive_claims:
        folder_name = claim.get('folder_name', '')
        if folder_name:
            folder_count[normalize_text(folder_name)].append(claim)

    for folder_name, claims in folder_count.items():
        if len(claims) > 1:
            duplicates['onedrive_duplicates'].append({
                'folder_name': folder_name,
                'count': len(claims),
                'claims': claims
            })

    return duplicates


def compare_claims(encircle_claims, onedrive_claims):
    """Simple matching: compare contractor ID against folder name."""
    valid_encircle = [c for c in encircle_claims if _is_valid_claim(c)]
    valid_onedrive = [c for c in onedrive_claims if _is_valid_folder(c)]

    results = {
        'summary': {
            'total_encircle': len(encircle_claims),
            'total_onedrive': len(onedrive_claims),
            'matches': 0,
            'encircle_only': 0,
            'onedrive_only': 0,
            'encircle_test_data': len([c for c in encircle_claims if not _is_valid_claim(c)]),
            'onedrive_test_data': len([c for c in onedrive_claims if not _is_valid_folder(c)]),
            'match_breakdown': {'high_confidence': 0, 'medium_confidence': 0, 'low_confidence': 0}
        },
        'matched_pairs': [],
        'encircle_missing_onedrive': [],
        'onedrive_extra': [],
        'encircle_test_data': [c for c in encircle_claims if not _is_valid_claim(c)],
        'onedrive_test_data': [c for c in onedrive_claims if not _is_valid_folder(c)],
        'duplicates': find_duplicates(valid_encircle, valid_onedrive)
    }

    matched_encircle = set()
    matched_onedrive = set()
    MATCH_THRESHOLD = 0.65

    for encircle_claim in valid_encircle:
        if encircle_claim['id'] in matched_encircle:
            continue
        contractor_id = encircle_claim.get('contractor_identifier', '').strip()
        if not contractor_id:
            continue
        best_match = None
        best_score = 0

        for onedrive_claim in valid_onedrive:
            if onedrive_claim['folder_id'] in matched_onedrive:
                continue
            score = calculate_match_score(contractor_id, onedrive_claim.get('folder_name', '').strip())
            if score > best_score:
                best_score = score
                best_match = onedrive_claim

        if best_match and best_score >= MATCH_THRESHOLD:
            confidence_level = "High" if best_score >= 0.8 else "Medium" if best_score >= 0.65 else "Low"
            results['matched_pairs'].append({
                'encircle': encircle_claim,
                'onedrive': best_match,
                'match_type': f'Fuzzy Match ({confidence_level})',
                'confidence': f'{int(best_score * 100)}%'
            })
            matched_encircle.add(encircle_claim['id'])
            matched_onedrive.add(best_match['folder_id'])
            results['summary']['matches'] += 1
            if best_score >= 0.8:
                results['summary']['match_breakdown']['high_confidence'] += 1
            elif best_score >= 0.65:
                results['summary']['match_breakdown']['medium_confidence'] += 1
            else:
                results['summary']['match_breakdown']['low_confidence'] += 1

    for encircle_claim in valid_encircle:
        if encircle_claim['id'] not in matched_encircle:
            results['encircle_missing_onedrive'].append(encircle_claim)
            results['summary']['encircle_only'] += 1

    for onedrive_claim in valid_onedrive:
        if onedrive_claim['folder_id'] not in matched_onedrive:
            results['onedrive_extra'].append(onedrive_claim)
            results['summary']['onedrive_only'] += 1

    return results
