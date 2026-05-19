"""
Management command to import client data from 01-INFO Excel files in OneDrive.
Reads from Documents/01-current-ins-estimates/<client folder>/01-INFO-<name>.xlsx
Specifically imports the jobinfo(2) tab, matching Column B labels to Column C data.

Usage:
    python manage.py import_from_excel
    python manage.py import_from_excel --folder-name "John Doe@123 Main St"
    python manage.py import_from_excel --dry-run
"""

from django.core.management.base import Command
from django.db import transaction
from docsAppR.models import Client
from docsAppR.file_manager import OneDriveManager  # TODO: Rename class
import openpyxl
from io import BytesIO
import re


# Field mapping from Excel labels (Column B) to Django model fields
FIELD_MAPPING = {
    # Customer fields
    'Property Owner': 'pOwner',
    'Property Address': 'pAddress',
    'City, State, ZIP': 'pCityStateZip',
    'Email': 'cEmail',
    'Phone': 'cPhone',
    'Co-Owner': 'coOwner2',
    'Co-Owner Phone': 'cPhone2',
    'Co-Owner Address': 'cAddress2',
    'Co-Owner City/State/ZIP': 'cCityStateZip2',
    'Co-Owner Email': 'cEmail2',

    # Claim fields
    'Cause of Loss': 'causeOfLoss',
    'Date of Loss': 'dateOfLoss',
    'Claim Number': 'claimNumber',
    'Policy Number': 'policyNumber',

    # Insurance fields
    'Insurance Company': 'insuranceCo_Name',
    'Insurance Address': 'insAddressOvernightMail',
    'Insurance City/State/ZIP': 'insCityStateZip',
    'Desk Adjuster': 'deskAdjusterDA',
    'DA Email': 'DAEmail',
    'DA Phone': 'DAPhone',
    'Field Adjuster': 'fieldAdjusterName',
    'FA Email': 'fieldAdjEmail',
    'FA Phone': 'phoneFieldAdj',

    # Mortgage fields
    'Mortgage Company': 'mortgageCo',
    'Mortgage Account': 'mortgageAccountCo',
    'Mortgage Contact': 'mortgageContactPerson',
    'Mortgage Email': 'mortgageEmail',
    'Mortgage Phone': 'mortgagePhoneContact',

    # Contractor fields
    'Contractor Name': 'coName',
    'Contractor Website': 'coWebsite',
    'Contractor Address': 'coAddress',
    'Contractor City/State': 'coCityState',
    'Contractor Phone': 'coRepPH',
    'Contractor Email': 'coREPEmail',

    # ALE fields
    'Loss of Use/ALE': 'lossOfUseALE',
    'Tenant/Lessee': 'ale_lessee_name',
}


class Command(BaseCommand):
    help = 'Import client data from 01-INFO Excel files in OneDrive'

    def add_arguments(self, parser):
        parser.add_argument(
            '--folder-name',
            type=str,
            help='Specific folder to import (e.g., "John Doe@123 Main St")',
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Show what would be imported without making changes',
        )
        parser.add_argument(
            '--force',
            action='store_true',
            help='Overwrite existing client data',
        )

    def handle(self, *args, **options):
        folder_name = options.get('folder_name')
        dry_run = options.get('dry_run', False)
        force = options.get('force', False)

        self.stdout.write(self.style.SUCCESS('=' * 70))
        self.stdout.write(self.style.SUCCESS('01-INFO Excel Import'))
        self.stdout.write(self.style.SUCCESS('=' * 70))

        if dry_run:
            self.stdout.write(self.style.WARNING('DRY RUN MODE - No changes will be made'))

        try:
            # Initialize OneDrive manager
            self.stdout.write('Initializing OneDrive manager...')
            manager = OneDriveManager()

            # Get Documents folder ID
            docs_folder_id = manager.get_documents_folder()
            if not docs_folder_id:
                raise Exception("Documents folder not found in OneDrive")

            # Navigate to 01-current-ins-estimates
            estimates_folder = self.find_folder(manager, docs_folder_id, '01-current-ins-estimates')
            if not estimates_folder:
                raise Exception("01-current-ins-estimates folder not found in Documents/")

            if folder_name:
                # Import specific folder
                self.import_folder(manager, estimates_folder['id'], folder_name, dry_run, force)
            else:
                # Import all folders
                self.import_all_folders(manager, estimates_folder['id'], dry_run, force)

            self.stdout.write(self.style.SUCCESS('\n' + '=' * 70))
            self.stdout.write(self.style.SUCCESS('Import completed successfully!'))
            self.stdout.write(self.style.SUCCESS('=' * 70))

        except Exception as e:
            self.stdout.write(self.style.ERROR(f'\nError: {str(e)}'))
            raise

    def find_folder(self, manager, parent_id, folder_name):
        """Find a folder by name within parent"""
        items = manager.list_folder_contents(parent_id)
        for item in items:
            if item.get('folder') and item.get('name') == folder_name:
                return item
        return None

    def import_all_folders(self, manager, parent_id, dry_run, force):
        """Import all client folders"""
        self.stdout.write('\nScanning for client folders...')

        items = manager.list_folder_contents(parent_id)
        client_folders = [item for item in items if item.get('folder')]

        self.stdout.write(f'Found {len(client_folders)} folders')

        for folder in client_folders:
            folder_name = folder.get('name')
            folder_id = folder.get('id')

            self.stdout.write(f'\n{"-" * 70}')
            self.stdout.write(f'Processing: {folder_name}')

            try:
                self.import_folder(manager, parent_id, folder_name, dry_run, force)
            except Exception as e:
                self.stdout.write(self.style.ERROR(f'Error importing {folder_name}: {str(e)}'))
                continue

    def import_folder(self, manager, parent_id, folder_name, dry_run, force):
        """Import a specific client folder"""
        # Find the folder
        folder = self.find_folder(manager, parent_id, folder_name)
        if not folder:
            self.stdout.write(self.style.WARNING(f'Folder not found: {folder_name}'))
            return

        folder_id = folder['id']

        # Find the 01-INFO Excel file
        excel_file = self.find_info_file(manager, folder_id, folder_name)
        if not excel_file:
            self.stdout.write(self.style.WARNING(f'01-INFO file not found in {folder_name}'))
            return

        self.stdout.write(f'Found Excel file: {excel_file["name"]}')

        # Download the file
        file_content = manager.download_file(excel_file['id'])
        if not file_content:
            self.stdout.write(self.style.ERROR('Failed to download file'))
            return

        # Parse the Excel file
        data = self.parse_excel(file_content)
        if not data:
            self.stdout.write(self.style.WARNING('No data extracted from Excel'))
            return

        self.stdout.write(f'Extracted {len(data)} fields from Excel')

        # Check if client exists
        client = self.find_existing_client(folder_name, data)

        if client and not force:
            self.stdout.write(self.style.WARNING(f'Client already exists: {client.pOwner}'))
            self.stdout.write('Use --force to overwrite existing data')
            return

        # Import/update the client
        if not dry_run:
            if client:
                self.update_client(client, data)
                self.stdout.write(self.style.SUCCESS(f'Updated client: {client.pOwner}'))
            else:
                client = self.create_client(data)
                self.stdout.write(self.style.SUCCESS(f'Created client: {client.pOwner}'))
        else:
            action = 'Update' if client else 'Create'
            self.stdout.write(f'[DRY RUN] Would {action} client with {len(data)} fields')

    def find_info_file(self, manager, folder_id, folder_name):
        """Find the 01-INFO Excel file in the folder"""
        items = manager.list_folder_contents(folder_id)

        # Look for files starting with "01-INFO" and ending with .xlsx
        for item in items:
            if item.get('file'):
                name = item.get('name', '')
                if name.startswith('01-INFO') and name.endswith('.xlsx'):
                    return item
        return None

    def parse_excel(self, file_content):
        """Parse the Excel file and extract data from jobinfo(2) tab"""
        data = {}

        try:
            # Load workbook
            wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)

            # Find the jobinfo(2) sheet
            sheet = None
            for sheet_name in wb.sheetnames:
                if 'jobinfo' in sheet_name.lower():
                    sheet = wb[sheet_name]
                    break

            if not sheet:
                self.stdout.write(self.style.WARNING('jobinfo sheet not found'))
                return data

            # Read data from columns B (labels) and C (values)
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
                label_cell = row[1]  # Column B
                value_cell = row[2]  # Column C

                if label_cell.value and value_cell.value:
                    label = str(label_cell.value).strip()
                    value = value_cell.value

                    # Map to Django field
                    field_name = FIELD_MAPPING.get(label)
                    if field_name:
                        # Clean and format value
                        if isinstance(value, str):
                            value = value.strip()
                        data[field_name] = value

            return data

        except Exception as e:
            self.stdout.write(self.style.ERROR(f'Error parsing Excel: {str(e)}'))
            return {}

    def find_existing_client(self, folder_name, data):
        """Find existing client by folder name or data"""
        # Try to extract owner name from folder (format: "Name@Address")
        if '@' in folder_name:
            owner_name = folder_name.split('@')[0].strip()
            clients = Client.objects.filter(pOwner__icontains=owner_name)
            if clients.count() == 1:
                return clients.first()

        # Try by claim number if available
        if 'claimNumber' in data:
            clients = Client.objects.filter(claimNumber=data['claimNumber'])
            if clients.exists():
                return clients.first()

        return None

    def create_client(self, data):
        """Create a new client from parsed data"""
        with transaction.atomic():
            client = Client.objects.create(**data)
            # OneDrive sync status removed
            client.save()
            return client

    def update_client(self, client, data):
        """Update existing client with parsed data"""
        with transaction.atomic():
            for field, value in data.items():
                setattr(client, field, value)
            client.save()
            return client
