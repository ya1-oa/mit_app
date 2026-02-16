"""
Django Imports
"""
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.core import serializers
from django.core.cache import cache
from django.core.files.base import ContentFile
from django.core.files.storage import FileSystemStorage, default_storage
from django.core.mail import EmailMessage
from django.core.management import call_command
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Avg, Case, Count, F, Q, When, IntegerField
from django.http import HttpResponse, JsonResponse, HttpRequest
from django.shortcuts import get_object_or_404, redirect, render
from django.template import Template, Context
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.views import View
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET

"""
Third Party App Imports
"""
from allauth.account.decorators import login_required

"""
Project Specific Imports
"""
from .config.excel_mappings import SCOPE_FORM_MAPPINGS
from .forms import ClientForm, CreateUserForm, UploadClientForm, UploadFilesForm, LandlordForm, EmailForm, EmailScheduleForm
from .models import ChecklistItem, Client, File, Document, Landlord, SentEmail, EmailSchedule, EmailOpenEvent, DocumentCategory, ReadingImage, Room, WorkType, RoomWorkTypeValue, Lease, LeaseDocument, LeaseActivity
from automations.tasks import RoomTemplateAutomation

"""
Python Standard Library
"""
import base64
import csv
import datetime as dt
import io
import json
import logging
import math
import os
import platform
import re
import shutil
import subprocess
import tempfile
import time
import zipfile
from collections import defaultdict
from decimal import Decimal, InvalidOperation
from difflib import SequenceMatcher
from io import BytesIO
from pathlib import Path
from urllib.parse import quote

"""
Third Party Libraries
"""
import openpyxl
import pandas as pd
import requests
from dateutil.parser import parse as parse_date
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from weasyprint import HTML
from xhtml2pdf import pisa

# Configure logging
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)


class EncircleExcelExporter:
    """
    Class to export Encircle claims data to an Excel workbook with multiple worksheets
    """
    
    def __init__(self):
        self.wb = Workbook()
        self.default_sheet = self.wb.active
        self.default_sheet.title = "Claims Summary"
        self.styles = self._define_styles()
    
    def _define_styles(self):
        """Define reusable cell styles"""
        return {
            'header': {
                'font': Font(bold=True, color="FFFFFF"),
                'fill': PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid"),
                'alignment': Alignment(horizontal="center", vertical="center", wrap_text=True),
                'border': Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
            },
            'data': {
                'alignment': Alignment(vertical="center", wrap_text=True),
                'border': Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
            }
        }
    
    def _apply_style(self, cell, style_name):
        """Apply a predefined style to a cell"""
        if style_name in self.styles:
            style = self.styles[style_name]
            if 'font' in style:
                cell.font = style['font']
            if 'fill' in style:
                cell.fill = style['fill']
            if 'alignment' in style:
                cell.alignment = style['alignment']
            if 'border' in style:
                cell.border = style['border']
    
    def _auto_adjust_columns(self, sheet):
        """Auto-adjust column widths based on content"""
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    cell_value = cell.value
                    if cell_value is not None:
                        cell_length = len(str(cell_value))
                        if cell_length > max_length:
                            max_length = cell_length
                except Exception:
                    pass
            
            # Set minimum width and maximum width limits
            adjusted_width = min(max(max_length + 2, 10), 50) * 1.2
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _safe_cell_value(self, value):
        """Safely convert value to Excel-compatible format"""
        if value is None:
            return 'N/A'
        elif isinstance(value, (list, tuple)):
            return ", ".join(str(item) for item in value)
        elif isinstance(value, dict):
            return str(value)
        else:
            return str(value)
    
    def export_claims(self, claims_data):
        """
        Main method to export all claims data to Excel
        """
        try:
            # Validate input data
            if not isinstance(claims_data, dict):
                raise ValueError("claims_data must be a dictionary")
            
            # Create worksheets
            claims = claims_data.get('claims', [])
            if claims:
                self._create_claims_sheet(claims)
            
            # If we have detailed claim data, create additional sheets
            if 'claim_details' in claims_data and claims_data['claim_details']:
                self._create_details_sheet(claims_data['claim_details'])
                
            if 'rooms' in claims_data and claims_data['rooms']:
                self._create_rooms_sheet(claims_data['rooms'])
                
            if 'floor_plan' in claims_data and claims_data['floor_plan']:
                self._create_floor_plan_sheet(claims_data['floor_plan'])
            
            # Remove default sheet if we created other sheets and it's empty
            if len(self.wb.worksheets) > 1:
                claims_sheet = self.wb["Claims Summary"]
                if claims_sheet.max_row == 1:  # Only headers
                    self.wb.remove(claims_sheet)
            
            # Generate timestamp for filename
            timestamp = dt.datetime().strftime("%Y%m%d_%H%M%S")
            filename = f"encircle_claims_export_{timestamp}.xlsx"
            
            # Save workbook to a BytesIO object
            output = BytesIO()
            self.wb.save(output)
            output.seek(0)
            
            return output, filename
            
        except Exception as e:
            raise Exception(f"Error generating Excel file: {str(e)}")
    
    def _create_claims_sheet(self, claims):
    
        """Create a vertical claims worksheet matching the exact template format."""
        sheet = self.wb["Claims Summary"]
        
        # Define the exact row structure based on your template
        template_rows = [
            "Insured Name",
            "Address",
            "City, State",
            "Email",
            "Phone",
            "TBD",  # Placeholder 1
            "TBD",  # Placeholder 2
            "TBD",  # Placeholder 3
            "TBD",  # Placeholder 4
            "TBD",  # Placeholder 5
            "Cause of Loss",
            "Date of Loss",
            "Type of Loss",
            "TBD",  # Placeholder 6
            "TBD",  # Placeholder 7
            "Coverage A Dwelling",
            "Coverage B Other Structures",
            "Coverage C Personal Property",
            "Coverage D Loss of Use",
            "Coverage E Liability",
            "Coverage F Medical",
            "Insurance Company",
            "Policy Number",
            "Claim Number",
            "Adjuster Email",
            "Adjuster Name",
            "Adjuster Phone",
            "Insurance Company Phone",
            "Insurance Company Fax",
            "Claims Email",
            "TBD",  # Placeholder 8
            "TBD",  # Placeholder 9
            "TBD",  # Placeholder 10
            "TBD",  # Placeholder 11
            "TBD",  # Placeholder 12
            "TBD",  # Placeholder 13
            "TBD",  # Placeholder 14
            "TBD",  # Placeholder 15
            "TBD",  # Placeholder 16
            "TBD",  # Placeholder 17
            "TBD",  # Placeholder 18
            "TBD",  # Placeholder 19
            "TBD",  # Placeholder 20
            "TBD",  # Placeholder 21
            "TBD",  # Placeholder 22
            "TBD",  # Placeholder 23
            "TBD",  # Placeholder 24
            "TBD",  # Placeholder 25
            "TBD",  # Placeholder 26
            "TBD",  # Placeholder 27
            "TBD",  # Placeholder 28
            "TBD",  # Placeholder 29
            "TBD",  # Placeholder 30
            "TBD",  # Placeholder 31
            "TBD",  # Placeholder 32
            "TBD",  # Placeholder 33
            "TBD",  # Placeholder 34
            "TBD",  # Placeholder 35
            "TBD",  # Placeholder 36
            "TBD",  # Placeholder 37
            "TBD",  # Placeholder 38
            "TBD",  # Placeholder 39
            "TBD",  # Placeholder 40
            "TBD",  # Placeholder 41
            "TBD",  # Placeholder 42
            "TBD",  # Placeholder 43
            "TBD",  # Placeholder 44
            "TBD",  # Placeholder 45
            "TBD",  # Placeholder 46
            "TBD",  # Placeholder 47
            "TBD",  # Placeholder 48
            "TBD",  # Placeholder 49
            "TBD",  # Placeholder 50
            "TBD",  # Placeholder 51
            "TBD",  # Placeholder 52
            "TBD",  # Placeholder 53
            "TBD",  # Placeholder 54
            "TBD",  # Placeholder 55
            "TBD",  # Placeholder 56
            "TBD",  # Placeholder 57
            "TBD",  # Placeholder 58
            "TBD",  # Placeholder 59
            "TBD",  # Placeholder 60
            "TBD",  # Placeholder 61
            "TBD",  # Placeholder 62
            "TBD",  # Placeholder 63
            "TBD",  # Placeholder 64
            "TBD",  # Placeholder 65
            "TBD",  # Placeholder 66
            "TBD",  # Placeholder 67
            "TBD",  # Placeholder 68
            "TBD",  # Placeholder 69
            "TBD",  # Placeholder 70
            "TBD",  # Placeholder 71
            "TBD",  # Placeholder 72
            "TBD",  # Placeholder 73
            "TBD",  # Placeholder 74
            "TBD",  # Placeholder 75
            "TBD",  # Placeholder 76
            "TBD",  # Placeholder 77
            "TBD",  # Placeholder 78
            "TBD",  # Placeholder 79
            "TBD",  # Placeholder 80
            "TBD",  # Placeholder 81
            "TBD",  # Placeholder 82
            "TBD",  # Placeholder 83
            "TBD",  # Placeholder 84
            "TBD",  # Placeholder 85
            "TBD",  # Placeholder 86
            "TBD",  # Placeholder 87
            "TBD",  # Placeholder 88
            "TBD",  # Placeholder 89
            "TBD",  # Placeholder 90
            "TBD",  # Placeholder 91
            "TBD",  # Placeholder 92
            "TBD",  # Placeholder 93
            "TBD",  # Placeholder 94
            "TBD",  # Placeholder 95
            "TBD",  # Placeholder 96
            "TBD",  # Placeholder 97
            "TBD",  # Placeholder 98
            "TBD",  # Placeholder 99
            "TBD",  # Placeholder 100
            "TBD",  # Placeholder 101
            "TBD",  # Placeholder 102
            "TBD",  # Placeholder 103
            "TBD",  # Placeholder 104
            "TBD",  # Placeholder 105
            "TBD",  # Placeholder 106
            "TBD",  # Placeholder 107
            "TBD",  # Placeholder 108
            "TBD",  # Placeholder 109
            "TBD",  # Placeholder 110
            "TBD",  # Placeholder 111
            "TBD",  # Placeholder 112
            "TBD",  # Placeholder 113
            "TBD",  # Placeholder 114
            "TBD",  # Placeholder 115
            "TBD",  # Placeholder 116
            "TBD",  # Placeholder 117
            "TBD",  # Placeholder 118
            "TBD",  # Placeholder 119
            "TBD",  # Placeholder 120
            "TBD",  # Placeholder 121
            "TBD",  # Placeholder 122
            "TBD",  # Placeholder 123
            "TBD",  # Placeholder 124
            "TBD",  # Placeholder 125
            "TBD",  # Placeholder 126
            "TBD",  # Placeholder 127
            "TBD",  # Placeholder 128
            "TBD",  # Placeholder 129
            "TBD",  # Placeholder 130
            "TBD",  # Placeholder 131
            "TBD",  # Placeholder 132
            "TBD",  # Placeholder 133
            "TBD",  # Placeholder 134
            "TBD",  # Placeholder 135
            "TBD",  # Placeholder 136
            "TBD",  # Placeholder 137
            "TBD",  # Placeholder 138
            "TBD",  # Placeholder 139
            "TBD",  # Placeholder 140
            "TBD",  # Placeholder 141
            "TBD",  # Placeholder 142
            "TBD",  # Placeholder 143
            "TBD",  # Placeholder 144
            "TBD",  # Placeholder 145
            "TBD",  # Placeholder 146
            "TBD",  # Placeholder 147
            "TBD",  # Placeholder 148
            "TBD",  # Placeholder 149
            "TBD",  # Placeholder 150
            "TBD",  # Placeholder 151
            "TBD",  # Placeholder 152
            "TBD",  # Placeholder 153
            "TBD",  # Placeholder 154
            "TBD",  # Placeholder 155
            "TBD",  # Placeholder 156
            "TBD",  # Placeholder 157
            "TBD",  # Placeholder 158
            "TBD",  # Placeholder 159
            "TBD",  # Placeholder 160
            "TBD",  # Placeholder 161
            "TBD",  # Placeholder 162
            "TBD",  # Placeholder 163
            "TBD",  # Placeholder 164
            "TBD",  # Placeholder 165
            "TBD",  # Placeholder 166
            "TBD",  # Placeholder 167
            "TBD",  # Placeholder 168
            "TBD",  # Placeholder 169
            "TBD",  # Placeholder 170
            "TBD",  # Placeholder 171
            "TBD",  # Placeholder 172
            "TBD",  # Placeholder 173
            "TBD",  # Placeholder 174
            "TBD",  # Placeholder 175
            "TBD",  # Placeholder 176
            "TBD",  # Placeholder 177
            "TBD",  # Placeholder 178
            "TBD",  # Placeholder 179
            "TBD",  # Placeholder 180
            "TBD",  # Placeholder 181
            "TBD",  # Placeholder 182
            "TBD",  # Placeholder 183
            "TBD",  # Placeholder 184
            "TBD",  # Placeholder 185
            "TBD",  # Placeholder 186
            "TBD",  # Placeholder 187
            "TBD",  # Placeholder 188
            "TBD",  # Placeholder 189
            "TBD",  # Placeholder 190
            "TBD",  # Placeholder 191
            "TBD",  # Placeholder 192
            "TBD",  # Placeholder 193
            "TBD",  # Placeholder 194
            "TBD",  # Placeholder 195
            "TBD",  # Placeholder 196
            "TBD",  # Placeholder 197
            "TBD",  # Placeholder 198
            "TBD",  # Placeholder 199
            "TBD",  # Placeholder 200
        ]
    
        # Write template headers in column A
        for row_num, header in enumerate(template_rows, start=1):
            sheet.cell(row=row_num, column=1, value=header)
        
        # Write claim data starting from column B
        for col_num, claim in enumerate(claims, start=2):
            if not isinstance(claim, dict):
                continue
                
            # Map claim data to specific rows based on your template
            data_mapping = {
                1: claim.get('policyholder_name', 'TBD'),  # Insured Name
                2: claim.get('full_address', 'TBD'),       # Address
                3: f"{claim.get('city', '')}, {claim.get('state', '')}".strip(', '),  # City, State
                4: claim.get('email', 'TBD'),              # Email
                5: claim.get('phone', 'TBD'),             # Phone
                11: claim.get('cause_of_loss', 'TBD'),    # Cause of Loss
                12: claim.get('date_of_loss', 'TBD'),     # Date of Loss
                22: claim.get('insurance_company_name', 'TBD'),  # Insurance Company
                23: claim.get('policy_number', 'TBD'),    # Policy Number
                24: claim.get('claim_number', 'TBD'),     # Claim Number
                25: claim.get('adjuster_email', 'TBD'),   # Adjuster Email
                26: claim.get('adjuster_name', 'TBD'),    # Adjuster Name
                27: claim.get('adjuster_phone', 'TBD'),  # Adjuster Phone
                # Add more mappings as needed for your specific template
            }
            
            # Write the data to the appropriate rows
            for row_num, value in data_mapping.items():
                safe_value = self._safe_cell_value(value)
                sheet.cell(row=row_num, column=col_num, value=safe_value)
        
        # Adjust column widths for better readability
        self._auto_adjust_columns(sheet)

    def _create_details_sheet(self, claim_details):
        """Create a worksheet for detailed claim information"""
        if not isinstance(claim_details, dict):
            return
            
        sheet = self.wb.create_sheet("Claim Details")
        
        # Group details into logical sections
        sections = {
            "Basic Information": [
                ("Claim ID", claim_details.get('id')),
                ("Brand ID", claim_details.get('brand_id')),
                ("Organization ID", claim_details.get('organization_id')),
                ("Date Created", claim_details.get('date_claim_created')),
                ("Permalink URL", claim_details.get('permalink_url'))
            ],
            "Property Information": [
                ("Full Address", claim_details.get('full_address')),
                ("Policyholder", claim_details.get('policyholder_name')),
                ("Phone", claim_details.get('policyholder_phone_number')),
                ("Email", claim_details.get('policyholder_email_address'))
            ],
            "Insurance Information": [
                ("Insurance Company", claim_details.get('insurance_company_name')),
                ("Policy Number", claim_details.get('policy_number')),
                ("Insurer ID", claim_details.get('insurer_identifier')),
                ("Broker/Agent", claim_details.get('broker_or_agent_name'))
            ],
            "Loss Information": [
                ("Date of Loss", claim_details.get('date_of_loss')),
                ("Type of Loss", claim_details.get('type_of_loss')),
                ("Loss Details", claim_details.get('loss_details')),
                ("CAT Code", claim_details.get('cat_code'))
            ],
            "Adjuster Information": [
                ("Adjuster Name", claim_details.get('adjuster_name')),
                ("Project Manager", claim_details.get('project_manager_name')),
                ("Contractor ID", claim_details.get('contractor_identifier')),
                ("Assignment ID", claim_details.get('assignment_identifier'))
            ],
            "Financial Information": [
                ("Repair Estimate", claim_details.get('repair_estimate')),
                ("Contents Estimate", claim_details.get('contents_estimate')),
                ("Emergency Estimate", claim_details.get('emergency_estimate')),
                ("Sales Tax", claim_details.get('sales_tax')),
                ("Default Depreciation", claim_details.get('default_depreciation')),
                ("Max Depreciation", claim_details.get('max_depreciation'))
            ]
        }
        
        row_num = 1
        for section_name, items in sections.items():
            # Write section header
            if sheet.max_column >= 2:
                sheet.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
            cell = sheet.cell(row=row_num, column=1, value=section_name)
            cell.font = Font(bold=True, size=14)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            row_num += 1
            
            # Write section items
            for label, value in items:
                label_cell = sheet.cell(row=row_num, column=1, value=label)
                label_cell.font = Font(bold=True)
                
                value_cell = sheet.cell(row=row_num, column=2, value=self._safe_cell_value(value))
                row_num += 1
            
            # Add empty row between sections
            row_num += 1
        
        self._auto_adjust_columns(sheet)
    
    def _create_rooms_sheet(self, rooms):
        """Create a worksheet for room data"""
        if not isinstance(rooms, list) or not rooms:
            return
            
        sheet = self.wb.create_sheet("Rooms")
        
        headers = [
            "Claim ID", "Claim Name", "Room Name", "Type", "Floor", 
            "Ceiling Height", "Length", "Width",
            "Area", "Perimeter", "Damage Present",
            "Contents Affected"
        ]
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            self._apply_style(cell, 'header')
        
        # Write room data
        for row_num, room in enumerate(rooms, 2):
            if not isinstance(room, dict):
                continue
                
            row_data = [
                room.get('claim_id'),
                room.get('claim_name'),
                room.get('name'),
                room.get('room_type'),
                room.get('floor_name'),
                room.get('ceiling_height'),
                room.get('length'),
                room.get('width'),
                room.get('area'),
                room.get('perimeter'),
                "Yes" if room.get('damage_present') else "No",
                "Yes" if room.get('contents_affected') else "No"
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = sheet.cell(row=row_num, column=col_num, value=self._safe_cell_value(value))
                self._apply_style(cell, 'data')
        
        self._auto_adjust_columns(sheet)
    
    def _create_floor_plan_sheet(self, floor_plan):
        """Create a worksheet for floor plan dimensional data"""
        if not isinstance(floor_plan, dict) or not floor_plan:
            return
            
        sheet = self.wb.create_sheet("Floor Plans")
        
        headers = [
            "Floor", "Room Name", "Primary Length", 
            "Primary Width", "Bounding Width", "Bounding Height",
            "Area", "Ceiling Height"
        ]
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            self._apply_style(cell, 'header')
        
        row_num = 2
        for floor_name, rooms in floor_plan.items():
            if not isinstance(rooms, dict):
                continue
                
            for room_name, dimensions in rooms.items():
                if not isinstance(dimensions, dict):
                    continue
                    
                primary_dims = dimensions.get('primary_dimensions', {})
                bounding_box = dimensions.get('bounding_box', {})
                
                row_data = [
                    floor_name.replace('Floor_', 'Floor '),
                    room_name,
                    primary_dims.get('length') if isinstance(primary_dims, dict) else None,
                    primary_dims.get('width') if isinstance(primary_dims, dict) else None,
                    bounding_box.get('width') if isinstance(bounding_box, dict) else None,
                    bounding_box.get('height') if isinstance(bounding_box, dict) else None,
                    dimensions.get('area'),
                    dimensions.get('ceiling_height')
                ]
                
                for col_num, value in enumerate(row_data, 1):
                    cell = sheet.cell(row=row_num, column=col_num, value=self._safe_cell_value(value))
                    self._apply_style(cell, 'data')
                    
                row_num += 1
        
        self._auto_adjust_columns(sheet)

class EncircleAPIClient:
    """
    Centralized API client for Encircle API operations
    Following Single Responsibility Principle
    """
    
    def __init__(self):
        self.api_key = "367382d2-0b2d-4b01-9d06-8f18fd492f5e"
        self.base_url = "https://api.encircleapp.com/v1"
        self.v2_base_url = "https://api.encircleapp.com/v2"
        self.headers = {"Authorization": f"Bearer {self.api_key}"}
        
    def _make_request(self, endpoint, params=None):
        """
        Generic method to make API requests with error handling
        """
        try:
            url = f"{self.base_url}/{endpoint}"
            response = requests.get(url, headers=self.headers, params=params)
            response.raise_for_status()
            print(response.json())
            return response.json()
        except requests.exceptions.RequestException as e:
            logger.error(f"API request failed for {endpoint}: {str(e)}")
            raise Exception(f"API request failed: {str(e)}")

    def _make_v2_request(self, endpoint, params=None):
        """
        Generic method to make API requests with error handling
        """
        try:
            url = f"{self.v2_base_url}/{endpoint}"
            response = requests.get(url, headers=self.headers, params=params)

            if response.status_code != 200:
                logger.error(f"API request failed for {endpoint}: {response.status_code} - {response.text}")
                raise Exception(f"API request failed: {response.status_code} - {response.text}")
            return response.json()
        except requests.exceptions.RequestException as e:
            logger.error(f"API request failed for {endpoint}: {str(e)}")
            raise Exception(f"API request failed: {str(e)}")

    def get_all_claims(self, limit=100, order='newest'):
        """
        Fetch all property claims using cursor-based pagination with 'after' parameter
        
        Args:
            limit (int): Number of results per page (max 100, default 100)
            order (str): Order of results ('newest' or 'oldest', default 'newest')
        
        Returns:
            list: All property claims across all pages
        """
        all_claims = {'list': []}
        after_cursor = None
        page_count = 0
        
        while True:
            page_count += 1
            
            # Build parameters for the request
            params = {
                'limit': min(limit, 100),  # Ensure we don't exceed the max limit of 100
                'order': order
            }
            
            # Add 'after' parameter if we have a cursor (for subsequent pages)
            if after_cursor:
                params['after'] = after_cursor
            
            try:
                print(f"Fetching page {page_count} of claims (limit: {params['limit']})")
                response = self._make_request("property_claims", params=params)
                
                claims = response.get('list', [])
                # Look for the 'after' cursor in the response
                after_cursor = response.get('cursor')['after']
                
                if not claims:
                    print("No more claims found")
                    break
                
                all_claims['list'].extend(claims)
                print(f"Retrieved {len(claims)} claims from page {page_count}. Total so far: {len(all_claims)}")
                
                # If there's no 'after' cursor, we've reached the last page
                if not after_cursor or len(claims) < params['limit']:
                    if not after_cursor:
                        print("No 'after' cursor found - reached last page")
                    else:
                        print(f"Retrieved {len(claims)} claims (less than limit of {params['limit']}) - likely last page")
                    break
                
            except Exception as e:
                print(f"Error fetching page {page_count}: {str(e)}")
                break
        
        print(f"Successfully retrieved {len(all_claims)} total claims across {page_count} pages") 
        return all_claims
    
    def get_claim_details(self, claim_id):
        """
        Fetch detailed information for a specific claim
        """
        return self._make_request(f"property_claims/{claim_id}")
    

    def get_claim_structures(self, claim_id):
        """
        Fetch structures for a specific claim
        """
        return self._make_request(f"property_claims/{claim_id}/structures")
    
    
    def get_claim_rooms(self, claim_id, structure_id):
        """
        Fetch room list for a specific claim
        """
        return self._make_request(f"property_claims/{claim_id}/structures/{structure_id}/rooms")
    
    def get_claim_floor_plan(self, claim_id):
        """
        Fetch floor plan dimensions for a specific claim
        """

        return self._make_v2_request(f"property_claims/{claim_id}/floor_plan_dimensions")

    # ── Write methods (GreenField / create) ───────────────────────────────────

    def _make_post_request(self, endpoint, payload):
        """Generic POST to v1 API."""
        try:
            url = f"{self.base_url}/{endpoint}"
            headers = {**self.headers, "Content-Type": "application/json"}
            response = requests.post(url, headers=headers, json=payload)
            if not response.ok:
                try:
                    body = response.json()
                except Exception:
                    body = response.text
                logger.error(f"POST request failed for {endpoint}: {response.status_code} — {body}")
                logger.error(f"POST payload was: {payload}")
                response.raise_for_status()
            return response.json()
        except requests.exceptions.RequestException as e:
            logger.error(f"POST request failed for {endpoint}: {str(e)}")
            raise Exception(f"POST request failed: {str(e)}")

    def get_account_ids(self):
        """
        Fetch organization_id and brand_id from the first existing claim.
        Both are per-organization constants required when creating new claims.
        Returns dict with 'organization_id' and 'brand_id' (both may be None).
        """
        result = {'organization_id': None, 'brand_id': None}
        try:
            resp = self._make_request("property_claims", params={"limit": 1})
            claims = resp.get("list", []) if isinstance(resp, dict) else []
            if claims:
                c = claims[0]
                result['organization_id'] = str(c['organization_id']) if c.get('organization_id') else None
                result['brand_id'] = str(c['brand_id']) if c.get('brand_id') else None
                logger.info(f"get_account_ids: organization_id={result['organization_id']}, brand_id={result['brand_id']}")
            else:
                logger.warning("get_account_ids: no existing claims to extract IDs from")
        except Exception as exc:
            logger.warning(f"get_account_ids: failed — {exc}", exc_info=True)
        return result

    def _make_patch_request(self, endpoint, payload):
        """Generic PATCH to v1 API."""
        try:
            url = f"{self.base_url}/{endpoint}"
            headers = {**self.headers, "Content-Type": "application/json"}
            response = requests.patch(url, headers=headers, json=payload)
            response.raise_for_status()
            return response.json()
        except requests.exceptions.RequestException as e:
            logger.error(f"PATCH request failed for {endpoint}: {str(e)}")
            raise Exception(f"PATCH request failed: {str(e)}")

    def create_claim(self, claim_data):
        """
        Create a new property claim in Encircle.
        Automatically injects brand_id (required by API) and ensures
        type_of_loss is set (required by API).
        Returns the newly created claim dict (includes 'id').
        """
        payload = dict(claim_data)
        # organization_id and brand_id are required — fetch from an existing claim
        if not payload.get('organization_id') or not payload.get('brand_id'):
            ids = self.get_account_ids()
            if ids['organization_id'] and not payload.get('organization_id'):
                payload['organization_id'] = ids['organization_id']
            if ids['brand_id'] and not payload.get('brand_id'):
                payload['brand_id'] = ids['brand_id']
            if not payload.get('organization_id'):
                logger.error("create_claim: organization_id could not be resolved — request will likely fail")
        # type_of_loss is required — default to 'Other' if not provided
        if not payload.get('type_of_loss'):
            payload['type_of_loss'] = 'Other'
        logger.info(f"create_claim: final payload = {payload}")
        return self._make_post_request("property_claims", payload)

    def get_or_create_default_structure(self, encircle_claim_id):
        """
        Return the first structure on a claim (Encircle always creates a
        'Main Building' structure by default).  Falls back to creating one
        if none exist.
        Returns a structure dict (includes 'id').
        """
        structures_resp = self.get_claim_structures(encircle_claim_id)
        structures = structures_resp.get("list", structures_resp) if isinstance(structures_resp, dict) else structures_resp
        if structures:
            return structures[0]
        # No structures yet – create one
        return self._make_post_request(
            f"property_claims/{encircle_claim_id}/structures",
            {"name": "Main Building"}
        )

    def create_room(self, encircle_claim_id, structure_id, room_payload):
        """
        Create a room inside a structure.
        room_payload keys: name (required), description (optional)
        Returns the newly created room dict.
        """
        return self._make_post_request(
            f"property_claims/{encircle_claim_id}/structures/{structure_id}/rooms",
            room_payload
        )

class EncircleDataProcessor:
    """
    Data processing utilities for Encircle API responses
    Following Single Responsibility Principle
    """
    
    @staticmethod
    def _get_polygon_dimensions(coordinates):
        """
        Calculate width and length from polygon coordinates.
        Assumes rectangular rooms and finds the bounding box dimensions.
        """
        # Extract all x and y coordinates
        points = coordinates[0]  # First ring of the polygon
        x_coords = [point[0] for point in points]
        y_coords = [point[1] for point in points]
        
        # Calculate bounding box dimensions
        max_x_coords = max(x_coords) / 30.48
        min_x_coords = min(x_coords) / 30.48
        max_y_coords = max(y_coords) / 30.48
        min_y_coords = min(y_coords) / 30.48
        
        width = max_x_coords - min_x_coords
        height = max_y_coords - min_y_coords
        
        return width, height
    
    @staticmethod
    def _get_edge_lengths(coordinates):
        """
        Calculate actual edge lengths of the polygon.
        More accurate for irregular shapes.
        """
        points = coordinates[0]  # First ring of the polygon
        edge_lengths = []
        
        for i in range(len(points) - 1):
            x1, y1 = points[i]
            x2, y2 = points[i + 1]
            length = math.sqrt((x2 - x1)**2 + (y2 - y1)**2) / 30.48
            edge_lengths.append(length)
        
        return edge_lengths
    
    @staticmethod
    def process_floor_plan_data(raw_floor_plan_data):
        """
        Process floor plan data to calculate room dimensions
        """
        if not raw_floor_plan_data or 'list' not in raw_floor_plan_data:
            return {}
            
        room_dimensions = {}
        
        # Iterate through all floors
        for floor_idx, floor in enumerate(raw_floor_plan_data['list'][0]['floors']):
            floor_rooms = {}
            
            # Iterate through all features (rooms) in this floor
            for feature in floor['features']:
                room_name = feature['properties']['name']
                coordinates = feature['geometry']['coordinates']
                
                # Calculate dimensions using both methods
                width, height = EncircleDataProcessor._get_polygon_dimensions(coordinates)
                edge_lengths = EncircleDataProcessor._get_edge_lengths(coordinates)
                
                # For rectangular rooms, take the two longest edges as length/width
                #edge_lengths_sorted = sorted(edge_lengths, reverse=True)
                
                floor_rooms[room_name] = {
                    'bounding_box': {
                        'width': round(width, 2),
                        'height': round(height, 2)
                    },
                    #'primary_dimensions': {
                    #    'length': round(edge_lengths_sorted[0], 2) if edge_lengths_sorted else 0,
                    #    'width': round(edge_lengths_sorted[1], 2) if len(edge_lengths_sorted) > 1 else 0
                    #},
                    'ceiling_height': feature['properties'].get('ceiling_height', 'N/A'),
                    'area': round(width * height, 2)  # Add area calculation
                }
            
            room_dimensions[f'Floor_{floor_idx + 1}'] = floor_rooms
        
        return room_dimensions
    
    @staticmethod
    def get_room_dimensions_simple(room_feature):
        """
        Simple function to get width and length for a single room feature.
        
        Args:
            room_feature: Single feature from the JSON (contains geometry and properties)
        
        Returns:
            dict: Dictionary containing width, length, and area
        """
        coordinates = room_feature['geometry']['coordinates'][0]
        
        # Get bounding box
        x_coords = [point[0] for point in coordinates]
        y_coords = [point[1] for point in coordinates]
        
        width = (max(x_coords) - min(x_coords)) / 12
        height = (max(y_coords) - min(y_coords)) / 12
        
        return {
            'width': round(width, 2),
            'length': round(height, 2),
            'area': round(width * height, 2)
        }
    
    @staticmethod
    def process_claims_list(raw_claims_data):
        """
        Process and structure the claims list data to match API response
        """
        print(raw_claims_data)
        if not raw_claims_data or 'list' not in raw_claims_data:
            return []
        
        processed_claims = []
        for claim in raw_claims_data['list']:
            processed_claim = {
                'id': claim.get('id'),
                'brand_id': claim.get('brand_id'),
                'organization_id': claim.get('organization_id'),
                'permalink_url': claim.get('permalink_url'),
                'type_of_loss': claim.get('type_of_loss'),
                'full_address': claim.get('full_address'),
                'date_claim_created': claim.get('date_claim_created'),
                'date_of_loss': claim.get('date_of_loss'),
                'policyholder_name': claim.get('policyholder_name'),
                'insurance_company_name': claim.get('insurance_company_name'),
                'policy_number': claim.get('policy_number'),
                'adjuster_name': claim.get('adjuster_name'),
                'project_manager_name': claim.get('project_manager_name'),
                'total_rooms': 0,  # Will be populated later
                'room_types': []   # Will be populated later
            }
            processed_claims.append(processed_claim)
        
        return processed_claims
    
    @staticmethod
    def process_claim_rooms(raw_rooms_data):
        """
        Process room data for a specific claim
        """
        if not raw_rooms_data or 'list' not in raw_rooms_data:
            return []
        
        processed_rooms = []
        room_types = set()
        
        for room in raw_rooms_data['list']:
            room_data = {
                'id': room.get('id'),
                'name': room.get('name', 'Unnamed Room'),
                'room_type': room.get('room_type', 'Unknown'),
                'floor_name': room.get('floor_name', 'Unknown Floor'),
                'ceiling_height': room.get('ceiling_height', 0),
                'length': room.get('length', 0),
                'width': room.get('width', 0),
                'area': room.get('area', 0),
                'perimeter': room.get('perimeter', 0),
                'damage_present': room.get('damage_present', False),
                'contents_affected': room.get('contents_affected', False)
            }
            processed_rooms.append(room_data)
            room_types.add(room_data['room_type'])
        
        return processed_rooms, list(room_types)
    
    @staticmethod
    def process_claim_details(raw_claim_data):
        """
        Process detailed claim information to match API response structure
        """
        if not raw_claim_data:
            return {}
        
        return {
            'id': raw_claim_data.get('id'),
            'brand_id': raw_claim_data.get('brand_id'),
            'organization_id': raw_claim_data.get('organization_id'),
            'permalink_url': raw_claim_data.get('permalink_url'),
            'type_of_loss': raw_claim_data.get('type_of_loss'),
            'locale': raw_claim_data.get('locale'),
            'adjuster_name': raw_claim_data.get('adjuster_name'),
            'assignment_identifier': raw_claim_data.get('assignment_identifier'),
            'cat_code': raw_claim_data.get('cat_code'),
            'contents_estimate': raw_claim_data.get('contents_estimate'),
            'contractor_identifier': raw_claim_data.get('contractor_identifier'),
            'date_claim_created': raw_claim_data.get('date_claim_created'),
            'date_of_loss': raw_claim_data.get('date_of_loss'),
            'default_depreciation': raw_claim_data.get('default_depreciation'),
            'emergency_estimate': raw_claim_data.get('emergency_estimate'),
            'full_address': raw_claim_data.get('full_address'),
            'insurer_identifier': raw_claim_data.get('insurer_identifier'),
            'loss_details': raw_claim_data.get('loss_details'),
            'max_depreciation': raw_claim_data.get('max_depreciation'),
            'policyholder_email_address': raw_claim_data.get('policyholder_email_address'),
            'policyholder_name': raw_claim_data.get('policyholder_name'),
            'policyholder_phone_number': raw_claim_data.get('policyholder_phone_number'),
            'project_manager_name': raw_claim_data.get('project_manager_name'),
            'repair_estimate': raw_claim_data.get('repair_estimate'),
            'sales_tax': raw_claim_data.get('sales_tax'),
            'broker_or_agent_name': raw_claim_data.get('broker_or_agent_name'),
            'insurance_company_name': raw_claim_data.get('insurance_company_name'),
            'policy_number': raw_claim_data.get('policy_number')
        }

class EncircleMediaDownloader:
    """
    Downloads media files from Encircle API and organizes them by room labels.
    Files are numbered sequentially and saved in room-specific folders.
    """
    
    def __init__(self, api_client, target_rooms=None):
        self.api_client = api_client
        self.target_rooms = target_rooms or []  # Ensure it's always a list
        self.downloaded_files = 0
        self.failed_downloads = 0
        self.base_dir = "encircle_media_downloads"
        os.makedirs(self.base_dir, exist_ok=True)
        
        # Ensure base directory exists
        os.makedirs(self.base_dir, exist_ok=True)
        
    def download_claim_media(self, property_claim_id):
        """
        Main method to download all media for a specific claim
        """
        print(f"\nStarting media download for claim ID: {property_claim_id}")
        print("="*60)
        
        try:
            # Get media list from API
            media_list = self._get_media_list(property_claim_id)
            
            if not media_list:
                print("No media found for this claim.")
                return
            
            # Process each media item
            for idx, media_item in enumerate(media_list, 1):
                if self._should_download(media_item):
                    self._process_media_item(media_item, idx, len(media_list))
                else:
                    print(f"Skipping {media_item['filename']} - not in target rooms")

        except Exception as e:
            logging.error(f"Error downloading media for claim {property_claim_id}: {str(e)}")
            raise
            
        print("\nDownload Summary:")
        print(f"- Successfully downloaded: {self.downloaded_files}")
        print(f"- Failed downloads: {self.failed_downloads}")
    
    def _should_download(self, media_item):
        """Check if media belongs to target room(s)"""
        if not self.target_rooms:
            return True  # Download all if no filter
        
        labels = media_item.get('labels', [])
        return any(room.lower() in [label.lower() for label in labels] 
               for room in self.target_rooms)

    def _get_media_list(self, property_claim_id):
        """Fetch media list from API with pagination support"""
        all_media = []
        after_cursor = None
        
        while True:
            params = {'limit': 100}
            if after_cursor:
                params['after'] = after_cursor
                
            endpoint = f"property_claims/{property_claim_id}/media"
            response = self.api_client._make_request(endpoint, params=params)
            
            if not response or 'list' not in response:
                break
                
            all_media.extend(response['list'])
            after_cursor = response.get('cursor', {}).get('after')
            
            if not after_cursor:
                break
                
        return all_media
        
    def _process_media_item(self, media_item, current_index, total_items):
        """Handle a single media item download"""
        try:
            # Determine folder based on labels
            folder_name = self._get_folder_name(media_item)
            file_extension = self._get_file_extension(media_item['content_type'])
            
            # Create safe filename with sequential number
            seq_num = str(current_index).zfill(len(str(total_items)))
            clean_filename = f"{seq_num}_{media_item['filename']}"
            clean_filename = self._sanitize_filename(clean_filename)
            
            # Ensure proper file extension
            if not clean_filename.lower().endswith(file_extension.lower()):
                clean_filename += file_extension
                
            # Prepare full path
            folder_path = os.path.join(self.base_dir, folder_name)
            os.makedirs(folder_path, exist_ok=True)
            file_path = os.path.join(folder_path, clean_filename)
            
            # Download the file
            print(f"\nDownloading {current_index}/{total_items}: {clean_filename}")
            print(f"Type: {media_item['content_type']}")
            print(f"Labels: {', '.join(media_item.get('labels', ['No labels']))}")
            print(f"Saving to: {file_path}")
            
            self._download_file(media_item['download_uri'], file_path)
            self.downloaded_files += 1
            
            # Add metadata file
            self._create_metadata_file(media_item, file_path)
            
        except Exception as e:
            self.failed_downloads += 1
            logging.error(f"Failed to process media item: {str(e)}")
            print(f"Error processing item {current_index}: {str(e)}")
            
    def _get_folder_name(self, media_item):
        """Create nested folder structure from all labels, handling edge cases"""
        labels = media_item.get('labels', [])
        
        if not labels:
            return "unlabeled_media"
        
        # Handle cases where labels might be empty strings
        valid_labels = [label.strip() for label in labels if label.strip()]
        if not valid_labels:
            return "unlabeled_media"
        
        # Limit folder depth to prevent overly long paths
        max_depth = 3  # Main_Building/Sub_Room/Area
        truncated_labels = valid_labels[:max_depth]
        
        # Sanitize and join
        return os.path.join(*[self._sanitize_folder_name(label) for label in truncated_labels])
        
    def _sanitize_filename(self, filename):
        """Remove invalid characters from filenames"""
        invalid_chars = '<>:"/\\|?*'
        for char in invalid_chars:
            filename = filename.replace(char, '_')
        return filename
        
    def _sanitize_folder_name(self, foldername):
        """Sanitize folder names and truncate if too long"""
        invalid_chars = '<>:"/\\|?*'
        for char in invalid_chars:
            foldername = foldername.replace(char, '_')
        return foldername[:50]  # Prevent too long folder names
        
    def _get_file_extension(self, content_type):
        """Map content type to file extension"""
        extension_map = {
            'image/jpeg': '.jpg',
            'image/png': '.png',
            'application/pdf': '.pdf',
            'video/mp4': '.mp4',
            'application/vnd.openxmlformats-officedocument.wordprocessingml.document': '.docx',
        }
        return extension_map.get(content_type, '.bin')
        
    def _download_file(self, url, file_path):
        """Download the actual file from the URI"""
        try:
            response = requests.get(url, stream=True)
            response.raise_for_status()
            
            with open(file_path, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    f.write(chunk)
                    
            print("Download completed successfully.")
            
        except requests.exceptions.RequestException as e:
            raise Exception(f"Download failed: {str(e)}")
            
    def _create_metadata_file(self, media_item, file_path):
        """Create a sidecar file with metadata information"""
        metadata = {
            'original_filename': media_item['filename'],
            'content_type': media_item['content_type'],
            'source_type': media_item['source']['type'],
            'source_id': media_item['source']['primary_id'],
            'creator': media_item['creator']['actor_identifier'],
            'created_date': media_item['primary_server_created'],
            'labels': media_item.get('labels', []),
            'download_time': dt.now().isoformat()
        }
        
        metadata_path = f"{file_path}.meta.json"
        with open(metadata_path, 'w') as f:
            json.dump(metadata, f, indent=2)
import re
from collections import defaultdict
from difflib import SequenceMatcher

# ---------------------------
# Simple Normalization & Token Extraction
# ---------------------------

def normalize_text(text):
    """Basic normalization: uppercase, strip extra spaces"""
    if not text:
        return ""
    return ' '.join(text.upper().split())

def extract_tokens(text):
    """
    Extract meaningful tokens from text.
    Returns a set of normalized words (3+ chars) excluding common noise.
    """
    if not text:
        return set()
    
    # Split by @ to separate main part from address
    parts = text.split('@')
    main_part = parts[0] if parts else text
    address_part = parts[1] if len(parts) > 1 else ""
    
    # Extract alphanumeric tokens
    tokens = re.findall(r'\b[A-Z0-9]{2,}\b', main_part.upper())
    
    # Add address tokens if present
    if address_part:
        address_tokens = re.findall(r'\b[A-Z0-9]{2,}\b', address_part.upper())
        tokens.extend(address_tokens)
    
    # Filter out noise words
    noise_words = {'LLC', 'INC', 'THE', 'AND', 'FOR', 'CLAIM', 'EST', 'FIRE', 'WATER', 'STORM'}
    tokens = [t for t in tokens if t not in noise_words and len(t) >= 2]
    
    return set(tokens)

def extract_location_code(text):
    """
    Extract location codes like GA22, OH24, GA22A, OH24-900
    """
    if not text:
        return None
    
    # Pattern: 2-3 letters + 2-3 digits + optional letters/numbers
    match = re.search(r'\b([A-Z]{2,3}\d{2,3}[A-Z0-9\-]*)\b', text.upper())
    return match.group(1) if match else None

# ---------------------------
# Simple Fuzzy Matching
# ---------------------------

def calculate_match_score(encircle_contractor, folder_name):
    """
    Simple fuzzy matching between contractor ID and folder name.
    Returns score 0-1.
    """
    if not encircle_contractor or not folder_name:
        return 0.0
    
    # Normalize both
    contractor_norm = normalize_text(encircle_contractor)
    folder_norm = normalize_text(folder_name)
    
    score = 0.0
    
    # 1. Check if contractor ID is substring of folder (40 points)
    if contractor_norm in folder_norm or folder_norm in contractor_norm:
        score += 0.4
    
    # 2. Location code match (30 points)
    contractor_location = extract_location_code(contractor_norm)
    folder_location = extract_location_code(folder_norm)
    
    if contractor_location and folder_location:
        # Exact match
        if contractor_location == folder_location:
            score += 0.3
        # Fuzzy match (e.g., GA22 vs GA22A)
        elif contractor_location[:4] == folder_location[:4]:
            score += 0.2
    
    # 3. Token overlap (30 points)
    contractor_tokens = extract_tokens(contractor_norm)
    folder_tokens = extract_tokens(folder_norm)
    
    if contractor_tokens and folder_tokens:
        common_tokens = contractor_tokens.intersection(folder_tokens)
        token_score = len(common_tokens) / max(len(contractor_tokens), len(folder_tokens))
        score += token_score * 0.3
    
    return min(score, 1.0)

# ---------------------------
# Main Comparison Function
# ---------------------------

def compare_claims(encircle_claims, onedrive_claims):
    """
    Simple matching: compare contractor ID against folder name
    """
    # Filter out test data
    valid_encircle = [c for c in encircle_claims if _is_valid_claim(c)]
    valid_onedrive = [c for c in onedrive_claims if _is_valid_folder(c)]
    
    encircle_test_data = [c for c in encircle_claims if not _is_valid_claim(c)]
    onedrive_test_data = [c for c in onedrive_claims if not _is_valid_folder(c)]
    
    results = {
        'summary': {
            'total_encircle': len(encircle_claims),
            'total_onedrive': len(onedrive_claims),
            'matches': 0,
            'encircle_only': 0,
            'onedrive_only': 0,
            'encircle_test_data': len(encircle_test_data),
            'onedrive_test_data': len(onedrive_test_data),
            'match_breakdown': {
                'high_confidence': 0,
                'medium_confidence': 0,
                'low_confidence': 0,
            }
        },
        'matched_pairs': [],
        'encircle_missing_onedrive': [],
        'onedrive_extra': [],
        'encircle_test_data': encircle_test_data,
        'onedrive_test_data': onedrive_test_data,
        'duplicates': find_duplicates(valid_encircle, valid_onedrive)
    }
    
    matched_encircle = set()
    matched_onedrive = set()
    
    print("=" * 80)
    print("SIMPLIFIED CLAIM MATCHING")
    print("=" * 80)
    
    # Match each encircle claim to best onedrive folder
    for encircle_claim in valid_encircle:
        if encircle_claim['id'] in matched_encircle:
            continue
        
        contractor_id = encircle_claim.get('contractor_identifier', '').strip()
        
        if not contractor_id:
            continue
        
        best_match = None
        best_score = 0
        
        # Check against all onedrive folders
        for onedrive_claim in valid_onedrive:
            if onedrive_claim['folder_id'] in matched_onedrive:
                continue
            
            folder_name = onedrive_claim.get('folder_name', '').strip()
            
            # Calculate match score
            score = calculate_match_score(contractor_id, folder_name)
            
            # Print analysis
            if score >= 0.5:  # Only print potential matches
                print(f"\nPOTENTIAL MATCH:")
                print(f"  Encircle: {contractor_id}")
                print(f"  OneDrive: {folder_name}")
                print(f"  Score: {score:.2f} ({int(score*100)}%)")
            
            if score > best_score:
                best_score = score
                best_match = onedrive_claim
        
        # Accept matches above threshold
        MATCH_THRESHOLD = 0.65
        if best_match and best_score >= MATCH_THRESHOLD:
            confidence_level = "High" if best_score >= 0.8 else "Medium" if best_score >= 0.65 else "Low"
            
            results['matched_pairs'].append({
                'encircle': encircle_claim,
                'onedrive': best_match,
                'match_type': f'Fuzzy Match ({confidence_level})',
                'confidence': f'{int(best_score * 100)}%'
            })
            
            matched_encircle.add(encircle_claim['id'])
            matched_onedrive.add(best_match['folder_id'])
            results['summary']['matches'] += 1
            
            # Update breakdown
            if best_score >= 0.8:
                results['summary']['match_breakdown']['high_confidence'] += 1
            elif best_score >= 0.65:
                results['summary']['match_breakdown']['medium_confidence'] += 1
            else:
                results['summary']['match_breakdown']['low_confidence'] += 1
            
            print(f"  ✓ MATCHED! ({confidence_level} confidence)")
    
    # Collect unmatched claims
    for encircle_claim in valid_encircle:
        if encircle_claim['id'] not in matched_encircle:
            results['encircle_missing_onedrive'].append(encircle_claim)
            results['summary']['encircle_only'] += 1
    
    for onedrive_claim in valid_onedrive:
        if onedrive_claim['folder_id'] not in matched_onedrive:
            results['onedrive_extra'].append(onedrive_claim)
            results['summary']['onedrive_only'] += 1
    
    print("\n" + "=" * 80)
    print(f"MATCHING COMPLETE")
    print(f"Matches: {results['summary']['matches']}")
    print(f"Encircle Only: {results['summary']['encircle_only']}")
    print(f"OneDrive Only: {results['summary']['onedrive_only']}")
    print("=" * 80)
    
    return results

# ---------------------------
# Filter Functions (from original code)
# ---------------------------

_TEST_EXCLUDE_PATTERNS = [
    'HOW2', 'TEST', 'TEMPLATE', 'SAMPLE', 'ROOMLISTS', 'READINGS',
    'TMPL', 'CHECKLIST', 'TRAILER', 'WAREHOUSE', 'DEFAULT', 'TEMP',
    'PLACEHOLDER', 'EXAMPLE', 'DEMO', 'XXXX', 'AAA', '===', 'BACKEND', 'TUTORIAL'
]

def _is_valid_claim(claim):
    """Filter out test/placeholder claims from Encircle."""
    if not claim.get('policyholder_name') and not claim.get('contractor_identifier'):
        return False
    policyholder = (claim.get('policyholder_name') or '').upper()
    contractor = (claim.get('contractor_identifier') or '').upper()
    return not any(p in policyholder or p in contractor for p in _TEST_EXCLUDE_PATTERNS)

def _is_valid_folder(folder_claim):
    """Filter out test/placeholder folders from OneDrive."""
    folder_name = (folder_claim.get('folder_name') or '').upper()
    exclude_patterns = _TEST_EXCLUDE_PATTERNS + [
        'CLOSED CLAIMS', 'PROOF OF LOSS', 'DRAWINGS', 'APPRAISALS',
        'FOLDER', 'TEXT'
    ]
    if any(p in folder_name for p in exclude_patterns):
        return False
    clean_alpha = re.sub(r'[^A-Z]', '', folder_name)
    return len(clean_alpha) >= 3

def find_duplicates(encircle_claims, onedrive_claims):
    """Find duplicate claims in both systems"""
    duplicates = {'encircle_duplicates': [], 'onedrive_duplicates': []}
    
    # Encircle duplicates by contractor ID
    contractor_count = defaultdict(list)
    for claim in encircle_claims:
        contractor_id = claim.get('contractor_identifier', '')
        if contractor_id and contractor_id.strip():
            contractor_count[contractor_id].append(claim)
    
    for contractor_id, claims in contractor_count.items():
        if len(claims) > 1:
            duplicates['encircle_duplicates'].append({
                'contractor_id': contractor_id, 
                'count': len(claims), 
                'claims': claims
            })
    
    # OneDrive duplicates by normalized folder name
    folder_count = defaultdict(list)
    for claim in onedrive_claims:
        folder_name = claim.get('folder_name', '')
        if folder_name:
            normalized = normalize_text(folder_name)
            folder_count[normalized].append(claim)
    
    for folder_name, claims in folder_count.items():
        if len(claims) > 1:
            duplicates['onedrive_duplicates'].append({
                'folder_name': folder_name, 
                'count': len(claims), 
                'claims': claims
            })
    
    return duplicates

@csrf_exempt
def create_room_template_from_excel(request):
    """
    Handle user pasted room data with room names and LOS, execute automation with base Excel template
    
    Expected POST data:
    - email: User email
    - password: User password
    - room_data: Pasted room data in format "Room Name, LOS" (one per line)
    - selected_templates: List of template IDs to create
    """
    try:
        # Validate required fields
        required_fields = ['room_data']
        missing_fields = [field for field in required_fields if field not in request.POST or not request.POST[field].strip()]
        
        if missing_fields:
            return JsonResponse(
                {
                    "status": "failed",
                    "error": f"Missing or empty required fields: {', '.join(missing_fields)}"
                },
                status=400
            )
        
        # Extract and validate room data
        room_data_raw = request.POST.get('room_data', '').strip()
        room_entries = parse_room_data(room_data_raw)
        
        if not room_entries:
            return JsonResponse(
                {
                    "status": "failed",
                    "error": "No valid room data found. Please paste room data in format 'Room Name, LOS' (one per line)."
                },
                status=400
            )
        
        # Extract other form data
        email = request.POST.get('email')
        password = request.POST.get('password')
        selected_templates = request.POST.getlist('selected_templates')
        
        if not selected_templates:
            return JsonResponse(
                {
                    "status": "failed",
                    "error": "No templates selected. Please select at least one template to generate."
                },
                status=400
            )
        
        # Update base Excel file directly with room data
        base_excel_path = None
        try:
            base_excel_path = find_base_excel_template()
            update_base_excel_with_room_data(base_excel_path, room_entries)
        except Exception as e:
            return JsonResponse(
                {
                    "status": "failed",
                    "error": f"Failed to update base Excel file: {str(e)}",
                    "stage": "excel_update",
                    "debug_info": get_template_debug_info()
                },
                status=500
            )
        
        try:
            # Initialize automation
            automation = RoomTemplateAutomation(headless=True)
                
            # Run automation with selected templates using the updated base Excel file
            results = automation.run_automation(
                excel_file_path=base_excel_path,
                selected_template_ids=selected_templates,
                delete_existing=True
            )
            
            # Add additional processing metadata
            results['file_processing'] = {
                'rooms_processed': len(room_entries),
                'room_entries': [
                    {
                        'room_name': entry['room_name'],
                        'los': entry['los']
                    } for entry in room_entries[:10]  # Show first 10 for confirmation
                ],
                'total_rooms': len(room_entries),
                'base_file_updated': True,
                'base_file_path': base_excel_path,
                'templates_requested': len(selected_templates),
                'columns_populated': ['Q (Room Names)', 'R (LOS)'],
                'formulas_recalculated': True
            }
            
            # Add truncation notice if more than 10 rooms
            if len(room_entries) > 10:
                results['file_processing']['room_list_note'] = f"Showing first 10 of {len(room_entries)} room entries"
            
            return JsonResponse(results)
                        
        except Exception as e:
            return JsonResponse(
                {
                    "status": "failed",
                    "error": str(e),
                    "stage": "automation_execution",
                    "rooms_attempted": len(room_entries),
                    "base_file_path": base_excel_path
                },
                status=500
            )
        
    except json.JSONDecodeError:
        return JsonResponse(
            {
                "status": "failed",
                "error": "Invalid JSON data"
            },
            status=400
        )
    except Exception as e:
        return JsonResponse(
            {
                "status": "failed",
                "error": str(e),
                "stage": "request_processing"
            },
            status=500
        )


def parse_room_data(room_data_text):
    """
    Parse room data text into separate room names and LOS values
    
    Args:
        room_data_text (str): Raw text with room data in format "Room Name, LOS"
        
    Returns:
        list: List of dictionaries with 'room_name', 'los', and metadata
    """
    lines = [line.strip() for line in room_data_text.split('\n') if line.strip()]
    parsed_data = []
    
    for line_num, line in enumerate(lines, 1):
        if ',' in line:
            # Split on the first comma only to handle LOS descriptions with commas
            parts = line.split(',', 1)
            room_name = parts[0].strip()
            los = parts[1].strip()
            
            if room_name and los:
                parsed_data.append({
                    'room_name': room_name,
                    'los': los,
                    'original_line': line,
                    'line_number': line_num
                })
            else:
                print(f"Warning: Line {line_num} has empty room name or LOS: '{line}'")
        else:
            print(f"Warning: Line {line_num} missing comma separator: '{line}'")
    
    return parsed_data


def find_base_excel_template():
    """
    Find the base Excel template file by checking multiple possible locations and filenames
    
    Returns:
        str: Path to the found template file
        
    Raises:
        FileNotFoundError: If no template file is found
    """
    # Primary template filename and location
    primary_template = '01-INFO.xlsx'
    
    # Possible directories (relative to BASE_DIR) - prioritizing the known location
    possible_dirs = [
        'docsAppR/templates/excel/',  # Known correct location
        'templates/excel/',
        'docsAppR/templates/',
        'templates/',
        'media/templates/',
        'static/templates/',
        'app/templates/excel/',
        'app/docsAppR/templates/excel/',
        'excel_templates/',
        '.'  # Current directory
    ]
    
    # Fallback template filenames if primary not found
    fallback_filenames = [
        'base_room_template.xlsx',
        'INFO.xlsx',
        'template.xlsx',
        'base_template.xlsx'
    ]
    
    # First, search for the primary template in all directories
    for directory in possible_dirs:
        dir_path = os.path.join(settings.BASE_DIR, directory)
        if os.path.exists(dir_path):
            template_path = os.path.join(dir_path, primary_template)
            if os.path.exists(template_path):
                print(f"Found Excel template at: {template_path}")
                return template_path
    
    # If primary template not found, search for fallback templates
    for directory in possible_dirs:
        dir_path = os.path.join(settings.BASE_DIR, directory)
        if os.path.exists(dir_path):
            for filename in fallback_filenames:
                template_path = os.path.join(dir_path, filename)
                if os.path.exists(template_path):
                    print(f"Found fallback Excel template at: {template_path}")
                    return template_path
    
    # If not found, list what we have for debugging
    debug_info = get_template_debug_info()
    raise FileNotFoundError(
        f"Excel template '{primary_template}' not found. Searched in: {possible_dirs} "
        f"Also searched for fallback files: {fallback_filenames}. "
        f"Debug info: {debug_info}"
    )


def get_template_debug_info():
    """
    Get debugging information about available files and directories
    
    Returns:
        dict: Debug information about file system
    """
    debug_info = {
        'base_dir': settings.BASE_DIR,
        'base_dir_exists': os.path.exists(settings.BASE_DIR),
        'directories': {},
        'current_working_dir': os.getcwd(),
        'target_file': '01-INFO.xlsx',
        'expected_location': '/app/excel/'
    }
    
    # Check common directories, including the known location
    search_dirs = [
        'app/excel',
        'docsAppR',
        'docsAppR/templates',
        'docsAppR/templates/excel',
        'templates',
        'templates/excel',
        'media',
        'static',
        'app',
        'app/docsAppR',
        'app/docsAppR/templates',
        '.'
    ]
    
    for dir_name in search_dirs:
        dir_path = os.path.join(settings.BASE_DIR, dir_name)
        debug_info['directories'][dir_name] = {
            'exists': os.path.exists(dir_path),
            'full_path': dir_path
        }
        
        if os.path.exists(dir_path):
            try:
                files = os.listdir(dir_path)
                excel_files = [f for f in files if f.endswith(('.xlsx', '.xls'))]
                debug_info['directories'][dir_name]['files'] = files[:10]  # First 10 files
                debug_info['directories'][dir_name]['excel_files'] = excel_files
                
                # Check specifically for 01-INFO.xlsx
                if '01-INFO.xlsx' in files:
                    debug_info['directories'][dir_name]['has_target_file'] = True
                    
            except Exception as e:
                debug_info['directories'][dir_name]['error'] = str(e)
    
    return debug_info


def update_base_excel_with_room_data(base_excel_path, room_entries):
    """
    Update the base Excel file directly with room data
    
    Args:
        base_excel_path (str): Path to the base Excel template file
        room_entries (list): List of dictionaries with 'room_name' and 'los' keys
        
    Raises:
        Exception: If Excel manipulation fails
    """
    
    try:
        print(f"Updating base Excel file: {base_excel_path}")
        print(f"Processing {len(room_entries)} room entries...")
        
        # Create a backup of the original file
        backup_path = base_excel_path + '.backup'
        if os.path.exists(backup_path):
            os.remove(backup_path)
        
        # Load the workbook and find the ROOMS# sheet
        workbook = load_workbook(base_excel_path)
        
        # Try different possible sheet names for ROOMS#
        possible_sheet_names = ['ROOMS#', 'ROOMS #', 'Rooms#', 'Rooms #', 'ROOMS', 'Rooms', 'Sheet1', 'INFO']
        target_sheet = None
        
        print(f"Available sheets in workbook: {workbook.sheetnames}")
        
        for sheet_name in possible_sheet_names:
            if sheet_name in workbook.sheetnames:
                target_sheet = workbook[sheet_name]
                print(f"Using sheet '{sheet_name}' for room data")
                break
        
        if target_sheet is None:
            # If no exact match, look for sheets containing 'rooms' or use first sheet
            for sheet_name in workbook.sheetnames:
                if 'rooms' in sheet_name.lower():
                    target_sheet = workbook[sheet_name]
                    print(f"Using sheet '{sheet_name}' as ROOMS# equivalent")
                    break
            
            # If still no match, use the first sheet
            if target_sheet is None:
                target_sheet = workbook.worksheets[0]
                print(f"Using first sheet '{target_sheet.title}' as default")
        
        # Clear existing data in columns Q and R (starting from row 3 to preserve headers)
        max_row_to_clear = 1000  # Conservative limit
        
        print(f"Clearing existing data in columns Q and S...")
        
        # Clear column Q (Room Names) and column R (LOS)
        cleared_rows = 0
        for row in range(3, max_row_to_clear + 1):
            q_cell = target_sheet[f'Q{row}']
            r_cell = target_sheet[f'S{row}']
            
            if q_cell.value is not None or r_cell.value is not None:
                q_cell.value = None
                r_cell.value = None
                cleared_rows += 1
            else:
                # If we hit several consecutive empty cells in both columns, stop clearing
                empty_count = 0
                for check_row in range(row, min(row + 5, max_row_to_clear + 1)):
                    q_check = target_sheet[f'Q{check_row}'].value
                    r_check = target_sheet[f'R{check_row}'].value
                    if q_check is None and r_check is None:
                        empty_count += 1
                
                if empty_count >= 5:  # If 5 consecutive empty pairs, stop clearing
                    break
        
        print(f"Cleared {cleared_rows} existing rows")
        
        # Insert room data into columns Q and R starting from row 3
        print(f"Inserting {len(room_entries)} room entries...")
        for idx, entry in enumerate(room_entries, start=4):
            # Column Q: Room Name
            target_sheet[f'Q{idx}'] = entry['room_name']
            # Column R: LOS
            target_sheet[f'S{idx}'] = entry['los']
            
            if idx <= 7:  # Log first few entries
                print(f"  Row {idx}: Q='{entry['room_name']}', S='{entry['los']}'")
        

        # Save the workbook with updated data
        workbook.save(base_excel_path)
        workbook.close()
        
        print(f"Successfully saved updated data to base Excel file")
        
        print(f"Successfully updated base Excel file with {len(room_entries)} room entries:")
        print(f"  - Column Q: Room Names")
        print(f"  - Column S: LOS Data")
        print(f"  - Sheet: '{target_sheet.title}' starting at row 3")
        print(f"  - Base file: {base_excel_path}")
        
    except Exception as e:
        raise Exception(f"Failed to update base Excel file: {str(e)}")


def get_base_excel_template_info():
    """
    Utility function to get information about the base Excel template
    
    Returns:
        dict: Information about the base template file including column structure
    """
    try:
        base_excel_path = find_base_excel_template()
        
        info = {
            'path': base_excel_path,
            'exists': True,
            'sheets': [],
            'size': 0,
            'columns': {
                'Q': 'Room Names',
                'S': 'LOS Data'
            },
            'direct_update': True,
            'temp_files_used': False
        }
        
        try:
            info['size'] = os.path.getsize(base_excel_path)
            workbook = load_workbook(base_excel_path, read_only=True)
            info['sheets'] = workbook.sheetnames
            
            # Try to get column headers from any sheet that might contain room data
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                try:
                    # Check what's in the header rows for columns Q and R
                    q_header_1 = sheet['Q1'].value
                    q_header_2 = sheet['Q2'].value
                    r_header_1 = sheet['S1'].value
                    r_header_2 = sheet['S2'].value
                    
                    if q_header_1 or q_header_2 or r_header_1 or r_header_2:
                        info['column_headers'] = {
                            'sheet': sheet_name,
                            'Q': {
                                'row_1': q_header_1,
                                'row_2': q_header_2
                            },
                            'S': {
                                'row_1': r_header_1,
                                'row_2': r_header_2
                            }
                        }
                        
                        # Check if there are existing formulas that reference these columns
                        formula_count = 0
                        for row in sheet.iter_rows(min_row=1, max_row=100):
                            for cell in row:
                                if cell.data_type == 'f' and cell.value:
                                    if 'Q' in str(cell.value) or 'S' in str(cell.value):
                                        formula_count += 1
                        
                        info['formulas_referencing_QR'] = formula_count
                        break
                        
                except Exception as header_error:
                    info['header_read_error'] = str(header_error)
                    continue
                    
            workbook.close()
            
        except Exception as e:
            info['error'] = str(e)
            
    except FileNotFoundError:
        info = {
            'path': None,
            'exists': False,
            'error': 'Template file not found',
            'debug_info': get_template_debug_info(),
            'direct_update': False,
            'temp_files_used': False
        }
    
    return info


def validate_room_data_format(request):
    """
    API endpoint to validate room data format without processing
    
    Returns:
        JsonResponse: Validation results including parsed data preview
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    room_data_raw = request.POST.get('room_data', '').strip()
    
    if not room_data_raw:
        return JsonResponse({
            'valid': False,
            'error': 'No room data provided',
            'entries': [],
            'total_lines': 0,
            'valid_entries': 0
        })
    
    # Parse the data
    parsed_entries = parse_room_data(room_data_raw)
    total_lines = len([line for line in room_data_raw.split('\n') if line.strip()])
    
    # Create preview of parsed data
    preview_entries = []
    for entry in parsed_entries[:5]:  # Show first 5
        preview_entries.append({
            'room_name': entry['room_name'],
            'los': entry['los'],
            'line_number': entry['line_number']
        })
    
    return JsonResponse({
        'valid': len(parsed_entries) > 0,
        'total_lines': total_lines,
        'valid_entries': len(parsed_entries),
        'invalid_entries': total_lines - len(parsed_entries),
        'entries_preview': preview_entries,
        'truncated': len(parsed_entries) > 5,
        'columns_target': {
            'Q': 'Room Names',
            'R': 'LOS Data'
        },
        'template_info': get_base_excel_template_info(),
        'update_method': 'direct_base_file_update'
    })


# Debug endpoint to help troubleshoot template issues
def debug_template_files(request):
    """
    Debug endpoint to show what template files are available
    """
    return JsonResponse({
        'template_info': get_base_excel_template_info(),
        'debug_info': get_template_debug_info(),
        'update_method': 'direct_base_file_update',
        'recalculation': 'libreoffice_headless'
    })
import io
import zipfile

def download_media_view(request):
    if request.method == 'POST':
        claim_id = request.POST.get('claim_id')
        room_filter = request.POST.get('room_filter', '')
        target_rooms = [r.strip() for r in room_filter.split(',') if r.strip()]
        
        try:
            # Create a zip file in memory
            zip_buffer = io.BytesIO()
            
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                api_client = EncircleAPIClient()
                downloader = ZipMediaDownloader(api_client, target_rooms, zip_file)
                downloader.download_claim_media(int(claim_id))
            
            # Prepare the response with the zip file
            response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
            response['Content-Disposition'] = f'attachment; filename="claim_{claim_id}_media.zip"'
            
            msg = f"Prepared media download for {claim_id}"
            if target_rooms:
                msg += f" (filtered by rooms: {', '.join(target_rooms)})"
            messages.success(request, msg)
            
            return response
            
        except Exception as e:
            messages.error(request, f"Error: {str(e)}")
            return redirect('download_media')
    
    return render(request, 'account/download_media.html')

class ZipMediaDownloader:
    """
    Downloads media files from Encircle API and organizes them in a zip file by room labels.
    """
    
    def __init__(self, api_client, target_rooms=None, zip_file=None):
        self.api_client = api_client
        self.target_rooms = target_rooms or []
        self.downloaded_files = 0
        self.failed_downloads = 0
        self.zip_file = zip_file
        
    def download_claim_media(self, property_claim_id):
        """
        Main method to download all media for a specific claim into the zip file
        """
        print(f"\nStarting media download for claim ID: {property_claim_id}")
        print("="*60)
        
        try:
            # Get media list from API
            media_list = self._get_media_list(property_claim_id)
            
            if not media_list:
                print("No media found for this claim.")
                return
            
            # Process each media item
            for idx, media_item in enumerate(media_list, 1):
                if self._should_download(media_item):
                    self._process_media_item(media_item, idx, len(media_list))
                else:
                    print(f"Skipping {media_item['filename']} - not in target rooms")

        except Exception as e:
            logging.error(f"Error downloading media for claim {property_claim_id}: {str(e)}")
            raise
            
        print("\nDownload Summary:")
        print(f"- Successfully downloaded: {self.downloaded_files}")
        print(f"- Failed downloads: {self.failed_downloads}")
    
    def _should_download(self, media_item):
        """Check if media belongs to target room(s)"""
        if not self.target_rooms:
            return True  # Download all if no filter
        
        labels = media_item.get('labels', [])
        return any(room.lower() in [label.lower() for label in labels] 
               for room in self.target_rooms)

    def _get_media_list(self, property_claim_id):
        """Fetch media list from API with pagination support"""
        all_media = []
        after_cursor = None
        
        while True:
            params = {'limit': 100}
            if after_cursor:
                params['after'] = after_cursor
                
            endpoint = f"property_claims/{property_claim_id}/media"
            response = self.api_client._make_request(endpoint, params=params)
            
            if not response or 'list' not in response:
                break
                
            all_media.extend(response['list'])
            after_cursor = response.get('cursor', {}).get('after')
            
            if not after_cursor:
                break
                
        return all_media
        
    def _process_media_item(self, media_item, current_index, total_items):
        """Handle a single media item download into the zip file"""
        try:
            # Determine folder based on labels
            folder_name = self._get_folder_name(media_item)
            file_extension = self._get_file_extension(media_item['content_type'])
            
            # Create safe filename with sequential number
            seq_num = str(current_index).zfill(len(str(total_items)))
            clean_filename = f"{seq_num}_{media_item['filename']}"
            clean_filename = self._sanitize_filename(clean_filename)
            
            # Ensure proper file extension
            if not clean_filename.lower().endswith(file_extension.lower()):
                clean_filename += file_extension
                
            # Prepare full path in zip
            zip_path = os.path.join(folder_name, clean_filename)
            
            # Download the file
            print(f"\nDownloading {current_index}/{total_items}: {clean_filename}")
            print(f"Type: {media_item['content_type']}")
            print(f"Labels: {', '.join(media_item.get('labels', ['No labels']))}")
            print(f"Adding to zip at: {zip_path}")
            
            file_content = self._download_file_content(media_item['download_uri'])
            self.zip_file.writestr(zip_path, file_content)
            self.downloaded_files += 1
            
            # Add metadata file
            self._add_metadata_file(media_item, zip_path)
            
        except Exception as e:
            self.failed_downloads += 1
            logging.error(f"Failed to process media item: {str(e)}")
            print(f"Error processing item {current_index}: {str(e)}")
            
    def _get_folder_name(self, media_item):
        """Create nested folder structure from all labels, handling edge cases"""
        labels = media_item.get('labels', [])
        
        if not labels:
            return "unlabeled_media"
        
        # Handle cases where labels might be empty strings
        valid_labels = [label.strip() for label in labels if label.strip()]
        if not valid_labels:
            return "unlabeled_media"
        
        # Limit folder depth to prevent overly long paths
        max_depth = 3  # Main_Building/Sub_Room/Area
        truncated_labels = valid_labels[:max_depth]
        
        # Sanitize and join
        return os.path.join(*[self._sanitize_folder_name(label) for label in truncated_labels])
        
    def _sanitize_filename(self, filename):
        """Remove invalid characters from filenames"""
        invalid_chars = '<>:"/\\|?*'
        for char in invalid_chars:
            filename = filename.replace(char, '_')
        return filename
        
    def _sanitize_folder_name(self, foldername):
        """Sanitize folder names and truncate if too long"""
        invalid_chars = '<>:"/\\|?*'
        for char in invalid_chars:
            foldername = foldername.replace(char, '_')
        return foldername[:50]  # Prevent too long folder names
        
    def _get_file_extension(self, content_type):
        """Map content type to file extension"""
        extension_map = {
            'image/jpeg': '.jpg',
            'image/png': '.png',
            'application/pdf': '.pdf',
            'video/mp4': '.mp4',
            'application/vnd.openxmlformats-officedocument.wordprocessingml.document': '.docx',
        }
        return extension_map.get(content_type, '.bin')
        
    def _download_file_content(self, url):
        """Download the actual file content from the URI"""
        try:
            response = requests.get(url, stream=True)
            response.raise_for_status()
            return response.content
        except requests.exceptions.RequestException as e:
            raise Exception(f"Download failed: {str(e)}")
            
    def _add_metadata_file(self, media_item, file_path):
        """Create a metadata file in the zip"""
        metadata = {
            'original_filename': media_item['filename'],
            'content_type': media_item['content_type'],
            'source_type': media_item['source']['type'],
            'source_id': media_item['source']['primary_id'],
            'creator': media_item['creator']['actor_identifier'],
            'created_date': media_item['primary_server_created'],
            'labels': media_item.get('labels', []),
            'download_time': dt.datetime.now().isoformat()
        }
        
        metadata_path = f"{file_path}.meta.json"
        self.zip_file.writestr(metadata_path, json.dumps(metadata, indent=2))

@login_required
def export_claims_to_excel(request, claim_id=None):
    """
    View to export claims data to Excel
    """
    try:
        api_client = EncircleAPIClient()
        processor = EncircleDataProcessor()
        exporter = EncircleExcelExporter()
        
        if claim_id:
            # Export single claim with all details
            raw_claim_details = api_client.get_claim_details(claim_id)
            claim_details = processor.process_claim_details(raw_claim_details)

            structures_response = api_client.get_claim_structures(claim_id)
            if not structures_response or not structures_response.get('list'):
                raise ValueError(f"No structures found for claim {claim_id}")
            
            structure_id = structures_response['list'][0]['id']

            raw_rooms_data = api_client.get_claim_rooms(claim_id, structure_id)
            rooms_data, room_types = processor.process_claim_rooms(raw_rooms_data)

            floor_plan_data = None
            try:
                raw_floor_plan = api_client.get_claim_floor_plan(claim_id)
                floor_plan_data = processor.process_floor_plan_data(raw_floor_plan)
            except Exception as e:
                logger.warning(f"Could not fetch floor plan: {str(e)}")

            export_data = {
                'claim_details': claim_details,
                'rooms': rooms_data,
                'floor_plan': floor_plan_data
            }
        else:
            # Export all claims with consolidated rooms and floor plans
            raw_claims = api_client.get_all_claims()
            processed_claims = processor.process_claims_list(raw_claims)
            processed_claims = processed_claims
            logger.info(f"Processing {len(processed_claims)} claims for detailed export")

            # Collect rooms and floor plan data from ALL claims
            all_rooms_data = []
            all_floor_plan_data = {}
            
            # Process each claim with better error handling
            for i, claim in enumerate(processed_claims):
                current_claim_id = claim.get('id')
                if not current_claim_id:
                    logger.warning(f"Claim at index {i} has no ID, skipping")
                    continue

                logger.info(f"Processing claim {current_claim_id} ({i+1}/{len(processed_claims)})")
                
                try:
                    # Get structures for this claim
                    structures_response = api_client.get_claim_structures(current_claim_id)
                    if not structures_response or not structures_response.get('list'):
                        logger.warning(f"No structures found for claim {current_claim_id}")
                        continue

                    structure_id = structures_response['list'][0]['id']

                    # Get rooms for this claim
                    try:
                        raw_rooms_data = api_client.get_claim_rooms(current_claim_id, structure_id)
                        rooms_data, room_types = processor.process_claim_rooms(raw_rooms_data)

                        # Add claim_id to each room for identification
                        if rooms_data:  # Check if rooms_data is not empty
                            for room in rooms_data:
                                room['claim_id'] = current_claim_id
                                room['claim_name'] = claim.get('policyholder_name', 'Unknown')
                            all_rooms_data.extend(rooms_data)
                            logger.info(f"Added {len(rooms_data)} rooms for claim {current_claim_id}")
                    except Exception as e:
                        logger.warning(f"Could not fetch rooms for claim {current_claim_id}: {str(e)}")

                    # Get floor plan for this claim
                    try:
                        raw_floor_plan = api_client.get_claim_floor_plan(current_claim_id)
                        if raw_floor_plan:  # Check if data exists
                            floor_plan_data = processor.process_floor_plan_data(raw_floor_plan)
                            if floor_plan_data:
                                # Namespace floor plan data by claim ID
                                all_floor_plan_data[f"Claim_{current_claim_id}"] = floor_plan_data
                                logger.info(f"Added floor plan for claim {current_claim_id}")
                    except Exception as e:
                        logger.warning(f"Could not fetch floor plan for claim {current_claim_id}: {str(e)}")

                except Exception as e:
                    logger.error(f"Critical error processing claim {current_claim_id}: {str(e)}")
                    continue

            logger.info(f"Collected {len(all_rooms_data)} total rooms and {len(all_floor_plan_data)} floor plans")

            export_data = {
                'claims': processed_claims,
                'rooms': all_rooms_data if all_rooms_data else [],  # Use empty list instead of None
                'floor_plan': all_floor_plan_data if all_floor_plan_data else {}  # Use empty dict instead of None
            }
        
        # Debug: Log export data structure
        logger.info("Export data structure:")
        logger.info(f"- Claims: {len(export_data.get('claims', []))}")
        logger.info(f"- Rooms: {len(export_data.get('rooms', []))}")
        logger.info(f"- Floor plans: {len(export_data.get('floor_plan', {}))}")
        logger.info(f"- Claim details: {'Yes' if export_data.get('claim_details') else 'No'}")
        
        # Generate Excel file
        excel_file, filename = exporter.export_claims(export_data)
        
        # Validate Excel file was created
        if not excel_file:
            raise ValueError("Excel file generation returned None")
        
        # Get the actual bytes content
        excel_content = excel_file.getvalue()
        
        # Validate content size
        if len(excel_content) == 0:
            raise ValueError("Generated Excel file is empty")
        
        logger.info(f"Generated Excel file: {filename}, Size: {len(excel_content)} bytes")
        
        # Create response with proper headers
        response = HttpResponse(
            excel_content,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response['Content-Length'] = len(excel_content)

        return response
        
    except Exception as e:
        logger.error(f"Error exporting to Excel: {str(e)}")
        logger.error(f"Exception type: {type(e).__name__}")
        import traceback
        logger.error(f"Full traceback: {traceback.format_exc()}")
        
        # Return simple error response for all request types
        messages.error(request, f"Error generating Excel file: {str(e)}")
        return redirect('encircle_claims_dashboard')

# Django Views
def encircle_claims_dashboard(request):
    """
    Main dashboard view for displaying all claims
    """
    return render(request, 'account/encircle_dashboard.html')


def fetch_all_claims_api(request):
    """
    API endpoint to fetch all claims with basic information
    """
    try:
        api_client = EncircleAPIClient()
        processor = EncircleDataProcessor()
        
        # Fetch all claims
        raw_claims = api_client.get_all_claims()
        processed_claims = processor.process_claims_list(raw_claims)
        
        # Enhance each claim with room count and types
        for claim in processed_claims:
            try:
                # Get structures with validation
                structures_response = api_client.get_claim_structures(claim['id'])
                if not structures_response or not isinstance(structures_response, dict) or 'list' not in structures_response:
                    logger.warning(f"Invalid structures data for claim {claim['id']}")
                    claim['total_rooms'] = 'N/A'
                    claim['room_types'] = []
                    continue
                    
                structures_list = structures_response.get('list', [])
                if not structures_list:
                    claim['total_rooms'] = 0
                    claim['room_types'] = []
                    continue
                
                # Get rooms for the first structure only (to avoid too many API calls)
                first_structure_id = structures_list[0].get('id')
                if not first_structure_id:
                    logger.warning(f"No valid structure ID found for claim {claim['id']}")
                    claim['total_rooms'] = 'N/A'
                    claim['room_types'] = []
                    continue
                
                rooms_response = api_client.get_claim_rooms(claim['id'], first_structure_id)
                if not rooms_response or not isinstance(rooms_response, dict) or 'list' not in rooms_response:
                    logger.warning(f"Invalid rooms data for claim {claim['id']}")
                    claim['total_rooms'] = 'N/A'
                    claim['room_types'] = []
                    continue
                
                processed_data, processed_types = processor.process_claim_rooms(rooms_response)
                claim['total_rooms'] = len(processed_data)
                claim['room_types'] = processed_types[:5]  # Limit to first 5 types for overview
            except Exception as e:
                logger.warning(f"Error processing rooms for claim {claim['id']}: {str(e)}")
                claim['total_rooms'] = 'N/A'
                claim['room_types'] = []
            break   
        
        return JsonResponse({
            'success': True,
            'claims': processed_claims,
            'total_claims': len(processed_claims)
        })
        
    except Exception as e:
        logger.error(f"Error fetching claims: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e),
            'claims': [],
            'total_claims': 0
        }, status=500)


def fetch_claim_details_api(request, claim_id):
    """
    API endpoint to fetch detailed information for a specific claim
    """
    try:
        api_client = EncircleAPIClient()
        processor = EncircleDataProcessor()
        
        # Fetch claim details
        raw_claim_details = api_client.get_claim_details(claim_id)
        claim_details = processor.process_claim_details(raw_claim_details)
        
        structures_response = api_client.get_claim_structures(claim_id)
        structure_id = structures_response['list'][0]['id']
        
        # Fetch rooms for this claim
        raw_rooms_data = api_client.get_claim_rooms(claim_id, structure_id)
        rooms_data, room_types = processor.process_claim_rooms(raw_rooms_data)
        
        # Fetch and process floor plan data
        floor_plan_data = None
        try:
            raw_floor_plan = api_client.get_claim_floor_plan(claim_id)
            floor_plan_data = processor.process_floor_plan_data(raw_floor_plan)
        except Exception as e:
            logger.warning(f"Could not fetch floor plan for claim {claim_id}: {str(e)}")
        
        return JsonResponse({
            'success': True,
            'claim_details': claim_details,
            'rooms': rooms_data,
            'room_types': room_types,
            'total_rooms': len(rooms_data),
            'floor_plan': floor_plan_data
        })
        
    except Exception as e:
        logger.error(f"Error fetching claim details for {claim_id}: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


def fetch_claim_rooms_api(request, claim_id, structure_id):
    """
    API endpoint to fetch rooms data for a specific claim
    """
    try:
        api_client = EncircleAPIClient()
        processor = EncircleDataProcessor()
        
        raw_rooms_data = api_client.get_claim_rooms(claim_id, structure_id)
        rooms_data, room_types = processor.process_claim_rooms(raw_rooms_data)
        
        return JsonResponse({
            'success': True,
            'rooms': rooms_data,
            'room_types': room_types,
            'total_rooms': len(rooms_data)
        })
        
    except Exception as e:
        logger.error(f"Error fetching rooms for claim {claim_id}: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

def recursive_dir_list(dir, dic):
    for f in os.listdir(dir):
        path = os.path.join(dir, f)
        
        if os.path.isdir(path):
            dic[f] = {}
            recursive_dir_list(path,dic[f])
        else:
            dic[f] = f
            
    return dic

def home(request):
    
      # request to server to return files in document folder as object
      # send this object to the template for display
      # reload page to update display
      # once on template, load object and display.
      # start from the top most objects and display those filenames
      # for each folder object make it clickable
      # once clicked index the object for all children
      # and repalce entire display with jsut children of that folder
      # display path to that location in the top
      # the true path to the file on server should be calculated somewhere
      # probably not on the page tho to not expose paths

    #dic_of_files = recursive_dir_list(settings.MEDIA_ROOT + "/uploads/documents/", {})
    ## handling upload
    #if request == "POST":
    #    form = UploadFilesForm(request.POST, request.FILES)
    #    if form.is_valid():
    #        # function that handles file
    #        return


    if request.user.is_authenticated:
        return render(request, 'account/dashboard.html')
    else:
        return render(request, 'account/login.html')

def logout_view(request):
    logout(request)
    return redirect('/')

@login_required
def get_dimensions(request):
    
    return render(request, 'account/encircle.html')

def fetch_dimensions_API(request, claim_id):
    import requests

    try: 
        api_key = "367382d2-0b2d-4b01-9d06-8f18fd492f5e"

        url = f"https://api.encircleapp.com/v2/property_claims/{claim_id}/floor_plan_dimensions"
    
        headers = {"Authorization" : f"Bearer {api_key}"}

        response = requests.get(url, headers=headers)

        #print(response.json())
        raw_data = response.json()
        processed_data = process_floor_data(raw_data)
        
        return JsonResponse(processed_data)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def fetch_from_encircle_api(claim_id):

    return response.json()

def process_floor_data(raw_data):
    
    result = {
        'floors': [],
        'summary': {
            'totalFloors': 0,
            'totalRooms': 0,
            'roomsByType': {}
        }
    }
    
    # Check if we have valid data
    if not raw_data or 'list' not in raw_data or not raw_data['list']:
        return result
    
    floor_names = ['Basement', 'Main Floor', 'Second Floor', 'Third Floor', 'Attic']
    floor_count = 0
    room_count = 0
    room_types = {}
    
    # Process each floor group
    for floor_group in raw_data['list']:
        if 'floors' not in floor_group:
            continue
        
        # Process each floor in the group
        for i, floor in enumerate(floor_group['floors']):
            if 'features' not in floor:
                continue
                
            floor_name = floor_names[floor_count] if floor_count < len(floor_names) else f"Floor {floor_count + 1}"
            floor_data = {
                'id': floor_count,
                'name': floor_name,
                'rooms': [],
                'totalRooms': 0
            }
            
            # Process rooms (features)
            for feature in floor['features']:
                if feature['type'] == 'Feature' and 'properties' in feature and 'geometry' in feature:
                    room_name = feature['properties'].get('name', 'Unnamed Room')
                    
                    # Update room type counts for summary
                    if room_name in room_types:
                        room_types[room_name] += 1
                    else:
                        room_types[room_name] = 1
                    
                    # Calculate room area (approximate)
                    area = calculate_polygon_area(feature['geometry']['coordinates'][0]) if feature['geometry']['type'] == 'Polygon' else 0
                    
                    room = {
                        'id': room_count,
                        'name': room_name,
                        'ceilingHeight': feature['properties'].get('ceiling_height', 0),
                        'area': round(area, 2),  # Area in square units
                        'coordinates': feature['geometry']['coordinates'][0] if feature['geometry']['type'] == 'Polygon' else []
                    }
                    
                    floor_data['rooms'].append(room)
                    room_count += 1
            
            floor_data['totalRooms'] = len(floor_data['rooms'])
            result['floors'].append(floor_data)
            floor_count += 1
    
    # Update summary information
    result['summary']['totalFloors'] = floor_count
    result['summary']['totalRooms'] = room_count
    result['summary']['roomsByType'] = room_types
    
    return result

def calculate_polygon_area(coordinates):
    """
    Calculate the area of a polygon using the Shoelace formula
    
    Args:
        coordinates: List of [x, y] coordinate pairs forming a polygon
        
    Returns:
        float: Area of the polygon
    """
    n = len(coordinates)
    area = 0.0
    
    for i in range(n):
        j = (i + 1) % n
        area += coordinates[i][0] * coordinates[j][1]
        area -= coordinates[j][0] * coordinates[i][1]
    
    area = abs(area) / 2.0
    return area


from .forms import UploadClientForm

@login_required
def create(request):
    if request.method == "POST":
        if 'excel_file' in request.FILES:
            return handle_excel_import(request)
        else:
            form = UploadClientForm(request.POST)
            if form.is_valid():
                form.save()
                messages.success(request, 'Client created successfully!')
                return redirect('dashboard')
            return render(request, 'account/create.html', {'form': form})
    
    form = UploadClientForm()
    return render(request, 'account/create.html', {'form': form})


"""

API FOR TEMPLATE CREATOR SENDS EACH 01-ROOMS (old 01 INFO) TO THE SERVER AFTER RUNNING 
SYSTEM AUTOMATICALLY LOADS IT IN THE SYSTEM
"""

import json
import pandas as pd
from io import BytesIO
from django.shortcuts import render, redirect
from django.contrib import messages
from django.utils import timezone
from .models import Client
from .forms import ClientForm
from .models import Client, Room, WorkType, RoomWorkTypeValue

import openpyxl
from openpyxl.utils import get_column_letter

def import_client_from_info_file(excel_file):
    """
    Extract client data from 01-INFO.xlsx file from jobinfo(2) tab
    WITH FORMULA SUPPORT - gets calculated values, not formula text
    """
    try:
        # Use openpyxl to read Excel with formula values
        excel_file.seek(0)
        wb = openpyxl.load_workbook(BytesIO(excel_file.read()), data_only=True)  # data_only=True gets calculated values
        
        # Check for jobinfo(2) tab
        sheet_name = 'jobinfo(2)'
        if sheet_name not in wb.sheetnames:
            # Try alternative names
            possible_sheets = [s for s in wb.sheetnames if 'jobinfo' in s.lower()]
            if possible_sheets:
                sheet_name = possible_sheets[0]
                print(f"ℹ️ Using alternative sheet: {sheet_name}")
            else:
                raise ValueError(f"No jobinfo sheet found in INFO file. Available sheets: {', '.join(wb.sheetnames)}")
        
        ws = wb[sheet_name]
        
        print(f"✅ Loaded jobinfo sheet: {ws.max_row} rows, {ws.max_column} columns")
        
        # Extract client data using openpyxl
        client_data = extract_client_data_from_jobinfo_openpyxl(ws)
        wb.close()  # Close workbook to free memory
        
        return client_data
        
    except Exception as e:
        raise Exception(f"Failed to process INFO file: {str(e)}")


def extract_client_data_from_jobinfo_openpyxl(worksheet):
    """
    Extract client data from jobinfo worksheet using openpyxl
    Gets calculated values from formulas
    """
    client_data = {}
    
    # Constants for jobinfo structure
    HEADER_COLUMN = 1  # Column A (openpyxl uses 1-indexed)
    DATA_COLUMN = 2    # Column B
    
    print("🔍 Scanning jobinfo sheet for client data with formula support...")
    
    for row in range(1, worksheet.max_row + 1):
        header_cell = worksheet.cell(row=row, column=HEADER_COLUMN)
        value_cell = worksheet.cell(row=row, column=DATA_COLUMN)
        
        header = str(header_cell.value).strip() if header_cell.value is not None else None
        value = value_cell.value  # This gets the calculated value, even if it's a formula
        
        if not header or header in ['None', 'nan', ''] or value is None:
            continue
            
        # Use your proven field normalization
        field_name = (header.lower()
                     .replace(' ', '_')
                     .replace('/', '_')
                     .replace('\\', '_')
                     .replace('.', '_')
                     .replace('-', '_')
                     .replace(':', '_')
                     .replace('__', '_')
                     .replace('#', 'num')
                     .strip('_'))
        
        # Debug output for key fields
        if any(keyword in field_name for keyword in ['owner', 'address', 'email', 'phone', 'claim']):
            print(f"📝 Key field: '{header}' → '{field_name}' = '{value}' (type: {type(value).__name__})")
        
        # Special handling for date fields
        if any(term in field_name for term in ['date', 'dol']):
            parsed_date = parse_excel_date_openpyxl(value)
            if parsed_date:
                print(f"📅 Parsed date: {header} → {parsed_date}")
                if 'loss' in field_name:
                    client_data['date_of_loss'] = parsed_date
                else:
                    client_data[field_name] = parsed_date
            else:
                client_data[field_name] = value  # Store original if date parsing fails
            continue
        
        # Handle boolean fields
        if isinstance(value, str) and value.lower() in ('yes', 'no', 'true', 'false', 'y', 'n'):
            value = value.lower() in ('yes', 'true', 'y')
            print(f"⚡ Converted boolean: {header} → {value}")
        elif isinstance(value, (int, float)) and value in (0, 1):
            # Handle 1/0 as boolean
            value = bool(value)
            print(f"⚡ Converted numeric boolean: {header} → {value}")
        
        # Handle numeric fields
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            # It's already a number, keep as is
            pass
        elif isinstance(value, str) and value.replace('.', '').replace(',', '').isdigit():
            try:
                # Remove commas and convert to float
                value = float(value.replace(',', ''))
                print(f"🔢 Converted string to numeric: {header} → {value}")
            except ValueError:
                pass
        
        client_data[field_name] = value
    
    print(f"✅ Extracted {len(client_data)} client data fields from jobinfo (with formula support)")
    return client_data



def import_rooms_from_rooms_file(excel_file, client):
    """
    Extract room data from 01-ROOMS.xlsm file from ROOMS# tab
    WITH FORMULA SUPPORT - gets calculated values from formulas
    """
    try:
        # Use openpyxl to read Excel with formula values
        excel_file.seek(0)
        wb = openpyxl.load_workbook(BytesIO(excel_file.read()), data_only=True)  # data_only=True gets calculated values
        
        if 'ROOMS#' not in wb.sheetnames:
            raise ValueError(f"ROOMS# sheet not found in ROOMS file. Available sheets: {', '.join(wb.sheetnames)}")
        
        ws = wb['ROOMS#']
        
        print(f"✅ Loaded ROOMS# sheet: {ws.max_row} rows, {ws.max_column} columns")
        print(f"📊 Columns available: A-{get_column_letter(ws.max_column)}")
        
        rooms_data = extract_room_data_from_rooms_sheet_openpyxl(ws)
        wb.close()  # Close workbook to free memory
        
        if not rooms_data:
            raise ValueError("No room data found in ROOMS# sheet")
        
        # Create rooms in database
        rooms_created, wt_values_created = create_rooms_for_client(client, rooms_data)
        
        return {
            'rooms_processed': rooms_created,
            'work_type_values_created': wt_values_created,
            'total_rooms_found': len(rooms_data),
            'rooms_sample': rooms_data[:3] if rooms_data else []
        }
        
    except Exception as e:
        raise Exception(f"Failed to process ROOMS file: {str(e)}")

def extract_room_data_from_rooms_sheet_openpyxl(worksheet):
    """
    Enhanced room data extraction using openpyxl with formula support
    Columns U to BT, with each work type section being 5 columns apart
    """
    rooms_data = []
    
    # Define the work types and their column ranges (1-indexed for openpyxl)
    work_type_sections = {
        100: range(21, 26),   # U-Y (columns 21-25)
        200: range(26, 31),   # Z-AD (columns 26-30)  
        300: range(31, 36),   # AE-AI (columns 31-35)
        400: range(36, 41),   # AJ-AN (columns 36-40)
        500: range(41, 46),   # AO-AS (columns 41-45)
        800: range(46, 51),   # AT-AX (columns 46-50)
        6100: range(51, 56),  # AY-BC (columns 51-55) - DAY 1
        6200: range(56, 61),  # BD-BH (columns 56-60) - DAY 2  
        6300: range(61, 66),  # BI-BM (columns 61-65) - DAY 3
        6400: range(66, 71),  # BN-BR (columns 66-70) - DAY 4
    }
    
    print(f"🔍 Scanning for room data in columns U to BT with formula support...")
    print(f"📋 Work types configured: {list(work_type_sections.keys())}")
    
    rooms_found = 0
    max_rows_to_scan = min(worksheet.max_row, 200)  # Limit scanning to first 200 rows for performance
    
    # Iterate through rows to find room data
    for row in range(1, max_rows_to_scan + 1):
        # Check if this row contains room data by looking at first work type section
        first_section_cols = work_type_sections[100]
        room_name_parts = []
        
        for col in first_section_cols:
            if col <= worksheet.max_column:
                cell = worksheet.cell(row=row, column=col)
                if cell.value is not None and str(cell.value).strip():
                    room_name_parts.append(str(cell.value).strip())
        
        # If we have room name parts, this is a room row
        if room_name_parts:
            room_name = ' '.join(room_name_parts).strip()
            # Filter out header rows and empty names
            if (room_name and 
                room_name not in ['', 'Room Name', 'Room', 'ROOM NAME', 'ROOM'] and
                not any(keyword in room_name.upper() for keyword in ['HEADER', 'TITLE', 'DESCRIPTION'])):
                
                rooms_found += 1
                room_data = {
                    'room_name': room_name,
                    'sequence': len(rooms_data) + 1,
                    'work_type_values': {}
                }
                
                if rooms_found <= 3:  # Log first 3 rooms for debugging
                    print(f"🚪 Found room {rooms_found}: '{room_name}'")
                if value_col <= worksheet.max_column:
                    cell = worksheet.cell(row=row, column=value_col)
    
                # Extract work type values for ALL 10 work types
                work_types_found = 0
                for wt_id, col_range in work_type_sections.items():
                    # Use the LAST column of each 5-column section for the value
                    value_col = list(col_range)[4]  # Last column of the 5-column range
                    
                    if cell.value is not None and str(cell.value).strip():
                        value_str = str(cell.value).strip().upper()
                        value_type = determine_los_travel_value_enhanced(value_str)
                            
                        if value_type != 'NA':  # Only store if it has a meaningful value
                            room_data['work_type_values'][wt_id] = value_type
                            work_types_found += 1
                
                if work_types_found > 0 and rooms_found <= 3:
                    print(f"   📋 Work types with values: {work_types_found}")
                    # Show sample work type values for first few rooms
                    sample_wt = list(room_data['work_type_values'].items())[:2]
                    for wt_id, val_type in sample_wt:
                        print(f"     - {wt_id}: {val_type}")
                
                rooms_data.append(room_data)
    
    print(f"✅ Found {len(rooms_data)} rooms with work type data (formula values extracted)")
    
    # Show detailed sample of what was extracted
    if rooms_data:
        print("📊 Room data sample (with formula values):")
        for i, room in enumerate(rooms_data[:2]):
            wt_details = []
            for wt_id, val_type in list(room['work_type_values'].items())[:4]:  # Show first 4 work types
                wt_details.append(f"{wt_id}:{val_type}")
            print(f"   {i+1}. '{room['room_name']}' - Work types: {', '.join(wt_details)}")
    
    return rooms_data
def parse_excel_date_openpyxl(value):
    """
    Enhanced date parser for openpyxl values
    Handles datetime objects, Excel serial numbers, and strings
    """
    if value is None:
        return None
    
    try:
        # If it's already a datetime object
        if isinstance(value, (dt.datetime, dt.date)):
            return value.date() if isinstance(value, dt.datetime) else value
        
        # If it's an Excel serial number (float)
        if isinstance(value, (int, float)):
            print(f"🔢 Parsing Excel date number: {value}")
            try:
                # Excel's date system starts from 1899-12-30
                if value < 0:
                    return None
                if value == 0:  # Excel's zero date
                    return None
                if value == 60:  # Excel's 1900-02-29 (non-existent)
                    return dt.date(1900, 2, 28)
                if value < 60:  # Adjust for Excel's leap year bug
                    value += 1
                return (dt.datetime(1899, 12, 30) + dt.timedelta(days=value)).date()
            except Exception:
                return None
        
        # If it's a string, try parsing
        if isinstance(value, str):
            value = value.strip()
            if not value or value.upper() in ('TBD', 'NA', 'N/A', 'UNKNOWN'):
                return None
            
            # Clean the string
            value = (value.replace('  ', ' ')
                     .replace('Sept', 'Sep')
                     .replace('Febr', 'Feb')
                     .replace('Dece', 'Dec')
                     .split(' ')[0])  # Take only first part if space separated
            
            # Try parsing as datetime string
            date_formats = [
                '%Y-%m-%d %H:%M:%S',  # 2024-12-02 00:00:00
                '%Y-%m-%d',           # 2024-12-02
                '%m/%d/%Y',           # 1/20/2024
                '%d-%b-%y',           # 24-Jan-24
                '%d-%b-%Y',           # 20-Jan-2024
                '%b %d, %Y',          # Dec 23, 2022
                '%B %d, %Y',          # December 23, 2022
                '%d-%m-%Y',           # 20-01-2024
                '%d %b %Y',           # 20 Jan 2024
                '%d %B %Y'            # 20 January 2024
            ]
            
            for fmt in date_formats:
                try:
                    parsed_dt = dt.datetime.strptime(value, fmt)
                    return parsed_dt.date()
                except ValueError:
                    continue
        
        return None
        
    except Exception as e:
        print(f"❌ Date parsing error for value '{value}': {str(e)}")
        return None

def determine_los_travel_value_enhanced(value_str):
    """
    Enhanced value determination for LOS/TRAVEL/NA
    Handles numeric values from formulas properly
    """
    if value_str is None:
        return 'NA'
    
    if isinstance(value_str, (int, float)):
        # Numeric value from formula - treat as LOS
        return 'LOS'
    
    value_clean = str(value_str).upper().strip()
    
    if not value_clean or value_clean in ['', 'N/A', 'NA', 'NULL', 'NONE', 'NAN']:
        return 'NA'
    
    # Handle TRAVEL indicators
    if any(travel_indicator in value_clean for travel_indicator in ['TRAVEL', 'TRVL', 'TRV', 'TRAV', "TRAVEL/AREA"]):
        return 'TRAVEL'
    
    # Handle LOS indicators
    if any(los_indicator in value_clean for los_indicator in ['LOS', 'LINE OF SIGHT', 'SIGHT', 'L.O.S']):
        return 'LOS'
    
    # If it's a pure number (integer or decimal), treat as LOS
    try:
        # Remove any non-numeric characters except decimal point and negative sign
        numeric_value = ''.join(c for c in value_clean if c.isdigit() or c in '.-')
        if numeric_value and numeric_value != '-' and numeric_value != '.':
            float(numeric_value)
            return 'LOS'
    except ValueError:
        pass
    
    # Check for numeric values with text or symbols
    if any(char.isdigit() for char in value_clean):
        return 'LOS'
    
    # Default to NA if we can't determine
    return 'NA'

def ensure_work_types_exist():
    """
    Make sure all work types are created in the database
    ENHANCED with all work types (100-700 + MC readings)
    """
    work_types = {}

    # Define all work types
    work_type_definitions = [
        (100, 'Overview'),
        (200, 'Source'),
        (300, 'CPS'),
        (400, 'PPR'),
        (500, 'Demo'),
        (600, 'Mitigation'),
        (700, 'HMR'),
        (6100, 'DAY 1 MC Readings'),
        (6200, 'DAY 2 MC Readings'),
        (6300, 'DAY 3 MC Readings'),
        (6400, 'DAY 4 MC Readings'),
    ]
    
    print("🔧 Ensuring work types exist in database...")
    
    for wt_id, wt_name in work_type_definitions:
        work_type, created = WorkType.objects.get_or_create(
            work_type_id=wt_id,
            defaults={'name': wt_name}
        )
        if created:
            print(f"✅ Created new work type: {wt_id} - {wt_name}")
        else:
            print(f"✓ Work type exists: {wt_id} - {wt_name}")
        work_types[wt_id] = work_type
    
    print(f"✅ All {len(work_types)} work types verified")
    return work_types

def create_rooms_for_client(client, rooms_data):
    """
    Create room and work type data for a client
    ENHANCED with better logging
    """
    work_types = ensure_work_types_exist()
    
    # Delete existing rooms for this client
    existing_rooms_count = client.rooms.count()
    if existing_rooms_count > 0:
        print(f"🗑️ Removing {existing_rooms_count} existing rooms for client {client.pOwner}")
        client.rooms.all().delete()
    
    rooms_created = 0
    work_type_values_created = 0
    
    print(f"🏗️ Creating {len(rooms_data)} rooms for client {client.pOwner}...")
    
    for room_info in rooms_data:
        room = Room.objects.create(
            client=client,
            room_name=room_info['room_name'],
            sequence=room_info['sequence']
        )
        rooms_created += 1
        
        # Create work type values for this room
        for wt_id, value_type in room_info.get('work_type_values', {}).items():
            RoomWorkTypeValue.objects.create(
                room=room,
                work_type=work_types[wt_id],
                value_type=value_type
            )
            work_type_values_created += 1
        
        # Log first few rooms
        if rooms_created <= 3:
            print(f"   {rooms_created}. {room.room_name} - {len(room_info.get('work_type_values', {}))} work types")
    
    print(f"✅ Created {rooms_created} rooms with {work_type_values_created} work type values")
    return rooms_created, work_type_values_created

def create_or_update_client(client_data):
    """
    Create or update client based on extracted data
    INTEGRATED with your API response structure
    """
    # Map client data to your model fields
    mapped_data = map_client_data_to_model(client_data)
    
    # Extract the owner name to check for existing client
    owner_name = mapped_data.get('pOwner', '')
    if not owner_name:
        raise ValueError("No property owner name found in INFO file")
    
    existing_client = Client.objects.filter(pOwner=owner_name).first()
    
    if existing_client:
        # UPDATE EXISTING CLIENT
        update_fields = {}
        for field, new_value in mapped_data.items():
            if hasattr(existing_client, field) and new_value is not None:
                current_value = getattr(existing_client, field)
                if current_value != new_value:
                    update_fields[field] = new_value
        
        if update_fields:
            Client.objects.filter(pk=existing_client.pk).update(**update_fields)
            print(f"🔄 Updated client {owner_name}: {len(update_fields)} fields changed")
            client_data['_action'] = 'updated'
        else:
            print(f"➖ No changes for client {owner_name}")
            client_data['_action'] = 'unchanged'
        
        # Refresh to get updated instance
        existing_client.refresh_from_db()
        return existing_client
    else:
        # CREATE NEW CLIENT
        client = Client.objects.create(**mapped_data)
        print(f"🆕 Created new client {owner_name}")
        client_data['_action'] = 'created'
        return client

def import_from_master_insurer_file(excel_file):
    """
    Import multiple clients from MASTER-Insurer format
    Structure: Column C = field names, Columns D onward = client data (one client per column)
    """
    try:
        # Use openpyxl to read Excel
        excel_file.seek(0)
        wb = openpyxl.load_workbook(BytesIO(excel_file.read()), data_only=True)
        ws = wb.active
        
        print(f"📊 Processing MASTER-Insurer file: {ws.max_row} rows, {ws.max_column} columns")
        
        # Map excel column letters to column numbers
        # Column C = 3, Column D = 4, etc.
        HEADER_COLUMN = 3  # Column C contains field names
        
        # Find all field names from Column C
        field_mapping = {}  # {row_number: field_name}
        
        for row in range(1, ws.max_row + 1):
            header_cell = ws.cell(row=row, column=HEADER_COLUMN)
            if header_cell.value and str(header_cell.value).strip():
                raw_header = str(header_cell.value).strip()
                
                # Normalize field name (same logic as your other function)
                field_name = (raw_header.lower()
                            .replace(' ', '_')
                            .replace('-', '_')
                            .replace('#', 'num')
                            .replace(':', '')
                            .replace('  ', '_')
                            .replace('__', '_')
                            .strip('_'))
                
                # Special handling for specific fields
                if 'property_owner' in field_name:
                    field_name = 'pOwner'  # Map to your model field
                elif 'claim_num' in field_name or 'claim#' in field_name:
                    field_name = 'claim_num'
                elif 'insurance_co' in field_name:
                    field_name = 'ins_company'
                
                field_mapping[row] = field_name
                print(f"📝 Field mapping: row {row} '{raw_header}' → '{field_name}'")
        
        # Process each client column (starting from Column D = 4)
        clients_data = []
        client_columns_processed = 0
        
        for col in range(4, ws.max_column + 1):  # Start from Column D
            client_data = {}
            has_data = False
            
            for row, field_name in field_mapping.items():
                value_cell = ws.cell(row=row, column=col)
                value = value_cell.value
                
                if value is not None and str(value).strip():
                    has_data = True
                    
                    # Process value based on field type
                    if any(date_term in field_name for date_term in ['date', 'dol']):
                        parsed_date = parse_excel_date_openpyxl(value)
                        if parsed_date:
                            client_data[field_name] = parsed_date
                            print(f"📅 Parsed date for {field_name}: {value} → {parsed_date}")
                        else:
                            client_data[field_name] = value
                    elif isinstance(value, str) and value.upper() in ('Y', 'N', 'YES', 'NO'):
                        client_data[field_name] = value.upper() in ('Y', 'YES')
                    else:
                        client_data[field_name] = value
            
            # Only add client if we found data (not an empty column)
            if has_data:
                # Extract room areas from specific fields
                rooms = []
                for i in range(1, 26):  # Room/Area 1 to 25
                    room_field = f'room_area_{i}'
                    if room_field in client_data:
                        room_name = client_data.pop(room_field)
                        if room_name and str(room_name).strip():
                            rooms.append({
                                'room_name': str(room_name).strip(),
                                'sequence': i
                            })
                
                if rooms:
                    client_data['_rooms'] = rooms
                    print(f"🚪 Found {len(rooms)} rooms for client")
                
                clients_data.append(client_data)
                client_columns_processed += 1
                
                # Show first client for debugging
                if client_columns_processed == 1:
                    print(f"👤 First client sample data:")
                    print(f"  - Owner: {client_data.get('pOwner', 'Unknown')}")
                    print(f"  - Address: {client_data.get('property_address_street', 'Unknown')}")
                    print(f"  - Claim #: {client_data.get('claim_num', 'Unknown')}")
        
        print(f"✅ Processed {client_columns_processed} clients from MASTER file")
        wb.close()
        
        return clients_data
        
    except Exception as e:
        raise Exception(f"Failed to process MASTER-Insurer file: {str(e)}")

def map_client_data_to_model(raw_data):
    """
    Map extracted data to Client model fields - FIXED FOR YOUR MODEL
    """
    # Helper functions
    def get_val(key, default=''):
        value = raw_data.get(key)
        if value is None or pd.isna(value):
            return default
        if isinstance(value, str) and not value.strip():
            return default
        return value
    
    def get_bool(key, default=False):
        value = get_val(key)
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            return value.upper() in ['Y', 'YES', 'TRUE', '1']
        if isinstance(value, (int, float)):
            return bool(value)
        return default
    
    def get_date(key):
        """Convert to datetime for DateTimeField"""
        value = get_val(key)
        if not value:
            return None

        # If it's already a date or datetime object, use it
        if isinstance(value, dt.datetime):
            return value
        if isinstance(value, dt.date):
            return dt.datetime.combine(value, dt.time.min)

        # Otherwise, parse it
        date_val = parse_excel_date(value)
        if date_val:
            # Convert date to datetime with midnight time
            return dt.datetime.combine(date_val, dt.time.min)
        return None
    
    return {
        # Property Owner Information - ALL GOOD
        'pOwner': get_val('property_owner_name'),
        'pAddress': get_val('property_address_street'),
        'pCityStateZip': get_val('property_city_state_zip'),
        'cEmail': get_val('customer_email'),
        'cPhone': get_val('cst_owner_phonenum'),
        
        # Co-Owner Information - ALL GOOD
        'coOwner2': get_val('co_owner_cst2'),
        'cPhone2': get_val('cst_ph_num_2'),
        'cAddress2': get_val('cst_address_num_2'),
        'cCityStateZip2': get_val('cst_city_state_zip_2'),
        'cEmail2': get_val('email_cst_num_2'),
        
        # Claim Information - FIXED DATES
        'causeOfLoss': get_val('cause_of_loss_2'),
        'dateOfLoss': get_date('date_of_loss_2') or timezone.now(),
        'rebuildType1': get_val('rebuild_type_1'),
        'rebuildType2': get_val('rebuild_type_2'),
        'rebuildType3': get_val('rebuild_type_3'),
        'demo': get_bool('demo'),
        'mitigation': get_bool('mitigation'),
        'otherStructures': get_bool('other_structures'),
        'replacement': get_bool('replacement'),
        'CPSCLNCONCGN': get_bool('cps_cln_con_cgn'),
        'yearBuilt': get_val('year_built'),
        'contractDate': get_date('contract_date') or timezone.now(),  # DateTimeField
        # 'lossOfUse': get_val('loss_of_use_ale'),  # REMOVED - not in model
        'breathingIssue': get_val('breathing_issue'),
        'hazardMaterialRemediation': get_val('hmr'),
        
        # Insurance Information - ALL GOOD
        'insuranceCo_Name': get_val('insurance_co_name'),
        'claimNumber': get_val('claim_num'),
        'policyNumber': get_val('policy_num'),
        'emailInsCo': get_val('email_ins_co'),
        'deskAdjusterDA': get_val('desk_adjuster_da'),
        'DAPhone': get_val('da_phone'),
        'DAPhExt': get_val('da_ph_ext_num'),
        'DAEmail': get_val('da_email'),
        'fieldAdjusterName': get_val('field_adjuster_name'),
        'phoneFieldAdj': get_val('phone_num_field_adj'),
        'fieldAdjEmail': get_val('field_adj_email'),
        'adjContents': get_val('adj_contents'),
        'adjCpsPhone': get_val('adj_cps_phone_num'),
        'adjCpsEmail': get_val('adj_cps_email'),
        'emsAdj': get_val('tmp_adj'),
        'emsAdjPhone': get_val('tmp_adj_phone_num'),
        'emsTmpEmail': get_val('adj_tmp_email'),
        'attLossDraftDept': get_val('att_loss_draft_dept'),
        'insAddressOvernightMail': get_val('address_ins_overnight_mail'),
        'insCityStateZip': get_val('city_state_zip_ins'),
        'insuranceCoPhone': get_val('insurance_co_phone'),
        'insWebsite': get_val('website_ins_co'),
        'insMailingAddress': get_val('mailing_address_ins'),
        'insMailCityStateZip': get_val('mail_city_state_zip_ins'),
        'mortgageCoFax': get_val('fax_ins_co'),
        
        # Rooms Information - ALL GOOD
        'newCustomerID': get_val('new_customer_num'),
        'roomID': get_val('room_id'),
        
        # Mortgage Information - ALL GOOD
        'mortgageCo': get_val('mortgage_co'),
        'mortgageAccountCo': get_val('account_num_mtge_co'),
        'mortgageContactPerson': get_val('contact_person_mtge'),
        'mortgagePhoneContact': get_val('phone_num_mtge_contact'),
        'mortgagePhoneExtContact': get_val('ph_ext_mtge_contact'),
        'mortgageAttnLossDraftDept': get_val('attn_loss_draft_dept'),
        'mortgageOverNightMail': get_val('mtge_ovn_mail'),
        'mortgageCityStZipOVN': get_val('city_st_zip_mtge_ovn'),
        'mortgageEmail': get_val('email_mtge'),
        'mortgageWebsite': get_val('mtge_website'),
        'mortgageCoFax': get_val('mtge_co_fax_num'),
        'mortgageMailingAddress': get_val('mailing_address_mtge'),
        'mortgageInitialOfferPhase1ContractAmount': get_val('initial_offer_phase_1_contract_amount'),
        
        # Cash Flow - ALL GOOD
        'drawRequest': get_val('draw_request'),
        # 'custId': get_val('cust_id'),  # REMOVED - not in model
        
        # Contractor Information - ALL GOOD
        'coName': get_val('co_name'),
        'coWebsite': get_val('co_website'),
        'coEmailstatus': get_val('co_emailstatus'),
        'coAddress': get_val('co_adress'),
        'coCityState': get_val('co_city_state'),
        'coAddress2': get_val('co_address_2'),
        'coCityState2': get_val('co_city_state_2'),
        'coCityState3': get_val('co_city_state_3'),
        'coLogo1': get_val('co_logo_1'),
        'coLogo2': get_val('co_logo_2'),
        'coLogo3': get_val('co_logo_3'),
        'coRepPH': get_val('co_rep_ph'),
        'coREPEmail': get_val('co_rep_email'),
        'coPhone2': get_val('co_ph_num_2'),
        'TinW9': get_val('tin_w9'),
        'fedExAccount': get_val('fedex_account_num'),

        # Claim Reporting - FIXED DATE
        'claimReportDate': get_date('claim_report_date') or timezone.now(),  # DateTimeField
        'insuranceCustomerServiceRep': get_val('co_represesntative'),
        'timeOfClaimReport': get_val('time_of_claim_report'),
        'phoneExt': get_val('phone_ext'),
        'tarpExtTMPOk': get_bool('tarp_ext_tmp_ok'),
        'IntTMPOk': get_bool('int_tmp_ok'),
        'DRYPLACUTOUTMOLDSPRAYOK': get_bool('drypla_cutout_mold_spray_ok'),
        
        # ALE Information - CORRECTED FOR YOUR MODEL
        'lossOfUseALE': get_val('ale_info'),  # This matches your model
        'ale_lessee_name': get_val('tenant_lesee'),
        'ale_lessee_home_address': get_val('property_address_street_ale'),
        'ale_lessee_city_state_zip': get_val('property_city_state_zip_ale'),
        'ale_lessee_email': get_val('customer_email_ale'),
        'ale_lessee_phone': get_val('cst_owner_phonenum_ale'),
        'ale_rental_bedrooms': get_val('bedrooms'),
        'ale_rental_months': get_val('months'),
        'ale_rental_start_date': get_date('start_date') or timezone.now(),  # DateField
        'ale_rental_end_date': get_date('end_date') or timezone.now(),  # DateField
        
        # Handle decimal for amount
        'ale_rental_amount_per_month': parse_decimal(get_val('terms_amount')),
        
        # Lessor Information - CORRECTED
        'ale_lessor_name': get_val('lessor'),
        'ale_lessor_leased_address': get_val('leased_address'),
        'ale_lessor_city_zip': get_val('city_zip_lessor'),
        'ale_lessor_phone': get_val('phone_lessor'),
        'ale_lessor_email': get_val('email_lessor'),
        'ale_lessor_mailing_address': get_val('lessor_mailing_address'),
        'ale_lessor_mailing_city_zip': get_val('city_zip_lessor_mail'),
        'ale_lessor_contact_person': get_val('lessor_contact_person'),
        
        # Real Estate Company - CORRECTED
        'ale_re_company_name': get_val('real_estate_company'),
        'ale_re_mailing_address': get_val('mailing_address_re'),
        'ale_re_city_zip': get_val('city_zip_re'),
        'ale_re_contact_person': get_val('contact_re'),
        'ale_re_phone': get_val('phone_re'),
        'ale_re_email': get_val('email_re'),
        'ale_re_owner_broker_name': get_val('owner_broker'),
        'ale_re_owner_broker_phone': get_val('phone_owner_broker'),
        'ale_re_owner_broker_email': get_val('email_owner_broker'),
    
    } 

def parse_decimal(value):
    """Parse string to decimal for monetary values"""
    if not value or pd.isna(value):
        return None
    
    try:
        # Handle various input types
        if isinstance(value, (int, float)):
            return Decimal(str(value))
        
        # Handle string values
        if isinstance(value, str):
            value = value.replace('$', '').replace(',', '').strip()
            if not value or value.upper() in ['NA', 'N/A', 'NULL', '']:
                return None
            return Decimal(value)
        
        # Try to convert anything else
        return Decimal(str(value))
    except (ValueError, TypeError, InvalidOperation):
        return None

# MAIN IMPORT FUNCTION
@csrf_exempt
def import_client_with_rooms_formula_support(request):
    """
    Main endpoint to import client from 01-INFO and rooms from 01-ROOMS
    WITH FORMULA SUPPORT - gets calculated values from Excel formulas
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    if 'info_file' not in request.FILES:
        return JsonResponse({'error': 'No INFO file provided'}, status=400)
    
    if 'rooms_file' not in request.FILES:
        return JsonResponse({'error': 'No ROOMS file provided'}, status=400)
    
    try:
        info_file = request.FILES['info_file']
        rooms_file = request.FILES['rooms_file']
        
        print("🚀 Starting complete client import process WITH FORMULA SUPPORT...")
        
        # Step 1: Import client data from INFO file (with formula support)
        client_data = import_client_from_info_file(excel_file=info_file)
        
        # Step 2: Create or update client
        client = create_or_update_client(client_data)
        
        # Step 3: Import room data from ROOMS file (with formula support)
        room_results = import_rooms_from_rooms_file(excel_file=rooms_file, client=client)
        
        response_data = {
            'status': 'success',
            'client': {
                'id': client.id,
                'name': client.pOwner,
                'action': client_data.get('_action', 'unknown')
            },
            'rooms': room_results,
            'processing_summary': {
                'client_data_processed': True,
                'rooms_processed': room_results['rooms_processed'],
                'work_type_values_created': room_results['work_type_values_created'],
                'total_rooms_found': room_results['total_rooms_found'],
                'formula_support': True  # Indicate that formula values were extracted
            }
        }
        
        print(f"✅ Import completed successfully with formula support: {response_data}")
        return JsonResponse(response_data)
        
    except Exception as e:
        error_msg = f"❌ Complete import error (formula support): {str(e)}"
        print(error_msg)
        import traceback
        print(f"🔍 Stack trace: {traceback.format_exc()}")
        
        return JsonResponse({
            'status': 'error',
            'message': str(e),
            'formula_support': True
        }, status=500)

# KEEP YOUR EXISTING parse_excel_date FUNCTION
def parse_excel_date(value):
    """Parse Excel date values to Python date"""
    if not value or pd.isna(value):
        return None
    
    try:
        # If it's already a datetime/date
        if isinstance(value, (datetime, date)):
            if isinstance(value, datetime):
                return value.date()
            return value
        
        # If it's an Excel serial number (float)
        if isinstance(value, (int, float)):
            # Excel base date is 1899-12-30 for Windows Excel
            base_date = date(1899, 12, 30)
            if value == 0:
                return None
            return base_date + dt.timedelta(days=value)
        
        # If it's a string
        if isinstance(value, str):
            value = value.strip()
            if not value:
                return None
            # Try various date formats
            for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y/%m/%d']:
                try:
                    return dt.datetime.strptime(value, fmt).date()
                except ValueError:
                    continue
        
        return None
    except Exception as e:
        print(f"⚠️ Error parsing date {value}: {e}")
        return None

def clean_session_data(data):
    """
    Recursively convert non-serializable objects to JSON-serializable formats.
    Handles datetime, date, time, pandas Timestamp, numpy types, and other common cases.
    """
    if isinstance(data, (dt.datetime, dt.date, dt.time)):
        return data.isoformat()
    elif isinstance(data, pd.Timestamp):
        return data.isoformat()
    elif isinstance(data, pd.Series):
        return data.to_dict()
    elif isinstance(data, pd.DataFrame):
        return data.to_dict(orient='records')
    elif isinstance(data, dict):
        return {k: clean_session_data(v) for k, v in data.items()}
    elif isinstance(data, (list, tuple, set)):
        return [clean_session_data(x) for x in data]
    elif hasattr(data, 'tolist'):  # numpy arrays and similar
        return data.tolist()
    elif hasattr(data, 'isoformat'):  # other datetime-like objects
        return data.isoformat()
    elif isinstance(data, (int, float, str, bool)) or data is None:
        return data
    else:
        return str(data)  # fallback to string representation

def handle_excel_import(request):
    # Debug: Check if file exists in request
    if 'excel_file' not in request.FILES:
        messages.error(request, 'No file was uploaded')
        return redirect('create')
    
    excel_file = request.FILES['excel_file']
    
    # Debug: Log file info
    print(f"📁 File received: {excel_file.name} (Size: {excel_file.size/1024:.2f} KB)")
    messages.info(request, f"Processing file: {excel_file.name}")

    if not excel_file.name.endswith(('.xlsx', '.xls', '.xlsm')):
        messages.error(request, 'Invalid file type. Please upload an Excel file (.xlsx, .xls, or .xlsm)')
        return redirect('create')
    
    try:
        # DETECT FILE TYPE AND ROUTE ACCORDINGLY
        file_name = excel_file.name.lower()
        
        if 'master' in file_name or 'insurer' in file_name:
            # Process MASTER-Insurer file format (columns D onward)
            print("🔍 Detected MASTER-Insurer file format")
            return process_master_file(request, excel_file)
        else:
            # Process INFO file format (01-INFO.xlsx with jobinfo(2) tab)
            print("🔍 Detected INFO file format")
            return process_info_file(request, excel_file)
            
    except Exception as e:
        error_msg = f"❌ File processing error: {str(e)}"
        print(error_msg)
        messages.error(request, error_msg)
        return redirect('create')

from decimal import Decimal, InvalidOperation

def process_master_file(request, excel_file):
    """Process 30-MASTER-Insurer.xlsx files with ALL sheet - COMPLETE"""
    try:
        # Debug: Verify sheet exists
        xls = pd.ExcelFile(BytesIO(excel_file.read()))
        
        # Try to find the right sheet
        sheet_to_use = None
        for sheet in xls.sheet_names:
            if sheet.upper() in ['ALL', 'MASTER', 'DATA', 'CLIENTS']:
                sheet_to_use = sheet
                break
        
        if not sheet_to_use:
            # Use first sheet as fallback
            sheet_to_use = xls.sheet_names[0]
        
        print(f"📄 Using sheet: '{sheet_to_use}'")
        
        # Reset file pointer after checking sheets
        excel_file.seek(0)
        df = pd.read_excel(BytesIO(excel_file.read()), sheet_name=sheet_to_use)
        
        # Debug: Show structure
        print("🔍 Excel Structure:")
        print(f"Columns: {df.columns.tolist()}")
        print(f"Shape: {df.shape}")
        print(f"First 5 rows of column C (headers):")
        for i in range(min(5, len(df))):
            print(f"  Row {i+1}: '{df.iloc[i, 2] if 2 < len(df.columns) else 'N/A'}'")
        
        # Constants for MASTER file structure
        HEADER_COLUMN = 2  # Column C (0-indexed)
        FIRST_DATA_COLUMN = 3  # Column D
        
        total_clients = len(df.columns) - FIRST_DATA_COLUMN
        messages.info(request, f"Found {total_clients} potential clients in file (columns D onward)")
        
        success_count = 0
        update_count = 0
        error_count = 0
        processing_details = []
        
        # Build header mapping ONCE
        header_mapping = {}
        print("\n🔍 Building header mapping from Column C:")
        for row_idx in range(len(df)):
            header = str(df.iloc[row_idx, HEADER_COLUMN]).strip() if pd.notna(df.iloc[row_idx, HEADER_COLUMN]) else ""
            if header:
                field_name = normalize_header_for_mapping(header)
                header_mapping[row_idx] = {
                    'original': header,
                    'field': field_name
                }
                print(f"  Row {row_idx+1}: '{header}' → '{field_name}'")
        
        # Process each client column (starting from Column D)
        for col_idx in range(FIRST_DATA_COLUMN, len(df.columns)):
            col_letter = chr(65 + col_idx)  # Excel column letter
            client_status = {
                'column': col_idx + 1,  # Excel column number
                'column_letter': col_letter,
                'name': None,
                'status': None,
                'message': None,
                'dates': [],
                'errors': []
            }

            try:
                print(f"\n{'='*60}")
                print(f"📊 Processing Client Column {col_letter}")
                print(f"{'='*60}")
                
                # Build client data from header mapping
                client_data = {}
                
                # First: Find property owner name
                claim_owner = None
                owner_row = None
                
                for row_idx, mapping in header_mapping.items():
                    if 'property_owner_name' in mapping['field'].lower():
                        value = df.iloc[row_idx, col_idx]
                        if pd.notna(value):
                            claim_owner = str(value).strip()
                            owner_row = row_idx
                            break
                
                if not claim_owner:
                    # Try to find any value in this column
                    for row_idx in range(min(10, len(df))):
                        value = df.iloc[row_idx, col_idx]
                        if pd.notna(value) and str(value).strip():
                            claim_owner = str(value).strip()
                            owner_row = row_idx
                            print(f"⚠️  Found owner name from row {row_idx+1}: '{claim_owner}'")
                            break
                
                if not claim_owner or claim_owner.upper() in ['NA', 'N/A', '']:
                    print(f"⏭️  Column {col_letter}: Empty column, skipping")
                    continue
                
                client_status['name'] = claim_owner
                print(f"👤 Client Name: '{claim_owner}'")
                
                # Collect ALL data for this client
                for row_idx, mapping in header_mapping.items():
                    original_header = mapping['original']
                    field_name = mapping['field']
                    value = df.iloc[row_idx, col_idx]
                    
                    # Skip if value is empty
                    if pd.isna(value) or (isinstance(value, str) and not value.strip()):
                        continue
                    
                    # Special handling for dates
                    if any(term in field_name for term in ['date', 'dol']):
                        parsed_date = parse_excel_date(value)
                        if parsed_date:
                            client_status['dates'].append(f"{original_header}: {parsed_date}")
                            client_data[field_name] = parsed_date
                        else:
                            client_data[field_name] = value
                        continue
                    
                    # Handle boolean fields (Y/N, Yes/No)
                    if isinstance(value, str):
                        value_upper = value.strip().upper()
                        if value_upper in ['Y', 'YES', 'TRUE']:
                            client_data[field_name] = True
                        elif value_upper in ['N', 'NO', 'FALSE']:
                            client_data[field_name] = False
                        else:
                            client_data[field_name] = value
                    else:
                        client_data[field_name] = value
                
                # Debug: Show what we collected
                print(f"📋 Collected {len(client_data)} data points")
                print(f"📅 Dates found: {len(client_status['dates'])}")
                
                # Extract rooms from room_area fields
                rooms_data = extract_rooms_from_master_data(client_data)
                if rooms_data:
                    print(f"🏠 Found {len(rooms_data)} rooms")
                    for room in rooms_data[:3]:  # Show first 3 rooms
                        print(f"  - {room['sequence']}. {room['room_name']}")
                
                # Map to model using COMPLETE mapping
                mapped_data = map_client_data_to_model(client_data)
                
                # Ensure we have the owner name
                if not mapped_data.get('pOwner'):
                    mapped_data['pOwner'] = claim_owner
                    print(f"⚠️  Added owner name to mapped data: '{claim_owner}'")
                
                # Debug: Check critical fields
                critical_fields = ['pOwner', 'pAddress', 'claimNumber', 'insuranceCo_Name']
                for field in critical_fields:
                    value = mapped_data.get(field, 'NOT FOUND')
                    print(f"  {field}: '{value}'")
                
                # Update or create client
                existing_client = Client.objects.filter(pOwner=mapped_data['pOwner']).first()
                
                if existing_client:
                    update_fields = {}
                    for field, new_value in mapped_data.items():
                        if hasattr(existing_client, field) and new_value is not None:
                            current_value = getattr(existing_client, field, None)
                            if current_value != new_value:
                                update_fields[field] = new_value
                    
                    if update_fields:
                        Client.objects.filter(pk=existing_client.pk).update(**update_fields)
                        existing_client.refresh_from_db()
                        client_status['status'] = 'updated'
                        client_status['message'] = f"Updated {len(update_fields)} fields"
                        update_count += 1
                        print(f"🔄 Updated client '{claim_owner}' ({len(update_fields)} fields)")
                    else:
                        client_status['status'] = 'unchanged'
                        client_status['message'] = "No changes needed"
                        print(f"➖ No changes for client '{claim_owner}'")
                    
                    client = existing_client
                else:
                    client = Client.objects.create(**mapped_data)
                    client_status['status'] = 'created'
                    client_status['message'] = "New client created"
                    success_count += 1
                    print(f"🆕 Created new client '{claim_owner}'")
                
                # Create rooms for this client
                if rooms_data:
                    print(f"🏗️ Creating {len(rooms_data)} rooms for '{claim_owner}'...")
                    rooms_created, wt_created = create_rooms_for_client(client, rooms_data)
                    print(f"✅ Created {rooms_created} rooms with {wt_created} work type values")
                    client_status['message'] += f", {rooms_created} rooms created"
                
                processing_details.append(client_status)
                
            except Exception as e:
                import traceback
                error_detail = traceback.format_exc()
                print(f"❌ Error processing column {col_letter}: {str(e)}")
                print(f"Error details:\n{error_detail}")
                
                client_status['status'] = 'failed'
                error_msg = f"Column {col_letter}: {str(e)}"
                client_status['errors'].append(error_msg)
                error_count += 1
                processing_details.append(client_status)
        
        # Prepare final results
        print(f"\n{'='*60}")
        print(f"✅ IMPORT COMPLETE")
        print(f"{'='*60}")
        print(f"   Successfully created: {success_count}")
        print(f"   Updated: {update_count}")
        print(f"   Failed: {error_count}")
        print(f"   Total columns processed: {len(df.columns) - FIRST_DATA_COLUMN}")
        
        # Store results in session
        session_data = {
            'success_count': success_count,
            'update_count': update_count,
            'error_count': error_count,
            'processing_details': processing_details[:20],
            'file_name': excel_file.name,
            'timestamp': timezone.now().isoformat()
        }
        
        request.session['import_results'] = clean_session_data(session_data)
        
        result_message = f"Import complete: {success_count} created, {update_count} updated, {error_count} errors"
        messages.success(request, result_message)
        
        return render(request, 'account/create.html', {
            'form': ClientForm(),
            'import_summary': {
                'success_count': success_count,
                'update_count': update_count,
                'error_count': error_count,
                'total_clients': total_clients,
                'rooms_imported': sum(len(d.get('rooms', [])) for d in processing_details if 'rooms' in d)
            }
        })
        
    except Exception as e:
        import traceback
        error_msg = f"❌ MASTER file processing error: {str(e)}"
        print(error_msg)
        print(f"Error details:\n{traceback.format_exc()}")
        messages.error(request, error_msg)
        return redirect('create')

def extract_rooms_from_master_data(client_data):
    """Extract rooms from MASTER file client data for Room model"""
    rooms_data = []
    room_name_counts = {}  # Track duplicate room names

    for room_num in range(1, 26):
        room_field = f'room_area_{room_num}'
        room_value = client_data.get(room_field)

        if room_value:
            if isinstance(room_value, str):
                room_value = room_value.strip()
            else:
                room_value = str(room_value).strip()

        # Check if it's a valid room name
        if (room_value and
            room_value not in ['', 'NA', 'N/A', 'None', 'nan'] and
            len(room_value) > 1):

            # Handle duplicate room names by appending a number
            original_name = room_value
            if original_name in room_name_counts:
                room_name_counts[original_name] += 1
                room_value = f"{original_name} ({room_name_counts[original_name]})"
            else:
                room_name_counts[original_name] = 1

            rooms_data.append({
                'room_name': room_value,
                'sequence': room_num,
                'work_type_values': {}  # Empty for MASTER files
            })

    return rooms_data

def normalize_header_for_mapping(header):
    """Normalize Excel header to field name for MASTER files"""
    if not header or pd.isna(header):
        return ""
    
    header_str = str(header).strip()
    
    # Special mapping for your exact headers
    header_mapping = {
        # Property Owner Information
        'Property-Owner Name': 'property_owner_name',
        'Property address: street': 'property_address_street',
        'Property city, state, zip': 'property_city_state_zip',
        'Customer Email': 'customer_email',
        'Cst-owner Phone#': 'cst_owner_phonenum',
        
        # Co-Owner Information
        'Co-Owner.cst#2': 'co_owner_cst2',
        'cst ph # 2': 'cst_ph_num_2',
        'Cst address # 2': 'cst_address_num_2',
        'city, state-cst#2': 'cst_city_state_zip_2',
        'email-cst #2': 'email_cst_num_2',
        
        # Claim Information
        'Cause of Loss': 'cause_of_loss',
        'date of loss': 'date_of_loss',
        'rebuild  type 1': 'rebuild_type_1',
        'rebuild  type 2': 'rebuild_type_2',
        'rebuild  type 3': 'rebuild_type_3',
        'DEMO': 'demo',
        'Mitigation': 'mitigation',
        'Other Structures': 'other_structures',
        'Replacement': 'replacement',
        'CPS / CLN / CON/ CGN': 'cps_cln_con_cgn',
        'Year Built': 'year_built',
        'Contract Date': 'contract_date',
        'Loss of use/ ALE': 'loss_of_use_ale',
        'Breathing issue': 'breathing_issue',
        'HMR': 'hmr',
        
        # Insurance Information
        'Insurance Co. Name': 'insurance_co_name',
        'Claim #': 'claim_num',
        'policy #': 'policy_num',
        'Email INS. co.': 'email_ins_co',
        'DESK Adjuster DA': 'desk_adjuster_da',
        'DA Phone': 'da_phone',
        'DA Ph. Ext. #': 'da_ph_ext_num',
        'DA Email': 'da_email',
        'Field Adjuster Name': 'field_adjuster_name',
        'Phone # field adj': 'phone_num_field_adj',
        'Field adj email': 'field_adj_email',
        'adj contents': 'adj_contents',
        'adj CPS phone #': 'adj_cps_phone_num',
        'adj CPS email': 'adj_cps_email',
        'TMP adj': 'tmp_adj',
        'TMP adj phone #': 'tmp_adj_phone_num',
        'adj TMP email': 'adj_tmp_email',
        'ATT: Loss Draft Dept.': 'att_loss_draft_dept',
        'address ins overnight mail': 'address_ins_overnight_mail',
        'city, state-zip ins': 'city_state_zip_ins',
        'Insurance Co. Phone': 'insurance_co_phone',
        'Website Ins Co.': 'website_ins_co',
        'Mailing   address INS': 'mailing_address_ins',
        'Mail city, state, zip INS': 'mail_city_state_zip_ins',
        'FAX Ins. Co': 'fax_ins_co',
        
        # Rooms Information
        'NEW CUSTOMER #': 'new_customer_num',
        'ROOM ID': 'room_id',
        
        # Room Areas (1-25)
        'Room/Area 1': 'room_area_1',
        'Room/Area 2': 'room_area_2',
        'Room/Area 3': 'room_area_3',
        'Room/Area 4': 'room_area_4',
        'Room/Area 5': 'room_area_5',
        'Room/Area 6': 'room_area_6',
        'Room/Area 7': 'room_area_7',
        'Room/Area 8': 'room_area_8',
        'Room/Area 9': 'room_area_9',
        'Room/Area 10': 'room_area_10',
        'Room/Area 11': 'room_area_11',
        'Room/Area 12': 'room_area_12',
        'Room/Area 13': 'room_area_13',
        'Room/Area 14': 'room_area_14',
        'Room/Area 15': 'room_area_15',
        'Room/Area 16': 'room_area_16',
        'Room/Area 17': 'room_area_17',
        'Room/Area 18': 'room_area_18',
        'Room/Area 19': 'room_area_19',
        'Room/Area 20': 'room_area_20',
        'Room/Area 21': 'room_area_21',
        'Room/Area 22': 'room_area_22',
        'Room/Area 23': 'room_area_23',
        'Room/Area 24': 'room_area_24',
        'Room/Area 25': 'room_area_25',
        
        # Mortgage Information
        'Mortgage co': 'mortgage_co',
        'Account# Mtge Co.': 'account_num_mtge_co',
        'Loan status': 'loan_status',
        'contact person mtge': 'contact_person_mtge',
        'Phone # MTGE contact': 'phone_num_mtge_contact',
        'Ph. Ext. Mtge contact': 'ph_ext_mtge_contact',
        'Attn.: Loss Draft Dept': 'attn_loss_draft_dept',
        'Mtge OVN mail': 'mtge_ovn_mail',
        'city, St., zip ,mtge OVN': 'city_st_zip_mtge_ovn',
        'Phone # MTGE co.': 'phone_num_mtge_co',
        'email mtge': 'email_mtge',
        'mtge website': 'mtge_website',
        'MTGE co. Fax #': 'mtge_co_fax_num',
        'Mailing   address mtge': 'mailing_address_mtge',
        'Mail city, state, zip mtge': 'mail_city_state_zip_mtge',
        'Initial Offer / phase 1 contract amount': 'initial_offer_phase_1_contract_amount',
        
        # Cash Flow
        'Draw Request': 'draw_request',
        'Cust id': 'cust_id',
        
        # Contractor Information
        'co name': 'co_name',
        'Co. website': 'co_website',
        'co. EMAIL/co. status': 'co_emailstatus',
        'co address': 'co_adress',
        'co. city state': 'co_city_state',
        'co. address 2': 'co_address_2',
        'co. city state 2': 'co_city_state_2',
        'co address 3': 'co_address_3',
        'co. city state 3': 'co_city_state_3',
        'Co. logo 1': 'co_logo_1',
        'Co. logo 2': 'co_logo_2',
        'Co. logo 3': 'co_logo_3',
        'Co. REP. / PH': 'co_rep_ph',
        'CO.REP. email': 'co_rep_email',
        'Co PH # 2': 'co_ph_num_2',
        'CO.REP. email 2': 'co_rep_email_2',
        'TIN W9': 'tin_w9',
        'FedEx     account #': 'fedex_account_num',
        
        # Claim Reporting
        'claim report date': 'claim_report_date',
        'Time OF CLAIM REPORT': 'time_of_claim_report',
        'co.represesntative': 'co_represesntative',
        'phone ext.': 'phone_ext',
        'Tarp ext. TMP ok': 'tarp_ext_tmp_ok',
        'Int TMP ok': 'int_tmp_ok',
        'DRY/PLA CUTOUT MOLD SPRAY  OK': 'drypla_cutout_mold_spray_ok',
        
        # ALE Information
        'ALE INFO  …            APC #': 'ale_info',
        'Lesse info / NAME': 'tenant_lesee',
        'HOME ADDRESS': 'property_address_street_ale',
        'Property city, state, zip': 'property_city_state_zip_ale',
        'Customer Email': 'customer_email_ale',
        'Customer Phone#': 'cst_owner_phonenum_ale',
        'RENTAL INFO': 'rental_info',
        'bedrooms': 'bedrooms',
        'months': 'months',
        'START DATE': 'start_date',
        'END DATE': 'end_date',
        'Amount / Month': 'terms_amount',
        
        # Lessor Information
        'LESSOR INFO / NAME': 'lessor',
        'Leased Address': 'leased_address',
        'city zip': 'city_zip_lessor',
        'phone #': 'phone_lessor',
        'Email lessor': 'email_lessor',
        'Lessor mailing Address': 'lessor_mailing_address',
        'city zip': 'city_zip_lessor_mail',
        'LESSOR CONTACT PERSON': 'lessor_contact_person',
        
        # Claim Info (duplicate section)
        'CLAIM INFO': 'claim_info_header',
        'Cause of Loss': 'cause_of_loss_2',
        'date of loss': 'date_of_loss_2',
        'Insurance Co. Name': 'insurance_co_name_2',
        'Claim #': 'claim_num_2',
        'policy #': 'policy_num_2',
        'Email INS. co.': 'email_ins_co_2',
        'DESK Adjuster DA': 'desk_adjuster_da_2',
        'DA Phone': 'da_phone_2',
        'DA Ph. Ext. #': 'da_ph_ext_num_2',
        'DA Email': 'da_email_2',
        
        # Real Estate Company
        'REAL ESTATE COMPANY': 'real_estate_company',
        'MAILING ADDRESS': 'mailing_address_re',
        'city zip': 'city_zip_re',
        'CONTACT': 'contact_re',
        'phone #': 'phone_re',
        'Email': 'email_re',
        'OWNER/BROKER': 'owner_broker',
        'phone #': 'phone_owner_broker',
        'Email': 'email_owner_broker',
    }
    
    # Try exact match first
    if header_str in header_mapping:
        return header_mapping[header_str]
    
    # Fallback: generic normalization
    field_name = header_str.lower()
    field_name = field_name.replace(' ', '_').replace('-', '_').replace('#', 'num')
    field_name = field_name.replace(':', '').replace('.', '').replace(',', '')
    field_name = field_name.replace('__', '_').strip('_')
    
    return field_name



def process_info_file(request, excel_file):
    """Process 01-INFO.xlsx files with jobinfo(2) tab"""
    try:
        # Use your existing import_client_from_info_file function
        client_data = import_client_from_info_file(excel_file)
        
        if not client_data:
            messages.error(request, "No client data found in INFO file")
            return redirect('create')
        
        # Create or update client
        client = create_or_update_client(client_data)
        
        messages.success(request, f'Client {client.pOwner} imported successfully from INFO file!')
        return redirect('dashboard')
        
    except Exception as e:
        error_msg = f"❌ INFO file import failed: {str(e)}"
        print(error_msg)
        messages.error(request, error_msg)
        return redirect('create')


def normalize_header(header):
    """Normalize Excel header to field name"""
    if not header or pd.isna(header):
        return ""
    
    # Convert to string and clean
    header_str = str(header).strip()
    
    # Your existing normalization logic
    field_name = (header_str.lower()
                 .replace(' ', '_')
                 .replace('/', '_')
                 .replace('\\', '_')
                 .replace('.', '_')
                 .replace('-', '_')
                 .replace(':', '_')
                 .replace('__', '_')
                 .replace('#', 'num')
                 .strip('_'))
    
    return field_name
def parse_excel_date(value):
    """Robust date parser that handles all cases"""
    # Handle empty/None values

    try:
        if pd.isna(value) or value in ('', 'TBD', 'NA', 'N/A'):
            return None


        
        # Handle Excel serial numbers
        if isinstance(value, (int, float)):
            print(f"🔢 Parsing Excel date number: {value}")
            try:
                # Excel's date system starts from 1900-01-01 (with 1900 incorrectly treated as leap year)
                if value < 0:
                    return None
                if value == 0:  # Excel's zero date
                    return None
                if value == 60:  # Excel's 1900-02-29 (non-existent)
                    return dt(1900, 2, 28).date()
                if value < 60:  # Adjust for Excel's leap year bug
                    value += 1
                return (dt.datetime(1899, 12, 30) + pd.Timedelta(days=value)).date()
            except Exception:
                return None
        
        # Convert to string if not already
        if not isinstance(value, str):
            value = str(value).strip()
            print(f"🔢 Parsing Excel date number: {value}")
        else:
            value = value.strip()
        
        # Handle obvious non-dates
        if not value or value.upper() in ('TBD', 'NA', 'N/A', 'UNKNOWN'):
            return None
        
        # Clean the string
        value = (value.replace(' ', ' ')  # Non-breaking space
                 .replace('Sept', 'Sep')
                 .replace('Febr', 'Feb')
                 .replace('Dece', 'Dec')
                 .split(' ')[0])  # Take only first part if space separated
        
        # Try parsing as datetime string
        date_formats = [
            '%Y-%m-%d %H:%M:%S',  # 2024-12-02 00:00:00
            '%Y-%m-%d',           # 2024-12-02
            '%m/%d/%Y',           # 1/20/2024
            '%d-%b-%y',           # 24-Jan-24
            '%d-%b-%Y',           # 20-Jan-2024
            '%b %d, %Y',          # Dec 23, 2022
            '%B %d, %Y',          # December 23, 2022
            '%d-%m-%Y',           # 20-01-2024
            '%d %b %Y',           # 20 Jan 2024
            '%d %B %Y'            # 20 January 2024
        ]
        
        for fmt in date_formats:
            try:
                return dt.datetime.strptime(value, fmt).date()
            except ValueError:
                continue
        
        # Handle pandas Timestamp
        try:
            if isinstance(value, (pd.Timestamp, dt)):
                return value.date()
        except:
            pass
        
        # Final check for invalid dates
        if re.match(r'^\d{5,}', value) or re.match(r'.*\D.*\D.*', value):  # Long numbers or multiple non-digits
            return None
        
        return None
    except Exception as e:
        print(f"❌ Date parsing error for value '{value}': {str(e)}")
        return None

def generate_data_report(request):
    if 'import_results' not in request.session:
        messages.error(request, "No import data available to generate report")
        return redirect('create')
    
    import_data = request.session['import_results']
    
    try:
        df = pd.DataFrame.from_dict(import_data['excel_data'])
    except Exception as e:
        messages.error(request, f"Error loading import data: {str(e)}")
        return redirect('create')

    # Initialize report data with additional metadata
    report = {
        'metadata': {
            'report_generated': timezone.now().strftime("%Y-%m-%d %H:%M"),
            'source_file': import_data.get('file_name', 'Unknown')
        },
        'summary': {
            'total_clients': len(df.columns) - 3,  # Subtract ignored columns
            'processed': import_data['success_count'] + import_data['update_count'],
            'errors': import_data['error_count'],
            'success_rate': round((import_data['success_count'] + import_data['update_count']) / max(1, len(df.columns) - 3) * 100, 1)
        },
        'field_stats': {},
        'data_quality_issues': [],
        'most_problematic_fields': []
    }
    
    # Constants for our structure
    HEADER_COLUMN = 2  # Column C
    FIRST_DATA_COLUMN = 3  # Column D
    
    # Track most problematic fields
    field_problem_counts = []
    
    # Analyze each field
    for row_idx in range(len(df)):
        header = str(df.iloc[row_idx, HEADER_COLUMN]).strip().lower() if pd.notna(df.iloc[row_idx, HEADER_COLUMN]) else None
        if not header:
            continue
            
        # Normalize field name (keep original for display)
        field_name = header.replace(' ', '_').replace('/', '_').strip('_')
        display_name = header.title()  # For nicer display
        
        # Initialize field stats
        field_stats = {
            'display_name': display_name,
            'total': 0,
            'empty': 0,
            'tbd': 0,
            'na': 0,
            'invalid': 0,
            'completeness': 0,
            'examples': set()
        }
        
        # Check each client column
        for col_idx in range(FIRST_DATA_COLUMN, len(df.columns)):
            value = df.iloc[row_idx, col_idx]
            if pd.isna(value):
                field_stats['empty'] += 1
                continue
                
            value_str = str(value).strip().upper()
            field_stats['total'] += 1
            
            # Check for placeholder values
            if value_str in ('TBD', 'TO BE DETERMINED'):
                field_stats['tbd'] += 1
            elif value_str in ('NA', 'N/A', '#N/A'):
                field_stats['na'] += 1
            # Special validation for certain fields
            elif 'zip' in field_name and not any(c.isdigit() for c in value_str):
                field_stats['invalid'] += 1
            elif 'date' in field_name and parse_excel_date(value) is None:
                field_stats['invalid'] += 1
                
            # Collect sample values
            if len(field_stats['examples']) < 3:
                field_stats['examples'].add(str(value))
        
        # Calculate completeness percentage
        total_values = field_stats['total'] + field_stats['empty']
        if total_values > 0:
            field_stats['completeness'] = round((field_stats['total'] - field_stats['empty']) / total_values * 100, 1)
        
        # Only include fields with issues in the report
        if field_stats['empty'] or field_stats['tbd'] or field_stats['na'] or field_stats['invalid']:
            report['field_stats'][display_name] = field_stats
            problem_count = sum([field_stats['empty'], field_stats['tbd'], field_stats['na'], field_stats['invalid']])
            field_problem_counts.append((display_name, problem_count))
    
    # Generate quality issue summary
    quality_issues = []
    for field, stats in report['field_stats'].items():
        issues = []
        if stats['empty']:
            issues.append(f"{stats['empty']} empty")
        if stats['tbd']:
            issues.append(f"{stats['tbd']} TBD")
        if stats['na']:
            issues.append(f"{stats['na']} N/A")
        if stats['invalid']:
            issues.append(f"{stats['invalid']} invalid")
        
        quality_issues.append({
            'field': field,
            'issues': ', '.join(issues),
            'examples': ', '.join(stats['examples']),
            'completeness': stats['completeness']
        })
    
    # Sort by most problematic first
    report['data_quality_issues'] = sorted(quality_issues, 
                                         key=lambda x: (100 - x['completeness']), 
                                         reverse=True)
    
    # Identify top 5 most problematic fields
    field_problem_counts.sort(key=lambda x: x[1], reverse=True)
    report['most_problematic_fields'] = field_problem_counts[:5]
    
    return render(request, 'account/data_report.html', {
        'report': report,
        'import_summary': {
            'success_count': import_data['success_count'],
            'update_count': import_data['update_count'],
            'error_count': import_data['error_count'],
            'total_clients': len(df.columns) - 3
        }
    })
def checklist(request):
    labels = ["CLG", "LIT", "HVC", "MISC-1", "WAL", "ELE", "FLR", "BB", "MISC-2", "DOR", "OPEN", "WDW", "WDT"]
    activity = ["ALL", "QTY+", "CLN", "R&R", "D&R", "MSK", "MN", "S++", "PNT", "SND", "LF"]
    labelValues = [""]
    claims = Client.objects.all()
    
    # Get rooms for selected claim
    selected_claim_id = request.GET.get('claim')
    rooms = []
    
    if selected_claim_id:
        try:

            client = get_object_or_404(Client, pOwner=selected_claim_id)
            #create a temporary template file with this claims data in job info
            
            
            template_path = os.path.join(settings.BASE_DIR, 'docsAppR', 'templates', 'excel', 'templates', '60_scope_form.xlsx')
            destination_path = os.path.join(settings.BASE_DIR, 'docsAppR', 'templates', 'excel', 'custom templates', )
            
            shutil.copyfile("src", "dest")

            wb = load_workbook(destination_path, data_only=True)
            # Select the ScopeCHLST sheet
            ws = wb['jobinfo(2)']

            # Get all non-empty rooms
            for i in range(1, 26):
                room_attr = f'roomArea{i}'
                room_value = getattr(client, room_attr, None)
                
                if room_value and isinstance(room_value, str):
                    room_value = room_value.strip()
                    if room_value.lower() not in ['', 'tbd', 'n/a']:
                        rooms.append({
                            'id': room_attr,
                            'name': room_value
                        })
        except Client.DoesNotExist:
            rooms = []
            logger.error(f"Client not found: {selected_claim_id}")

    if request.method == 'POST':
        try:
            claim_id = request.POST.get('claim')
            room = request.POST.get('room')
            
            # Collect inspection data
            inspection_data = {
                label.lower(): request.POST.get(label.lower(), '')
                for label in labels
            }
            
            # Store inspection data in session
            request.session['inspection_data'] = {
                'claim_id': claim_id,
                'room': room,
                'inspection': inspection_data
            }
            
            # Generate and return PDF
            return generate_invoice_pdf(request, claim_id)
            
        except Exception as e:
            logger.error(f"Error in POST processing: {str(e)}")
            return HttpResponse(f"An error occurred while processing the form: {str(e)}", status=500)
    
    context = {
        'labels': labels,
        'claims': claims,
        'rooms': rooms,
        'selected_claim_id': selected_claim_id,
        'max_rooms' : 25
    }
    
    return render(request, 'account/checklist.html', context)

def process_master_insurer_file(request, excel_file):
    """Simplified MASTER file processor"""
    try:
        excel_file.seek(0)
        df = pd.read_excel(BytesIO(excel_file.read()), sheet_name='ALL')
        
        # Column C is headers (index 2), Column D onward is data (starting at index 3)
        HEADER_COLUMN_INDEX = 2  # Column C
        FIRST_DATA_COLUMN_INDEX = 3  # Column D
        
        print(f"📊 Processing {len(df.columns) - FIRST_DATA_COLUMN_INDEX} clients")
        
        # Build header mapping once
        header_map = {}
        for row_idx in range(len(df)):
            header = str(df.iloc[row_idx, HEADER_COLUMN_INDEX]).strip() if pd.notna(df.iloc[row_idx, HEADER_COLUMN_INDEX]) else ""
            if header:
                # Normalize header name
                field_name = normalize_header_to_field(header)
                header_map[row_idx] = field_name
        
        # Process each client column
        clients_processed = 0
        for col_idx in range(FIRST_DATA_COLUMN_INDEX, len(df.columns)):
            client_data = {}
            
            # Build client data from headers
            for row_idx, field_name in header_map.items():
                value = df.iloc[row_idx, col_idx]
                if pd.notna(value):
                    client_data[field_name] = value
            
            # Get owner name (should be in 'pOwner' field)
            owner_name = client_data.get('pOwner')
            if owner_name:
                # Map to your model
                mapped_data = map_client_data_to_model(client_data)
                
                # Create or update client
                existing = Client.objects.filter(pOwner=owner_name).first()
                if existing:
                    Client.objects.filter(pk=existing.pk).update(**mapped_data)
                else:
                    Client.objects.create(**mapped_data)
                
                clients_processed += 1
        
        messages.success(request, f"Processed {clients_processed} clients from MASTER file")
        return redirect('dashboard')
        
    except Exception as e:
        messages.error(request, f"MASTER file error: {str(e)}")
        return redirect('create')

def normalize_header_to_field(header):
    """Convert Excel header to field name"""
    # Your existing normalization logic
    field_name = (header.lower()
                 .replace(' ', '_')
                 .replace('-', '_')
                 .replace('#', 'num')
                 .replace(':', '')
                 .strip('_'))
    
    # Map specific headers to model fields
    header_mapping = {
        'property_owner_name': 'pOwner',
        'property_address_street': 'pAddress',
        'property_city_state_zip': 'pCityStateZip',
        'customer_email': 'cEmail',
        'cst_owner_phonenum': 'cPhone',
        # Add all your mappings from your map_client_data_to_model function
    }
    
    return header_mapping.get(field_name, field_name)

import os
import re
import math
import logging
import tempfile
import subprocess
import time
from pathlib import Path
from openpyxl import load_workbook
from django.core.files.base import ContentFile
from django.http import JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404, render
from django.conf import settings
from django.contrib.auth.decorators import login_required
from docsAppR.models import Client, File  # Update with your actual models

logger = logging.getLogger(__name__)

# Helper functions
def safe_filename(name, max_length=120):
    """Create filesystem-safe filename"""
    return re.sub(r'[<>:"/\\|?*]', '_', name)[:max_length]

def get_room_index_from_name(room_name):
    """Extract the room index from room name"""
    match = re.search(r'(\d+)', room_name)
    if match:
        return int(match.group(1))
    return None

def calculate_print_area(num_labels):
    """Calculate the print area based on number of labels requested"""
    if num_labels <= 0:
        return "A1:B4"  # Default to at least one label area
    
    labels_per_block = 2
    blocks_needed = math.ceil(num_labels / labels_per_block)
    end_row = blocks_needed * 4
    
    return f"A1:B{end_row}"

def create_excel_from_template(template_path, output_path, sheet_name, room_index, claim_id, client):
    """Creates an Excel file from template with client data populated"""
    try:
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb = load_workbook(template_path, data_only=False)  # Keep formulas as formulas
        
        if sheet_name not in wb.sheetnames:
            logger.error(f"Sheet '{sheet_name}' not found in template. Available sheets: {wb.sheetnames}")
            return False
        
        # Populate jobinfo sheet if exists
        jobinfo_sheet = None
        for sheet in wb.sheetnames:
            if 'jobinfo(2)' in sheet.lower():
                jobinfo_sheet = wb[sheet]
                break
        
        if jobinfo_sheet:
            try:
                # Create a mapping of field names to their column C positions
                # This should match how your template is structured
                field_mapping = {
                    # Primary Client Information
                    'pOwner': 1,
                    'pAddress': 2,
                    'pCityStateZip': 3,
                    'cEmail': 4,
                    'cPhone': 5,
                    'coOwner2': 6,
                    'cPhone2': 7,
                    'cAddress2': 8,
                    'cCityStateZip2': 9,
                    'cEmail2': 10,
                    
                    # Claim Information
                    'causeOfLoss': 11,
                    'dateOfLoss': 12,
                    'rebuildType1': 13,
                    'rebuildType2': 14,
                    'rebuildType3': 15,
                    'demo': 16,
                    'mitigation': 17,
                    'otherStructures': 18,
                    'replacement': 19,
                    'CPSCLNCONCGN': 20,
                    'yearBuilt': 21,
                    'contractDate': 22,
                    'lossOfUse': 23,
                    'breathingIssue': 24,
                    'hazardMaterialRemediation': 25,
                    
                    # Insurance Information
                    'insuranceCo_Name': 26,
                    'insAddressOvernightMail': 27,
                    'insCityStateZip': 28,
                    'insuranceCoPhone': 29,
                    'insWebsite': 30,
                    'insMailingAddress': 31,
                    'insMailCityStateZip': 32,
                    'claimNumber': 33,
                    'policyNumber': 34,
                    'emailInsCo': 35,
                    'deskAdjusterDA': 36,
                    'DAPhone': 37,
                    'DAPhExt': 38,
                    'DAEmail': 39,
                    'fieldAdjusterName': 40,
                    'phoneFieldAdj': 41,
                    'fieldAdjEmail': 42,
                    'adjContents': 43,
                    'adjCpsPhone': 44,
                    'adjCpsEmail': 45,
                    'emsAdj': 46,
                    'emsAdjPhone': 47,
                    'emsTmpEmail': 48,
                    'attLossDraftDept': 49,
                    
                    # Room Information
                    'newCustomerID': 50,
                    'roomID': 51,
                    'roomArea1': 52,
                    'roomArea2': 53,
                    'roomArea3': 54,
                    'roomArea4': 55,
                    'roomArea5': 56,
                    'roomArea6': 57,
                    'roomArea7': 58,
                    'roomArea8': 59,
                    'roomArea9': 60,
                    'roomArea10': 61,
                    'roomArea11': 62,
                    'roomArea12': 63,
                    'roomArea13': 64,
                    'roomArea14': 65,
                    'roomArea15': 66,
                    'roomArea16': 67,
                    'roomArea17': 68,
                    'roomArea18': 69,
                    'roomArea19': 70,
                    'roomArea20': 71,
                    'roomArea21': 72,
                    'roomArea22': 73,
                    'roomArea23': 74,
                    'roomArea24': 75,
                    'roomArea25': 76,
                    
                    # Mortgage Information
                    'mortgageCo': 77,
                    'mortgageAccountCo': 78,
                    'mortgageContactPerson': 79,
                    'mortgagePhoneContact': 80,
                    'mortgagePhoneExtContact': 81,
                    'mortgageAttnLossDraftDept': 82,
                    'mortgageOverNightMail': 83,
                    'mortgageCityStZipOVN': 84,
                    'mortgageEmail': 85,
                    'mortgageWebsite': 86,
                    'mortgageCoFax': 87,
                    'mortgageMailingAddress': 88,
                    'mortgageInitialOfferPhase1ContractAmount': 89,
                    
                    # Cash Flow
                    'drawRequest': 90,
                    
                    # Contractor Information
                    'coName': 91,
                    'coWebsite': 92,
                    'coEmailstatus': 93,
                    'coAddress': 94,
                    'coCityState': 95,
                    'coAddress2': 96,
                    'coCityState2': 97,
                    'coCityState3': 98,
                    'coLogo1': 99,
                    'coLogo2': 100,
                    'coLogo3': 101,
                    'coRepPH': 102,
                    'coREPEmail': 103,
                    'coPhone2': 104,
                    'TinW9': 105,
                    'fedExAccount': 106,
                    
                    # Claim Reporting
                    'claimReportDate': 107,
                    'insuranceCustomerServiceRep': 108,
                    'timeOfClaimReport': 109,
                    'phoneExt': 110,
                    'tarpExtTMPOk': 111,
                    'IntTMPOk': 112,
                    'DRYPLACUTOUTMOLDSPRAYOK': 113,
                    
                    # ALE Information
                    'lossOfUseALE': 114,
                    'tenantLesee': 115,
                    'propertyAddressStreet': 116,
                    'propertyCityStateZip': 117,
                    'customerEmail': 118,
                    'cstOwnerPhoneNumber': 119,
                    
                    # Duplicate/Additional Fields (appearing later in model)
                    'contractDate': 122,
                    'insuranceCoName': 123,
                    'claimNumber': 124,
                    'policyClaimNumber': 125,
                    'emailInsCo': 126,
                    'deskAdjusterDA': 127,
                    'DAPhone': 128,
                    'DAPhExtNumber': 129,
                    'DAEmail': 130,
                    'startDate': 131,
                    'endDate': 132,
                    'lessor': 133,
                    'propertyAddressStreet': 134,
                    'propertyCityStateZip': 135,
                    'customerEmail': 136,
                    'cstOwnerPhoneNumber': 137,
                    'bedrooms': 138,
                    'termsAmount': 139
                }
                
                datetime_fields = [
                    'dateOfLoss',
                    'contractDate',
                    'claimReportDate',
                    'startDate',
                    'endDate'
                ]

                # Populate the fields we know about
                for field_name, row_num in field_mapping.items():
                    cell_ref = f'C{row_num}'
                    try:
                        value = getattr(client, field_name, None)
                        
                        # Handle datetime fields
                        if field_name in datetime_fields and value is not None:
                            if value.tzinfo is not None:
                                # Remove timezone info
                                value = value.replace(tzinfo=None)
                            jobinfo_sheet[cell_ref] = value
                        
                        # Handle boolean fields
                        elif isinstance(value, bool):
                            jobinfo_sheet[cell_ref] = 'Yes' if value else 'No'
                        
                        # Handle normal fields
                        else:
                            jobinfo_sheet[cell_ref] = str(value) if value not in [None, ''] else 'TBD'
                    
                    except Exception as e:
                        logger.warning(f"Error setting {field_name} in {cell_ref}: {str(e)}")
                        jobinfo_sheet[cell_ref] = 'ERROR'

                # Refresh formulas (optional - only needed if you have formulas referencing these cells)
                for row in jobinfo_sheet.iter_rows():
                    for cell in row:
                        if cell.data_type == 'f':  # Formula
                            try:
                                jobinfo_sheet[cell.coordinate] = f'={cell.value}'
                            except Exception as e:
                                logger.warning(f"Could not refresh formula in {cell.coordinate}: {str(e)}")
                
                logger.info("Successfully populated client data in jobinfo sheet")
                
            except Exception as e:
                logger.warning(f"Could not populate client data: {str(e)}", exc_info=True)
        
        # Save the workbook
        wb.template = False  # Important: Set to False to save as regular workbook
        wb.save(output_path)
        wb.close()
        return True
        
    except Exception as e:
        logger.error(f"Failed to create Excel from template: {str(e)}", exc_info=True)
        return False

import platform

@login_required
def generate_combined_labels(request, claim_id):
    """
    Return a single combined PDF containing wall labels + box labels for every
    room in the given claim.  Used by the 'Download All Labels' button on the
    labels page.
    """
    import io
    from django.http import HttpResponse, Http404
    from .tasks import _create_combined_wall_labels_pdf, _create_combined_box_labels_pdf

    try:
        client = Client.objects.get(id=claim_id)
    except Client.DoesNotExist:
        raise Http404

    rooms = client.rooms.all().order_by('sequence')
    if not rooms.exists():
        from django.http import HttpResponse
        return HttpResponse("No rooms configured for this claim.", status=400)

    # ── Build wall-labels PDF ──────────────────────────────────────────────
    wall_buf = io.BytesIO()
    _create_combined_wall_labels_pdf(wall_buf, client, rooms)
    wall_buf.seek(0)

    # ── Build box-labels PDF ──────────────────────────────────────────────
    box_buf = io.BytesIO()
    _create_combined_box_labels_pdf(box_buf, client, rooms)
    box_buf.seek(0)

    # ── Merge into one PDF using PyPDF or simple concatenation ────────────
    try:
        from pypdf import PdfWriter, PdfReader
        writer = PdfWriter()
        for buf in (wall_buf, box_buf):
            reader = PdfReader(buf)
            for page in reader.pages:
                writer.add_page(page)
        merged_buf = io.BytesIO()
        writer.write(merged_buf)
        merged_buf.seek(0)
        combined_bytes = merged_buf.read()
    except ImportError:
        # pypdf not installed – fall back to returning just wall labels
        combined_bytes = wall_buf.read()

    safe_name = "".join(
        c for c in (client.pOwner or 'Claim') if c.isalnum() or c in (' ', '-', '_')
    ).strip().replace(' ', '_')

    response = HttpResponse(combined_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{safe_name}_All_Labels.pdf"'
    return response


@login_required
@login_required
def labels(request):
    logger.info(f"Labels function called - method: {request.method}")
    logger.info(f"User: {request.user}, Authenticated: {request.user.is_authenticated}")
    
    # GET request handling - show the form
    if request.method == 'GET':
        try:
            claims = Client.objects.all()
            selected_claim_id = request.GET.get('claim')
            rooms = []
            
            if selected_claim_id:
                try:
                    client = get_object_or_404(Client, pOwner=selected_claim_id)

                    # UPDATED: Get rooms from Room model instead of roomArea fields
                    for room in client.rooms.all().order_by('sequence'):
                        rooms.append({
                            'id': str(room.id),
                            'name': room.room_name,
                            'sequence': room.sequence
                        })

                    logger.info(f"Found {len(rooms)} rooms for claim {selected_claim_id}")
                    
                except Client.DoesNotExist:
                    rooms = []
                    logger.error(f"Client not found for pOwner: {selected_claim_id}")
                except Exception as e:
                    rooms = []
                    logger.error(f"Unexpected error loading rooms for claim {selected_claim_id}: {str(e)}")
                    logger.debug(f"Traceback: {traceback.format_exc()}")
            
            context = {
                'claims': claims,
                'rooms': rooms,
                'selected_claim_id': selected_claim_id
            }
            return render(request, 'account/labels.html', context)
            
        except Exception as e:
            logger.error(f"Error in GET request: {str(e)}", exc_info=True)
            return JsonResponse({'status': 'error', 'message': 'Error loading page'}, status=500)
    
    # POST request handling - generate PDFs
    elif request.method == 'POST':
        try:
            logger.info("=== STARTING LABEL GENERATION ===")
            
            # Initialize room_labels dictionary
            room_labels = {}
            claim_id = request.POST.get('claim', '').strip()
            logger.info(f"Claim ID from POST: '{claim_id}'")
            
            if not claim_id:
                logger.error("Missing claim ID in POST data")
                return JsonResponse({'status': 'error', 'message': 'Missing claim ID'}, status=400)

            # Parse room labels from POST data
            logger.info("Parsing room labels from POST data:")
            for key, value in request.POST.items():
                if key.startswith('room_labels['):
                    try:
                        room_name = key[len('room_labels['):-1]  # Extract room name
                        count = int(value)
                        if count > 0:
                            room_labels[room_name] = count
                            logger.info(f"  - {room_name}: {count} labels")
                    except ValueError as ve:
                        logger.warning(f"Invalid value for room label {key}: {value}")
                        continue

            logger.info(f"Total room labels parsed: {len(room_labels)}")
            
            if not room_labels:
                logger.info("No room labels requested")
                return JsonResponse({'status': 'success', 'message': 'No labels requested', 'pdfs': []})

            # Get client data
            logger.info(f"Looking up client with pOwner: '{claim_id}'")
            try:
                client = Client.objects.get(pOwner=claim_id)
                logger.info(f"Client found: {client.pOwner}")
            except Client.DoesNotExist:
                logger.error(f"Client not found for pOwner: {claim_id}")
                return JsonResponse({'status': 'error', 'message': 'Client not found'}, status=404)
            except Exception as e:
                logger.error(f"Error getting client: {str(e)}")
                return JsonResponse({'status': 'error', 'message': 'Error retrieving client data'}, status=500)

            # UPDATED: Create room index mapping from Room model
            room_indices = {}
            logger.info("Creating room index mapping:")
            for room in client.rooms.all().order_by('sequence'):
                # Use sequence + 1 to maintain compatibility with old 1-based numbering
                room_indices[room.room_name] = room.sequence + 1
                logger.info(f"  - Room {room.sequence + 1}: '{room.room_name}'")

            logger.info(f"Room indices mapping created with {len(room_indices)} entries")

            # Check if template exists
            template_path = os.path.join(settings.BASE_DIR, 'docsAppR', 'templates', 'excel', 'room_labels_template.xlsx')
            logger.info(f"Template path: {template_path}")
            
            if not os.path.exists(template_path):
                logger.error(f"Template file not found at: {template_path}")
                return JsonResponse({'status': 'error', 'message': 'Template file not found'}, status=500)
            else:
                logger.info("Template file found")

            # Start PDF generation
            logger.info("Starting PDF generation in temporary directory")
            with tempfile.TemporaryDirectory() as temp_dir:
                pdfs_info = []
                logger.info(f"Temporary directory created: {temp_dir}")
                
                for room_name, num_labels in room_labels.items():
                    try:
                        logger.info(f"--- Processing room: '{room_name}', labels: {num_labels} ---")
                        
                        # Get room index
                        room_index = room_indices.get(room_name.strip())
                        logger.info(f"Room index from mapping: {room_index}")
                        
                        if not room_index:
                            # Try to get room index from name using helper function
                            logger.info("Trying to get room index from name using helper function")
                            try:
                                room_index = get_room_index_from_name(room_name)
                                logger.info(f"Room index from helper function: {room_index}")
                            except Exception as e:
                                logger.error(f"Error in get_room_index_from_name: {str(e)}")
                                room_index = None
                            
                            if not room_index:
                                logger.warning(f"No index found for room: '{room_name}'. Skipping.")
                                continue

                        # Create safe filenames
                        safe_claim = safe_filename(claim_id) if callable(safe_filename) else claim_id.replace(' ', '_')
                        safe_room = safe_filename(room_name) if callable(safe_filename) else room_name.replace(' ', '_')
                        excel_filename = f"labels_{safe_claim}_{safe_room}.xlsx"
                        pdf_filename = f"labels_{safe_claim}_{safe_room}.pdf"
                        temp_excel_path = os.path.join(temp_dir, excel_filename)
                        temp_pdf_path = os.path.join(temp_dir, pdf_filename)
                        sheet_name = f"RM ({room_index})"

                        logger.info(f"File details:")
                        logger.info(f"  - Excel: {temp_excel_path}")
                        logger.info(f"  - PDF: {temp_pdf_path}")
                        logger.info(f"  - Sheet: {sheet_name}")

                        # Generate thermal printer PDF directly (bypassing Excel template)
                        logger.info("Generating thermal printer PDF...")
                        try:
                            create_room_label_pdf_thermal(
                                pdf_path=temp_pdf_path,
                                room_name=room_name,
                                claim_name=client.pOwner,
                                num_labels=num_labels
                            )
                            logger.info("PDF generation successful")
                        except Exception as e:
                            logger.error(f"PDF generation failed for {room_name}: {str(e)}")
                            continue

                        # 3. Store the PDF to server folder AND File model
                        if os.path.exists(temp_pdf_path):
                            logger.info(f"PDF file exists at: {temp_pdf_path}")
                            try:
                                with open(temp_pdf_path, 'rb') as pdf_file:
                                    pdf_content = pdf_file.read()

                                logger.info(f"PDF content size: {len(pdf_content)} bytes")

                                # UPDATED: Save PDF to server claim folder structure
                                from .claim_folder_utils import copy_file_to_claim_folder

                                # Get Templates folder name for this client
                                client_folder_name = f"{client.pOwner}@{client.pAddress}" if client.pOwner and client.pAddress else f"Client_{client.id}"
                                safe_folder_name = re.sub(r'[<>:"/\\|?*]', '_', client_folder_name)
                                destination_folder = f"Templates {safe_folder_name}"

                                # Copy PDF to server folder
                                try:
                                    server_pdf_path = copy_file_to_claim_folder(
                                        client=client,
                                        source_file_path=temp_pdf_path,
                                        destination_folder_type=destination_folder,
                                        new_filename=pdf_filename
                                    )
                                    logger.info(f"Saved PDF to server: {server_pdf_path}")
                                except Exception as copy_err:
                                    logger.warning(f"Could not copy to server folder: {str(copy_err)}")
                                    server_pdf_path = None

                                # Also create File object for download links
                                pdf_obj = File(filename=pdf_filename, size=len(pdf_content))
                                pdf_obj.file.save(pdf_filename, ContentFile(pdf_content))

                                pdfs_info.append({
                                    'room_name': room_name,
                                    'pdf_url': pdf_obj.file.url,
                                    'server_path': server_pdf_path,  # Track server location
                                    'num_labels': num_labels,
                                    'print_area': calculate_print_area(num_labels) if callable(calculate_print_area) else "unknown"
                                })

                                logger.info(f"Successfully generated PDF for {room_name} with {num_labels} labels")
                            except Exception as e:
                                logger.error(f"Error saving PDF file: {str(e)}")
                                continue
                        else:
                            logger.error(f"PDF file not found after conversion: {temp_pdf_path}")

                    except Exception as e:
                        logger.error(f"Error processing room {room_name}: {str(e)}", exc_info=True)
                        continue

                logger.info(f"=== PDF GENERATION COMPLETED ===")
                logger.info(f"Generated {len(pdfs_info)} PDFs out of {len(room_labels)} requested")
                
                from django.urls import reverse
                combined_url = reverse('generate_combined_labels', args=[claim_id])
                if pdfs_info:
                    return JsonResponse({
                        'status': 'success',
                        'message': f'Generated {len(pdfs_info)} PDF(s) successfully',
                        'pdfs': pdfs_info,
                        'combined_pdf_url': combined_url,
                    })
                else:
                    return JsonResponse({
                        'status': 'success',
                        'message': 'No valid labels generated',
                        'pdfs': [],
                        'combined_pdf_url': combined_url,
                    })

        except Exception as e:
            logger.error(f"=== LABEL GENERATION FAILED ===")
            logger.error(f"Error: {str(e)}", exc_info=True)
            return JsonResponse({
                'status': 'error',
                'message': 'Label generation failed. Please try again.'
            }, status=500)
    
    # Handle other HTTP methods (PUT, DELETE, etc.)
    else:
        return HttpResponseNotAllowed(['GET', 'POST'])


@login_required
def wall_labels(request):
    """
    Generate wall orientation labels for thermal printer (3x4 inch labels)
    Shows room name and wall orientation diagram
    """
    logger.info(f"Wall labels function called - method: {request.method}")

    # GET request handling - show the form (same as room labels)
    if request.method == 'GET':
        try:
            claims = Client.objects.all()
            selected_claim_id = request.GET.get('claim')
            rooms = []

            if selected_claim_id:
                try:
                    client = get_object_or_404(Client, pOwner=selected_claim_id)
                    for room in client.rooms.all().order_by('sequence'):
                        rooms.append({
                            'id': str(room.id),
                            'name': room.room_name,
                            'sequence': room.sequence
                        })
                    logger.info(f"Found {len(rooms)} rooms for claim {selected_claim_id}")
                except Client.DoesNotExist:
                    rooms = []
                    logger.error(f"Client not found for pOwner: {selected_claim_id}")
                except Exception as e:
                    rooms = []
                    logger.error(f"Unexpected error loading rooms for claim {selected_claim_id}: {str(e)}")

            context = {
                'claims': claims,
                'rooms': rooms,
                'selected_claim_id': selected_claim_id,
                'label_type': 'wall'  # Indicate this is wall labels
            }
            return render(request, 'account/wall_labels.html', context)
        except Exception as e:
            logger.error(f"Error in GET request: {str(e)}", exc_info=True)
            return JsonResponse({'status': 'error', 'message': 'Error loading page'}, status=500)

    # POST request handling - generate wall label PDFs
    elif request.method == 'POST':
        try:
            logger.info("=== STARTING WALL LABEL GENERATION ===")

            room_labels = {}
            claim_id = request.POST.get('claim', '').strip()
            logger.info(f"Claim ID from POST: '{claim_id}'")

            if not claim_id:
                logger.error("Missing claim ID in POST data")
                return JsonResponse({'status': 'error', 'message': 'Missing claim ID'}, status=400)

            # Parse room labels from POST data
            for key, value in request.POST.items():
                if key.startswith('room_labels['):
                    try:
                        room_name = key[len('room_labels['):-1]
                        count = int(value)
                        if count > 0:
                            room_labels[room_name] = count
                            logger.info(f"  - {room_name}: {count} wall labels")
                    except ValueError:
                        continue

            if not room_labels:
                return JsonResponse({'status': 'success', 'message': 'No labels requested', 'pdfs': []})

            # Get client data
            try:
                client = Client.objects.get(pOwner=claim_id)
                logger.info(f"Client found: {client.pOwner}")
            except Client.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'Client not found'}, status=404)

            # Generate PDFs
            with tempfile.TemporaryDirectory() as temp_dir:
                pdfs_info = []

                for room_name, num_labels in room_labels.items():
                    try:
                        logger.info(f"--- Processing wall label for room: '{room_name}', labels: {num_labels} ---")

                        # Create safe filenames
                        safe_claim = safe_filename(claim_id) if callable(safe_filename) else claim_id.replace(' ', '_')
                        safe_room = safe_filename(room_name) if callable(safe_filename) else room_name.replace(' ', '_')
                        pdf_filename = f"wall_label_{safe_claim}_{safe_room}.pdf"
                        temp_pdf_path = os.path.join(temp_dir, pdf_filename)

                        # Generate thermal printer PDF
                        create_wall_label_pdf(
                            pdf_path=temp_pdf_path,
                            room_name=room_name,
                            claim_name=client.pOwner,
                            num_labels=num_labels
                        )

                        # Store the PDF
                        if os.path.exists(temp_pdf_path):
                            with open(temp_pdf_path, 'rb') as pdf_file:
                                pdf_content = pdf_file.read()

                            # Save to server folder
                            from .claim_folder_utils import copy_file_to_claim_folder
                            client_folder_name = f"{client.pOwner}@{client.pAddress}" if client.pOwner and client.pAddress else f"Client_{client.id}"
                            safe_folder_name = re.sub(r'[<>:"/\\|?*]', '_', client_folder_name)
                            destination_folder = f"Templates {safe_folder_name}"

                            try:
                                server_pdf_path = copy_file_to_claim_folder(
                                    client=client,
                                    source_file_path=temp_pdf_path,
                                    destination_folder_type=destination_folder,
                                    new_filename=pdf_filename
                                )
                            except Exception as copy_err:
                                logger.warning(f"Could not copy to server folder: {str(copy_err)}")
                                server_pdf_path = None

                            # Create File object for download
                            pdf_obj = File(filename=pdf_filename, size=len(pdf_content))
                            pdf_obj.file.save(pdf_filename, ContentFile(pdf_content))

                            pdfs_info.append({
                                'room_name': room_name,
                                'pdf_url': pdf_obj.file.url,
                                'server_path': server_pdf_path,
                                'num_labels': num_labels
                            })

                            logger.info(f"Successfully generated wall label PDF for {room_name}")
                    except Exception as e:
                        logger.error(f"Error processing room {room_name}: {str(e)}", exc_info=True)
                        continue

                logger.info(f"=== WALL LABEL GENERATION COMPLETED ===")

                if pdfs_info:
                    return JsonResponse({
                        'status': 'success',
                        'message': f'Generated {len(pdfs_info)} wall label PDF(s) successfully',
                        'pdfs': pdfs_info
                    })
                else:
                    return JsonResponse({
                        'status': 'success',
                        'message': 'No valid labels generated',
                        'pdfs': []
                    })

        except Exception as e:
            logger.error(f"=== WALL LABEL GENERATION FAILED ===")
            logger.error(f"Error: {str(e)}", exc_info=True)
            return JsonResponse({
                'status': 'error',
                'message': 'Wall label generation failed. Please try again.'
            }, status=500)
    else:
        return HttpResponseNotAllowed(['GET', 'POST'])


def create_room_label_pdf_thermal(pdf_path, room_name, claim_name, num_labels):
    """
    Create thermal printer PDF for room/box labels (3x4 inch)
    Simple design with room name and claim name
    """
    from reportlab.lib.pagesizes import inch
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import inch as INCH
    from reportlab.lib import colors

    # 3x4 inch label size for thermal printer
    LABEL_WIDTH = 4 * INCH
    LABEL_HEIGHT = 3 * INCH

    c = canvas.Canvas(pdf_path, pagesize=(LABEL_WIDTH, LABEL_HEIGHT))

    # Generate the requested number of labels
    for label_num in range(num_labels):
        # Room Name (large, centered)
        c.setFont("Helvetica-Bold", 36)
        c.drawCentredString(LABEL_WIDTH / 2, LABEL_HEIGHT / 2 + 0.2 * INCH, room_name.upper())

        # Claim Name (smaller, below room name)
        c.setFont("Helvetica", 14)
        c.drawCentredString(LABEL_WIDTH / 2, LABEL_HEIGHT / 2 - 0.4 * INCH, claim_name)

        # Add decorative border
        c.setStrokeColor(colors.black)
        c.setLineWidth(2)
        c.rect(0.2 * INCH, 0.2 * INCH, LABEL_WIDTH - 0.4 * INCH, LABEL_HEIGHT - 0.4 * INCH)

        # Add page break if more labels are needed
        if label_num < num_labels - 1:
            c.showPage()

    c.save()
    logger.info(f"Created room label PDF: {pdf_path} with {num_labels} label(s)")


def create_wall_label_pdf(pdf_path, room_name, claim_name, num_labels):
    """
    Create thermal printer PDF for wall labels (3x4 inch)
    Shows room name and wall orientation diagram matching the FOYER structure
    """
    from reportlab.lib.pagesizes import inch
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import inch as INCH
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib import colors

    # 3x4 inch label size for thermal printer
    LABEL_WIDTH = 4 * INCH
    LABEL_HEIGHT = 3 * INCH

    c = canvas.Canvas(pdf_path, pagesize=(LABEL_WIDTH, LABEL_HEIGHT))

    # Generate the requested number of labels
    for label_num in range(num_labels):
        # Room Name (large, centered at top)
        c.setFont("Helvetica-Bold", 28)
        c.drawCentredString(LABEL_WIDTH / 2, LABEL_HEIGHT - 0.5 * INCH, room_name)

        # Orientation diagram - matching your FOYER image structure
        center_y = LABEL_HEIGHT / 2 - 0.1 * INCH
        center_x = LABEL_WIDTH / 2

        # Draw orientation boxes/sections
        # W=1 (left side)
        c.setFont("Helvetica", 10)
        c.drawCentredString(center_x - 1.2 * INCH, center_y, "W=1")

        # CENTER (middle, with arrow pointing up)
        c.setFont("Helvetica-Bold", 12)
        c.drawCentredString(center_x, center_y + 0.3 * INCH, "CENTER")
        # Draw up arrow
        c.line(center_x, center_y, center_x, center_y + 0.2 * INCH)
        c.line(center_x - 0.05 * INCH, center_y + 0.15 * INCH, center_x, center_y + 0.2 * INCH)
        c.line(center_x + 0.05 * INCH, center_y + 0.15 * INCH, center_x, center_y + 0.2 * INCH)

        # W=3 (right side)
        c.setFont("Helvetica", 10)
        c.drawCentredString(center_x + 1.2 * INCH, center_y, "W=3")

        # W=4 (bottom)
        c.drawCentredString(center_x, center_y - 0.5 * INCH, "W=4")

        # Draw circular arrows around the orientation (optional decoration)
        # This adds the curved arrow effect from your image
        c.setStrokeColor(colors.blue)
        c.setLineWidth(2)
        # Left curved arrow
        c.arc(center_x - 1.5 * INCH, center_y - 0.15 * INCH,
              center_x - 0.9 * INCH, center_y + 0.15 * INCH,
              startAng=30, extent=120)
        # Right curved arrow
        c.arc(center_x + 0.9 * INCH, center_y - 0.15 * INCH,
              center_x + 1.5 * INCH, center_y + 0.15 * INCH,
              startAng=30, extent=120)

        # Reset stroke color
        c.setStrokeColor(colors.black)
        c.setLineWidth(1)

        # Add dotted separator line (matching your image)
        c.setDash(3, 3)
        c.line(0.5 * INCH, LABEL_HEIGHT - 0.9 * INCH,
               LABEL_WIDTH - 0.5 * INCH, LABEL_HEIGHT - 0.9 * INCH)
        c.setDash()  # Reset to solid line

        # Add page break if more labels are needed
        if label_num < num_labels - 1:
            c.showPage()

    c.save()
    logger.info(f"Created wall label PDF: {pdf_path} with {num_labels} label(s)")


def convert_excel_to_pdf_with_pages(excel_path, pdf_path, sheet_name, room_name, p_owner, num_labels):
    """Convert Excel to PDF with proper print area, eliminating formula errors"""
    try:
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            temp_xlsx = tmp_file.name
        
        try:
            # 1. Load and prepare the workbook
            wb = load_workbook(excel_path)
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found in {excel_path}")
            
            # 2. Process the target sheet
            ws = wb[sheet_name]

            # 4. Remove other sheets (after processing formulas)
            for sheet in wb.sheetnames:
                if sheet != sheet_name:
                    wb.remove(wb[sheet])
            

            labels_per_page = 4
            total_labels = math.ceil(num_labels / labels_per_page) * labels_per_page
            total_rows = total_labels * 2 

            room_name_upper = room_name.upper()  # Convert to ALL CAPS
            for row_num in range(1, total_rows + 1):
                # Odd rows (1, 3, 5...) get room name
                if row_num % 2 == 1:
                    ws.cell(row=row_num, column=1, value=room_name_upper)  # A1, A3, A5...
                # Even rows (2, 4, 6...) get owner name
                else:
                    ws.cell(row=row_num, column=1, value=p_owner)  # A2, A4, A6...
            
            # 5. Set print area
            print_area = calculate_print_area(num_labels)
            if ':' in print_area:  # Validate print area format
                ws.print_area = print_area
                logger.info(f"Set print area for {sheet_name}: {print_area}")
            else:
                logger.warning(f"Invalid print area format: {print_area}")
            
            # 6. Save the cleaned workbook
            wb.template = False
            wb.security = None
            wb.save(temp_xlsx)
            wb.close()
            
            # 7. Convert to PDF using LibreOffice - FIXED PATH FOR LINUX
            temp_dir = os.path.dirname(temp_xlsx)
            
            # Detect operating system and set correct LibreOffice path
            if platform.system() == "Windows":
                libreoffice_path = 'C:\\Program Files\\LibreOffice\\program\\soffice.exe'
            else:
                libreoffice_path = '/usr/bin/libreoffice'  # Linux path
            
            cmd = [
                libreoffice_path,
                '--headless',
                '--convert-to', 'pdf',
                '--outdir', temp_dir,
                temp_xlsx
            ]
            
            # 8. Run conversion with error handling
            result = subprocess.run(
                cmd, 
                check=True, 
                timeout=120, 
                capture_output=True,
                text=True
            )
            logger.debug(f"LibreOffice output: {result.stdout}")
            
            # 9. Handle the generated PDF
            generated_pdf = os.path.splitext(temp_xlsx)[0] + '.pdf'
            max_wait = 30
            waited = 0
            
            while waited < max_wait:
                if os.path.exists(generated_pdf):
                    try:
                        with open(generated_pdf, 'rb') as f:
                            pdf_data = f.read()
                        with open(pdf_path, 'wb') as f:
                            f.write(pdf_data)
                        logger.info(f"PDF successfully generated at {pdf_path}")
                        return True
                    except PermissionError:
                        time.sleep(0.5)
                        waited += 0.5
                else:
                    time.sleep(0.5)
                    waited += 0.5
            
            raise RuntimeError(f"PDF generation timed out after {max_wait} seconds")
            
        finally:
            # 10. Cleanup temporary files
            cleanup_files = [temp_xlsx]
            temp_pdf = os.path.splitext(temp_xlsx)[0] + '.pdf'
            if os.path.exists(temp_pdf):
                cleanup_files.append(temp_pdf)
            
            for file_path in cleanup_files:
                try:
                    os.unlink(file_path)
                except Exception as e:
                    logger.warning(f"Could not delete temp file {file_path}: {e}")
    
    except subprocess.CalledProcessError as e:
        error_msg = f"LibreOffice conversion failed: {e.stderr}"
        logger.error(error_msg)
        raise RuntimeError(error_msg)
    except Exception as e:
        logger.error(f"PDF conversion error: {str(e)}", exc_info=True)
        raise


# views.py
from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Count, Avg, Case, When, IntegerField, F
from django.http import JsonResponse
from django.utils import timezone
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
import json
from .models import Client, ChecklistItem, Room, RoomWorkTypeValue, WorkType

@login_required
def dashboard(request):
    # Handle client selection
    selected_client_id = request.GET.get('selected_client')
    selected_client = None

    # Get all clients
    clients = Client.objects.all()

    # Apply search filter
    search = request.GET.get('search', '')
    if search:
        clients = clients.filter(
            Q(pOwner__icontains=search) |
            Q(claimNumber__icontains=search) |
            Q(pAddress__icontains=search) |
            Q(newCustomerID__icontains=search)
        )

    # Apply cause of loss filter
    cause_of_loss = request.GET.get('cause_of_loss', '')
    if cause_of_loss:
        clients = clients.filter(causeOfLoss__icontains=cause_of_loss)

    # Apply age filter
    age = request.GET.get('age', '')
    if age:
        cutoff_date = timezone.now() - dt.timedelta(days=int(age))
        clients = clients.filter(dateOfLoss__gte=cutoff_date)

    # Apply sorting
    sort_by = request.GET.get('sort', 'name')
    if sort_by == 'name':
        clients = clients.order_by('pOwner')
    elif sort_by == 'id':
        clients = clients.order_by('newCustomerID')
    elif sort_by == 'date':
        clients = clients.order_by('-dateOfLoss')

    # Apply type filtering
    filter_type = request.GET.get('type', 'all')
    if filter_type == 'CPS':
        clients = clients.filter(CPSCLNCONCGN=True)
    elif filter_type == 'MIT':
        clients = clients.filter(mitigation=True)
    elif filter_type == 'PPR':
        clients = clients.filter(replacement=True)
    
    # PAGINATION
    paginator = Paginator(clients, 10)  # Show 10 claims per page
    page = request.GET.get('page', 1)
    
    try:
        paginated_clients = paginator.page(page)
    except PageNotAnInteger:
        paginated_clients = paginator.page(1)
    except EmptyPage:
        paginated_clients = paginator.page(paginator.num_pages)
    
    # Set selected client
    if selected_client_id:
        try:
            selected_client = Client.objects.get(id=selected_client_id)
            # Ensure checklist items exist for selected client
            from .signals import create_checklist_items_for_client
            create_checklist_items_for_client(selected_client)
            selected_client.update_completion_stats()
        except Client.DoesNotExist:
            selected_client = None
    
    # Calculate dashboard statistics
    total_claims = clients.count()
    
    # Safely calculate average completion
    completion_avg = clients.aggregate(avg=Avg('completion_percent'))['avg']
    avg_completion = round(completion_avg, 1) if completion_avg is not None else 0
    
    # Claim type counts
    mit_count = clients.filter(mitigation=True).count()
    cps_count = clients.filter(CPSCLNCONCGN=True).count()
    ppr_count = clients.filter(replacement=True).count()
    
    # ORIGINAL AGE DISTRIBUTION THRESHOLDS
    now = timezone.now().date()
    age_categories = {
        "0-30 days": Q(dateOfLoss__isnull=False, dateOfLoss__gte=now - dt.timedelta(days=30)),
        "31-60 days": Q(dateOfLoss__isnull=False, dateOfLoss__lt=now - dt.timedelta(days=30)) & 
                     Q(dateOfLoss__gte=now - dt.timedelta(days=60)),
        "61-120 days": Q(dateOfLoss__isnull=False, dateOfLoss__lt=now - dt.timedelta(days=60)) & 
                      Q(dateOfLoss__gte=now - dt.timedelta(days=120)),
        "121-180 days": Q(dateOfLoss__isnull=False, dateOfLoss__lt=now - dt.timedelta(days=120)) & 
                       Q(dateOfLoss__gte=now - dt.timedelta(days=180)),
        "181-360 days": Q(dateOfLoss__isnull=False, dateOfLoss__lt=now - dt.timedelta(days=180)) & 
                        Q(dateOfLoss__gte=now - dt.timedelta(days=360)),
        "360+ days": Q(dateOfLoss__isnull=False, dateOfLoss__lt=now - dt.timedelta(days=360)),
        "No Date": Q(dateOfLoss__isnull=True)
    }
    
    age_distribution_data = {}
    for category, query in age_categories.items():
        count = clients.filter(query).count()
        age_distribution_data[category] = count
    
    # Cause of loss distribution with safe handling
    cause_of_loss_data = {}
    for client in clients:
        cause = client.causeOfLoss.strip() if client.causeOfLoss else 'Unknown'
        if not cause:
            cause = 'Unknown'
        cause_of_loss_data[cause] = cause_of_loss_data.get(cause, 0) + 1
    
    # Sort and get top causes
    top_causes = dict(sorted(cause_of_loss_data.items(), key=lambda x: x[1], reverse=True)[:5])
    
    # Work type distribution with safe handling
    work_type_stats = {
        'MIT': mit_count,
        'CPS': cps_count, 
        'PPR': ppr_count
    }
    
    # Document completion rates with safe handling
    document_completion_data = {'MIT': 0, 'CPS': 0, 'PPR': 0}
    try:
        for category in ['MIT', 'CPS', 'PPR']:
            items = ChecklistItem.objects.filter(document_category=category)
            total = items.count()
            if total > 0:
                completed = items.filter(is_completed=True).count()
                document_completion_data[category] = round((completed / total) * 100, 1)
    except Exception as e:
        print(f"Error calculating document completion: {e}")

    # Completion distribution (how many claims at each completion level)
    completion_distribution = {
        '0%': clients.filter(completion_percent=0).count(),
        '1-25%': clients.filter(completion_percent__gt=0, completion_percent__lte=25).count(),
        '26-50%': clients.filter(completion_percent__gt=25, completion_percent__lte=50).count(),
        '51-75%': clients.filter(completion_percent__gt=50, completion_percent__lte=75).count(),
        '76-99%': clients.filter(completion_percent__gt=75, completion_percent__lt=100).count(),
        '100%': clients.filter(completion_percent=100).count(),
    }

    # Category-specific completion stats
    category_stats = {}
    for category in ['MIT', 'CPS', 'PPR']:
        items = ChecklistItem.objects.filter(document_category=category)
        total = items.count()
        completed = items.filter(is_completed=True).count()
        category_stats[category] = {
            'total': total,
            'completed': completed,
            'pending': total - completed,
            'percent': round((completed / total) * 100, 1) if total > 0 else 0
        }

    # Checklist item completion by document type (top items needing attention)
    from django.db.models import Sum
    items_by_type = ChecklistItem.objects.values('document_type').annotate(
        total=Count('id'),
        completed=Count('id', filter=Q(is_completed=True))
    ).order_by('-total')[:10]

    items_completion = {}
    for item in items_by_type:
        doc_type = item['document_type']
        # Get display name
        display_name = dict(ChecklistItem.DOCUMENT_TYPES).get(doc_type, doc_type)
        items_completion[display_name] = {
            'total': item['total'],
            'completed': item['completed'],
            'percent': round((item['completed'] / item['total']) * 100, 1) if item['total'] > 0 else 0
        }

    context = {
        'total_claims': total_claims,
        'avg_completion': avg_completion,
        'mit_count': mit_count,
        'cps_count': cps_count,
        'ppr_count': ppr_count,
        # JSON serialized data for JavaScript
        'age_distribution_json': json.dumps(age_distribution_data),
        'cause_of_loss_json': json.dumps(top_causes),
        'work_type_stats_json': json.dumps(work_type_stats),
        'document_completion_json': json.dumps(document_completion_data),
        'completion_distribution_json': json.dumps(completion_distribution),
        'category_stats_json': json.dumps(category_stats),
        'items_completion_json': json.dumps(items_completion),
    }

    return render(request, 'account/dashboard.html', context)

@login_required
def update_checklist(request):
    if request.method == 'POST':
        try:
            client_id = request.POST.get('client_id')
            if not client_id:
                return JsonResponse({'success': False, 'error': 'No client ID provided'})
                
            client = get_object_or_404(Client, id=client_id)
            
            # Update all checklist items
            for item in client.checklist_items.all():
                field_name = f'item_{item.id}'
                item.is_completed = field_name in request.POST
                item.save()
            
            # Update completion stats
            client.update_completion_stats()
            
            return JsonResponse({
                'success': True,
                'completion_percent': client.completion_percent,
                'category_completion': getattr(client, 'category_completion', {})
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

@login_required
def api_client_details(request, client_id):
    try:
        from .models import OneDriveFolder, OneDriveFile, SyncLog
        client = Client.objects.get(id=client_id)

        # Ensure completion stats are up to date
        client.update_completion_stats()

        # Get all checklist items for this client
        checklist_categories = []
        for category_code, category_name in ChecklistItem.DOCUMENT_CATEGORIES:
            items = client.checklist_items.filter(document_category=category_code)
            if items.exists():
                category_data = {
                    'grouper': category_name,
                    'items': []
                }
                for item in items:
                    category_data['items'].append({
                        'id': item.id,
                        'document_type': item.document_type,
                        'document_type_display': item.get_document_type_display(),
                        'is_completed': item.is_completed,
                        'document_category': item.document_category
                    })
                checklist_categories.append(category_data)

        # Get OneDrive folders
        onedrive_folders = []
        for folder in client.onedrive_folders.filter(is_active=True):
            onedrive_folders.append({
                'folder_path': folder.folder_path,
                'onedrive_folder_id': folder.onedrive_folder_id,
                'last_synced': folder.last_synced.isoformat() if folder.last_synced else None
            })

        # Get OneDrive files
        onedrive_files = []
        for file in client.onedrive_files.all()[:20]:  # Limit to 20 most recent
            onedrive_files.append({
                'file_name': file.file_name,
                'file_type': file.file_type,
                'sync_status': file.get_sync_status_display(),
                'last_modified': file.last_modified_onedrive.isoformat() if file.last_modified_onedrive else None,
                'onedrive_file_id': file.onedrive_file_id
            })

        # Get recent sync logs
        sync_logs = []
        for log in client.sync_logs.all()[:5]:  # Last 5 sync activities
            sync_logs.append({
                'sync_direction': log.get_sync_direction_display(),
                'sync_status': log.get_sync_status_display(),
                'timestamp': log.timestamp.isoformat() if log.timestamp else None,
                'error_message': log.error_message
            })

        # Get rooms data
        rooms_data = client.get_rooms_data()

        # Prepare comprehensive client data
        client_data = {
            'id': client.id,
            # Customer Info
            'pOwner': client.pOwner or '',
            'pAddress': client.pAddress or '',
            'pCityStateZip': client.pCityStateZip or '',
            'cEmail': client.cEmail or '',
            'cPhone': client.cPhone or '',
            'coOwner2': client.coOwner2 or '',
            'cPhone2': client.cPhone2 or '',
            'cAddress2': client.cAddress2 or '',
            'cCityStateZip2': client.cCityStateZip2 or '',
            'cEmail2': client.cEmail2 or '',

            # Claim Info
            'causeOfLoss': client.causeOfLoss or '',
            'dateOfLoss': client.dateOfLoss.isoformat() if client.dateOfLoss else '',
            'yearBuilt': client.yearBuilt or '',
            'contractDate': client.contractDate.isoformat() if client.contractDate else '',
            'mitigation': client.mitigation,
            'CPSCLNCONCGN': client.CPSCLNCONCGN,
            'replacement': client.replacement,
            'demo': client.demo,
            'otherStructures': client.otherStructures,

            # Insurance Info
            'insuranceCo_Name': client.insuranceCo_Name or '',
            'claimNumber': client.claimNumber or '',
            'policyNumber': client.policyNumber or '',
            'deskAdjusterDA': client.deskAdjusterDA or '',
            'DAPhone': client.DAPhone or '',
            'DAEmail': client.DAEmail or '',
            'fieldAdjusterName': client.fieldAdjusterName or '',
            'phoneFieldAdj': client.phoneFieldAdj or '',
            'fieldAdjEmail': client.fieldAdjEmail or '',

            # Mortgage Info
            'mortgageCo': client.mortgageCo or '',
            'mortgageAccountCo': client.mortgageAccountCo or '',
            'mortgageContactPerson': client.mortgageContactPerson or '',
            'mortgagePhoneContact': client.mortgagePhoneContact or '',
            'mortgageEmail': client.mortgageEmail or '',

            # Contractor Info
            'coName': client.coName or '',
            'coWebsite': client.coWebsite or '',
            'coAddress': client.coAddress or '',
            'coPhone': client.cPhone or '',

            # OneDrive Info (status field removed)
            'onedrive_folder_id': client.onedrive_folder_id or '',
            'last_onedrive_sync': client.last_onedrive_sync.isoformat() if client.last_onedrive_sync else None,
            'onedrive_folders': onedrive_folders,
            'onedrive_files': onedrive_files,
            'sync_logs': sync_logs,

            # Completion Stats
            'completion_percent': client.completion_percent,
            'category_completion': getattr(client, 'category_completion', {}),
            'checklist_items': checklist_categories,

            # Rooms
            'rooms': rooms_data,
            'newCustomerID': client.newCustomerID or ''
        }

        return JsonResponse({'success': True, 'client': client_data})
    except Client.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Client not found'})
    except Exception as e:
        import traceback
        return JsonResponse({'success': False, 'error': str(e), 'traceback': traceback.format_exc()})

def calculate_percentage_change(old_value, new_value):
    if old_value == 0:
        return 0
    return ((new_value - old_value) / old_value) * 100


# LEASE GENERATION VIEWS
def client_list(request):
    # Get all clients from the database
    clients = Client.objects.all()
    documents = Document.objects.all()
    selected_client = None
    selected_documents = []  # Changed to handle multiple documents
    form = None

    if 'client_name' in request.GET:
        client_id = request.GET['client_name']
        if client_id:  # Only try to get client if an ID was provided
            selected_client = get_object_or_404(clients, pOwner=client_id)

            # Always select the 3 specific documents
            document_names = ['Engagement Agreement', 'Term Sheet', 'Month to Month Rental']
            selected_documents = Document.objects.filter(name__in=document_names)

            if selected_documents and selected_documents.first().document_type == 'lease':
                # Create a Landlord instance pre-populated with ALE data from the client
                landlord = Landlord()

                # Map Client ALE fields to Landlord fields
                # Lessor (Landlord) Information
                landlord.full_name = selected_client.ale_lessor_name or ''
                landlord.address = selected_client.ale_lessor_mailing_address or ''
                landlord.city = selected_client.ale_lessor_city_zip or ''  # Contains city+zip
                landlord.state = ''  # Not directly mapped
                landlord.zip_code = ''  # Part of city_zip
                landlord.phone = selected_client.ale_lessor_phone or ''
                landlord.email = selected_client.ale_lessor_email or ''
                landlord.contact_person_1 = selected_client.ale_lessor_contact_person or ''

                # Property Information (Leased Property)
                landlord.property_address = selected_client.ale_lessor_leased_address or ''
                landlord.property_city = ''  # Need to parse from ale_lessor_city_zip
                landlord.property_state = ''  # Need to parse
                landlord.property_zip = ''  # Need to parse

                # Rental Terms
                landlord.term_start_date = selected_client.ale_rental_start_date
                landlord.term_end_date = selected_client.ale_rental_end_date
                landlord.default_rent_amount = selected_client.ale_rental_amount_per_month or 0
                landlord.bedrooms = int(selected_client.ale_rental_bedrooms) if selected_client.ale_rental_bedrooms and selected_client.ale_rental_bedrooms.isdigit() else 1
                landlord.rental_months = int(selected_client.ale_rental_months) if selected_client.ale_rental_months and selected_client.ale_rental_months.isdigit() else 12

                # Real Estate Company Information
                landlord.real_estate_company = selected_client.ale_re_company_name or ''
                landlord.company_mailing_address = selected_client.ale_re_mailing_address or ''
                landlord.company_city = ''  # Need to parse from ale_re_city_zip
                landlord.company_state = ''  # Need to parse
                landlord.company_zip = ''  # Need to parse
                landlord.company_contact_person = selected_client.ale_re_contact_person or ''
                landlord.company_phone = selected_client.ale_re_phone or ''
                landlord.company_email = selected_client.ale_re_email or ''
                landlord.broker_name = selected_client.ale_re_owner_broker_name or ''
                landlord.broker_phone = selected_client.ale_re_owner_broker_phone or ''
                landlord.broker_email = selected_client.ale_re_owner_broker_email or ''

                # Create form with pre-populated landlord instance
                form = LandlordForm(instance=landlord)
    
    if selected_client:
        print(selected_client.__dict__)
    
    return render(request, "account/client_list.html", {
        "clients": clients,
        "documents": documents,
        "selected_client": selected_client,
        "selected_documents": selected_documents,  # Changed to plural
        'current_client_id': selected_client.id if selected_client else None,
        "form": form,
    })

import re
import os
import logging
import zipfile
from io import BytesIO
from django.shortcuts import get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.core.files.storage import default_storage
from django.template import Template, Context
from dateutil.parser import parse as parse_date
from xhtml2pdf import pisa
from weasyprint import HTML
from django.conf import settings
from .models import Client, Document, Landlord

logger = logging.getLogger(__name__)

def generate_all_documents(request):
    """Generate all 3 documents at once and return as ZIP or individual previews"""
    logger.debug("Batch document generation started")
    
    if request.method != 'POST':
        logger.error("Invalid request method")
        return HttpResponse("Only POST requests are allowed", status=405)

    
        # Debug logging for form data
    logger.debug(f"Form data received: {dict(request.POST)}")
    logger.debug(f"Exclude security deposit: {request.POST.get('exclude_security_deposit')}")
    logger.debug(f"Exclude inspection fee: {request.POST.get('exclude_inspection_fee')}")
    logger.debug(f"Real estate company: {request.POST.get('real_estate_company')}")
    logger.debug(f"Company mailing address: {request.POST.get('company_mailing_address')}")

    try:
        # Get required parameters
        client_name = request.POST.get('client_name')
        if not client_name:
            logger.error("Missing client_name")
            return HttpResponse("client_name is required", status=400)

        # Get models
        try:
            client = get_object_or_404(Client, pOwner=client_name)
            # Get all 3 specific documents
            document_names = ['Engagement Agreement', 'Term Sheet', 'Month to Month Rental']
            documents = Document.objects.filter(name__in=document_names)
            logger.debug(f"Found client {client_name} and {documents.count()} documents")
        except Exception as e:
            logger.error(f"Error fetching models: {str(e)}")
            return HttpResponse("Error loading client or documents", status=404)

        # Date formatting function
        def clean_and_format_date(date_str):
            if not date_str:
                return ""
            
            try:
                cleaned = re.sub(r'[^\d/-]', '', str(date_str))
                date_obj = parse_date(cleaned)
                
                if not date_obj:
                    for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%m-%d-%Y', '%m/%d/%Y', '%d-%m-%Y', '%d/%m/%Y'):
                        try:
                            date_obj = dt.datetime.strptime(cleaned, fmt).date()
                            break
                        except ValueError:
                            continue
                
                if date_obj:
                    try:
                        return date_obj.strftime('%B %d, %Y').replace(' 0', ' ')
                    except:
                        return date_obj.strftime('%B %d, %Y')
            except Exception as e:
                logger.warning(f"Date formatting failed for {date_str}: {str(e)}")
            return ""

        def format_agreement_date(date_str):
            """Format date as 'Xth day of Month Year' for legal documents"""
            if not date_str:
                return ""

            try:
                cleaned = re.sub(r'[^\d/-]', '', str(date_str))
                date_obj = parse_date(cleaned)

                if not date_obj:
                    for fmt in ('%Y-%m-%d', '%Y\\%m\\%d', '%m-%d-%Y', '%m\\%d\\%Y', '%d-%m-%Y', '%d\\%m\\%Y'):
                        try:
                            date_obj = dt.datetime.strptime(cleaned, fmt).date()
                            break
                        except ValueError:
                            continue

                if date_obj:
                    day = date_obj.day
                    # Add ordinal suffix (1st, 2nd, 3rd, 4th, etc.)
                    if 10 <= day % 100 <= 20:
                        suffix = 'th'
                    else:
                        suffix = {1: 'st', 2: 'nd', 3: 'rd'}.get(day % 10, 'th')

                    month_year = date_obj.strftime('%B %Y')
                    return f"{day}{suffix} day of {month_year}"
            except Exception as e:
                logger.warning(f"Agreement date formatting failed for {date_str}: {str(e)}")
            return ""

        # Handle preview request
        if request.POST.get('preview') == 'true':
            logger.debug("Generating HTML previews for all documents")
            previews = {}
            
            for document in documents:
                try:
                    # Read template content
                    if not document.file:
                        logger.error(f"No template file attached to document: {document.name}")
                        continue
                        
                    template_path = document.file.path
                    logger.debug(f"Attempting to read template from: {template_path}")
                    
                    if not os.path.exists(template_path):
                        logger.error(f"Template file not found at: {template_path}")
                        continue
                        
                    with open(template_path, 'r', encoding='utf-8') as template_file:
                        template_content = template_file.read()
                    
                    # Create template and context
                    template = Template(template_content)
                    context = Context({
                        'client': client,
                        'document': document,
                        'preview': True,
                        'today': dt.datetime.now().strftime('%B %d, %Y')  # Fixed datetime
                    })

                    # Process lease-specific data for ALL documents
                    # All 3 documents will receive the same landlord data
                    lease_agreement_date = request.POST.get('lease_agreement_date', '')
                    term_start_date = request.POST.get('term_start_date', '')
                    term_end_date = request.POST.get('term_end_date', '')
                    is_renewal = request.POST.get('is_renewal') == 'true'
                    exclude_security_deposit = request.POST.get('exclude_security_deposit') == 'true'
                    exclude_inspection_fee = request.POST.get('exclude_inspection_fee') == 'true'

                    context.update({
                        'formatted_agreement_date': format_agreement_date(lease_agreement_date),
                        'lease_agreement_date': lease_agreement_date,
                        'formatted_start_date': clean_and_format_date(term_start_date),
                        'formatted_end_date': clean_and_format_date(term_end_date),
                        'term_start_date': term_start_date,
                        'term_end_date': term_end_date,
                        'is_renewal': is_renewal,
                        'exclude_security_deposit': exclude_security_deposit,
                        'exclude_inspection_fee': exclude_inspection_fee,

                    })

                    # Add ALL landlord data from form
                    landlord_data = {
                        # Basic Information
                        'full_name': request.POST.get('full_name'),
                        'address': request.POST.get('address'),
                        'city': request.POST.get('city'),
                        'state': request.POST.get('state'),
                        'zip_code': request.POST.get('zip_code'),
                        'phone': request.POST.get('phone'),
                        'email': request.POST.get('email'),
                        
                        # Rental Property Information
                        'property_address': request.POST.get('property_address'),
                        'property_city': request.POST.get('property_city'),
                        'property_state': request.POST.get('property_state'),
                        'property_zip': request.POST.get('property_zip'),
                        
                        # Term dates
                        'term_start_date': request.POST.get('term_start_date'),
                        'term_end_date': request.POST.get('term_end_date'),

                        # Agreement Defaults
                        'default_rent_amount': request.POST.get('default_rent_amount', 0),
                        'default_security_deposit': request.POST.get('default_security_deposit', 0),
                        'default_rent_due_day': request.POST.get('default_rent_due_day', 1),
                        'default_late_fee': request.POST.get('default_late_fee', 0),
                        'default_late_fee_start_day': request.POST.get('default_late_fee_start_day', 5),
                        'default_eviction_day': request.POST.get('default_eviction_day', 10),
                        'default_nsf_fee': request.POST.get('default_nsf_fee', 0),
                        'default_max_occupants': request.POST.get('default_max_occupants', 10),
                        'default_parking_spaces': request.POST.get('default_parking_spaces', 2),
                        'default_parking_fee': request.POST.get('default_parking_fee', 0),
                        'default_inspection_fee': request.POST.get('default_inspection_fee', 300.00),
                        'bedrooms': request.POST.get('bedrooms', 1),
                        'rental_months': request.POST.get('rental_months'),
                        
                        # Additional Contact Persons
                        'contact_person_1': request.POST.get('contact_person_1'),
                        'contact_person_2': request.POST.get('contact_person_2'),
                        'contact_phone': request.POST.get('contact_phone'),
                        'contact_email': request.POST.get('contact_email'),
                        
                        # Real Estate Company Information
                        'real_estate_company': request.POST.get('real_estate_company'),
                        'company_mailing_address': request.POST.get('company_mailing_address'),
                        'company_city': request.POST.get('company_city'),
                        'company_state': request.POST.get('company_state'),
                        'company_zip': request.POST.get('company_zip'),
                        'company_contact_person': request.POST.get('company_contact_person'),
                        'company_phone': request.POST.get('company_phone'),
                        'company_email': request.POST.get('company_email'),
                        'broker_name': request.POST.get('broker_name'),
                        'broker_phone': request.POST.get('broker_phone'),
                        'broker_email': request.POST.get('broker_email'),

                        # Special Lease Instructions/Notes
                        'lease_special_notes': request.POST.get('lease_special_notes', ''),

                        'is_renewal': is_renewal,
                        'exclude_security_deposit': exclude_security_deposit,
                        'exclude_inspection_fee': exclude_inspection_fee,
                    }

                    # Convert numeric fields with proper error handling
                    numeric_fields = {
                        'default_rent_amount': 0,
                        'default_security_deposit': 0,
                        'default_late_fee': 50,
                        'default_nsf_fee': 35,
                        'default_inspection_fee': 300.00,
                        'bedrooms': 0,
                        'rental_months': 0,
                        'default_rent_due_day': 1,
                        'default_late_fee_start_day': 5,
                        'default_eviction_day': 10,
                        'default_max_occupants': 10,
                        'default_parking_spaces': 2,
                        'default_parking_fee': 0
                    }

                    for field, default in numeric_fields.items():
                        try:
                            value = request.POST.get(field, default)
                            if value:
                                if field in ['default_rent_amount', 'default_security_deposit', 
                                           'default_late_fee', 'default_nsf_fee', 'default_parking_fee', 
                                           'default_inspection_fee']:
                                    landlord_data[field] = float(value)
                                else:
                                    landlord_data[field] = int(float(value))  # Handle integer fields
                            else:
                                landlord_data[field] = default
                        except (ValueError, TypeError) as e:
                            logger.warning(f"Invalid number format for {field}: {str(e)}")
                            landlord_data[field] = default

                    context['landlord'] = landlord_data

                    # Render template
                    html_content = template.render(context)
                    previews[document.name] = html_content
                    
                except Exception as e:
                    logger.error(f"Error generating preview for {document.name}: {str(e)}")
                    previews[document.name] = f"<p>Error generating preview: {str(e)}</p>"

            # Generate Input Sheet preview
            try:
                input_sheet_path = os.path.join(
                    settings.BASE_DIR, 'docsAppR', 'templates', 'account', 'lease_input_sheet.html'
                )

                if os.path.exists(input_sheet_path):
                    with open(input_sheet_path, 'r', encoding='utf-8') as input_sheet_file:
                        input_sheet_content = input_sheet_file.read()

                    input_sheet_template = Template(input_sheet_content)

                    # Build context for input sheet
                    input_sheet_landlord_data = {
                        'full_name': request.POST.get('full_name'),
                        'address': request.POST.get('address'),
                        'city': request.POST.get('city'),
                        'state': request.POST.get('state'),
                        'zip_code': request.POST.get('zip_code'),
                        'phone': request.POST.get('phone'),
                        'email': request.POST.get('email'),
                        'property_address': request.POST.get('property_address'),
                        'property_city': request.POST.get('property_city'),
                        'property_state': request.POST.get('property_state'),
                        'property_zip': request.POST.get('property_zip'),
                        'default_rent_amount': request.POST.get('default_rent_amount', 0),
                        'default_security_deposit': request.POST.get('default_security_deposit', 0),
                        'default_rent_due_day': request.POST.get('default_rent_due_day', 1),
                        'default_late_fee': request.POST.get('default_late_fee', 0),
                        'default_late_fee_start_day': request.POST.get('default_late_fee_start_day', 5),
                        'default_eviction_day': request.POST.get('default_eviction_day', 10),
                        'default_nsf_fee': request.POST.get('default_nsf_fee', 0),
                        'default_max_occupants': request.POST.get('default_max_occupants', 10),
                        'default_parking_spaces': request.POST.get('default_parking_spaces', 2),
                        'default_parking_fee': request.POST.get('default_parking_fee', 0),
                        'default_inspection_fee': request.POST.get('default_inspection_fee', 300.00),
                        'bedrooms': request.POST.get('bedrooms', 1),
                        'rental_months': request.POST.get('rental_months'),
                        'contact_person_1': request.POST.get('contact_person_1'),
                        'contact_person_2': request.POST.get('contact_person_2'),
                        'contact_phone': request.POST.get('contact_phone'),
                        'contact_email': request.POST.get('contact_email'),
                        'real_estate_company': request.POST.get('real_estate_company'),
                        'company_mailing_address': request.POST.get('company_mailing_address'),
                        'company_city': request.POST.get('company_city'),
                        'company_state': request.POST.get('company_state'),
                        'company_zip': request.POST.get('company_zip'),
                        'company_contact_person': request.POST.get('company_contact_person'),
                        'company_phone': request.POST.get('company_phone'),
                        'company_email': request.POST.get('company_email'),
                        'broker_name': request.POST.get('broker_name'),
                        'broker_phone': request.POST.get('broker_phone'),
                        'broker_email': request.POST.get('broker_email'),
                        'lease_special_notes': request.POST.get('lease_special_notes', ''),
                    }

                    # Convert numeric fields
                    for field in ['default_rent_amount', 'default_security_deposit', 'default_late_fee',
                                 'default_nsf_fee', 'default_parking_fee', 'default_inspection_fee']:
                        try:
                            input_sheet_landlord_data[field] = float(input_sheet_landlord_data[field] or 0)
                        except (ValueError, TypeError):
                            input_sheet_landlord_data[field] = 0

                    for field in ['default_rent_due_day', 'default_late_fee_start_day', 'default_eviction_day',
                                 'default_max_occupants', 'default_parking_spaces', 'bedrooms', 'rental_months']:
                        try:
                            input_sheet_landlord_data[field] = int(float(input_sheet_landlord_data[field] or 0))
                        except (ValueError, TypeError):
                            input_sheet_landlord_data[field] = 0

                    input_sheet_context = Context({
                        'client': client,
                        'landlord': input_sheet_landlord_data,
                        'today': dt.datetime.now().strftime('%B %d, %Y'),
                        'formatted_agreement_date': format_agreement_date(lease_agreement_date),
                        'formatted_start_date': clean_and_format_date(term_start_date),
                        'formatted_end_date': clean_and_format_date(term_end_date),
                        'is_renewal': is_renewal,
                        'exclude_security_deposit': exclude_security_deposit,
                        'exclude_inspection_fee': exclude_inspection_fee,
                    })

                    input_sheet_html = input_sheet_template.render(input_sheet_context)
                    previews['Input Sheet'] = input_sheet_html

            except Exception as e:
                logger.error(f"Error generating Input Sheet preview: {str(e)}")
                previews['Input Sheet'] = f"<p>Error generating Input Sheet preview: {str(e)}</p>"

            # Return JSON with all previews
            return JsonResponse({'previews': previews})

        # Generate PDFs for download
        else:
            logger.debug("Starting PDF generation for all documents")
            
            # Create in-memory ZIP file
            zip_buffer = BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                generated_count = 0
                
                for document in documents:
                    try:
                        # Read template content
                        if not document.file:
                            logger.error(f"No template file attached to document: {document.name}")
                            continue
                            
                        template_path = document.file.path
                        
                        if not os.path.exists(template_path):
                            logger.error(f"Template file not found at: {template_path}")
                            continue
                            
                        with open(template_path, 'r', encoding='utf-8') as template_file:
                            template_content = template_file.read()
                        
                        # Create template and context
                        template = Template(template_content)
                        context = Context({
                            'client': client,
                            'document': document,
                            'preview': False,
                            'today': dt.datetime.now().strftime('%B %d, %Y')  # Fixed datetime
                        })

                        # Process lease-specific data (same as preview)
                        lease_agreement_date = request.POST.get('lease_agreement_date', '')
                        term_start_date = request.POST.get('term_start_date', '')
                        term_end_date = request.POST.get('term_end_date', '')
                        is_renewal = request.POST.get('is_renewal') == 'true'
                        exclude_security_deposit = request.POST.get('exclude_security_deposit') == 'true'
                        exclude_inspection_fee = request.POST.get('exclude_inspection_fee') == 'true'


                        context.update({
                            'formatted_agreement_date': format_agreement_date(lease_agreement_date),
                            'lease_agreement_date': lease_agreement_date,
                            'formatted_start_date': clean_and_format_date(term_start_date),
                            'formatted_end_date': clean_and_format_date(term_end_date),
                            'term_start_date': term_start_date,
                            'term_end_date': term_end_date,
                            'is_renewal': is_renewal,
                            'exclude_security_deposit': exclude_security_deposit,
                            'exclude_inspection_fee': exclude_inspection_fee,
                        })

                        # Add ALL landlord data from form (same structure as preview)
                        landlord_data = {
                            # Basic Information
                            'full_name': request.POST.get('full_name'),
                            'address': request.POST.get('address'),
                            'city': request.POST.get('city'),
                            'state': request.POST.get('state'),
                            'zip_code': request.POST.get('zip_code'),
                            'phone': request.POST.get('phone'),
                            'email': request.POST.get('email'),
                            
                            # Rental Property Information
                            'property_address': request.POST.get('property_address'),
                            'property_city': request.POST.get('property_city'),
                            'property_state': request.POST.get('property_state'),
                            'property_zip': request.POST.get('property_zip'),
                            
                            # Term dates
                            'term_start_date': request.POST.get('term_start_date'),
                            'term_end_date': request.POST.get('term_end_date'),

                            # Agreement Defaults
                            'default_rent_amount': request.POST.get('default_rent_amount', 0),
                            'default_security_deposit': request.POST.get('default_security_deposit', 0),
                            'default_rent_due_day': request.POST.get('default_rent_due_day', 1),
                            'default_late_fee': request.POST.get('default_late_fee', 0),
                            'default_late_fee_start_day': request.POST.get('default_late_fee_start_day', 5),
                            'default_eviction_day': request.POST.get('default_eviction_day', 10),
                            'default_nsf_fee': request.POST.get('default_nsf_fee', 0),
                            'default_max_occupants': request.POST.get('default_max_occupants', 10),
                            'default_parking_spaces': request.POST.get('default_parking_spaces', 2),
                            'default_parking_fee': request.POST.get('default_parking_fee', 0),
                            'default_inspection_fee': request.POST.get('default_inspection_fee', 300.00),
                            'bedrooms': request.POST.get('bedrooms', 1),
                            'rental_months': request.POST.get('rental_months'),
                            
                            # Additional Contact Persons
                            'contact_person_1': request.POST.get('contact_person_1'),
                            'contact_person_2': request.POST.get('contact_person_2'),
                            'contact_phone': request.POST.get('contact_phone'),
                            'contact_email': request.POST.get('contact_email'),
                            
                            # Real Estate Company Information
                            'real_estate_company': request.POST.get('real_estate_company'),
                            'company_mailing_address': request.POST.get('company_mailing_address'),
                            'company_city': request.POST.get('company_city'),
                            'company_state': request.POST.get('company_state'),
                            'company_zip': request.POST.get('company_zip'),
                            'company_contact_person': request.POST.get('company_contact_person'),
                            'company_phone': request.POST.get('company_phone'),
                            'company_email': request.POST.get('company_email'),
                            'broker_name': request.POST.get('broker_name'),
                            'broker_phone': request.POST.get('broker_phone'),
                            'broker_email': request.POST.get('broker_email'),

                            # Special Lease Instructions/Notes
                            'lease_special_notes': request.POST.get('lease_special_notes', ''),

                            'is_renewal': is_renewal,
                            'exclude_security_deposit': exclude_security_deposit,
                            'exclude_inspection_fee': exclude_inspection_fee,

                        }

                        # Convert numeric fields (same as preview)
                        numeric_fields = {
                            'default_rent_amount': 0,
                            'default_security_deposit': 0,
                            'default_late_fee': 50,
                            'default_nsf_fee': 35,
                            'default_inspection_fee': 300.00,
                            'bedrooms': 0,
                            'rental_months': 0,
                            'default_rent_due_day': 1,
                            'default_late_fee_start_day': 5,
                            'default_eviction_day': 10,
                            'default_max_occupants': 10,
                            'default_parking_spaces': 2,
                            'default_parking_fee': 0
                        }

                        for field, default in numeric_fields.items():
                            try:
                                value = request.POST.get(field, default)
                                if value:
                                    if field in ['default_rent_amount', 'default_security_deposit', 
                                               'default_late_fee', 'default_nsf_fee', 'default_parking_fee', 
                                               'default_inspection_fee']:
                                        landlord_data[field] = float(value)
                                    else:
                                        landlord_data[field] = int(float(value))
                                else:
                                    landlord_data[field] = default
                            except (ValueError, TypeError) as e:
                                logger.warning(f"Invalid number format for {field}: {str(e)}")
                                landlord_data[field] = default

                        context['landlord'] = landlord_data

                        # Generate PDF
                        html_string = template.render(context)
                        pdf_bytes = HTML(
                            string=html_string,
                            base_url=request.build_absolute_uri('/')
                        ).write_pdf()

                        # Add to ZIP
                        filename = f"{document.name.replace(' ', '_')}_{client_name}.pdf"
                        zip_file.writestr(filename, pdf_bytes)

                        # Save PDF to disk for later viewing
                        lease_docs_dir = os.path.join(settings.MEDIA_ROOT, 'lease_documents', client_name.replace(' ', '_'))
                        os.makedirs(lease_docs_dir, exist_ok=True)
                        saved_pdf_path = os.path.join(lease_docs_dir, filename)
                        with open(saved_pdf_path, 'wb') as pdf_file:
                            pdf_file.write(pdf_bytes)

                        # Store the relative path for later reference
                        if not hasattr(request, '_generated_pdf_paths'):
                            request._generated_pdf_paths = {}
                        request._generated_pdf_paths[document.name] = f"lease_documents/{client_name.replace(' ', '_')}/{filename}"

                        generated_count += 1

                    except Exception as e:
                        logger.error(f"Error generating PDF for {document.name}: {str(e)}")
                        continue

                # Generate Input Sheet after all lease documents
                try:
                    logger.debug("Generating Lease Input Sheet")
                    input_sheet_path = os.path.join(
                        settings.BASE_DIR, 'docsAppR', 'templates', 'account', 'lease_input_sheet.html'
                    )

                    if os.path.exists(input_sheet_path):
                        with open(input_sheet_path, 'r', encoding='utf-8') as input_sheet_file:
                            input_sheet_content = input_sheet_file.read()

                        input_sheet_template = Template(input_sheet_content)

                        # Build context for input sheet (reuse last landlord_data and context values)
                        lease_agreement_date = request.POST.get('lease_agreement_date', '')
                        term_start_date = request.POST.get('term_start_date', '')
                        term_end_date = request.POST.get('term_end_date', '')
                        is_renewal = request.POST.get('is_renewal') == 'true'
                        exclude_security_deposit = request.POST.get('exclude_security_deposit') == 'true'
                        exclude_inspection_fee = request.POST.get('exclude_inspection_fee') == 'true'

                        # Rebuild landlord_data for input sheet
                        input_sheet_landlord_data = {
                            'full_name': request.POST.get('full_name'),
                            'address': request.POST.get('address'),
                            'city': request.POST.get('city'),
                            'state': request.POST.get('state'),
                            'zip_code': request.POST.get('zip_code'),
                            'phone': request.POST.get('phone'),
                            'email': request.POST.get('email'),
                            'property_address': request.POST.get('property_address'),
                            'property_city': request.POST.get('property_city'),
                            'property_state': request.POST.get('property_state'),
                            'property_zip': request.POST.get('property_zip'),
                            'default_rent_amount': request.POST.get('default_rent_amount', 0),
                            'default_security_deposit': request.POST.get('default_security_deposit', 0),
                            'default_rent_due_day': request.POST.get('default_rent_due_day', 1),
                            'default_late_fee': request.POST.get('default_late_fee', 0),
                            'default_late_fee_start_day': request.POST.get('default_late_fee_start_day', 5),
                            'default_eviction_day': request.POST.get('default_eviction_day', 10),
                            'default_nsf_fee': request.POST.get('default_nsf_fee', 0),
                            'default_max_occupants': request.POST.get('default_max_occupants', 10),
                            'default_parking_spaces': request.POST.get('default_parking_spaces', 2),
                            'default_parking_fee': request.POST.get('default_parking_fee', 0),
                            'default_inspection_fee': request.POST.get('default_inspection_fee', 300.00),
                            'bedrooms': request.POST.get('bedrooms', 1),
                            'rental_months': request.POST.get('rental_months'),
                            'contact_person_1': request.POST.get('contact_person_1'),
                            'contact_person_2': request.POST.get('contact_person_2'),
                            'contact_phone': request.POST.get('contact_phone'),
                            'contact_email': request.POST.get('contact_email'),
                            'real_estate_company': request.POST.get('real_estate_company'),
                            'company_mailing_address': request.POST.get('company_mailing_address'),
                            'company_city': request.POST.get('company_city'),
                            'company_state': request.POST.get('company_state'),
                            'company_zip': request.POST.get('company_zip'),
                            'company_contact_person': request.POST.get('company_contact_person'),
                            'company_phone': request.POST.get('company_phone'),
                            'company_email': request.POST.get('company_email'),
                            'broker_name': request.POST.get('broker_name'),
                            'broker_phone': request.POST.get('broker_phone'),
                            'broker_email': request.POST.get('broker_email'),
                            'lease_special_notes': request.POST.get('lease_special_notes', ''),
                        }

                        # Convert numeric fields
                        for field in ['default_rent_amount', 'default_security_deposit', 'default_late_fee',
                                     'default_nsf_fee', 'default_parking_fee', 'default_inspection_fee']:
                            try:
                                input_sheet_landlord_data[field] = float(input_sheet_landlord_data[field] or 0)
                            except (ValueError, TypeError):
                                input_sheet_landlord_data[field] = 0

                        for field in ['default_rent_due_day', 'default_late_fee_start_day', 'default_eviction_day',
                                     'default_max_occupants', 'default_parking_spaces', 'bedrooms', 'rental_months']:
                            try:
                                input_sheet_landlord_data[field] = int(float(input_sheet_landlord_data[field] or 0))
                            except (ValueError, TypeError):
                                input_sheet_landlord_data[field] = 0

                        input_sheet_context = Context({
                            'client': client,
                            'landlord': input_sheet_landlord_data,
                            'today': dt.datetime.now().strftime('%B %d, %Y'),
                            'formatted_agreement_date': format_agreement_date(lease_agreement_date),
                            'formatted_start_date': clean_and_format_date(term_start_date),
                            'formatted_end_date': clean_and_format_date(term_end_date),
                            'is_renewal': is_renewal,
                            'exclude_security_deposit': exclude_security_deposit,
                            'exclude_inspection_fee': exclude_inspection_fee,
                        })

                        input_sheet_html = input_sheet_template.render(input_sheet_context)
                        input_sheet_pdf = HTML(
                            string=input_sheet_html,
                            base_url=request.build_absolute_uri('/')
                        ).write_pdf()

                        # Add input sheet to ZIP (with 00_ prefix so it appears first)
                        input_sheet_filename = f"00_Input_Sheet_{client_name}.pdf"
                        zip_file.writestr(input_sheet_filename, input_sheet_pdf)

                        # Save input sheet PDF to disk for later viewing
                        lease_docs_dir = os.path.join(settings.MEDIA_ROOT, 'lease_documents', client_name.replace(' ', '_'))
                        os.makedirs(lease_docs_dir, exist_ok=True)
                        saved_input_sheet_path = os.path.join(lease_docs_dir, f"Input_Sheet_{client_name}.pdf")
                        with open(saved_input_sheet_path, 'wb') as pdf_file:
                            pdf_file.write(input_sheet_pdf)

                        # Store the relative path for later reference
                        if not hasattr(request, '_generated_pdf_paths'):
                            request._generated_pdf_paths = {}
                        request._generated_pdf_paths['Input Sheet'] = f"lease_documents/{client_name.replace(' ', '_')}/Input_Sheet_{client_name}.pdf"

                        logger.debug("Lease Input Sheet generated successfully")

                except Exception as e:
                    logger.error(f"Error generating Input Sheet: {str(e)}")
                    # Continue without input sheet if it fails

            if generated_count == 0:
                return HttpResponse("No documents could be generated", status=400)

            # Create Lease record with all form data and link documents to it
            try:
                from dateutil.parser import parse as date_parse

                # Parse dates
                start_date = None
                end_date = None
                agreement_date = None
                try:
                    if term_start_date:
                        start_date = date_parse(term_start_date).date()
                    if term_end_date:
                        end_date = date_parse(term_end_date).date()
                    if lease_agreement_date:
                        agreement_date = date_parse(lease_agreement_date).date()
                except:
                    pass

                # Helper to safely get numeric values
                def safe_decimal(value, default=0):
                    try:
                        return float(value) if value else default
                    except (ValueError, TypeError):
                        return default

                def safe_int(value, default=0):
                    try:
                        return int(float(value)) if value else default
                    except (ValueError, TypeError):
                        return default

                # Create the Lease record with all form data
                lease = Lease.objects.create(
                    client=client,
                    # Lessor Information
                    lessor_name=request.POST.get('full_name', ''),
                    lessor_address=request.POST.get('address', ''),
                    lessor_city=request.POST.get('city', ''),
                    lessor_state=request.POST.get('state', ''),
                    lessor_zip=request.POST.get('zip_code', ''),
                    lessor_phone=request.POST.get('phone', ''),
                    lessor_email=request.POST.get('email', ''),
                    lessor_contact_person_1=request.POST.get('contact_person_1', ''),
                    lessor_contact_person_2=request.POST.get('contact_person_2', ''),
                    lessor_contact_phone=request.POST.get('contact_phone', ''),
                    lessor_contact_email=request.POST.get('contact_email', ''),
                    # Property Information
                    property_address=request.POST.get('property_address', ''),
                    property_city=request.POST.get('property_city', ''),
                    property_state=request.POST.get('property_state', ''),
                    property_zip=request.POST.get('property_zip', ''),
                    bedrooms=safe_int(request.POST.get('bedrooms'), 1),
                    # Rental Terms
                    lease_start_date=start_date,
                    lease_end_date=end_date,
                    lease_agreement_date=agreement_date,
                    rental_months=safe_int(request.POST.get('rental_months'), 12),
                    monthly_rent=safe_decimal(request.POST.get('default_rent_amount'), 0),
                    security_deposit=safe_decimal(request.POST.get('default_security_deposit'), 0),
                    rent_due_day=safe_int(request.POST.get('default_rent_due_day'), 1),
                    late_fee=safe_decimal(request.POST.get('default_late_fee'), 50),
                    late_fee_start_day=safe_int(request.POST.get('default_late_fee_start_day'), 5),
                    eviction_day=safe_int(request.POST.get('default_eviction_day'), 10),
                    nsf_fee=safe_decimal(request.POST.get('default_nsf_fee'), 35),
                    max_occupants=safe_int(request.POST.get('default_max_occupants'), 10),
                    parking_spaces=safe_int(request.POST.get('default_parking_spaces'), 2),
                    parking_fee=safe_decimal(request.POST.get('default_parking_fee'), 0),
                    inspection_fee=safe_decimal(request.POST.get('default_inspection_fee'), 300),
                    # Real Estate Company Info
                    real_estate_company=request.POST.get('real_estate_company', ''),
                    company_mailing_address=request.POST.get('company_mailing_address', ''),
                    company_city=request.POST.get('company_city', ''),
                    company_state=request.POST.get('company_state', ''),
                    company_zip=request.POST.get('company_zip', ''),
                    company_contact_person=request.POST.get('company_contact_person', ''),
                    company_phone=request.POST.get('company_phone', ''),
                    company_email=request.POST.get('company_email', ''),
                    broker_name=request.POST.get('broker_name', ''),
                    broker_phone=request.POST.get('broker_phone', ''),
                    broker_email=request.POST.get('broker_email', ''),
                    # Special Notes
                    special_notes=request.POST.get('lease_special_notes', ''),
                    # Flags
                    is_renewal=is_renewal,
                    exclude_security_deposit=exclude_security_deposit,
                    exclude_inspection_fee=exclude_inspection_fee,
                    # Status
                    status='generated',
                    generated_at=timezone.now(),
                    # User tracking
                    created_by=request.user if request.user.is_authenticated else None,
                    last_modified_by=request.user if request.user.is_authenticated else None,
                )

                # Create LeaseDocument records for each document
                document_type_map = {
                    'Engagement Agreement': 'engagement_agreement',
                    'Term Sheet': 'term_sheet',
                    'Month to Month Rental': 'month_to_month_rental',
                }

                # Create LeaseDocument records and log activity for each document
                pdf_paths = getattr(request, '_generated_pdf_paths', {})
                for document in documents:
                    doc_type = document_type_map.get(document.name, 'engagement_agreement')
                    file_path = pdf_paths.get(document.name, '')
                    LeaseDocument.objects.create(
                        lease=lease,
                        document_type=doc_type,
                        document_name=f"{document.name} - {client_name}",
                        file_path=file_path,
                    )
                    # Log activity for each document created
                    LeaseActivity.objects.create(
                        lease=lease,
                        activity_type='document_created',
                        description=f'Created {document.name} for {client_name}',
                        performed_by=request.user if request.user.is_authenticated else None
                    )

                # Create input sheet document record
                input_sheet_path = pdf_paths.get('Input Sheet', '')
                LeaseDocument.objects.create(
                    lease=lease,
                    document_type='input_sheet',
                    document_name=f"Input Sheet - {client_name}",
                    file_path=input_sheet_path,
                )
                # Log activity for input sheet
                LeaseActivity.objects.create(
                    lease=lease,
                    activity_type='document_created',
                    description=f'Created Input Sheet for {client_name}',
                    performed_by=request.user if request.user.is_authenticated else None
                )

                # Log the overall lease generation activity
                LeaseActivity.objects.create(
                    lease=lease,
                    activity_type='generated',
                    description=f'Generated lease package for {client_name} (${lease.monthly_rent}/month, {start_date} to {end_date})',
                    performed_by=request.user if request.user.is_authenticated else None
                )

                logger.debug(f"Successfully created Lease record with {generated_count + 1} documents")

            except Exception as track_error:
                logger.error(f"Failed to create Lease record: {str(track_error)}")
                import traceback
                logger.error(traceback.format_exc())
                print(f"LEASE CREATION ERROR: {str(track_error)}")
                print(traceback.format_exc())
                # Don't fail the whole request if tracking fails

            # Return ZIP file
            zip_buffer.seek(0)
            response = HttpResponse(zip_buffer.read(), content_type='application/zip')
            response['Content-Disposition'] = f'attachment; filename="{client_name}_documents.zip"'
            logger.debug(f"Successfully generated {generated_count} documents")
            return response

    except Exception as e:
        logger.error(f"Unexpected error in batch generation: {str(e)}", exc_info=True)
        return HttpResponse(f"An unexpected error occurred: {str(e)}", status=500)

import re
import os
import logging
from io import BytesIO
from django.shortcuts import get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.core.files.storage import default_storage
from django.template import Template, Context
from dateutil.parser import parse as parse_date
from xhtml2pdf import pisa
from .models import Client, Document, Landlord

logger = logging.getLogger(__name__)
def generate_document_from_html(request):
    logger.debug("Document generation started")
    print("STATIC_ROOT:", settings.STATIC_ROOT)
    print("Files in static:", os.listdir(settings.STATIC_ROOT))
    
    
    if request.method != 'POST':
        logger.error("Invalid request method")
        return HttpResponse("Only POST requests are allowed", status=405)

    try:
        logger.debug(f"POST data received: {dict(request.POST)}")
        
        # Get required parameters
        document_name = request.POST.get('document_name')
        if not document_name:
            logger.error("Missing document_name")
            return HttpResponse("document_name is required", status=400)

        client_name = request.POST.get('client_name')
        if not client_name:
            logger.error("Missing client_name")
            return HttpResponse("client_name is required", status=400)

        # Get models
        try:
            client = get_object_or_404(Client, pOwner=client_name)
            document = get_object_or_404(Document, name=document_name)
            logger.debug(f"Found client {client_name} and document {document_name}")
        except Exception as e:
            logger.error(f"Error fetching models: {str(e)}")
            return HttpResponse("Error loading client or document", status=404)

        # Date formatting function
        def clean_and_format_date(date_str):
            if not date_str:
                return ""
            
            try:
                cleaned = re.sub(r'[^\d/-]', '', str(date_str))
                date_obj = parse_date(cleaned)
                
                if not date_obj:
                    for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%m-%d-%Y', '%m/%d/%Y', '%d-%m-%Y', '%d/%m/%Y'):
                        try:
                            date_obj = dt.datetime.strptime(cleaned, fmt).date()
                            break
                        except ValueError:
                            continue
                
                if date_obj:
                    try:
                        return date_obj.strftime('%B %d, %Y').replace(' 0', ' ')
                    except:
                        return date_obj.strftime('%B %d, %Y')
            except Exception as e:
                logger.warning(f"Date formatting failed for {date_str}: {str(e)}")
            return ""

        # Read template content
        try:
            if not document.file:
                logger.error("No template file attached to document")
                return HttpResponse("Document template file is missing", status=400)
                
            template_path = document.file.path
            logger.debug(f"Attempting to read template from: {template_path}")
            
            if not os.path.exists(template_path):
                logger.error(f"Template file not found at: {template_path}")
                return HttpResponse("Template file not found", status=404)
                
            with open(template_path, 'r', encoding='utf-8') as template_file:
                template_content = template_file.read()
            logger.debug("Successfully read template file")
            
            # Create a temporary template from the content
            template = Template(template_content)
            
        except Exception as e:
            logger.error(f"Error reading template file: {str(e)}")
            return HttpResponse(f"Error loading template: {str(e)}", status=500)

        # Prepare context
        context = Context({
            'client': client,
            'document': document,
            'preview': request.POST.get('preview') == 'true',
            'today': dt.datetime.now().strftime('%B %d, %Y')
        })

        # Process lease-specific data
        if document.document_type == 'lease':
            logger.debug("Processing lease document")
            try:
                term_start_date = request.POST.get('term_start_date', '')
                term_end_date = request.POST.get('term_end_date', '')


                print(clean_and_format_date(term_start_date))
                print(clean_and_format_date(term_end_date))
                context.update({
                    'formatted_start_date': clean_and_format_date(term_start_date),
                    'formatted_end_date': clean_and_format_date(term_end_date),
                    'term_start_date': term_start_date,
                    'term_end_date': term_end_date,
                })

                landlord_data = {
                # Basic Information
                'full_name': request.POST.get('full_name'),
                'address': request.POST.get('address'),
                'city': request.POST.get('city'),
                'state': request.POST.get('state'),
                'zip_code': request.POST.get('zip_code'),
                'phone': request.POST.get('phone'),
                'email': request.POST.get('email'),
                
                # Rental Property Information
                'property_address': request.POST.get('property_address'),
                'property_city': request.POST.get('property_city'),
                'property_state': request.POST.get('property_state'),
                'property_zip': request.POST.get('property_zip'),
                
                #term start and end
                'term_start_date': request.POST.get('term_start_date'),
                'term_end_date': request.POST.get('term_end_date'),

                # Agreement Defaults
                'default_rent_amount': request.POST.get('default_rent_amount', 0),
                'default_security_deposit': request.POST.get('default_security_deposit', 0),
                'default_rent_due_day': request.POST.get('default_rent_due_day', 1),
                'default_late_fee': request.POST.get('default_late_fee', 0),
                'default_late_fee_start_day': request.POST.get('default_late_fee_start_day', 5),
                'default_eviction_day': request.POST.get('default_eviction_day', 10),
                'default_nsf_fee': request.POST.get('default_nsf_fee', 0),
                'default_max_occupants': request.POST.get('default_max_occupants', 10),
                'default_parking_spaces': request.POST.get('default_parking_spaces', 2),
                'default_parking_fee': request.POST.get('default_parking_fee', 0),
                'default_inspection_fee': request.POST.get('default_inspection_fee', 300.00),
                'bedrooms': request.POST.get('bedrooms', 1),
                'rental_months': request.POST.get('rental_months'),
                # Additional Contact Persons
                'contact_person_1': request.POST.get('contact_person_1'),
                'contact_person_2': request.POST.get('contact_person_2'),
                'contact_phone': request.POST.get('contact_phone'),
                'contact_email': request.POST.get('contact_email'),
                
                # Real Estate Company Information
                'real_estate_company': request.POST.get('real_estate_company'),
                'company_mailing_address': request.POST.get('company_mailing_address'),
                'company_city': request.POST.get('company_city'),
                'company_state': request.POST.get('company_state'),
                'company_zip': request.POST.get('company_zip'),
                'company_contact_person': request.POST.get('company_contact_person'),
                'company_phone': request.POST.get('company_phone'),
                'company_email': request.POST.get('company_email'),
                'broker_name': request.POST.get('broker_name'),
                'broker_phone': request.POST.get('broker_phone'),
                'broker_email': request.POST.get('broker_email'),
                }


                numeric_fields = {
                    'default_rent_amount': 0,
                    'default_security_deposit': 0,
                    'default_late_fee': 50,
                    'default_nsf_fee': 35,
                    'default_inspection_fee': 300.00,
                    'bedrooms': 0,
                    'rental_months': 0
                }

                for field, default in numeric_fields.items():
                    try:
                        value = request.POST.get(field, default)
                        landlord_data[field] = float(value) if value else default
                    except (ValueError, TypeError) as e:
                        logger.warning(f"Invalid number format for {field}: {str(e)}")
                        landlord_data[field] = default

                context['landlord'] = landlord_data
                logger.debug("Lease context prepared successfully")

            except Exception as e:
                logger.error(f"Error processing lease data: {str(e)}")
                return HttpResponse(f"Error processing lease data: {str(e)}", status=400)

            # Handle preview request
            if request.POST.get('preview') == 'true':
                logger.debug("Generating HTML preview")
                try:
                    html_content = template.render(context)
                    return HttpResponse(html_content)
                except Exception as e:
                    logger.error(f"Template rendering failed: {str(e)}")
                    return HttpResponse(f"Error generating preview: {str(e)}", status=500)

            # Generate PDF
            logger.debug("Starting PDF generation")
            try:
                html_string = template.render(context)
    
                # Generate PDF with WeasyPrint
                pdf_bytes = HTML(
                    string=html_string,
                    base_url=request.build_absolute_uri('/')  # For resolving relative URLs
                ).write_pdf()
                
                response = HttpResponse(pdf_bytes, content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="{document_name}_{client_name}.pdf"'
                logger.debug("PDF generated successfully")
                return response
                
            except Exception as e:
                logger.error(f"PDF generation error: {str(e)}")
                return HttpResponse(f"Error generating PDF: {str(e)}", status=500)

    except Exception as e:
        logger.error(f"Unexpected error: {str(e)}", exc_info=True)
        return HttpResponse(f"An unexpected error occurred: {str(e)}", status=500)

def save_landlord(request):
    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        try:
            landlord_data = {
                # Basic Information
                'full_name': request.POST.get('full_name'),
                'address': request.POST.get('address'),
                'city': request.POST.get('city'),
                'state': request.POST.get('state'),
                'zip_code': request.POST.get('zip_code'),
                'phone': request.POST.get('phone'),
                'email': request.POST.get('email'),
                
                # Rental Property Information
                'property_address': request.POST.get('property_address'),
                'property_city': request.POST.get('property_city'),
                'property_state': request.POST.get('property_state'),
                'property_zip': request.POST.get('property_zip'),
                
                #term start and end
                'term_start_date': request.POST.get('term_start_date'),
                'term_end_date': request.POST.get('term_end_date'),

                # Agreement Defaults
                'default_rent_amount': request.POST.get('default_rent_amount', 0),
                'default_security_deposit': request.POST.get('default_security_deposit', 0),
                'default_rent_due_day': request.POST.get('default_rent_due_day', 1),
                'default_late_fee': request.POST.get('default_late_fee', 0),
                'default_late_fee_start_day': request.POST.get('default_late_fee_start_day', 5),
                'default_eviction_day': request.POST.get('default_eviction_day', 10),
                'default_nsf_fee': request.POST.get('default_nsf_fee', 0),
                'default_max_occupants': request.POST.get('default_max_occupants', 10),
                'default_parking_spaces': request.POST.get('default_parking_spaces', 2),
                'default_parking_fee': request.POST.get('default_parking_fee', 0),
                'default_inspection_fee': request.POST.get('default_inspection_fee', 300.00),
                'bedrooms': request.POST.get('bedrooms', 1),
                'rental_months': request.POST.get('rental_months'),
                # Additional Contact Persons
                'contact_person_1': request.POST.get('contact_person_1'),
                'contact_person_2': request.POST.get('contact_person_2'),
                'contact_phone': request.POST.get('contact_phone'),   
                'contact_email': request.POST.get('contact_email'),
                
                # Real Estate Company Information
                'real_estate_company': request.POST.get('real_estate_company'),
                'company_mailing_address': request.POST.get('company_mailing_address'),
                'company_city': request.POST.get('company_city'),
                'company_state': request.POST.get('company_state'),
                'company_zip': request.POST.get('company_zip'),
                'company_contact_person': request.POST.get('company_contact_person'),
                'company_phone': request.POST.get('company_phone'),
                'company_email': request.POST.get('company_email'),
                'broker_name': request.POST.get('broker_name'),
                'broker_phone': request.POST.get('broker_phone'),
                'broker_email': request.POST.get('broker_email'),
            }

            if landlord_data['term_start_date']:
                 landlord_data['term_start_date'] = parse_date(landlord_data['term_start_date'])
            if landlord_data['term_end_date']:
                landlord_data['term_end_date'] = parse_date(landlord_data['term_end_date'])

            # Convert empty strings to None for non-required fields
            for field in landlord_data:
                if landlord_data[field] == '':
                    landlord_data[field] = None
            
            # Validate required fields
            required_fields = [
                'full_name', 
                'address',
                'city',
                'state',
                'zip_code',
                'phone',
                'property_address',
                'property_city',
                'property_state',
                'property_zip'
            ]
            
            missing_fields = [field for field in required_fields if not landlord_data.get(field)]
            if missing_fields:
                return JsonResponse({
                    'success': False, 
                    'error': f'Missing required fields: {", ".join(missing_fields)}'
                })
            
            # Convert numeric fields
            numeric_fields = [
                'default_rent_amount',
                'default_security_deposit',
                'default_late_fee',
                'default_nsf_fee',
                'default_rent_due_day',
                'default_late_fee_start_day',
                'default_eviction_day',
                'default_max_occupants',
                'default_parking_spaces',
                'default_parking_fee',
                'default_inspection_fee'
            ]
            
            for field in numeric_fields:
                if landlord_data[field] is not None:
                    try:
                        if field in ['default_rent_amount', 'default_security_deposit', 
                                    'default_late_fee', 'default_nsf_fee', 'default_inspection_fee']:
                            landlord_data[field] = float(landlord_data[field])
                        else:
                            landlord_data[field] = int(landlord_data[field])
                    except (ValueError, TypeError):
                        return JsonResponse({
                            'success': False,
                            'error': f'Invalid value for {field.replace("_", " ").title()}'
                        })
            
            # Save to database
            landlord, created = Landlord.objects.update_or_create(
                property_address=landlord_data['property_address'],
                defaults=landlord_data
            )
            
            return JsonResponse({
                'success': True, 
                'created': created,
                'landlord_id': landlord.id
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False, 
                'error': str(e),
                'type': type(e).__name__
            })
    
    return JsonResponse({
        'success': False, 
        'error': 'Invalid request method or not AJAX'
    })
def convert_excel_to_pdf(excel_path, pdf_path):
    """Convert specific Excel sheet to PDF using the appropriate method for the OS"""
    if platform.system() == 'Windows':
        try:
            import win32com.client
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            wb = excel.Workbooks.Open(excel_path)
            wb.ExportAsFixedFormat(0, pdf_path)  # Export only the selected sheet
            wb.Close()
            excel.Quit()
        except Exception as e:
            logger.error(f"Error converting with Excel: {str(e)}")
            raise
    else:
        # For Linux using LibreOffice
        try:
            import subprocess
            
            # Get the directory of the output file
            output_dir = os.path.dirname(pdf_path)
            
            # Ensure the directory exists
            os.makedirs(output_dir, exist_ok=True)
            
            # First try unoconv if available
            try:
                subprocess.run(['which', 'unoconv'], check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                
                # Use unoconv for direct conversion
                subprocess.run([
                    'unoconv',
                    '-f', 'pdf',
                    '-o', pdf_path,
                    excel_path
                ], check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                
            except (subprocess.SubprocessError, FileNotFoundError):
                # Fall back to LibreOffice if unoconv not available
                subprocess.run([
                    'libreoffice',
                    '--headless',
                    '--convert-to', 'pdf',
                    '--outdir', output_dir,
                    excel_path
                ], check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                
                # LibreOffice will create a file with the same name but .pdf extension
                libreoffice_output = os.path.splitext(os.path.basename(excel_path))[0] + '.pdf'
                libreoffice_output_path = os.path.join(output_dir, libreoffice_output)
                
                # Rename to desired output name if necessary
                if os.path.exists(libreoffice_output_path) and libreoffice_output_path != pdf_path:
                    os.rename(libreoffice_output_path, pdf_path)
                
        except Exception as e:
            logger.error(f"Error converting with LibreOffice: {str(e)}")
            raise

def generate_invoice_pdf(request, client_id):
    try:
        # Fetch the client data
        client = get_object_or_404(Client, pOwner=client_id)
        logger.info(f"Generating Excel for client: {client_id}")
        
        # Get rooms data from POST
        rooms_data = json.loads(request.POST.get('rooms_data', '{}'))
        
        # Load the template Excel file
        template_path = os.path.join(settings.BASE_DIR, 'docsAppR', 'templates', 'excel', '60_scope_form.xlsx')
        wb = load_workbook(template_path, data_only=True)
        
        # Select the ScopeCHLST sheet
        ws = wb['ScopeCHLST']

        # Map inspection checklist data - column mappings
        checklist_mappings = {
            'clg': 'C',  # Ceiling
            'lit': 'D',  # Lighting
            'hvc': 'E',  # HVAC
            'wal': 'F',  # Walls
            'ele': 'G',  # Electrical
            'flr': 'H',  # Floor
            'bb': 'I',   # Baseboards
            'dor': 'J',  # Doors
            'wdw': 'K',  # Windows
            'wdt': 'L',  # Water Damage
        }
        
        # Create a mapping of room IDs to their row numbers
        room_rows = {}
        for row in range(2, ws.max_row + 1):  # Start from row 2
            room_cell = ws[f'B{row}'].value
            if room_cell:
                room_rows[str(room_cell).strip()] = row
        print(room_rows)
        # Precise data placement
        print(rooms_data.items())
        for room_id, room_data in rooms_data.items():
            if room_id in room_rows:
                row_number = room_rows[room_id]
                
                for field, column in checklist_mappings.items():
                    cell_value = room_data.get(field, '')
                    ws[f'{column}{row_number}'] = cell_value or 'N/A'
        
        # Generate filename
        filename = f"scope_form_{client_id}_all_rooms.xlsx"
        
        # Create temporary directory for file conversion
        with tempfile.TemporaryDirectory() as temp_dir:
            # Save Excel file to temp directory
            temp_excel_path = os.path.join(temp_dir, filename)
            wb.save(temp_excel_path)
            
            # Update JobInfo with claim/client info
            


            # Create PDF filename
            pdf_filename = f"scope_form_{client_id}.pdf"
            temp_pdf_path = os.path.join(temp_dir, pdf_filename)
            
            # Convert Excel to PDF
            convert_excel_to_pdf(temp_excel_path, temp_pdf_path)
            
            # Read the generated PDF
            with open(temp_pdf_path, 'rb') as pdf_file:
                pdf_content = pdf_file.read()
            
            # Save both Excel and PDF to File model
            excel_obj = File(
                filename=filename,
                size=os.path.getsize(temp_excel_path)
            )
            excel_obj.file.save(filename, ContentFile(open(temp_excel_path, 'rb').read()), save=True)
            
            pdf_obj = File(
                filename=pdf_filename,
                size=len(pdf_content)
            )
            pdf_obj.file.save(pdf_filename, ContentFile(pdf_content), save=True)
            
            # Generate response with PDF
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{pdf_filename}"'
            response.write(pdf_content)
            
            # Clear session data
            if 'inspection_data' in request.session:
                del request.session['inspection_data']
            
            return response
        
    except Exception as e:
        logger.error(f"Error generating files: {str(e)}")
        return HttpResponse(f"An error occurred while generating the files: {str(e)}", status=500)

@login_required
@login_required
def emails(request):
    # Get filter parameters
    category_id = request.GET.get('category')
    client_name = request.GET.get('client')
    date_range = request.GET.get('date_range', 'recent')

    # Base queryset - use Document model
    documents = Document.objects.filter(created_by=request.user)
    
    # Apply filters
    if category_id:
        documents = documents.filter(category_id=category_id)
    if client_name:
        documents = documents.filter(client_name__icontains=client_name)
    
    # Date range filters
    if date_range == 'today':
        documents = documents.filter(created_at__date=timezone.now().date())
    elif date_range == 'week':
        week_ago = timezone.now() - timezone.timedelta(days=7)
        documents = documents.filter(created_at__gte=week_ago)
    elif date_range == 'month':
        month_ago = timezone.now() - timezone.timedelta(days=30)
        documents = documents.filter(created_at__gte=month_ago)
    else:  # recent - last 50
        documents = documents.order_by('-created_at')[:50]
    
    categories = DocumentCategory.objects.all()
    sent_emails = SentEmail.objects.filter(sent_by=request.user)[:20]
    schedules = EmailSchedule.objects.filter(created_by=request.user, is_active=True)
    
    if request.method == 'POST':
        form = EmailForm(request.POST)
        # Set queryset to user's documents
        form.fields['documents'].queryset = documents
        if form.is_valid():
            try:
                # Get selected documents
                selected_docs = form.cleaned_data['documents']
                # Recipients is already a list from clean_recipients()
                recipients = form.cleaned_data['recipients']

                # Check if user pasted Excel data
                pasted_excel_data = request.POST.get('pasted_excel_data', '').strip()

                # Check if user wants to embed Excel as HTML (from request POST)
                embed_excel = request.POST.get('embed_excel', False)

                # Create email
                email = EmailMessage(
                    subject=form.cleaned_data['subject'],
                    body=form.cleaned_data['body'],
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    to=recipients,
                )

                # Handle pasted Excel data (highest priority)
                excel_embedded = False
                if pasted_excel_data:
                    try:
                        # Convert pasted Excel data to HTML
                        excel_html = convert_pasted_excel_to_html(pasted_excel_data)
                        email.body = excel_html
                        email.content_subtype = "html"
                        excel_embedded = True
                        messages.success(request, 'Pasted Excel data converted to HTML email body.')
                    except Exception as e:
                        messages.warning(request, f'Could not convert pasted data: {str(e)}')

                # Handle Excel file embedding or attachment
                if not excel_embedded:
                    for doc in selected_docs:
                        # Check if document is Excel and user wants it embedded
                        if embed_excel and doc.file.name.endswith(('.xlsx', '.xls')):
                            try:
                                # Convert Excel to HTML and use as email body
                                excel_html = convert_excel_to_html(doc.file.path)
                                email.body = excel_html
                                email.content_subtype = "html"
                                excel_embedded = True
                                # Don't attach the Excel file if embedded
                                continue
                            except Exception as e:
                                messages.warning(request, f'Could not embed Excel as HTML: {str(e)}. Attaching instead.')

                        # Attach all other documents (or Excel if embedding failed)
                        email.attach_file(doc.file.path)

                # Create SentEmail record for tracking
                sent_email = SentEmail.objects.create(
                    subject=form.cleaned_data['subject'],
                    body=form.cleaned_data['body'],
                    recipients=recipients,
                    sent_by=request.user,
                    notify_on_open=form.cleaned_data['notify_on_open'],
                    admin_notification_email=form.cleaned_data['admin_notification_email'] or request.user.email,
                    scheduled_send_time=timezone.now() if form.cleaned_data['send_now'] else form.cleaned_data['scheduled_time']
                )
                sent_email.documents.set(selected_docs)

                # Add tracking pixel to email
                tracking_url = request.build_absolute_uri(
                    f'/emails/track/{sent_email.tracking_pixel_id}/'
                )

                if email.content_subtype == "html" or excel_embedded:
                    # Add tracking pixel to HTML email
                    email.body += f'<img src="{tracking_url}" width="1" height="1" />'
                else:
                    # Convert plain text to HTML with tracking pixel
                    html_body = f'<div style="white-space: pre-wrap;">{form.cleaned_data["body"]}</div>'
                    html_body += f'<img src="{tracking_url}" width="1" height="1" />'
                    email.body = html_body
                    email.content_subtype = "html"
                
                # Send email
                if form.cleaned_data['send_now'] or not form.cleaned_data['scheduled_time']:
                    email.send()
                    messages.success(request, 'Email sent successfully!')
                else:
                    # For scheduled emails, you'd typically use Celery
                    messages.success(request, 'Email scheduled successfully!')
                
            except Exception as e:
                import traceback
                error_details = traceback.format_exc()
                logger.error(f"Error sending email: {str(e)}\n{error_details}")
                messages.error(request, f'Error sending email: {str(e)}')

            return redirect('emails')
        else:
            # Form validation failed - show errors
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = EmailForm()
        # Set queryset to user's documents for GET request
        form.fields['documents'].queryset = documents
    
    context = {
        'documents': documents,
        'categories': categories,
        'sent_emails': sent_emails,
        'schedules': schedules,
        'form': form,
        'current_filters': {
            'category_id': category_id,
            'client_name': client_name,
            'date_range': date_range,
        }
    }
    
    return render(request, 'account/emails.html', context)


def convert_pasted_excel_to_html(pasted_data):
    """
    Convert pasted Excel data (tab-separated values) to styled HTML table

    Args:
        pasted_data: String containing tab-separated values from Excel copy-paste

    Returns:
        Styled HTML string
    """
    import io

    try:
        # Split into rows
        rows = [line.split('\t') for line in pasted_data.strip().split('\n')]

        if not rows:
            return "<p>No data provided</p>"

        # Build HTML table
        html_table = "<table class='email-table'>\n"

        # First row as header
        html_table += "  <thead>\n    <tr>\n"
        for cell in rows[0]:
            html_table += f"      <th>{cell.strip()}</th>\n"
        html_table += "    </tr>\n  </thead>\n"

        # Rest as body
        html_table += "  <tbody>\n"
        for row in rows[1:]:
            html_table += "    <tr>\n"
            for cell in row:
                html_table += f"      <td>{cell.strip()}</td>\n"
            html_table += "    </tr>\n"
        html_table += "  </tbody>\n"
        html_table += "</table>"

        # Add professional CSS styling
        styled_html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                body {{
                    font-family: Arial, Helvetica, sans-serif;
                    background-color: #f4f4f4;
                    padding: 20px;
                }}
                .email-container {{
                    max-width: 900px;
                    margin: 0 auto;
                    background-color: white;
                    padding: 20px;
                    border-radius: 8px;
                    box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                }}
                .email-table {{
                    border-collapse: collapse;
                    width: 100%;
                    font-size: 14px;
                    margin: 20px 0;
                }}
                .email-table thead {{
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                    color: white;
                }}
                .email-table th {{
                    padding: 12px 15px;
                    text-align: left;
                    font-weight: 600;
                    border: 1px solid #ddd;
                }}
                .email-table td {{
                    padding: 10px 15px;
                    border: 1px solid #ddd;
                }}
                .email-table tbody tr:nth-child(even) {{
                    background-color: #f8f9fa;
                }}
                .email-table tbody tr:hover {{
                    background-color: #e9ecef;
                }}
            </style>
        </head>
        <body>
            <div class="email-container">
                {html_table}
            </div>
        </body>
        </html>
        """

        return styled_html

    except Exception as e:
        return f"<p>Error converting pasted data to HTML: {str(e)}</p>"


def convert_excel_to_html(excel_file_path):
    """
    Convert Excel file to styled HTML table for email body

    Args:
        excel_file_path: Path to Excel file

    Returns:
        Styled HTML string
    """
    import pandas as pd

    try:
        # Read Excel file (first sheet by default)
        df = pd.read_excel(excel_file_path, engine='openpyxl')

        # Convert to HTML with styling
        html_table = df.to_html(
            index=False,
            classes='email-table',
            border=0,
            escape=False,
            na_rep=''
        )

        # Add professional CSS styling
        styled_html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                body {{
                    font-family: Arial, Helvetica, sans-serif;
                    background-color: #f4f4f4;
                    padding: 20px;
                }}
                .email-container {{
                    max-width: 900px;
                    margin: 0 auto;
                    background-color: white;
                    padding: 20px;
                    border-radius: 8px;
                    box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                }}
                .email-table {{
                    border-collapse: collapse;
                    width: 100%;
                    font-size: 14px;
                    margin: 20px 0;
                }}
                .email-table thead {{
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                    color: white;
                }}
                .email-table th {{
                    padding: 12px 15px;
                    text-align: left;
                    font-weight: 600;
                    border: 1px solid #ddd;
                }}
                .email-table td {{
                    padding: 10px 15px;
                    border: 1px solid #ddd;
                }}
                .email-table tbody tr:nth-child(even) {{
                    background-color: #f8f9fa;
                }}
                .email-table tbody tr:hover {{
                    background-color: #e9ecef;
                }}
            </style>
        </head>
        <body>
            <div class="email-container">
                {html_table}
            </div>
        </body>
        </html>
        """

        return styled_html

    except Exception as e:
        return f"<p>Error converting Excel to HTML: {str(e)}</p>"


@login_required
def track_email_open(request, tracking_pixel_id):
    try:
        sent_email = SentEmail.objects.get(tracking_pixel_id=tracking_pixel_id)
        sent_email.is_opened = True
        sent_email.opened_at = timezone.now()
        sent_email.save()
        
        # Create open event
        EmailOpenEvent.objects.create(
            sent_email=sent_email,
            ip_address=get_client_ip(request),
            user_agent=request.META.get('HTTP_USER_AGENT', '')
        )
        
        # Send notification if enabled
        if sent_email.notify_on_open and sent_email.admin_notification_email:
            try:
                notification_email = EmailMessage(
                    subject=f"Email Opened: {sent_email.subject}",
                    body=f"""
                    Your email has been opened!
                    
                    Email: {sent_email.subject}
                    Opened at: {timezone.now()}
                    Recipient: {', '.join(sent_email.recipients)}
                    """,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    to=[sent_email.admin_notification_email],
                )
                notification_email.send()
            except Exception as e:
                # Log error but don't break the tracking
                print(f"Error sending notification: {e}")
        
    except SentEmail.DoesNotExist:
        pass
    
    # Return a 1x1 transparent GIF
    from django.http import HttpResponse
    response = HttpResponse(
        base64.b64decode(b'R0lGODlhAQABAIAAAAAAAP///yH5BAEAAAAALAAAAAABAAEAAAIBRAA7'),
        content_type='image/gif'
    )
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response

def get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip

@login_required
def create_schedule(request):
    if request.method == 'POST':
        form = EmailScheduleForm(request.POST)
        if form.is_valid():
            schedule = form.save(commit=False)
            schedule.created_by = request.user
            schedule.recipients = form.cleaned_data['recipients']
            schedule.save()
            form.save_m2m()  # Save many-to-many relationships
            
            messages.success(request, 'Email schedule created successfully!')
            return redirect('emails')
    else:
        form = EmailScheduleForm()
    
    return render(request, 'account/email_schedule_form.html', {'form': form})

@login_required
def document_list_api(request):
    """API endpoint for document filtering"""
    category_id = request.GET.get('category')
    client_name = request.GET.get('client')
    search = request.GET.get('search', '')

    documents = Document.objects.filter(created_by=request.user)
    
    if category_id:
        documents = documents.filter(category_id=category_id)
    if client_name:
        documents = documents.filter(client_name__icontains=client_name)
    if search:
        documents = documents.filter(
            Q(filename__icontains=search) | 
            Q(description__icontains=search) |
            Q(client_name__icontains=search)
        )
    
    documents = documents.order_by('-created_at')[:50]
    
    data = []
    for doc in documents:
        data.append({
            'id': str(doc.id),
            'filename': doc.filename,
            'client_name': doc.client_name,
            'category': doc.category.name if doc.category else '',
            'created_at': doc.created_at.strftime('%Y-%m-%d %H:%M'),
            'description': doc.description,
        })
    
    return JsonResponse({'documents': data})
    


def reading_browser(request):
    """Main view for the reading browser"""
    images = ReadingImage.objects.all()
    return render(request, 'account/browser.html', {'images': images})

@csrf_exempt
def upload_readings(request):
    """Handle image uploads"""
    if request.method == 'POST' and request.FILES.getlist('images'):
        uploaded_files = request.FILES.getlist('images')
        results = {
            'success': [],
            'errors': [],
            'duplicates': []
        }
        
        for uploaded_file in uploaded_files:
            # Check if file already exists
            if ReadingImage.objects.filter(filename=uploaded_file.name).exists():
                results['duplicates'].append(uploaded_file.name)
                continue
            
            try:
                # Create new ReadingImage
                reading_image = ReadingImage(
                    filename=uploaded_file.name,
                    size=uploaded_file.size,
                    file=uploaded_file
                )
                reading_image.save()
                results['success'].append(uploaded_file.name)
            except Exception as e:
                results['errors'].append(f"{uploaded_file.name}: {str(e)}")
        
        return JsonResponse(results)
    
    return JsonResponse({'error': 'No files provided'}, status=400)

# views.py - Update get_sorted_readings to include MC
def get_sorted_readings(request):
    """Get sorted images based on criteria"""
    sort_by = request.GET.get('sort_by', 'filename')  # Default to filename sort
    order = request.GET.get('order', 'asc')
    
    images = ReadingImage.objects.all()
    
    # Apply sorting
    if sort_by == 'rh':
        field = 'rh_value'
    elif sort_by == 't':
        field = 't_value'
    elif sort_by == 'gpp':
        field = 'gpp_value'
    elif sort_by == 'mc':
        field = 'mc_value'
    elif sort_by == 'filename':
        field = 'filename'
    else:
        field = 'filename'  # Default to filename sorting
    
    if order == 'desc':
        field = f'-{field}'
    
    images = images.order_by(field)
    
    # Prepare data for JSON response
    image_data = []
    for image in images:
        image_data.append({
            'id': image.id,
            'filename': image.filename,
            'url': image.file.url,
            'rh': image.rh_value,
            't': image.t_value,
            'gpp': image.gpp_value,
            'mc': image.mc_value, 
            'size': image.get_file_size_display()
        })
    
    return JsonResponse({'images': image_data})

# views.py - Improved export_readings function
@csrf_exempt
def export_readings(request):
    """Export selected images as zip file with better error handling"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            folder_structure = data.get('folders', {})
            
            if not folder_structure:
                return JsonResponse({'error': 'No folders provided'}, status=400)
            
            # Validate that we have some images to export
            total_images = sum(len(images) for images in folder_structure.values())
            if total_images == 0:
                return JsonResponse({'error': 'No images in folders to export'}, status=400)
            
            # Create zip file in memory
            response = HttpResponse(content_type='application/zip')
            response['Content-Disposition'] = 'attachment; filename="reading_images.zip"'
            
            try:
                with zipfile.ZipFile(response, 'w') as zip_file:
                    exported_count = 0
                    missing_files = []
                    
                    for folder_name, image_data_list in folder_structure.items():
                        for image_data in image_data_list:
                            try:
                                image_id = image_data.get('id')
                                image = ReadingImage.objects.get(id=image_id)
                                
                                if image.file and os.path.exists(image.file.path):
                                    # Add to folder in zip
                                    zip_path = os.path.join(folder_name, image.filename)
                                    zip_file.write(image.file.path, zip_path)
                                    exported_count += 1
                                else:
                                    missing_files.append(image.filename)
                                    
                            except ReadingImage.DoesNotExist:
                                missing_files.append(f"Image ID {image_id}")
                            except Exception as e:
                                missing_files.append(f"{image_data.get('filename', 'Unknown')}: {str(e)}")
                    
                    if exported_count == 0:
                        return JsonResponse({
                            'error': f'No files could be exported. Missing files: {missing_files}'
                        }, status=404)
                    
                    if missing_files:
                        print(f"Warning: {len(missing_files)} files could not be exported: {missing_files}")
                
                return response
                
            except zipfile.BadZipFile:
                return JsonResponse({'error': 'Error creating zip file'}, status=500)
            except OSError as e:
                return JsonResponse({'error': f'File system error: {str(e)}'}, status=500)
            
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)
        except Exception as e:
            return JsonResponse({'error': f'Unexpected error: {str(e)}'}, status=500)
    
    return JsonResponse({'error': 'Invalid request method'}, status=400)


# views.py - Updated rename_reading function
# views.py - COMPLETELY FIXED rename_reading function
@csrf_exempt
def rename_reading(request, image_id):
    """Rename a reading image and properly update the file field"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            new_filename = data.get('filename')
            
            if not new_filename:
                return JsonResponse({'error': 'No filename provided'}, status=400)
            
            image = ReadingImage.objects.get(id=image_id)
            
            # Get the current file information
            if not image.file:
                return JsonResponse({'error': 'Image file not found'}, status=404)
            
            old_file_path = image.file.path
            old_filename = image.filename
            
            print(f"Renaming from: {old_filename} to: {new_filename}")
            print(f"Old file path: {old_file_path}")
            
            # Validate that old file exists
            if not os.path.exists(old_file_path):
                return JsonResponse({
                    'error': f'Original file not found: {old_filename}'
                }, status=404)
            
            # Ensure the new filename has proper extension
            old_ext = os.path.splitext(old_filename)[1]
            new_ext = os.path.splitext(new_filename)[1]
            
            if not new_ext:
                new_filename += old_ext
            elif new_ext.lower() != old_ext.lower():
                return JsonResponse({
                    'error': f'Cannot change file extension from {old_ext} to {new_ext}'
                }, status=400)
            
            # Generate new file path
            file_dir = os.path.dirname(old_file_path)
            new_file_path = os.path.join(file_dir, new_filename)
            
            # Check if new filename already exists (and it's not the same file)
            if os.path.exists(new_file_path) and new_file_path != old_file_path:
                return JsonResponse({
                    'error': f'Filename already exists: {new_filename}'
                }, status=400)
            
            # Rename the file in storage
            try:
                os.rename(old_file_path, new_file_path)
                print(f"File renamed successfully on disk: {new_file_path}")
            except OSError as e:
                return JsonResponse({
                    'error': f'File system error: {str(e)}'
                }, status=500)
            
            # CRITICAL FIX: Update the file field to point to the new path
            # Get the relative path from the media root
            from django.conf import settings
            media_root = settings.MEDIA_ROOT
            relative_new_path = os.path.relpath(new_file_path, media_root)
            
            # Update both filename AND file field
            image.filename = new_filename
            image.file.name = relative_new_path  # This updates the FileField path!
            
            # Re-extract values from new filename
            image.extract_values_from_filename()
            
            # Save the model (this updates both fields in database)
            image.save()
            
            print(f"Database updated - filename: {image.filename}, file field: {image.file.name}")
            
            return JsonResponse({
                'success': True, 
                'message': 'Image renamed successfully',
                'new_filename': new_filename,
                'file_url': image.file.url  # Return the updated URL
            })
            
        except ReadingImage.DoesNotExist:
            return JsonResponse({'error': 'Image not found'}, status=404)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)
        except Exception as e:
            return JsonResponse({'error': f'Unexpected error: {str(e)}'}, status=500)
    
    return JsonResponse({'error': 'Invalid request method'}, status=400)

@csrf_exempt
def delete_reading(request, image_id):
    """Delete a reading image"""
    if request.method == 'DELETE':
        try:
            image = ReadingImage.objects.get(id=image_id)
            image.delete()
            return JsonResponse({'success': True, 'message': 'Image deleted successfully'})
        except ReadingImage.DoesNotExist:
            return JsonResponse({'error': 'Image not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Invalid request method'}, status=400)

from django.db.models import Count, Avg, Case, When, IntegerField, F, Q
from django.utils import timezone
import json
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from .models import Client, ChecklistItem

@login_required
def statistics(request):
    # Calculate basic metrics
    clients = Client.objects.all()
    total_claims = clients.count()
    avg_completion = clients.aggregate(avg=Avg('completion_percent'))['avg'] or 0
    
    # Get oldest claim age in days
    oldest_claim = clients.order_by('dateOfLoss').first()
    oldest_claim_age = 0
    if oldest_claim and oldest_claim.dateOfLoss:
        oldest_claim_age = (timezone.now().date() - oldest_claim.dateOfLoss).days
    
    # Claim type counts
    mit_count = clients.filter(mitigation=True).count()
    cps_count = clients.filter(CPSCLNCONCGN=True).count()
    ppr_count = clients.filter(replacement=True).count()
    
    # Calculate trends (month-over-month comparison)
    last_month = timezone.now() - dt.timedelta(days=30)
    claim_growth = calculate_percentage_change(
        clients.filter(created_at__lt=last_month).count(),
        total_claims
    )
    completion_trend = calculate_percentage_change(
        clients.filter(created_at__lt=last_month).aggregate(avg=Avg('completion_percent'))['avg'] or 0,
        avg_completion
    )
    age_trend = calculate_trend_days(oldest_claim_age, last_month)
    mit_trend = calculate_percentage_change(
        clients.filter(mitigation=True, created_at__lt=last_month).count(),
        mit_count
    )
    cps_trend = calculate_percentage_change(
        clients.filter(CPSCLNCONCGN=True, created_at__lt=last_month).count(),
        cps_count
    )
    ppr_trend = calculate_percentage_change(
        clients.filter(replacement=True, created_at__lt=last_month).count(),
        ppr_count
    )
    
    # Enhanced Age Distribution
    now = timezone.now().date()
    age_categories = {
        "0-30 days": Q(dateOfLoss__gte=now - dt.timedelta(days=30)),
        "31-60 days": Q(dateOfLoss__lt=now - dt.timedelta(days=30)) & 
                     Q(dateOfLoss__gte=now - dt.timedelta(days=60)),
        "61-120 days": Q(dateOfLoss__lt=now - dt.timedelta(days=60)) & 
                      Q(dateOfLoss__gte=now - dt.timedelta(days=120)),
        "121-180 days": Q(dateOfLoss__lt=now - dt.timedelta(days=120)) & 
                       Q(dateOfLoss__gte=now - dt.timedelta(days=180)),
        "181-360 days": Q(dateOfLoss__lt=now - dt.timedelta(days=180)) & 
                        Q(dateOfLoss__gte=now - dt.timedelta(days=360)),
        "360+ days": Q(dateOfLoss__lt=now - dt.timedelta(days=360))
    }
    
    # Age distribution by claim type
    age_distribution = {
        'all': {category: clients.filter(query).count() for category, query in age_categories.items()},
        'MIT': {category: clients.filter(query, mitigation=True).count() for category, query in age_categories.items()},
        'CPS': {category: clients.filter(query, CPSCLNCONCGN=True).count() for category, query in age_categories.items()},
        'PPR': {category: clients.filter(query, replacement=True).count() for category, query in age_categories.items()}
    }
    
    # Age completion data
    age_completion_data = {
        category: clients.filter(query).aggregate(
            avg=Avg('completion_percent')
        )['avg'] or 0
        for category, query in age_categories.items()
    }
    
    # Document completion rates by type
    document_stats = {
        'all': list(ChecklistItem.objects.values('document_type').annotate(
            total=Count('id'),
            completed=Count(Case(When(is_completed=True, then=1), output_field=IntegerField()))
        ).annotate(
            completion_rate=100.0 * F('completed') / F('total')
        ).order_by('-completion_rate')[:10]),
        'MIT': list(ChecklistItem.objects.filter(document_category='MIT').values('document_type').annotate(
            total=Count('id'),
            completed=Count(Case(When(is_completed=True, then=1), output_field=IntegerField()))
        ).annotate(
            completion_rate=100.0 * F('completed') / F('total')
        ).order_by('-completion_rate')[:10]),
        'CPS': list(ChecklistItem.objects.filter(document_category='CPS').values('document_type').annotate(
            total=Count('id'),
            completed=Count(Case(When(is_completed=True, then=1), output_field=IntegerField()))
        ).annotate(
            completion_rate=100.0 * F('completed') / F('total')
        ).order_by('-completion_rate')[:10]),
        'PPR': list(ChecklistItem.objects.filter(document_category='PPR').values('document_type').annotate(
            total=Count('id'),
            completed=Count(Case(When(is_completed=True, then=1), output_field=IntegerField()))
        ).annotate(
            completion_rate=100.0 * F('completed') / F('total')
        ).order_by('-completion_rate')[:10])
    }
    
    # Client completion data for bar chart
    client_completion_data = [
        {'client_name': c.pOwner, 'completion_percent': c.completion_percent}
        for c in clients.order_by('-completion_percent')[:10]  # Top 10 clients
    ]
    
    # Recent activity
    recent_activity = [
        {
            'user_initials': request.user.get_initials() if hasattr(request.user, 'get_initials') else 'SY',
            'message': f"Updated claim for {c.pOwner}",
            'timestamp': c.updated_at,
            'type': 'Update'
        }
        for c in clients.order_by('-updated_at')[:5]
    ]
    
    recent_uploads = UploadActivity.objects.all().order_by('-uploaded_at')[:10]
    
    context.update({
        'recent_uploads': recent_uploads,
        'upload_stats': {
            'total_uploads': UploadActivity.objects.count(),
            'successful_uploads': UploadActivity.objects.filter(status='SUCCESS').count(),
            'today_uploads': UploadActivity.objects.filter(
                uploaded_at__date=timezone.now().date()
            ).count(),
        }
    })
        # Add room statistics
    total_rooms = Room.objects.count()
    rooms_per_client = Client.objects.annotate(room_count=models.Count('rooms'))
    avg_rooms_per_client = rooms_per_client.aggregate(avg=models.Avg('room_count'))['avg'] or 0
    
    # Work type distribution
    work_type_stats = {}
    for wt in WorkType.objects.all():
        wt_count = RoomWorkTypeValue.objects.filter(work_type=wt).exclude(value_type='NA').count()
        work_type_stats[wt.work_type_id] = wt_count
    
    context.update({
        'total_rooms': total_rooms,
        'avg_rooms_per_client': round(avg_rooms_per_client, 1),
        'work_type_stats': work_type_stats,
    })
    
    context = {
        'total_claims': total_claims,
        'avg_completion': round(avg_completion, 1),
        'oldest_claim_age': oldest_claim_age,
        'mit_count': mit_count,
        'cps_count': cps_count,
        'ppr_count': ppr_count,
        
        'claim_growth': round(claim_growth, 1),
        'completion_trend': round(completion_trend, 1),
        'age_trend': age_trend,
        'mit_trend': round(mit_trend, 1),
        'cps_trend': round(cps_trend, 1),
        'ppr_trend': round(ppr_trend, 1),
        
        'age_distribution': json.dumps(age_distribution),
        'age_completion_data': json.dumps(age_completion_data),
        'document_stats': json.dumps(document_stats),
        'client_completion_data': json.dumps(client_completion_data),
        
        'recent_activity': recent_activity,
    }
    
    return render(request, 'account/statistics.html', context)

def calculate_percentage_change(old_value, new_value):
    if old_value == 0:
        return 0
    return ((new_value - old_value) / old_value) * 100

def calculate_trend_days(current_days, comparison_date):
    oldest_last_month = Client.objects.filter(created_at__lt=comparison_date) \
                                     .order_by('dateOfLoss').first()
    if not oldest_last_month or not oldest_last_month.dateOfLoss:
        return 0
    
    last_month_days = (comparison_date.date() - oldest_last_month.dateOfLoss).days
    return current_days - last_month_days

import os
import re
import base64
import requests
from urllib.parse import quote
from dotenv import load_dotenv
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_GET
from django.template import Template, Context
from collections import defaultdict
import logging

load_dotenv()
logger = logging.getLogger(__name__)

CLIENT_ID     = os.getenv("GRAPH_CLIENT_ID")
REDIRECT_URI  = os.getenv("GRAPH_REDIRECT_URI", "https://login.microsoftonline.com/common/oauth2/nativeclient")
SCOPE         = os.getenv("GRAPH_SCOPE", "offline_access Files.ReadWrite.All")
TENANT        = os.getenv("GRAPH_TENANT", "consumers")
REFRESH_TOKEN = os.getenv("GRAPH_REFRESH_TOKEN")
SHARED_ROOT_LINK = os.getenv("SHARED_ROOT_LINK")
TOKEN_URL = f"https://login.microsoftonline.com/{TENANT}/oauth2/v2.0/token"
GRAPH = "https://graph.microsoft.com/v1.0"

IMG_EXT = {".jpg", ".jpeg", ".png", ".gif", ".bmp", ".webp", ".tif", ".tiff"}

# --- Token Management ---
def _access_token_from_refresh():
    if not REFRESH_TOKEN:
        raise RuntimeError("GRAPH_REFRESH_TOKEN is missing.")
    data = {
        "client_id": CLIENT_ID,
        "grant_type": "refresh_token",
        "refresh_token": REFRESH_TOKEN,
        "scope": SCOPE,
        "redirect_uri": REDIRECT_URI
    }
    r = requests.post(TOKEN_URL, data=data, timeout=30)
    if r.status_code >= 400:
        raise RuntimeError(f"Token refresh failed: {r.status_code} {r.text}")
    return r.json()["access_token"]

def _share_id_from_url(shared_url: str) -> str:
    b = shared_url.encode("utf-8")
    s = base64.urlsafe_b64encode(b).decode("ascii").rstrip("=")
    return "u!" + s

# --- OneDrive Navigation ---
def _encode_path_segments(relative_path: str) -> str:
    segments = relative_path.split('/')
    encoded = ':/' + '/'.join(f"'{quote(seg, safe='')}'" for seg in segments)
    return encoded

def _get_shared_root_item(token: str, share_id: str):
    headers = {"Authorization": f"Bearer {token}"}
    url = f"{GRAPH}/shares/{share_id}/driveItem"
    r = requests.get(url, headers=headers, timeout=30)
    r.raise_for_status()
    return r.json()

def _list_children_by_path(token: str, drive_id: str, item_id: str, relative_path: str = None):
    headers = {"Authorization": f"Bearer {token}"}
    if relative_path:
        encoded_path = _encode_path_segments(relative_path)
        url = f"{GRAPH}/drives/{drive_id}/items/{item_id}{encoded_path}:/children"
    else:
        url = f"{GRAPH}/drives/{drive_id}/items/{item_id}/children"
    url += "?$select=id,name,folder,file,@microsoft.graph.downloadUrl"
    r = requests.get(url, headers=headers, timeout=30)
    if r.status_code == 404:
        return []
    r.raise_for_status()
    return r.json().get("value", [])

# --- Helpers ---
def _is_image(name: str) -> bool:
    lower = name.lower()
    return any(lower.endswith(ext) for ext in IMG_EXT)

def _day_label(name: str):
    if "rht" not in name.lower():
        return None
    m = re.search(r"day\s*([1-4])", name, flags=re.IGNORECASE)
    return int(m.group(1)) if m else None

def _num_prefix(filename: str):
    m = re.match(r"^\s*(\d+)", filename)
    return int(m.group(1)) if m else None

def _find_estimates_folder(token: str, drive_id: str, item_id: str):
    try:
        children = _list_children_by_path(token, drive_id, item_id)
        logger.info(f"Found {len(children)} children in folder")
        for child in children:
            name = child.get('name', '')
            if not child.get('folder'):
                continue
            name_lower = name.lower()
            if '01-' in name_lower and ('insurance' in name_lower or 'estimates' in name_lower or 'current-ins' in name_lower):
                logger.info(f"✓ Found estimates folder: {name}")
                return child
        for child in children:
            if not child.get('folder'):
                continue
            name_lower = child.get('name', '').lower()
            if name_lower in ['documents', 'claims', 'projects']:
                logger.info(f"Recursing into {child.get('name')}...")
                result = _find_estimates_folder(token, drive_id, child.get('id'))
                if result:
                    return result
        return None
    except Exception as e:
        logger.error(f"Error searching for estimates folder: {str(e)}")
        return None

# --- MAIN ENDPOINTS ---
@require_GET
def list_claim_images_shared(request):

    logger.info(f"SHARED_ROOT_LINK in env: {os.getenv('SHARED_ROOT_LINK')}")
    logger.info(f"All env vars: {dict(os.environ)}")

    client_id = (request.GET.get("clientId") or "").strip()
    if not SHARED_ROOT_LINK:
        return JsonResponse({"error": "SHARED_ROOT_LINK not set in env."}, status=500)
    try:
        token = _access_token_from_refresh()
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)
    share_id = _share_id_from_url(SHARED_ROOT_LINK)
    try:
        root_item = _get_shared_root_item(token, share_id)
        drive_id = root_item["parentReference"]["driveId"]
        root_item_id = root_item["id"]
    except Exception as e:
        return JsonResponse({"error": f"Failed to access shared root: {str(e)}"}, status=500)
    try:
        root_children = _list_children_by_path(token, drive_id, root_item_id)
    except Exception as e:
        return JsonResponse({"error": f"Failed to list shared root: {str(e)}"}, status=500)
    mit_item = None
    for it in root_children:
        if it.get("folder") and it.get("name", "").strip().lower() == "mit":
            mit_item = it
            break
    if not mit_item:
        return JsonResponse({"images": []})
    try:
        mit_children = _list_children_by_path(token, drive_id, mit_item["id"])
    except Exception as e:
        return JsonResponse({"error": f"Failed to list MIT: {str(e)}"}, status=500)
    candidates = []
    for it in mit_children:
        name = it.get("name", "")
        if not it.get("folder"):
            continue
        day = _day_label(name)
        if day is None:
            continue
        if client_id:
            if client_id.lower() in name.lower():
                candidates.append((day, it))
        else:
            candidates.append((day, it))
    if client_id and not candidates:
        for it in mit_children:
            name = it.get("name", "")
            if not it.get("folder"):
                continue
            day = _day_label(name)
            if day is not None:
                candidates.append((day, it))
    images = []
    seen = set()
    for day, rht_item in sorted(candidates, key=lambda x: x[0]):
        try:
            files = _list_children_by_path(token, drive_id, rht_item["id"])
        except Exception:
            continue
        for f in files:
            if not f.get("file"):
                continue
            nm = f.get("name", "")
            if not _is_image(nm):
                continue
            dl = f.get("@microsoft.graph.downloadUrl")
            if not dl:
                continue
            key = (nm, dl)
            if key in seen:
                continue
            seen.add(key)
            images.append({"name": nm, "downloadUrl": dl})
    def sort_key(item):
        num = _num_prefix(item["name"])
        return (num if num is not None else 10**9, item["name"].lower())
    images.sort(key=sort_key)
    return JsonResponse({"images": images})


# --- ENCIRCLE SYNC ---
@csrf_exempt
def generate_room_entries_from_configs(request):
    """
    Generate properly formatted room entries for each template
    
    Format: {number} …. {room_name} {description} {los_value}
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        # Get the JSON data from the form
        room_data_json = request.POST.get('room_data_json')
        selected_templates = request.POST.getlist('selected_templates')
        selected_work_types = request.POST.getlist('selected_work_types')  # Optional - only used with 'basic' template

        # CRITICAL: Reorder templates so job_types is ALWAYS processed last (appears first in Encircle)
        if 'job_types' in selected_templates:
            selected_templates = [t for t in selected_templates if t != 'job_types']
            selected_templates.append('job_types')

        if not room_data_json:
            return JsonResponse({'error': 'No room data provided'}, status=400)

        room_data = json.loads(room_data_json)
        rooms = room_data.get('rooms', [])
        configs = room_data.get('configs', {})

        if not rooms:
            return JsonResponse({'error': 'No rooms provided'}, status=400)

        # Check if at least one template is selected
        if not selected_templates:
            return JsonResponse({'error': 'No templates selected'}, status=400)

        # Work type static descriptions - NEW FORMAT (100-700)
        work_type_descs = {
            100: "= … JOB/ROOMS OVERVIEW PICS ..",
            200: "….. SOURCE of LOSS PICS …..",
            300: "….. C.P.S. …...",
            400: "….. PPR …..",
            500: "…… DMO = DEMOLITION …....",
            600: "… WTR MITIGATION EQUIPMENT & W.I.P . ...",
            700: "… HMR = HAZARDOUS MATERIALS ..."
        }

        # Section labels - NEW FORMAT (100-700)
        section_labels = {
            100: "100 .... = ... JOB/ROOMS OVERVIEW PICS .. ==========================",
            200: "200 .... ..... SOURCE of LOSS PICS ..... ===========================",
            300: "300 .... ..... C.P.S. ...... =======================================",
            400: "400 .... PPR ===================================================",
            500: "500 .... ...... DMO = DEMOLITION ....... ===========================",
            600: "600 . WTR MITIGATION EQUIPMENT & W.I.P. ============================",
            700: "700 . HMR = HAZARDOUS MATERIALS ====================================",
        }

        room_entries_by_template = {}

        # ===== TEMPLATE: Basic (100-700s) with optional work type deselection =====
        if 'basic' in selected_templates:
            # If user selected specific work types, use those. Otherwise use all.
            if selected_work_types:
                work_types_to_include = sorted([int(wt) for wt in selected_work_types])
            else:
                work_types_to_include = [100, 200, 300, 400, 500, 600, 700]

            basic_entries = []

            # Add the job types header entries at the top (0.0001-0.0004 and 1997-1999)
            basic_entries.extend([
                "0.0001 ….. JOBSITE VERIFICATION",
                "0.0002 . MECHANICALS = WATER METER READING & PLUMBING REPORT/INVOICE",
                "0.0003 . MECHANICALS = ELECTRICAL HAZARDS",
                "0.0004 . EXT DAMAGE IF APPLICABLE ROOF TARPS",
                "1997 . LEAD & HMR TESTING LAB RESULTS",
                "1998 . KITCHEN CABINETS SIZES U & L =LF/ CT = SF; APPLIANCES",
                "1999 . BATHROOM FIXTURES CAB SIZE & FIXTURES & TYPE",
            ])

            for work_type in work_types_to_include:
                # ADD SECTION LABEL FIRST
                basic_entries.append(section_labels[work_type])

                for idx, room_name in enumerate(rooms):
                    room_number = work_type + idx + 1

                    # ALWAYS USE 100s CONFIG FOR ALL WORK TYPES
                    room_config = configs.get(room_name, {})
                    config_value = None

                    # Get from 100s config
                    if 100 in room_config:
                        config_value = room_config[100]
                    elif '100' in room_config:
                        config_value = room_config['100']
                    else:
                        config_value = '.'

                    # ORIGINAL DISPLAY LOGIC
                    display_value = "…........." if config_value == "." else config_value

                    # ORIGINAL FORMAT: {number} …. {room_name} {description} {los_value}
                    entry = f"{room_number} {display_value} …. {room_name} {work_type_descs[work_type]}"
                    basic_entries.append(entry)

                # Add job type specific entries after 300s and 400s
                if work_type == 300:
                    basic_entries.extend([
                        "3222 . CPS DAY2 WIP OVERVIEW WIP BOXES PACKOUT PICS",
                        "3322 . CPS3 DAY3 STORAGE OVERVIEW STORAGE MOVE OUT PICS",
                        "3444 . CPS4 DAY4 PACKBACK OVERVIEW PACK-BACK / RESET PICS",
                    ])
                elif work_type == 400:
                    basic_entries.extend([
                        "4111.1 . REPLACEMENT 1 CON OVERVIEW DAY PICS",
                        "4222.2 . REPLACEMENT 2 CON WIP",
                        "4333.3 . REPLACEMENT 3 CON STORAGE",
                        "4444.4 . REPLACEMENT 4 CON DISPOSAL",
                    ])

            # Add the rebuild entries at the end (9998-9999)
            basic_entries.extend([
                "9998.0 . REBUILD OVERVIEW WORK IN PROGRESS.......WIP",
                "9999.0 . REBUILD INTERIOR COMPLETED WORK",
            ])

            room_entries_by_template['basic'] = basic_entries

        # ===== TEMPLATE: extended (400s only) =====
        if 'extended' in selected_templates:
            extended_entries = []

            # ADD SECTION LABEL FIRST
            extended_entries.append("400 .... NON SALVAGEABLE ITEMS =====================================")

            for idx, room_name in enumerate(rooms):
                room_number = 400 + idx + 1

                # ALWAYS USE 100s CONFIG
                room_config = configs.get(room_name, {})
                config_value = None

                if 100 in room_config:
                    config_value = room_config[100]
                elif '100' in room_config:
                    config_value = room_config['100']
                else:
                    config_value = '.'

                display_value = "…........." if config_value == "." else config_value

                entry = f"{room_number} …. {room_name} {work_type_descs[400]} {display_value}"
                extended_entries.append(entry)

            room_entries_by_template['extended'] = extended_entries

        # ===== TEMPLATE: readings (6000-7000s) =====
        if 'readings' in selected_templates:
            readings_entries = generate_8000_9000_entries(rooms, configs)
            room_entries_by_template['readings'] = readings_entries
        
        # ===== TEMPLATE: readings default (70000s) =====
        if 'readings default' in selected_templates:
            readings_default_entries = generate_70000_entries(rooms, configs)
            room_entries_by_template['readings default'] = readings_default_entries

        # ===== TEMPLATE: Job Types (0.0000-9999.0) - CREATED LAST = APPEARS FIRST IN ENCIRCLE =====
        if 'job_types' in selected_templates:
            job_types_entries = generate_job_types_entries()
            room_entries_by_template['job_types'] = job_types_entries

        # Sort selected_templates by priority (lower priority number = created first = appears last in Encircle)
        template_priority = {
            'readings': 1,        # 8000-9000 MC Day Readings - created first
            'extended': 2,        # 400 PPR List
            'basic': 3,           # 100-700 Base List - created last = appears first
            'readings default': 5,
            'job_types': 0,       # Job Default - created very first
        }
        selected_templates = sorted(selected_templates, key=lambda x: template_priority.get(x, 99))

        # Initialize automation
        automation = RoomTemplateAutomation(headless=True)

        # Run automation with the generated room entries
        results = automation.run_automation_with_room_data(
            room_entries=room_entries_by_template,
            selected_template_ids=selected_templates,
            delete_existing=True
        )
        
        # Return the automation results
        return JsonResponse({
            'overall_status': results.get('overall_status', 'unknown'),
            'templates_successful': results.get('templates_successful', 0),
            'templates_failed': results.get('templates_failed', 0),
            'templates_processed': results.get('templates_processed', []),
            'login_status': results.get('login_status'),
            'navigation_status': results.get('navigation_status'),
            'deletion_results': results.get('deletion_results'),
            'processed_rooms': len(rooms),
            'selected_templates': selected_templates,
            'room_entries_generated': {template: len(entries) for template, entries in room_entries_by_template.items()}
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON data'}, status=400)
    except ImportError as e:
        return JsonResponse({'error': f'Automation module not available: {str(e)}'}, status=500)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def generate_8000_9000_entries(rooms, configs):
    """
    Generate 6000-7000s Readings List entries
    ALWAYS USES 100s CONFIG FOR LOS VALUES
    """
    entries = []
    
    # Section labels for readings - 8000-9000 FORMAT
    readings_section_labels = {
        8100: "8100.0 . ..... DAY 1 MC READINGS ..... ======================",
        8200: "8200.0 . ..... DAY 2 MC READINGS ..... ======================",
        8300: "8300.0 . ..... DAY 3 MC READINGS ..... ======================",
        8400: "8400.0 . ..... DAY 4 MC READINGS ..... ======================",
    }
    
    # 6000s sections (use user room data with 100s LOS values)
    for work_type in [8100, 8200, 8300, 8400]:
        # ADD SECTION LABEL FIRST
        entries.append(readings_section_labels[work_type])
        
        for idx, room_name in enumerate(rooms):
            room_number = work_type + idx + 1
            
            # ALWAYS USE 100s CONFIG FOR LOS VALUES
            room_config = configs.get(room_name, {})
            config_value = None
            
            if 100 in room_config:
                config_value = room_config[100]
            elif '100' in room_config:
                config_value = room_config['100']
            else:
                config_value = '.'
            
            display_value = "…........." if config_value == "." else config_value
            
            # Format for 6000s - ORIGINAL FORMAT
            if work_type == 8100:
                desc = "   ...  DAY1    MC READINGS .. "
            elif work_type == 8200:
                desc = "  ...  DAY2    MC READINGS .."
            elif work_type == 8300:
                desc = "  ...  DAY3    MC READINGS .."
            else:  # 6400
                desc = "  ...  DAY4    MC READINGS"
            
            # ORIGINAL FORMAT: {number} …. {room_name} {description} {los_value}
            entry = f"{room_number} {display_value} .{str(work_type)[3]} …. {room_name} {desc}"
            entries.append(entry)
    
    # 7000s section label - ORIGINAL FORMAT
    entries.append("9000 RH &T & GPP  DRY CHAMBERS [DC] . READINGS ==================")
    
    # 7000s section (static entries - dry chambers) - ORIGINAL FORMAT
    static_7000s_entries = [
        "9100.0 RH &T & GPP  DRY CHAMBERS [DC] . READINGS  =========== ….. DAY 1 ….. ",
        "9100.0 …. EXTERIOR & UNAFFECTED AREA  ….. DAY 1 ….. ",
        "9101.0 …. DRY CHAMBER # 1 ….. DAY 1 …..  RH &T & GPP ",
        "9102.0 …. DRY CHAMBER # 2 ….. DAY 1 …..  RH &T & GPP ",
        "9103.0 …. DRY CHAMBER # 3 ….. DAY 1 …..  RH &T & GPP ",
        "9104.0 …. DRY CHAMBER # 4 ….. DAY 1 …..  RH &T & GPP ",
        "9200.0 RH &T & GPP  DRY CHAMBERS [DC] . READINGS  =========== ….. DAY 2 ….. ",
        "9200.2 …. EXTERIOR & UNAFFECTED AREA ….. DAY 2 ….. ",
        "9201.2 …. DRY CHAMBER # 1 ….. DAY 2 …..  RH &T & GPP ",
        "9202.2 …. DRY CHAMBER # 2 ….. DAY 2 …..  RH &T & GPP ",
        "9203.2 …. DRY CHAMBER # 3 ….. DAY 2 …..  RH &T & GPP ",
        "9204.2 …. DRY CHAMBER # 4 ….. DAY 2 …..  RH &T & GPP ",
        "9205.2 …. DRY CHAMBER # 5 ….. DAY 2 …..  RH &T & GPP ",
        "9300.0 RH &T & GPP  DRY CHAMBERS [DC] . READINGS  =========== ….. DAY 3 ….. ",
        "9300.0 …. EXTERIOR & UNAFFECTED AREA ….. DAY 3 ….. ",
        "9301.0 …. DRY CHAMBER # 1 ….. DAY 3 …..  RH &T & GPP ",
        "9302.0 …. DRY CHAMBER # 2 ….. DAY 3 …..  RH &T & GPP ",
        "9303.0 …. DRY CHAMBER # 3 ….. DAY 3 …..  RH &T & GPP ",
        "9304.0 …. DRY CHAMBER # 4 ….. DAY 3 …..  RH &T & GPP ",
        "9400.0 RH &T & GPP  DRY CHAMBERS [DC] . READINGS  =========== ….. DAY 4 ….. ",
        "9400.0 …. EXTERIOR & UNAFFECTED AREA ….. DAY 4 ….. ",
        "9401.0 …. DRY CHAMBER # 1 ….. DAY 4 …..  RH &T & GPP ",
        "9402.0 …. DRY CHAMBER # 2 ….. DAY 4 …..  RH &T & GPP ",
        "9403.0 …. DRY CHAMBER # 3 ….. DAY 4 …..  RH &T & GPP ",
        "9404.0 …. DRY CHAMBER # 4 ….. DAY 4 …..  RH &T & GPP "
    ]
    
    entries.extend(static_7000s_entries)
    return entries


def generate_70000_entries(rooms, configs):
    """
    Generate 70000s Stabilization Readings entries
    ALWAYS USES 100s CONFIG FOR LOS VALUES
    """
    entries = []
    
    # Add section label
    entries.append("70000 ….. ======= DAY # 0  …..  MC READINGS STABILIZATION ===============")
    
    for idx, room_name in enumerate(rooms):
        room_number = 70101 + idx
        
        # ALWAYS USE 100s CONFIG FOR LOS VALUES
        room_config = configs.get(room_name, {})
        config_value = None
        
        if 100 in room_config:
            config_value = room_config[100]
        elif '100' in room_config:
            config_value = room_config['100']
        else:
            config_value = '.'
        
        display_value = "….........." if config_value == "." else config_value
        
        # ORIGINAL FORMAT: {number} …. {room_name} DAY 0 MOISTURE READINGS {los_value}
        entry = f"{room_number} …. {room_name} … DAY # 0  … MC READINGS STABILIZATION … {display_value}"
        entries.append(entry)
    
    return entries


def generate_job_types_entries():
    """
    Generate static job types template entries (0.0000-9999.0)
    This template has HIGHEST PRIORITY and doesn't use user room data
    Format uses decimal numbers with simple space-separated descriptions for Encircle compatibility
    """
    entries = [
        "0.0001 ….. JOBSITE VERIFICATION",
        "0.0002 . MECHANICALS = WATER METER READING & PLUMBING REPORT/INVOICE",
        "0.0003 . MECHANICALS = ELECTRICAL HAZARDS",
        "0.0004 . EXT DAMAGE IF APPLICABLE ROOF TARPS",
        "1997 . LEAD & HMR TESTING LAB RESULTS",
        "1998 . KITCHEN CABINETS SIZES U & L =LF/ CT = SF; APPLIANCES",
        "1999 . BATHROOM FIXTURES CAB SIZE & FIXTURES & TYPE",
        "3222 . CPS DAY2 WIP OVERVIEW WIP BOXES PACKOUT PICS",
        "3322 . CPS3 DAY3 STORAGE OVERVIEW STORAGE MOVE OUT PICS",
        "3444 . CPS4 DAY4 PACKBACK OVERVIEW PACK-BACK / RESET PICS",
        "4111.1 . REPLACEMENT 1 CON OVERVIEW DAY PICS",
        "4222.2 . REPLACEMENT 2 CON WIP",
        "4333.3 . REPLACEMENT 3 CON STORAGE",
        "4444.4 . REPLACEMENT 4 CON DISPOSAL",
        "9998.0 . REBUILD OVERVIEW WORK IN PROGRESS.......WIP",
        "9999.0 . REBUILD INTERIOR COMPLETED WORK",
    ]
    return entries


@csrf_exempt
def get_all_clients(request):
    """
    Get all clients for dropdown selection
    """
    try:
        clients = Client.objects.all().order_by('-created_at')[:100]  # Limit to 100 most recent
        clients_data = []

        for client in clients:
            clients_data.append({
                'id': client.id,
                'pOwner': client.pOwner or 'Unknown',
                'pAddress': client.pAddress or '',
                'claimNumber': client.claimNumber or ''
            })

        return JsonResponse({'clients': clients_data})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
@csrf_exempt
def send_room_list_email(request):
    """
    Send room list email with dynamic claim data and room list
    Includes PDF attachment and HTML email body with user instructions
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    try:
        data = json.loads(request.body)
        recipients = data.get('recipients', [])
        room_data = data.get('room_data', {})
        claim_name = data.get('claim_name', 'CLAIM')
        claim_address = data.get('claim_address', '')

        if not recipients:
            return JsonResponse({'error': 'No recipients provided'}, status=400)

        if not room_data.get('rooms'):
            return JsonResponse({'error': 'No room data provided'}, status=400)

        # Get email version (default to 'table' for backward compatibility)
        email_version = data.get('email_version', 'table')

        # Generate email HTML with user instructions
        html_content = generate_room_list_email_html(claim_name, claim_address, room_data, version=email_version)

        # Generate PDF attachment with the same format
        pdf_buffer = generate_room_list_pdf(claim_name, claim_address, room_data, format_type=email_version)

        # Send email
        from django.core.mail import EmailMessage
        from django.conf import settings

        email = EmailMessage(
            subject=f'[ROOM LIST] {claim_name} — Worktype Documentation',
            body=html_content,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=recipients,
        )
        email.content_subtype = 'html'

        # Attach PDF
        pdf_filename = f"{claim_name.replace(' ', '_')}_Room_List_{email_version}.pdf"
        email.attach(pdf_filename, pdf_buffer.getvalue(), 'application/pdf')

        email.send() 

        return JsonResponse({
            'success': True,
            'recipients_count': len(recipients),
            'message': f'Email sent successfully to {len(recipients)} recipient(s) with PDF attachment',
            'format': email_version
        })

    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    
    
def generate_room_list_pdf(claim_name, claim_address, room_data):
    """
    Generate a PDF of the room list table

    Args:
        claim_name: Name of the claim
        claim_address: Address of the claim
        room_data: Dictionary with 'rooms' and 'configs' keys

    Returns:
        BytesIO buffer containing the PDF
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from io import BytesIO

    rooms = room_data.get('rooms', [])
    configs = room_data.get('configs', {})

    # Create PDF buffer
    buffer = BytesIO()

    # Create the PDF document in landscape mode for better table viewing
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter),
                           rightMargin=0.5*inch, leftMargin=0.5*inch,
                           topMargin=0.5*inch, bottomMargin=0.5*inch)

    # Container for the 'Flowable' objects
    elements = []

    # Define styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1e88e5'),
        spaceAfter=12,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )

    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=12,
        textColor=colors.HexColor('#555555'),
        spaceAfter=20,
        alignment=TA_CENTER
    )

    # Add title
    title = Paragraph(f"{claim_name} — Worktype Documentation", title_style)
    elements.append(title)

    # Add address
    subtitle = Paragraph(f"@ {claim_address}", subtitle_style)
    elements.append(subtitle)
    elements.append(Spacer(1, 0.2*inch))

    # Build table data
    table_data = []

    # Header row
    header = ['Room', '100\nOverview', 'LOS/\nTravel', '200\nSource', 'LOS/\nTravel',
              '300\nCPS', 'LOS/\nTravel', '400\nPPR', 'LOS/\nTravel',
              '500\nDemo', 'LOS/\nTravel', '600\nWTR', 'LOS/\nTravel',
              '700\nHMR', 'LOS/\nTravel']
    table_data.append(header)

    # Room rows
    for idx, room in enumerate(rooms):
        base_num = idx + 1

        # Get LOS/Travel value
        room_config = configs.get(room, {})
        config_value = room_config.get('100', room_config.get(100, '.'))
        los_cell_value = '' if config_value == '.' else config_value

        row = [
            room,
            f'1{base_num:02d}', los_cell_value,
            f'2{base_num:02d}', los_cell_value,
            f'3{base_num:02d}', los_cell_value,
            f'4{base_num:02d}', los_cell_value,
            f'5{base_num:02d}', los_cell_value,
            f'6{base_num:02d}', los_cell_value,
            f'7{base_num:02d}', los_cell_value
        ]
        table_data.append(row)

    # Create table
    col_widths = [1.8*inch] + [0.5*inch] * 14  # Room name wider, rest narrower
    table = Table(table_data, colWidths=col_widths)

    # Style the table
    table_style = TableStyle([
        # Header row
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e3f2fd')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),

        # Data cells
        ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),

        # Room name column
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),

        # Alternating row colors
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),

        # Work type columns (100, 200, 300, 400, 500, 600, 700)
        ('BACKGROUND', (1, 1), (1, -1), colors.HexColor('#fff8c6')),  # 100
        ('BACKGROUND', (3, 1), (3, -1), colors.HexColor('#f0f4f8')),  # 200
        ('BACKGROUND', (5, 1), (5, -1), colors.HexColor('#fff8c6')),  # 300
        ('BACKGROUND', (7, 1), (7, -1), colors.HexColor('#f0f4f8')),  # 400
        ('BACKGROUND', (9, 1), (9, -1), colors.HexColor('#fff8c6')),  # 500
        ('BACKGROUND', (11, 1), (11, -1), colors.HexColor('#f0f4f8')), # 600
        ('BACKGROUND', (13, 1), (13, -1), colors.HexColor('#fff8c6')), # 700

        # LOS/Travel columns
        ('BACKGROUND', (2, 1), (2, -1), colors.HexColor('#ffe6e6')),
        ('BACKGROUND', (4, 1), (4, -1), colors.HexColor('#ffe6e6')),
        ('BACKGROUND', (6, 1), (6, -1), colors.HexColor('#ffe6e6')),
        ('BACKGROUND', (8, 1), (8, -1), colors.HexColor('#ffe6e6')),
        ('BACKGROUND', (10, 1), (10, -1), colors.HexColor('#ffe6e6')),
        ('BACKGROUND', (12, 1), (12, -1), colors.HexColor('#ffe6e6')),
        ('BACKGROUND', (14, 1), (14, -1), colors.HexColor('#ffe6e6')),
        ('FONTNAME', (2, 1), (14, -1), 'Helvetica-Bold'),

        # Grid
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#d0d0d0')),
    ])

    table.setStyle(table_style)
    elements.append(table)

    # Build PDF
    doc.build(elements)

    # Reset buffer position
    buffer.seek(0)

    return buffer


def generate_room_list_email_html(claim_name, claim_address, room_data, version='table'):
    """
    Generate HTML email content for room list with dynamic data and user instructions

    Args:
        claim_name: Name of the claim
        claim_address: Address of the claim
        room_data: Dictionary with 'rooms' and 'configs' keys
        version: 'table' for original table format, 'list' for sequential list format

    Returns:
        HTML string for email content
    """
    rooms = room_data.get('rooms', [])
    configs = room_data.get('configs', {})

    # DEBUG: Log the configs to see what we're receiving
    print(f"[EMAIL DEBUG] Configs received: {configs}")
    print(f"[EMAIL DEBUG] Email version: {version}")

    if version == 'list':
        # Generate the list format
        return _generate_list_format_email(claim_name, claim_address, rooms, configs)
    else:
        # Generate the table format (original)
        return _generate_table_format_email(claim_name, claim_address, rooms, configs)


def _generate_table_format_email(claim_name, claim_address, rooms, configs):
    """Generate the original table format email"""
    # Build the room table rows with all work types (100, 200, 300, 400, 500, 600, 700)
    room_rows_html = ''
    for idx, room in enumerate(rooms):
        base_num = idx + 1

        # Get LOS/Travel value (100s value applies to all work types)
        # Handle both string '100' and integer 100 keys from JavaScript
        room_config = configs.get(room, {})
        config_value = room_config.get('100', room_config.get(100, '.'))
        print(f"[EMAIL DEBUG] Room: {room}, Room Config: {room_config}, Config Value: {config_value}")
        # Display the actual value (LOS, TRVL, etc.) or leave empty if '.'
        los_cell_value = '' if config_value == '.' else config_value

        # Alternating row backgrounds
        row_style = '' if idx % 2 == 0 else ''

        room_rows_html += f'''
        <tr{row_style}>
          <td>{room}</td>
          <td style="background:#fff8c6;">1{base_num:02d}</td>
          <td style="background:#ffe6e6; font-weight:bold;">{los_cell_value}</td>
          <td style="background:#f0f4f8;">2{base_num:02d}</td>
          <td style="background:#ffe6e6; font-weight:bold;">{los_cell_value}</td>
          <td style="background:#fff8c6;">3{base_num:02d}</td>
          <td style="background:#ffe6e6; font-weight:bold;">{los_cell_value}</td>
          <td style="background:#f0f4f8;">4{base_num:02d}</td>
          <td style="background:#ffe6e6; font-weight:bold;">{los_cell_value}</td>
          <td style="background:#fff8c6;">5{base_num:02d}</td>
          <td style="background:#ffe6e6; font-weight:bold;">{los_cell_value}</td>
          <td style="background:#f0f4f8;">6{base_num:02d}</td>
          <td style="background:#ffe6e6; font-weight:bold;">{los_cell_value}</td>
          <td style="background:#fff8c6;">7{base_num:02d}</td>
          <td style="background:#ffe6e6; font-weight:bold;">{los_cell_value}</td>
        </tr>
        '''

    html_content = f"""
<div style="font-family: Arial, sans-serif; background:#f5f7fa; padding:30px;">

  <!-- ========================= -->
  <!-- HEADER / BRANDING BAR -->
  <!-- ========================= -->
  <div style="
      background: linear-gradient(90deg, #1e88e5, #42a5f5);
      color:white;
      padding:20px 25px;
      border-radius:8px;
      font-size:22px;
      font-weight:bold;
      margin-bottom:25px;
      box-shadow:0 4px 12px rgba(0,0,0,0.15);
  ">
    {claim_name} — Worktype Documentation
  </div>

  <!-- ========================= -->
  <!-- USER INSTRUCTIONS -->
  <!-- ========================= -->
  <div style="
      background:white;
      border-radius:10px;
      padding:25px;
      margin-bottom:25px;
      box-shadow:0 3px 12px rgba(0,0,0,0.12);
      border-left: 5px solid #28a745;
  ">
    <h2 style="margin-top:0; color:#28a745; font-size:20px;">📋 How to Use This Email</h2>

    <p style="font-size:15px; color:#333; line-height:1.6; margin-bottom:15px;">
      This email contains the room list for <strong>{claim_name}</strong>. You have two ways to view and use this information:
    </p>

    <div style="margin-bottom:20px;">
      <h3 style="color:#1e88e5; font-size:16px; margin-bottom:10px;">📧 Option 1: View in Email</h3>
      <p style="font-size:14px; color:#555; line-height:1.6; margin-left:20px;">
        Simply scroll down in this email to see the complete room list table below. Perfect for quick reference!
      </p>
    </div>

    <div style="margin-bottom:20px;">
      <h3 style="color:#1e88e5; font-size:16px; margin-bottom:10px;">📎 Option 2: Download & Print the PDF</h3>
      <p style="font-size:14px; color:#555; line-height:1.6; margin-left:20px; margin-bottom:15px;">
        A PDF file is attached to this email. Use it to print a physical copy or save it for your records.
      </p>

      <div style="background:#f8f9fa; padding:15px; border-radius:6px; margin-left:20px;">
        <p style="font-size:14px; color:#333; font-weight:600; margin:0 0 10px 0;">🖨️ Printing Instructions:</p>

        <div style="margin-bottom:15px;">
          <p style="font-size:13px; color:#555; margin:0 0 8px 0; font-weight:600;">📱 On Mobile Phone (iPhone/Android):</p>
          <ol style="font-size:13px; color:#555; line-height:1.8; margin:0; padding-left:25px;">
            <li>Tap the PDF attachment at the top or bottom of this email</li>
            <li>The PDF will open in your phone's viewer</li>
            <li>Tap the Share icon (box with arrow) or Menu (three dots)</li>
            <li>Select "Print" from the menu</li>
            <li>Choose your printer (make sure your phone and printer are on the same WiFi network)</li>
            <li>Tap "Print"</li>
          </ol>
          <p style="font-size:12px; color:#777; margin:8px 0 0 0; font-style:italic;">
            💡 Tip: If you don't see your printer, make sure it's turned on and connected to WiFi!
          </p>
        </div>

        <div>
          <p style="font-size:13px; color:#555; margin:0 0 8px 0; font-weight:600;">💻 On Computer (PC/Mac):</p>
          <ol style="font-size:13px; color:#555; line-height:1.8; margin:0; padding-left:25px;">
            <li>Click the PDF attachment to download it</li>
            <li>Open the downloaded PDF file (usually in your Downloads folder)</li>
            <li>Click "File" then "Print" (or press Ctrl+P on PC / Cmd+P on Mac)</li>
            <li>Select your printer from the list</li>
            <li>Choose "Landscape" orientation for best results</li>
            <li>Click "Print"</li>
          </ol>
          <p style="font-size:12px; color:#777; margin:8px 0 0 0; font-style:italic;">
            💡 Tip: Use Landscape orientation so the table fits better on the page!
          </p>
        </div>
      </div>
    </div>

    <div style="background:#fff3cd; padding:12px; border-radius:6px; border-left:4px solid #ffc107; margin-top:15px;">
      <p style="font-size:13px; color:#856404; margin:0; line-height:1.6;">
        <strong>Need Help?</strong> If you have trouble opening the PDF or printing, please reply to this email and we'll assist you!
      </p>
    </div>
  </div>


  <!-- ========================= -->
  <!-- CARD: REFERENCE INDEX -->
  <!-- ========================= -->
  <div style="
      background:white;
      border-radius:10px;
      padding:25px;
      margin-bottom:35px;
      box-shadow:0 3px 12px rgba(0,0,0,0.12);
  ">
    <h2 style="margin-top:0; color:#1e88e5;">Reference Index — Worktype Codes</h2>

    <table cellspacing="0" cellpadding="8" border="1"
      style="border-collapse: collapse; width:100%; font-size:14px; border-color:#d0d0d0;">

      <tr>
        <th style="background:#e3f2fd; font-weight:bold;">Code</th>
        <th style="background:#e3f2fd; font-weight:bold;">Description</th>
      </tr>

      <tr><td>0.0001</td><td>Jobsite Verification</td></tr>
      <tr><td>0.0002</td><td>Mechanicals – Water Meter Reading & Plumbing Report/Invoice</td></tr>
      <tr><td>0.0003</td><td>Mechanicals – Electrical Hazards</td></tr>
      <tr><td>0.0004</td><td>Exterior Damage If Applicable Roof Tarps</td></tr>
      <tr><td>1997</td><td>Lead & HMR Testing Lab Results</td></tr>
      <tr><td>1998</td><td>Kitchen Cabinets Sizes U & L =LF/ CT = SF; Appliances</td></tr>
      <tr><td>1999</td><td>Bathroom Fixtures Cab Size & Fixtures & Type</td></tr>

      <tr><td>100</td><td>Rooms Overview</td></tr>
      <tr><td>200</td><td>Source of Loss</td></tr>
      <tr><td>300</td><td>CPS</td></tr>
      <tr><td>3222</td><td>CPS DAY2 WIP overview WIP boxes packout pics</td></tr>
      <tr><td>3322</td><td>CPS3 DAY3 storage overview storage MOVE OUT pics</td></tr>
      <tr><td>3444</td><td>CPS4 DAY4 packback overview pack-back / reset pics</td></tr>

      <tr><td>400</td><td>PPR</td></tr>
      <tr><td>4111.1</td><td>Replacement 1 CON overview day pics</td></tr>
      <tr><td>4222.2</td><td>Replacement 2 CON WIP</td></tr>
      <tr><td>4333.3</td><td>Replacement 3 CON storage</td></tr>
      <tr><td>4444.4</td><td>Replacement 4 CON disposal</td></tr>

      <tr><td>500</td><td>DMO Demo</td></tr>
      <tr><td>600</td><td>WTR Mitigation Equipment & W.I.P</td></tr>
      <tr><td>700</td><td>HMR</td></tr>

      <tr><td>9998.0</td><td>Rebuild overview work in progress.......WIP</td></tr>
      <tr><td>9999.0</td><td>Rebuild interior completed work</td></tr>
    </table>
  </div>



  <!-- ========================= -->
  <!-- CARD: ROOM LIST -->
  <!-- ========================= -->
  <div style="
      background:white;
      border-radius:10px;
      padding:25px;
      margin-bottom:35px;
      box-shadow:0 3px 12px rgba(0,0,0,0.12);
  ">
    <h2 style="color:#1e88e5; margin-top:0;">{claim_name} Worktype Room List</h2>
    <h3 style="color:#555; font-weight:normal; margin-top:5px;">
      @ {claim_address}
    </h3>

    <!-- MOBILE SCROLL HINT -->
    <div class="mobile-scroll-hint" style="display:none; background:#fff3cd; padding:10px; border-radius:6px; margin-bottom:10px; text-align:center; border:1px solid #ffc107;">
      <span style="color:#856404; font-size:13px;">📱 Scroll horizontally to see all columns →</span>
    </div>
    <style>
      @media only screen and (max-width: 768px) {{
        .mobile-scroll-hint {{ display:block !important; }}
      }}
    </style>

    <!-- MOBILE SAFE SCROLL WRAPPER -->
    <div style="width:100%; overflow-x:auto; -webkit-overflow-scrolling:touch;">

      <table cellspacing="0" cellpadding="8" border="1"
        style="border-collapse: collapse; width:100%; min-width:650px; table-layout:auto; font-size:14px; border-color:#d0d0d0;">

        <tr style="font-weight:bold;">
          <th style="background:#e3f2fd;">Room</th>
          <th style="background:#fff8c6;">100<br>Overview</th>
          <th style="background:#ffe6e6;">LOS/<br>Travel</th>
          <th style="background:#f0f4f8;">200<br>Source</th>
          <th style="background:#ffe6e6;">LOS/<br>Travel</th>
          <th style="background:#fff8c6;">300<br>CPS</th>
          <th style="background:#ffe6e6;">LOS/<br>Travel</th>
          <th style="background:#f0f4f8;">400<br>PPR</th>
          <th style="background:#ffe6e6;">LOS/<br>Travel</th>
          <th style="background:#fff8c6;">500<br>Demo</th>
          <th style="background:#ffe6e6;">LOS/<br>Travel</th>
          <th style="background:#f0f4f8;">600<br>WTR Equip</th>
          <th style="background:#ffe6e6;">LOS/<br>Travel</th>
          <th style="background:#fff8c6;">700<br>HMR</th>
          <th style="background:#ffe6e6;">LOS/<br>Travel</th>
        </tr>

        <!-- ROWS -->
        {room_rows_html}

      </table>
    </div>
  </div>



  <!-- ========================= -->
  <!-- FOOTER -->
  <!-- ========================= -->
  <div style="
      text-align:center;
      padding:15px;
      color:#777;
      font-size:12px;
      margin-top:20px;
  ">
    {claim_name} report | Powered by Claimet Email System
  </div>

</div>
    """

    return html_content


def _generate_list_format_email(claim_name, claim_address, rooms, configs):
    """Generate the sequential list format email"""

    # Define work type configurations with their descriptions
    work_types = [
        (100, "JOB/ROOMS OVERVIEW PICS", "...", "=========================="),
        (200, "SOURCE of LOSS PICS", ".....", "==========================="),
        (300, "C.P.S.", ".....", "======================================="),
        (400, "PPR", "", "============================================="),
        (500, "DMO = DEMOLITION", "......", "==========================="),
        (600, "WTR MITIGATION EQUIPMENT & W.I.P", "", "================================"),
        (700, "HMR = HAZARDOUS MATERIALS", "", "===================================="),
    ]

    # Build the list items
    list_items_html = ''

    # Add default codes at the beginning
    default_codes = [
        ("0.0001", "JOBSITE VERIFICATION", "....."),
        ("0.0002", "MECHANICALS = WATER METER READING & PLUMBING REPORT/INVOICE", "."),
        ("0.0003", "MECHANICALS = ELECTRICAL HAZARDS", "."),
        ("0.0004", "EXT DAMAGE IF APPLICABLE ROOF TARPS", "."),
        ("1997", "LEAD & HMR TESTING LAB RESULTS", "."),
        ("1998", "KITCHEN CABINETS SIZES U & L =LF/ CT = SF; APPLIANCES", "."),
        ("1999", "BATHROOM FIXTURES CAB SIZE & FIXTURES & TYPE", "."),
    ]

    for code, description, dots in default_codes:
        list_items_html += f'''
        <div style="padding:8px 0; border-bottom:1px solid #e0e0e0; font-family:monospace; font-size:14px;">
          <span style="display:inline-block; width:80px; font-weight:bold; color:#1e88e5;">{code}</span>
          <span style="color:#555;">{dots} {description}</span>
        </div>
        '''

    # Process each work type
    for work_type_num, work_type_desc, dots, separator in work_types:
        # Add the header for this work type
        list_items_html += f'''
        <div style="padding:10px 0; border-bottom:2px solid #1e88e5; font-family:monospace; font-size:14px; background:#e3f2fd; margin-top:10px;">
          <span style="display:inline-block; width:80px; font-weight:bold; color:#1e88e5;">{work_type_num}</span>
          <span style="font-weight:bold; color:#1e88e5;">{dots} = {work_type_desc} {separator}</span>
        </div>
        '''

        # Add each room for this work type
        for idx, room in enumerate(rooms):
            base_num = idx + 1
            room_code = f"{work_type_num // 100}{base_num:02d}"

            # Get LOS/Travel value for this room
            room_config = configs.get(room, {})
            config_value = room_config.get(str(work_type_num), room_config.get(work_type_num, '.'))

            # Display the config value (TRVL, LOS, etc.) or dots if empty
            display_value = config_value if config_value and config_value != '.' else '............'

            list_items_html += f'''
            <div style="padding:8px 0; border-bottom:1px solid #e0e0e0; font-family:monospace; font-size:14px;">
              <span style="display:inline-block; width:80px; font-weight:bold; color:#1e88e5;">{room_code}</span>
              <span style="display:inline-block; width:150px; color:#333;">{room}</span>
              <span style="color:#555;">{dots} {work_type_desc} {dots}</span>
              <span style="font-weight:bold; color:#d32f2f; margin-left:10px;">{display_value}</span>
            </div>
            '''

        # Add special codes after certain work types
        if work_type_num == 300:
            special_codes_300 = [
                ("3222", "CPS DAY2 WIP OVERVIEW WIP BOXES PACKOUT PICS", "."),
                ("3333", "CPS3 DAY3 STORAGE OVERVIEW STORAGE MOVE OUT PICS", "."),
                ("3444", "CPS4 DAY4 PACKBACK OVERVIEW PACK-BACK / RESET PICS", "."),
            ]
            for code, description, dots in special_codes_300:
                list_items_html += f'''
                <div style="padding:8px 0; border-bottom:1px solid #e0e0e0; font-family:monospace; font-size:14px;">
                  <span style="display:inline-block; width:80px; font-weight:bold; color:#1e88e5;">{code}</span>
                  <span style="color:#555;">{dots} {description}</span>
                </div>
                '''

        if work_type_num == 400:
            special_codes_400 = [
                ("4111.1", "REPLACEMENT 1 CON OVERVIEW DAY PICS", "."),
                ("4222.2", "REPLACEMENT 2 CON WIP", "."),
                ("4333.3", "REPLACEMENT 3 CON STORAGE", "."),
                ("4444.4", "REPLACEMENT 4 CON DISPOSAL", "."),
            ]
            for code, description, dots in special_codes_400:
                list_items_html += f'''
                <div style="padding:8px 0; border-bottom:1px solid #e0e0e0; font-family:monospace; font-size:14px;">
                  <span style="display:inline-block; width:80px; font-weight:bold; color:#1e88e5;">{code}</span>
                  <span style="color:#555;">{dots} {description}</span>
                </div>
                '''

    # Add rebuild codes at the end
    rebuild_codes = [
        ("9998.0", "REBUILD OVERVIEW WORK IN PROGRESS.......", "WIP"),
        ("9999.0", "REBUILD INTERIOR COMPLETED WORK", ""),
    ]

    for code, description, suffix in rebuild_codes:
        list_items_html += f'''
        <div style="padding:8px 0; border-bottom:1px solid #e0e0e0; font-family:monospace; font-size:14px;">
          <span style="display:inline-block; width:80px; font-weight:bold; color:#1e88e5;">{code}</span>
          <span style="color:#555;">. {description} {suffix}</span>
        </div>
        '''

    html_content = f"""
<div style="font-family: Arial, sans-serif; background:#f5f7fa; padding:30px;">

  <!-- ========================= -->
  <!-- HEADER / BRANDING BAR -->
  <!-- ========================= -->
  <div style="
      background: linear-gradient(90deg, #1e88e5, #42a5f5);
      color:white;
      padding:20px 25px;
      border-radius:8px;
      font-size:22px;
      font-weight:bold;
      margin-bottom:25px;
      box-shadow:0 4px 12px rgba(0,0,0,0.15);
  ">
    {claim_name} — Worktype Documentation
  </div>

  <!-- ========================= -->
  <!-- USER INSTRUCTIONS -->
  <!-- ========================= -->
  <div style="
      background:white;
      border-radius:10px;
      padding:25px;
      margin-bottom:25px;
      box-shadow:0 3px 12px rgba(0,0,0,0.12);
      border-left: 5px solid #28a745;
  ">
    <h2 style="margin-top:0; color:#28a745; font-size:20px;">📋 How to Use This Email</h2>

    <p style="font-size:15px; color:#333; line-height:1.6; margin-bottom:15px;">
      This email contains the room list for <strong>{claim_name}</strong>. You have two ways to view and use this information:
    </p>

    <div style="margin-bottom:20px;">
      <h3 style="color:#1e88e5; font-size:16px; margin-bottom:10px;">📧 Option 1: View in Email</h3>
      <p style="font-size:14px; color:#555; line-height:1.6; margin-left:20px;">
        Simply scroll down in this email to see the complete room list below. Perfect for quick reference!
      </p>
    </div>

    <div style="margin-bottom:20px;">
      <h3 style="color:#1e88e5; font-size:16px; margin-bottom:10px;">📎 Option 2: Download & Print the PDF</h3>
      <p style="font-size:14px; color:#555; line-height:1.6; margin-left:20px; margin-bottom:15px;">
        A PDF file is attached to this email. Use it to print a physical copy or save it for your records.
      </p>

      <div style="background:#f8f9fa; padding:15px; border-radius:6px; margin-left:20px;">
        <p style="font-size:14px; color:#333; font-weight:600; margin:0 0 10px 0;">🖨️ Printing Instructions:</p>

        <div style="margin-bottom:15px;">
          <p style="font-size:13px; color:#555; margin:0 0 8px 0; font-weight:600;">📱 On Mobile Phone (iPhone/Android):</p>
          <ol style="font-size:13px; color:#555; line-height:1.8; margin:0; padding-left:25px;">
            <li>Tap the PDF attachment at the top or bottom of this email</li>
            <li>The PDF will open in your phone's viewer</li>
            <li>Tap the Share icon (box with arrow) or Menu (three dots)</li>
            <li>Select "Print" from the menu</li>
            <li>Choose your printer (make sure your phone and printer are on the same WiFi network)</li>
            <li>Tap "Print"</li>
          </ol>
          <p style="font-size:12px; color:#777; margin:8px 0 0 0; font-style:italic;">
            💡 Tip: If you don't see your printer, make sure it's turned on and connected to WiFi!
          </p>
        </div>

        <div>
          <p style="font-size:13px; color:#555; margin:0 0 8px 0; font-weight:600;">💻 On Computer (PC/Mac):</p>
          <ol style="font-size:13px; color:#555; line-height:1.8; margin:0; padding-left:25px;">
            <li>Click the PDF attachment to download it</li>
            <li>Open the downloaded PDF file (usually in your Downloads folder)</li>
            <li>Click "File" then "Print" (or press Ctrl+P on PC / Cmd+P on Mac)</li>
            <li>Select your printer from the list</li>
            <li>Choose "Landscape" orientation for best results</li>
            <li>Click "Print"</li>
          </ol>
          <p style="font-size:12px; color:#777; margin:8px 0 0 0; font-style:italic;">
            💡 Tip: Use Landscape orientation so the table fits better on the page!
          </p>
        </div>
      </div>
    </div>

    <div style="background:#fff3cd; padding:12px; border-radius:6px; border-left:4px solid #ffc107; margin-top:15px;">
      <p style="font-size:13px; color:#856404; margin:0; line-height:1.6;">
        <strong>Need Help?</strong> If you have trouble opening the PDF or printing, please reply to this email and we'll assist you!
      </p>
    </div>
  </div>

  <!-- ========================= -->
  <!-- CARD: ROOM LIST -->
  <!-- ========================= -->
  <div style="
      background:white;
      border-radius:10px;
      padding:25px;
      margin-bottom:35px;
      box-shadow:0 3px 12px rgba(0,0,0,0.12);
  ">
    <h2 style="color:#1e88e5; margin-top:0;">{claim_name} Worktype Room List</h2>
    <h3 style="color:#555; font-weight:normal; margin-top:5px;">
      @ {claim_address}
    </h3>

    <!-- LIST OF WORK TYPES AND ROOMS -->
    <div style="margin-top:20px;">
      {list_items_html}
    </div>
  </div>

  <!-- ========================= -->
  <!-- FOOTER -->
  <!-- ========================= -->
  <div style="
      text-align:center;
      padding:15px;
      color:#777;
      font-size:12px;
      margin-top:20px;
  ">
    {claim_name} report | Powered by Claimet Email System
  </div>

</div>
    """

    return html_content


def generate_room_list_pdf(claim_name, claim_address, room_data, format_type='list'):
    """
    Generate a PDF of the room list in either table or list format

    Args:
        claim_name: Name of the claim
        claim_address: Address of the claim
        room_data: Dictionary with 'rooms' and 'configs' keys
        format_type: 'table' or 'list'

    Returns:
        BytesIO buffer containing the PDF
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from io import BytesIO

    rooms = room_data.get('rooms', [])
    configs = room_data.get('configs', {})

    # Create PDF buffer
    buffer = BytesIO()

    # Use portrait mode for list format, landscape for table format
    if format_type == 'list':
        # Use portrait with narrow margins for compact list
        doc = SimpleDocTemplate(buffer, pagesize=letter,
                               rightMargin=0.3*inch, leftMargin=0.3*inch,
                               topMargin=0.3*inch, bottomMargin=0.3*inch)
    else:
        # Use landscape for table format
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter),
                               rightMargin=0.3*inch, leftMargin=0.3*inch,
                               topMargin=0.3*inch, bottomMargin=0.3*inch)

    # Container for the 'Flowable' objects
    elements = []

    # Define styles
    styles = getSampleStyleSheet()
    
    if format_type == 'list':
        return _generate_list_pdf(claim_name, claim_address, rooms, configs, doc, styles, elements, buffer)
    else:
        return _generate_table_pdf(claim_name, claim_address, rooms, configs, doc, styles, elements, buffer)


def _generate_list_pdf(claim_name, claim_address, rooms, configs, doc, styles, elements, buffer):
    """Generate compact list format PDF"""
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.platypus import Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.units import inch

    # Title style - smaller and compact
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=12,
        textColor=colors.HexColor('#1e88e5'),
        spaceAfter=6,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )

    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#555555'),
        spaceAfter=12,
        alignment=TA_CENTER
    )

    # Add title
    title = Paragraph(f"{claim_name}", title_style)
    elements.append(title)

    # Add address
    subtitle = Paragraph(f"{claim_address}", subtitle_style)
    elements.append(subtitle)
    elements.append(Spacer(1, 0.1*inch))

    # Define work types and their configurations
    work_types = [
        (100, "JOB/ROOMS OVERVIEW PICS"),
        (200, "SOURCE of LOSS PICS"),
        (300, "C.P.S."),
        (400, "PPR"),
        (500, "DMO = DEMOLITION"),
        (600, "WTR MITIGATION EQUIPMENT & W.I.P"),
        (700, "HMR = HAZARDOUS MATERIALS"),
    ]

    # Start building the compact list
    all_data = []
    
    # Add default codes at top - very compact
    default_codes = [
        ("0.0001", "JOBSITE VERIFICATION"),
        ("0.0002", "MECHANICALS = WATER METER READING & PLUMBING REPORT/INVOICE"),
        ("0.0003", "MECHANICALS = ELECTRICAL HAZARDS"),
        ("0.0004", "EXT DAMAGE IF APPLICABLE ROOF TARPS"),
        ("1997", "LEAD & HMR TESTING LAB RESULTS"),
        ("1998", "KITCHEN CABINETS SIZES U & L =LF/ CT = SF; APPLIANCES"),
        ("1999", "BATHROOM FIXTURES CAB SIZE & FIXTURES & TYPE"),
    ]
    
    for code, desc in default_codes:
        all_data.append([code, desc])
    
    # Add separator
    all_data.append([""] * 2)
    
    # Add each work type and its rooms
    for work_type_num, work_type_desc in work_types:
        # Add work type header
        header_bg = colors.HexColor('#e3f2fd')  # Light blue background
        all_data.append([f"  {work_type_num}", f"= {work_type_desc}"])
        
        # Add each room for this work type
        for idx, room in enumerate(rooms):
            base_num = idx + 1
            room_code = f"{work_type_num // 100}{base_num:02d}"
            
            # Get LOS/Travel value
            room_config = configs.get(room, {})
            config_value = room_config.get(str(work_type_num), room_config.get(work_type_num, ''))
            display_value = config_value if config_value else ''
            
            # Compact row: [Code, Room, LOS Value]
            room_info = f"{room}  [{display_value}]" if display_value else room
            all_data.append([f"    {room_code}", room_info])
        
        # Add special codes after certain work types
        if work_type_num == 300:
            special_codes = [
                ("3222", "CPS DAY2 WIP OVERVIEW WIP BOXES PACKOUT PICS"),
                ("3333", "CPS3 DAY3 STORAGE OVERVIEW STORAGE MOVE OUT PICS"),
                ("3444", "CPS4 DAY4 PACKBACK OVERVIEW PACK-BACK / RESET PICS"),
            ]
            for code, desc in special_codes:
                all_data.append([f"    {code}", desc])
        
        if work_type_num == 400:
            special_codes = [
                ("4111.1", "REPLACEMENT 1 CON OVERVIEW DAY PICS"),
                ("4222.2", "REPLACEMENT 2 CON WIP"),
                ("4333.3", "REPLACEMENT 3 CON STORAGE"),
                ("4444.4", "REPLACEMENT 4 CON DISPOSAL"),
            ]
            for code, desc in special_codes:
                all_data.append([f"    {code}", desc])
    
    # Add rebuild codes at the end
    all_data.append([""] * 2)
    rebuild_codes = [
        ("9998.0", "REBUILD OVERVIEW WORK IN PROGRESS.......WIP"),
        ("9999.0", "REBUILD INTERIOR COMPLETED WORK"),
    ]
    for code, desc in rebuild_codes:
        all_data.append([code, desc])
    
    # Create compact table with minimal styling
    col_widths = [1.2*inch, 4.5*inch]  # Narrow columns
    
    # Create table
    table = Table(all_data, colWidths=col_widths, hAlign='LEFT')
    
    # Ultra compact styling
    table_style = TableStyle([
        # Default cell styling
        ('FONTNAME', (0, 0), (-1, -1), 'Courier'),  # Monospace font
        ('FONTSIZE', (0, 0), (-1, -1), 7),  # Very small font
        ('TOPPADDING', (0, 0), (-1, -1), 1),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        
        # Default code column styling
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Courier-Bold'),
        
        # Default description column styling
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        
        # Work type headers
        ('FONTNAME', (0, len(default_codes)+1), (1, len(default_codes)+1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, len(default_codes)+1), (1, len(default_codes)+1), 8),
        ('BACKGROUND', (0, len(default_codes)+1), (1, len(default_codes)+1), colors.HexColor('#e3f2fd')),
        
        # Special codes styling
        ('FONTNAME', (0, 0), (0, len(default_codes)), 'Courier-Bold'),
        ('TEXTCOLOR', (0, 0), (0, len(default_codes)), colors.HexColor('#1e88e5')),
        
        # Remove grid lines for cleaner look
        # ('GRID', (0, 0), (-1, -1), 0.25, colors.lightgrey),
    ])
    
    # Apply row-specific styling for work type sections
    current_row = len(default_codes) + 2  # Start after default codes and separator
    
    for work_type_num, _ in work_types:
        # Style the work type header
        table_style.add('BACKGROUND', (0, current_row), (1, current_row), colors.HexColor('#f0f8ff'))
        table_style.add('FONTNAME', (0, current_row), (1, current_row), 'Helvetica-Bold')
        table_style.add('FONTSIZE', (0, current_row), (1, current_row), 8)
        current_row += 1
        
        # Style the rooms for this work type
        for idx, room in enumerate(rooms):
            # Alternate row colors for readability
            if idx % 2 == 0:
                table_style.add('BACKGROUND', (0, current_row), (1, current_row), colors.HexColor('#f9f9f9'))
            
            # Make LOS value stand out
            room_config = configs.get(room, {})
            config_value = room_config.get(str(work_type_num), room_config.get(work_type_num, ''))
            if config_value:
                table_style.add('FONTNAME', (1, current_row), (1, current_row), 'Helvetica-Bold')
                table_style.add('TEXTCOLOR', (1, current_row), (1, current_row), colors.HexColor('#d32f2f'))
            
            current_row += 1
        
        # Skip special codes if they exist
        if work_type_num == 300 or work_type_num == 400:
            special_count = 3 if work_type_num == 300 else 4
            for i in range(special_count):
                table_style.add('FONTNAME', (0, current_row), (0, current_row), 'Courier-Bold')
                table_style.add('TEXTCOLOR', (0, current_row), (0, current_row), colors.HexColor('#1e88e5'))
                current_row += 1
    
    table.setStyle(table_style)
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    
    # Reset buffer position
    buffer.seek(0)
    return buffer


def _generate_table_pdf(claim_name, claim_address, rooms, configs, doc, styles, elements, buffer):
    """Generate table format PDF (existing functionality)"""
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.platypus import Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.units import inch
    # Keep your existing table PDF generation code but make it more compact
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=10,  # Smaller
        textColor=colors.HexColor('#1e88e5'),
        spaceAfter=8,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )

    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=8,  # Smaller
        textColor=colors.HexColor('#555555'),
        spaceAfter=10,
        alignment=TA_CENTER
    )

    # Add title
    title = Paragraph(claim_name, title_style)
    elements.append(title)

    # Add address
    subtitle = Paragraph(claim_address, subtitle_style)
    elements.append(subtitle)
    elements.append(Spacer(1, 0.1*inch))

    # Build compact table data
    table_data = []
    
    # Header row - single line
    header = ['Room', '100', 'L/T', '200', 'L/T', '300', 'L/T', '400', 'L/T',
              '500', 'L/T', '600', 'L/T', '700', 'L/T']
    table_data.append(header)
    
    # Room rows - ultra compact
    for idx, room in enumerate(rooms):
        base_num = idx + 1
        room_config = configs.get(room, {})
        config_value = room_config.get('100', room_config.get(100, ''))
        los_value = config_value if config_value else ''
        
        # Truncate room name if too long
        display_room = room[:15] + '...' if len(room) > 18 else room
        
        row = [
            display_room,
            f'1{base_num:02d}', los_value,
            f'2{base_num:02d}', los_value,
            f'3{base_num:02d}', los_value,
            f'4{base_num:02d}', los_value,
            f'5{base_num:02d}', los_value,
            f'6{base_num:02d}', los_value,
            f'7{base_num:02d}', los_value
        ]
        table_data.append(row)
    
    # Create table with very narrow columns
    col_widths = [1.0*inch] + [0.35*inch] * 14  # Even narrower
    
    table = Table(table_data, colWidths=col_widths)
    
    # Ultra compact table style
    table_style = TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e3f2fd')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 7),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 4),
        
        # Data
        ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 6),  # Very small
        ('TOPPADDING', (0, 1), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 2),
        
        # Room column
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        
        # Alternating rows
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
        
        # Work type backgrounds
        ('BACKGROUND', (1, 1), (1, -1), colors.HexColor('#fff8c6')),
        ('BACKGROUND', (3, 1), (3, -1), colors.HexColor('#f0f4f8')),
        ('BACKGROUND', (5, 1), (5, -1), colors.HexColor('#fff8c6')),
        ('BACKGROUND', (7, 1), (7, -1), colors.HexColor('#f0f4f8')),
        ('BACKGROUND', (9, 1), (9, -1), colors.HexColor('#fff8c6')),
        ('BACKGROUND', (11, 1), (11, -1), colors.HexColor('#f0f4f8')),
        ('BACKGROUND', (13, 1), (13, -1), colors.HexColor('#fff8c6')),
        
        # LOS columns
        ('BACKGROUND', (2, 1), (2, -1), colors.HexColor('#ffe6e6')),
        ('BACKGROUND', (4, 1), (4, -1), colors.HexColor('#ffe6e6')),
        ('BACKGROUND', (6, 1), (6, -1), colors.HexColor('#ffe6e6')),
        ('BACKGROUND', (8, 1), (8, -1), colors.HexColor('#ffe6e6')),
        ('BACKGROUND', (10, 1), (10, -1), colors.HexColor('#ffe6e6')),
        ('BACKGROUND', (12, 1), (12, -1), colors.HexColor('#ffe6e6')),
        ('BACKGROUND', (14, 1), (14, -1), colors.HexColor('#ffe6e6')),
        
        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d0d0d0')),
    ])
    
    table.setStyle(table_style)
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    
    # Reset buffer position
    buffer.seek(0)
    return buffer


from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json
from django.shortcuts import render

import re
from collections import defaultdict
from difflib import SequenceMatcher

# ---------------------------
# Simple Normalization & Token Extraction
# ---------------------------

def normalize_text(text):
    """Basic normalization: uppercase, strip extra spaces"""
    if not text:
        return ""
    return ' '.join(text.upper().split())

def extract_tokens(text):
    """
    Extract meaningful tokens from text.
    Returns a set of normalized words (3+ chars) excluding common noise.
    """
    if not text:
        return set()
    
    # Split by @ to separate main part from address
    parts = text.split('@')
    main_part = parts[0] if parts else text
    address_part = parts[1] if len(parts) > 1 else ""
    
    # Extract alphanumeric tokens
    tokens = re.findall(r'\b[A-Z0-9]{2,}\b', main_part.upper())
    
    # Add address tokens if present
    if address_part:
        address_tokens = re.findall(r'\b[A-Z0-9]{2,}\b', address_part.upper())
        tokens.extend(address_tokens)
    
    # Filter out noise words
    noise_words = {'LLC', 'INC', 'THE', 'AND', 'FOR', 'CLAIM', 'EST', 'FIRE', 'WATER', 'STORM'}
    tokens = [t for t in tokens if t not in noise_words and len(t) >= 2]
    
    return set(tokens)

def extract_location_code(text):
    """
    Extract location codes like GA22, OH24, GA22A, OH24-900
    """
    if not text:
        return None
    
    # Pattern: 2-3 letters + 2-3 digits + optional letters/numbers
    match = re.search(r'\b([A-Z]{2,3}\d{2,3}[A-Z0-9\-]*)\b', text.upper())
    return match.group(1) if match else None

# ---------------------------
# Simple Fuzzy Matching
# ---------------------------

def calculate_match_score(encircle_contractor, folder_name):
    """
    Simple fuzzy matching between contractor ID and folder name.
    Returns score 0-1.
    Rule: If 2+ tokens match, it's automatically a strong match (0.8+)
    """
    if not encircle_contractor or not folder_name:
        return 0.0
    
    # Normalize both
    contractor_norm = normalize_text(encircle_contractor)
    folder_norm = normalize_text(folder_name)
    
    # Extract tokens
    contractor_tokens = extract_tokens(contractor_norm)
    folder_tokens = extract_tokens(folder_norm)
    
    # Count matching tokens
    common_tokens = contractor_tokens.intersection(folder_tokens)
    num_matches = len(common_tokens)
    
    # RULE: 2+ matching tokens = automatic match
    if num_matches >= 2:
        return 0.85  # High confidence match
    
    # If only 1 token matches, use detailed scoring
    score = 0.0
    
    # 1. Check if contractor ID is substring of folder (40 points)
    if contractor_norm in folder_norm or folder_norm in contractor_norm:
        score += 0.4
    
    # 2. Location code match (30 points)
    contractor_location = extract_location_code(contractor_norm)
    folder_location = extract_location_code(folder_norm)
    
    if contractor_location and folder_location:
        # Exact match
        if contractor_location == folder_location:
            score += 0.3
        # Fuzzy match (e.g., GA22 vs GA22A)
        elif contractor_location[:4] == folder_location[:4]:
            score += 0.2
    
    # 3. Single token match (30 points)
    if num_matches == 1:
        score += 0.3
    
    return min(score, 1.0)

# ---------------------------
# Main Comparison Function
# ---------------------------

def compare_claims(encircle_claims, onedrive_claims):
    """
    Simple matching: compare contractor ID against folder name
    """
    # Filter out test data
    valid_encircle = [c for c in encircle_claims if _is_valid_claim(c)]
    valid_onedrive = [c for c in onedrive_claims if _is_valid_folder(c)]
    
    encircle_test_data = [c for c in encircle_claims if not _is_valid_claim(c)]
    onedrive_test_data = [c for c in onedrive_claims if not _is_valid_folder(c)]
    
    results = {
        'summary': {
            'total_encircle': len(encircle_claims),
            'total_onedrive': len(onedrive_claims),
            'matches': 0,
            'encircle_only': 0,
            'onedrive_only': 0,
            'encircle_test_data': len(encircle_test_data),
            'onedrive_test_data': len(onedrive_test_data),
            'match_breakdown': {
                'high_confidence': 0,
                'medium_confidence': 0,
                'low_confidence': 0,
            }
        },
        'matched_pairs': [],
        'encircle_missing_onedrive': [],
        'onedrive_extra': [],
        'encircle_test_data': encircle_test_data,
        'onedrive_test_data': onedrive_test_data,
        'duplicates': find_duplicates(valid_encircle, valid_onedrive)
    }
    
    matched_encircle = set()
    matched_onedrive = set()
    
    print("=" * 80)
    print("SIMPLIFIED CLAIM MATCHING")
    print("=" * 80)
    
    # Match each encircle claim to best onedrive folder
    for encircle_claim in valid_encircle:
        if encircle_claim['id'] in matched_encircle:
            continue
        
        contractor_id = encircle_claim.get('contractor_identifier', '').strip()
        
        if not contractor_id:
            continue
        
        best_match = None
        best_score = 0
        
        # Check against all onedrive folders
        for onedrive_claim in valid_onedrive:
            if onedrive_claim['folder_id'] in matched_onedrive:
                continue
            
            folder_name = onedrive_claim.get('folder_name', '').strip()
            
            # Calculate match score
            score = calculate_match_score(contractor_id, folder_name)
            
            # Print analysis
            if score >= 0.5:  # Only print potential matches
                print(f"\nPOTENTIAL MATCH:")
                print(f"  Encircle: {contractor_id}")
                print(f"  OneDrive: {folder_name}")
                print(f"  Score: {score:.2f} ({int(score*100)}%)")
            
            if score > best_score:
                best_score = score
                best_match = onedrive_claim
        
        # Accept matches above threshold
        MATCH_THRESHOLD = 0.65
        if best_match and best_score >= MATCH_THRESHOLD:
            confidence_level = "High" if best_score >= 0.8 else "Medium" if best_score >= 0.65 else "Low"
            
            results['matched_pairs'].append({
                'encircle': encircle_claim,
                'onedrive': best_match,
                'match_type': f'Fuzzy Match ({confidence_level})',
                'confidence': f'{int(best_score * 100)}%'
            })
            
            matched_encircle.add(encircle_claim['id'])
            matched_onedrive.add(best_match['folder_id'])
            results['summary']['matches'] += 1
            
            # Update breakdown
            if best_score >= 0.8:
                results['summary']['match_breakdown']['high_confidence'] += 1
            elif best_score >= 0.65:
                results['summary']['match_breakdown']['medium_confidence'] += 1
            else:
                results['summary']['match_breakdown']['low_confidence'] += 1
            
            print(f"  ✓ MATCHED! ({confidence_level} confidence)")
    
    # Collect unmatched claims
    for encircle_claim in valid_encircle:
        if encircle_claim['id'] not in matched_encircle:
            results['encircle_missing_onedrive'].append(encircle_claim)
            results['summary']['encircle_only'] += 1
    
    for onedrive_claim in valid_onedrive:
        if onedrive_claim['folder_id'] not in matched_onedrive:
            results['onedrive_extra'].append(onedrive_claim)
            results['summary']['onedrive_only'] += 1
    
    print("\n" + "=" * 80)
    print(f"MATCHING COMPLETE")
    print(f"Matches: {results['summary']['matches']}")
    print(f"Encircle Only: {results['summary']['encircle_only']}")
    print(f"OneDrive Only: {results['summary']['onedrive_only']}")
    print("=" * 80)
    
    return results

# ---------------------------
# Filter Functions (from original code)
# ---------------------------

_TEST_EXCLUDE_PATTERNS = [
    'HOW2', 'TEST', 'TEMPLATE', 'SAMPLE', 'ROOMLISTS', 'READINGS',
    'TMPL', 'CHECKLIST', 'TRAILER', 'WAREHOUSE', 'DEFAULT', 'TEMP',
    'PLACEHOLDER', 'EXAMPLE', 'DEMO', 'XXXX', 'AAA', '===', 'BACKEND', 'TUTORIAL'
]

def _is_valid_claim(claim):
    """Filter out test/placeholder claims from Encircle."""
    if not claim.get('policyholder_name') and not claim.get('contractor_identifier'):
        return False
    policyholder = (claim.get('policyholder_name') or '').upper()
    contractor = (claim.get('contractor_identifier') or '').upper()
    return not any(p in policyholder or p in contractor for p in _TEST_EXCLUDE_PATTERNS)

def _is_valid_folder(folder_claim):
    """Filter out test/placeholder folders from OneDrive."""
    folder_name = (folder_claim.get('folder_name') or '').upper()
    exclude_patterns = _TEST_EXCLUDE_PATTERNS + [
        'CLOSED CLAIMS', 'PROOF OF LOSS', 'DRAWINGS', 'APPRAISALS',
        'FOLDER', 'TEXT'
    ]
    if any(p in folder_name for p in exclude_patterns):
        return False
    clean_alpha = re.sub(r'[^A-Z]', '', folder_name)
    return len(clean_alpha) >= 3

import re
import logging
from collections import defaultdict
from difflib import SequenceMatcher
from django.http import HttpResponse
from django.views.decorators.http import require_GET
from django.template import Template, Context
from django.core.cache import cache

logger = logging.getLogger(__name__)

# Import your existing API clients and settings
# from .encircle_client import EncircleAPIClient
# from .encircle_processor import EncircleDataProcessor
# from .onedrive_client import _access_token_from_refresh, _share_id_from_url, _get_shared_root_item, _find_estimates_folder, _list_children_by_path
# from .settings import SHARED_ROOT_LINK

# Cache keys
ENCIRCLE_CACHE_KEY = 'encircle_claims_data'
ONEDRIVE_CACHE_KEY = 'onedrive_claims_data'
CACHE_TIMEOUT = 3600 * 24  # 24 hours

# =============================================================================
# MAIN VIEWS
# =============================================================================

@require_GET
def sync_encircle_onedrive(request):
    """Main sync endpoint - uses cached data"""
    try:
        # Use cached data
        encircle_claims = get_encircle_claims(use_cache=True)
        onedrive_claims = get_onedrive_claims(use_cache=True)
        comparison_results = compare_claims(encircle_claims, onedrive_claims)
        html_report = generate_comparison_report(comparison_results)
        return HttpResponse(html_report, content_type='text/html')
    except Exception as e:
        logger.error(f"Error syncing Encircle and OneDrive: {str(e)}")
        return HttpResponse(f"<h1>Error</h1><p>{str(e)}</p>", content_type='text/html', status=500)

@require_GET
def sync_encircle_onedrive_refresh(request):
    """Refresh endpoint - forces fresh download and clears cache"""
    try:
        logger.info("Forcing fresh data download - clearing cache")
        # Clear cache
        cache.delete(ENCIRCLE_CACHE_KEY)
        cache.delete(ONEDRIVE_CACHE_KEY)
        
        # Fetch fresh data
        encircle_claims = get_encircle_claims(use_cache=False)
        onedrive_claims = get_onedrive_claims(use_cache=False)
        comparison_results = compare_claims(encircle_claims, onedrive_claims)
        html_report = generate_comparison_report(comparison_results, is_refresh=True)
        return HttpResponse(html_report, content_type='text/html')
    except Exception as e:
        logger.error(f"Error syncing Encircle and OneDrive: {str(e)}")
        return HttpResponse(f"<h1>Error</h1><p>{str(e)}</p>", content_type='text/html', status=500)

# =============================================================================
# DATA FETCHING WITH CACHE
# =============================================================================

def get_encircle_claims(use_cache=True):
    """Fetch Encircle claims with optional caching"""
    if use_cache:
        cached_data = cache.get(ENCIRCLE_CACHE_KEY)
        if cached_data:
            logger.info(f"Using cached Encircle data ({len(cached_data)} claims)")
            return cached_data
    
    try:
        logger.info("Fetching fresh Encircle data...")
        api_client = EncircleAPIClient()
        processor = EncircleDataProcessor()
        raw_claims = api_client.get_all_claims()
        processed_claims = processor.process_claims_list(raw_claims)
        
        for claim in processed_claims:
            try:
                claim_details = api_client.get_claim_details(claim['id'])
                detailed_claim = processor.process_claim_details(claim_details)
                claim['contractor_identifier'] = str(detailed_claim.get('contractor_identifier', '') or '').strip()
                claim['policyholder_name'] = str(detailed_claim.get('policyholder_name', '') or '').strip()
                claim['full_address'] = str(detailed_claim.get('full_address', '') or '').strip()
                claim['insurance_company_name'] = str(detailed_claim.get('insurance_company_name', '') or '').strip()
            except Exception as e:
                logger.warning(f"Could not fetch details for claim {claim['id']}: {str(e)}")
                claim['contractor_identifier'] = ''
                claim['policyholder_name'] = ''
                claim['full_address'] = ''
                claim['insurance_company_name'] = ''
        
        logger.info(f"Found {len(processed_claims)} claims in Encircle")
        
        # Cache the data
        cache.set(ENCIRCLE_CACHE_KEY, processed_claims, CACHE_TIMEOUT)
        logger.info(f"Cached Encircle data for {CACHE_TIMEOUT} seconds")
        
        return processed_claims
    except Exception as e:
        logger.error(f"Error fetching Encircle claims: {str(e)}")
        raise Exception(f"Failed to fetch Encircle claims: {str(e)}")

def get_onedrive_claims(use_cache=True):
    """Fetch OneDrive claims with optional caching"""
    if use_cache:
        cached_data = cache.get(ONEDRIVE_CACHE_KEY)
        if cached_data:
            logger.info(f"Using cached OneDrive data ({len(cached_data)} folders)")
            return cached_data
    
    try:
        logger.info("Fetching fresh OneDrive data...")
        token = _access_token_from_refresh()
        share_id = _share_id_from_url(SHARED_ROOT_LINK)
        root_item = _get_shared_root_item(token, share_id)
        drive_id = root_item["parentReference"]["driveId"]
        root_item_id = root_item["id"]
        logger.info(f"Shared root - Item ID: {root_item_id}, Drive ID: {drive_id}, Name: {root_item.get('name')}")
        
        estimates_item = _find_estimates_folder(token, drive_id, root_item_id)
        if not estimates_item:
            logger.error("Could not find estimates folder")
            return []
        
        estimates_id = estimates_item.get('id')
        logger.info(f"Using estimates folder ID: {estimates_id}")
        
        folders = _list_children_by_path(token, drive_id, estimates_id)
        logger.info(f"Found {len(folders)} items in estimates folder")
        
        claims = []
        for folder in folders:
            if folder.get('folder'):
                folder_name = folder.get('name', '')
                folder_id = folder.get('id')
                if not folder_id:
                    continue
                
                try:
                    folder_contents = _list_children_by_path(token, drive_id, folder_id)
                    has_info_file = any(
                        item.get('name', '').lower().startswith('01-info') and 
                        item.get('name', '').lower().endswith(('.xlsx', '.xls'))
                        for item in folder_contents
                    )
                    contractor_id = extract_contractor_id_from_folder(folder_name)
                    claims.append({
                        'folder_name': folder_name,
                        'folder_id': folder_id,
                        'has_info_file': has_info_file,
                        'contractor_identifier': contractor_id if contractor_id else '',
                        'normalized_name': normalize_claim_name(folder_name),
                        'file_count': len(folder_contents)
                    })
                except Exception as e:
                    logger.warning(f"Could not check folder contents for {folder_name}: {str(e)}")
                    contractor_id = extract_contractor_id_from_folder(folder_name)
                    claims.append({
                        'folder_name': folder_name,
                        'folder_id': folder_id,
                        'has_info_file': False,
                        'contractor_identifier': contractor_id if contractor_id else '',
                        'normalized_name': normalize_claim_name(folder_name),
                        'error': str(e),
                        'file_count': 0
                    })
        
        logger.info(f"Found {len(claims)} claim folders in OneDrive")
        
        # Cache the data
        cache.set(ONEDRIVE_CACHE_KEY, claims, CACHE_TIMEOUT)
        logger.info(f"Cached OneDrive data for {CACHE_TIMEOUT} seconds")
        
        return claims
    except Exception as e:
        logger.error(f"Error fetching OneDrive claims: {str(e)}")
        raise Exception(f"Failed to fetch OneDrive claims: {str(e)}")

# =============================================================================
# SIMPLE NORMALIZATION & TOKEN EXTRACTION
# =============================================================================

def normalize_text(text):
    """Basic normalization: uppercase, strip extra spaces"""
    if not text:
        return ""
    return ' '.join(text.upper().split())

def normalize_year_code(code):
    """
    Normalize location codes to match only the first 4 characters
    GA22, GA22A, GA22P all become GA22
    OH24, OH24A, OH24APC all become OH24
    """
    if not code:
        return ""
    
    # Extract first 4 characters (2 letters + 2 digits)
    # This handles: GA22, OH24, GA22A, OH24A, etc.
    match = re.match(r'^([A-Z]{2}\d{2})', code.upper())
    if match:
        return match.group(1)
    
    return code.upper()

def normalize_name(name):
    """
    Normalize names to handle spelling variations
    CLEGGETT -> CLEGGET, THORNTON -> THORTON, etc.
    """
    if not name:
        return ""
    
    name = name.upper().strip()
    
    # Common spelling variations
    variations = {
        'CLEGGETT': 'CLEGGET',
        'CLEGGET': 'CLEGGET',
        'GOLLATTE': 'GOLLATE',
        'GOLLATE': 'GOLLATE',
        'THORNTON': 'THORTON',
        'THORTON': 'THORTON',
    }
    
    for original, normalized in variations.items():
        if original in name:
            name = name.replace(original, normalized)
    
    return name

def extract_tokens(text):
    """
    Extract meaningful tokens from text.
    Returns a set of normalized words (2+ chars) excluding common noise.
    Includes normalized year codes and name variations.
    """
    if not text:
        return set()
    
    text = normalize_name(text)  # Apply name normalization
    
    # Split by @ to separate main part from address
    parts = text.split('@')
    main_part = parts[0] if parts else text
    address_part = parts[1] if len(parts) > 1 else ""
    
    # Extract alphanumeric tokens
    tokens = re.findall(r'\b[A-Z0-9]{2,}\b', main_part.upper())
    
    # Add address tokens if present
    if address_part:
        address_tokens = re.findall(r'\b[A-Z0-9]{2,}\b', address_part.upper())
        tokens.extend(address_tokens)
    
    # Add normalized year codes
    normalized_codes = set()
    for token in tokens:
        normalized = normalize_year_code(token)
        if normalized and normalized != token:
            normalized_codes.add(normalized)
    
    # Filter out noise words
    noise_words = {'LLC', 'INC', 'THE', 'AND', 'FOR', 'CLAIM', 'EST', 'FIRE', 'WATER', 'STORM', 'WTR', 'RFG', 'INT', 'CPS', 'USAA', 'DMO', 'SOL', 'FSR', 'MIT'}
    tokens = [t for t in tokens if t not in noise_words and len(t) >= 2]
    
    # Combine original tokens with normalized codes
    all_tokens = set(tokens) | normalized_codes
    
    return all_tokens

def extract_location_code(text):
    """
    Extract and normalize location codes like GA22, OH24, GA22A, OH24-900
    Returns normalized 4-character code (e.g., GA22, OH24)
    """
    if not text:
        return None
    
    # Pattern: 2 letters + 2 digits + optional letters/numbers/hyphens
    match = re.search(r'\b([A-Z]{2}\d{2}[A-Z0-9\-]*)\b', text.upper())
    if match:
        raw_code = match.group(1)
        # Normalize to first 4 characters only
        return normalize_year_code(raw_code)
    
    return None
    
def extract_contractor_id_from_folder(folder_name):
    """
    Extract a meaningful contractor 'code' token from a folder name.
    Examples to capture: OH25, OH24A, GA23A, OH25P, GA24APC, OH24-900
    Returns '' if nothing meaningful is found.
    """
    if not folder_name:
        return ''
    s = str(folder_name).strip().upper()

    # Prefer hyphenated numeric variants like OH24-900 first
    m = re.search(r'\b([A-Z]{2,3}\d{2,3}-\d{2,4})\b', s)
    if m:
        return m.group(1)

    # Common code forms: 2-3 letters + 2-3 digits + optional 1-3 letters (e.g., OH24A, GA23APC, OH25P)
    m = re.search(r'\b([A-Z]{2,3}\d{2,3}[A-Z]{0,3})\b', s)
    if m:
        return m.group(1)

    # Long numeric ids as a fallback
    m = re.search(r'\b(\d{6,})\b', s)
    if m:
        return m.group(1)

    # C-prefixed ids (e.g., C1234)
    m = re.search(r'\b(C\d{3,})\b', s)
    if m:
        return m.group(1)

    # 3-3 numeric like 123-456
    m = re.search(r'\b(\d{3}-\d{3,})\b', s)
    if m:
        return m.group(1)

    return ''

def normalize_claim_name(name):
    normalized = name.upper().strip()
    normalized = re.sub(r'\s+', ' ', normalized)
    remove_words = ['CLAIM', 'INSURANCE', 'ESTIMATE', 'PROJECT', 'RENOVATION']
    for word in remove_words:
        normalized = re.sub(r'\b' + re.escape(word) + r'\b', '', normalized)
    normalized = re.sub(r'\s+', ' ', normalized).strip()
    return normalized

# =============================================================================
# SIMPLE FUZZY MATCHING
# =============================================================================

def calculate_match_score(encircle_contractor, folder_name):
    """
    Simple fuzzy matching between contractor ID and folder name.
    Returns score 0-1.
    Rule: If 2+ tokens match, it's automatically a strong match (0.8+)
    """
    if not encircle_contractor or not folder_name:
        return 0.0
    
    # Normalize both
    contractor_norm = normalize_text(encircle_contractor)
    folder_norm = normalize_text(folder_name)
    
    # Extract tokens
    contractor_tokens = extract_tokens(contractor_norm)
    folder_tokens = extract_tokens(folder_norm)
    
    # Count matching tokens
    common_tokens = contractor_tokens.intersection(folder_tokens)
    num_matches = len(common_tokens)
    
    # RULE: 2+ matching tokens = automatic match
    if num_matches >= 2:
        return 0.85  # High confidence match
    
    # If only 1 token matches, use detailed scoring
    score = 0.0
    
    # 1. Check if contractor ID is substring of folder (40 points)
    if contractor_norm in folder_norm or folder_norm in contractor_norm:
        score += 0.4
    
    # 2. Location code match (30 points)
    contractor_location = extract_location_code(contractor_norm)
    folder_location = extract_location_code(folder_norm)
    
    if contractor_location and folder_location:
        # Exact match
        if contractor_location == folder_location:
            score += 0.3
        # Fuzzy match (e.g., GA22 vs GA22A)
        elif contractor_location[:4] == folder_location[:4]:
            score += 0.2
    
    # 3. Single token match (30 points)
    if num_matches == 1:
        score += 0.3
    
    return min(score, 1.0)

# =============================================================================
# TEST DATA FILTERS
# =============================================================================

_TEST_EXCLUDE_PATTERNS = [
    'HOW2', 'TEST', 'TEMPLATE', 'SAMPLE', 'ROOMLISTS', 'READINGS',
    'TMPL', 'CHECKLIST', 'TRAILER', 'WAREHOUSE', 'DEFAULT', 'TEMP',
    'PLACEHOLDER', 'EXAMPLE', 'DEMO', 'XXXX', 'AAA', '===', 'BACKEND', 'TUTORIAL'
]

def _is_valid_claim(claim):
    """Filter out test/placeholder claims from Encircle."""
    if not claim.get('policyholder_name') and not claim.get('contractor_identifier'):
        return False
    policyholder = (claim.get('policyholder_name') or '').upper()
    contractor = (claim.get('contractor_identifier') or '').upper()
    return not any(p in policyholder or p in contractor for p in _TEST_EXCLUDE_PATTERNS)

def _is_valid_folder(folder_claim):
    """Filter out test/placeholder folders from OneDrive."""
    folder_name = (folder_claim.get('folder_name') or '').upper()
    exclude_patterns = _TEST_EXCLUDE_PATTERNS + [
        'CLOSED CLAIMS', 'PROOF OF LOSS', 'DRAWINGS', 'APPRAISALS',
        'FOLDER', 'TEXT'
    ]
    if any(p in folder_name for p in exclude_patterns):
        return False
    clean_alpha = re.sub(r'[^A-Z]', '', folder_name)
    return len(clean_alpha) >= 3

# =============================================================================
# DUPLICATE DETECTION
# =============================================================================

def find_duplicates(encircle_claims, onedrive_claims):
    """Find duplicate claims in both systems"""
    duplicates = {'encircle_duplicates': [], 'onedrive_duplicates': []}
    
    # Encircle duplicates by contractor ID
    contractor_count = defaultdict(list)
    for claim in encircle_claims:
        contractor_id = claim.get('contractor_identifier', '')
        if contractor_id and contractor_id.strip():
            contractor_count[contractor_id].append(claim)
    
    for contractor_id, claims in contractor_count.items():
        if len(claims) > 1:
            duplicates['encircle_duplicates'].append({
                'contractor_id': contractor_id, 
                'count': len(claims), 
                'claims': claims
            })
    
    # OneDrive duplicates by normalized folder name
    folder_count = defaultdict(list)
    for claim in onedrive_claims:
        folder_name = claim.get('folder_name', '')
        if folder_name:
            normalized = normalize_text(folder_name)
            folder_count[normalized].append(claim)
    
    for folder_name, claims in folder_count.items():
        if len(claims) > 1:
            duplicates['onedrive_duplicates'].append({
                'folder_name': folder_name, 
                'count': len(claims), 
                'claims': claims
            })
    
    return duplicates

# =============================================================================
# MAIN COMPARISON LOGIC
# =============================================================================

def compare_claims(encircle_claims, onedrive_claims):
    """
    Simple one-to-one matching: each encircle claim finds its BEST onedrive match
    """
    # Filter out test data
    valid_encircle = [c for c in encircle_claims if _is_valid_claim(c)]
    valid_onedrive = [c for c in onedrive_claims if _is_valid_folder(c)]
    
    encircle_test_data = [c for c in encircle_claims if not _is_valid_claim(c)]
    onedrive_test_data = [c for c in onedrive_claims if not _is_valid_folder(c)]
    
    results = {
        'summary': {
            'total_encircle': len(encircle_claims),
            'total_onedrive': len(onedrive_claims),
            'valid_encircle': len(valid_encircle),
            'valid_onedrive': len(valid_onedrive),
            'matches': 0,
            'encircle_only': 0,
            'onedrive_only': 0,
            'encircle_test_data': len(encircle_test_data),
            'onedrive_test_data': len(onedrive_test_data),
            'match_breakdown': {
                'high_confidence': 0,
                'medium_confidence': 0,
                'low_confidence': 0,
            }
        },
        'matched_pairs': [],
        'encircle_missing_onedrive': [],
        'onedrive_extra': [],
        'encircle_test_data': encircle_test_data,
        'onedrive_test_data': onedrive_test_data,
        'duplicates': find_duplicates(valid_encircle, valid_onedrive)
    }
    
    matched_encircle = set()
    matched_onedrive = set()
    
    print("=" * 80)
    print("ONE-TO-ONE CLAIM MATCHING")
    print("=" * 80)
    
    # Find BEST match for each encircle claim
    for encircle_claim in valid_encircle:
        contractor_id = encircle_claim.get('contractor_identifier', '').strip()
        
        if not contractor_id:
            continue
        
        best_match = None
        best_score = 0
        
        # Find the best matching onedrive folder
        for onedrive_claim in valid_onedrive:
            if onedrive_claim['folder_id'] in matched_onedrive:
                continue  # Skip already matched folders
                
            folder_name = onedrive_claim.get('folder_name', '').strip()
            score = calculate_match_score(contractor_id, folder_name)
            
            if score > best_score:
                best_score = score
                best_match = onedrive_claim
        
        # Accept best match if above threshold
        MATCH_THRESHOLD = 0.65
        if best_match and best_score >= MATCH_THRESHOLD:
            confidence_level = "High" if best_score >= 0.8 else "Medium" if best_score >= 0.65 else "Low"
            
            results['matched_pairs'].append({
                'encircle': encircle_claim,
                'onedrive': best_match,
                'match_type': f'Fuzzy Match ({confidence_level})',
                'confidence': f'{int(best_score * 100)}%'
            })
            
            matched_encircle.add(encircle_claim['id'])
            matched_onedrive.add(best_match['folder_id'])
            results['summary']['matches'] += 1
            
            # Update breakdown
            if best_score >= 0.8:
                results['summary']['match_breakdown']['high_confidence'] += 1
            elif best_score >= 0.65:
                results['summary']['match_breakdown']['medium_confidence'] += 1
            else:
                results['summary']['match_breakdown']['low_confidence'] += 1
    
    # Collect unmatched
    for encircle_claim in valid_encircle:
        if encircle_claim['id'] not in matched_encircle:
            results['encircle_missing_onedrive'].append(encircle_claim)
            results['summary']['encircle_only'] += 1
    
    for onedrive_claim in valid_onedrive:
        if onedrive_claim['folder_id'] not in matched_onedrive:
            results['onedrive_extra'].append(onedrive_claim)
            results['summary']['onedrive_only'] += 1
    
    print(f"\nMATCHING COMPLETE")
    print(f"Valid Claims - Encircle: {results['summary']['valid_encircle']}, OneDrive: {results['summary']['valid_onedrive']}")
    print(f"Total Matches: {results['summary']['matches']}")
    print(f"Unmatched Encircle: {results['summary']['encircle_only']}")
    print(f"Unmatched OneDrive: {results['summary']['onedrive_only']}")
    print("=" * 80)
    
    return results

# =============================================================================
# REPORT GENERATION
# =============================================================================

def _coerce_report_defaults(results):
    """Ensure all keys required by the template exist with sensible defaults and precompute display fields."""
    results.setdefault('matched_pairs', [])
    results.setdefault('encircle_missing_onedrive', [])
    results.setdefault('onedrive_extra', [])
    results.setdefault('encircle_test_data', [])
    results.setdefault('onedrive_test_data', [])

    summary = results.setdefault('summary', {})
    summary.setdefault('total_encircle', 0)
    summary.setdefault('total_onedrive', 0)
    summary.setdefault('valid_encircle', 0)
    summary.setdefault('valid_onedrive', 0)
    summary.setdefault('matches', len(results['matched_pairs']))
    summary.setdefault('encircle_only', len(results['encircle_missing_onedrive']))
    summary.setdefault('onedrive_only', len(results['onedrive_extra']))
    summary.setdefault('encircle_test_data', len(results['encircle_test_data']))
    summary.setdefault('onedrive_test_data', len(results['onedrive_test_data']))
    summary.setdefault('match_breakdown', {
        'high_confidence': 0,
        'medium_confidence': 0,
        'low_confidence': 0,
    })

    # Precompute totals for display
    summary['issues_total'] = summary['encircle_only'] + summary['onedrive_only']
    summary['test_total'] = summary['encircle_test_data'] + summary['onedrive_test_data']
    
    # Count unique encircle claims that have matches
    unique_matched_encircle = len(set(pair['encircle']['id'] for pair in results['matched_pairs']))
    summary['unique_matches'] = unique_matched_encircle

    # Display header counts (real claims only)
    results['encircle_render_total'] = unique_matched_encircle + summary['encircle_only']
    results['onedrive_render_total'] = summary['matches'] + summary['onedrive_only']

    # DON'T precompute unmatched_rows - we'll show missing claims separately
    return results

def generate_comparison_report(comparison_results, is_refresh=False):
    template_content = """
<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Claims Sync Report</title>
<style>
* { margin:0; padding:0; box-sizing:border-box; }
body { font-family:'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background:#f5f5f5; color:#333; line-height:1.4; }
.container { max-width:1800px; margin:0 auto; background:white; }
.header { background:linear-gradient(135deg,#2c3e50 0%,#3498db 100%); color:white; padding:30px; text-align:center; position:relative; }
.header h1 { font-size:2em; margin-bottom:8px; }
.header p { font-size:1em; opacity:.9; }
.cache-badge { position:absolute; top:15px; right:20px; background:rgba(255,255,255,.2); padding:8px 15px; border-radius:20px; font-size:.85em; border:1px solid rgba(255,255,255,.3); }
.cache-badge.cached { background:#27ae60; }
.cache-badge.fresh { background:#e74c3c; }
.status-bar { display:grid; grid-template-columns:repeat(5,1fr); gap:12px; padding:20px; background:#ecf0f1; }
.status-card { background:white; padding:12px; border-radius:6px; text-align:center; box-shadow:0 1px 3px rgba(0,0,0,.1); }
.status-card .label { font-size:.75em; color:#666; margin-bottom:6px; font-weight:600; text-transform:uppercase; }
.status-card .value { font-size:1.6em; font-weight:bold; }
.status-card .sub-value { font-size:.7em; color:#999; margin-top:3px; }
.status-card.matched .value { color:#27ae60; }
.status-card.missing .value { color:#e74c3c; }
.status-card.encircle .value { color:#3498db; }
.status-card.onedrive .value { color:#2980b9; }
.status-card.test-data .value { color:#f39c12; }
.content { padding:20px; }
.section { margin-bottom:30px; }
.section-title { font-size:1.3em; font-weight:600; color:#2c3e50; margin-bottom:15px; padding-bottom:8px; border-bottom:2px solid #3498db; }
.claims-list { background:#f8f9fa; padding:15px; border-radius:6px; max-height:400px; overflow-y:auto; }
.claim-row { display:flex; align-items:center; padding:8px 12px; margin-bottom:5px; background:white; border-radius:4px; border-left:4px solid #e74c3c; font-size:.9em; }
.claim-row.matched { border-left-color:#27ae60; background:#f0f8ff; }
.claim-row.unmatched { border-left-color:#e74c3c; background:#fff5f5; }
.claim-row.extra { border-left-color:#f39c12; background:#fff9e6; }
.claim-name { flex:1; font-weight:600; }
.claim-details { flex:2; color:#666; font-size:.85em; }
.claim-id { flex:1; color:#999; font-size:.8em; }
.note { background:#fff9e6; padding:10px; border-radius:4px; margin:12px 0; border-left:4px solid #f39c12; font-size:.9em; }
.match-breakdown { display:grid; grid-template-columns:repeat(auto-fit,minmax(180px,1fr)); gap:8px; margin:12px 0; }
.breakdown-item { background:#f8f9fa; padding:8px; border-radius:4px; text-align:center; }
.breakdown-item .count { font-size:1.1em; font-weight:bold; color:#2c3e50; }
.breakdown-item .label { font-size:.75em; color:#666; }
.footer { text-align:center; padding:15px; color:#999; font-size:.8em; border-top:1px solid #ecf0f1; }
.refresh-links { text-align:center; padding:15px; background:#ecf0f1; }
.refresh-links a { display:inline-block; margin:0 10px; padding:10px 20px; background:#3498db; color:white; text-decoration:none; border-radius:4px; font-weight:600; }
.refresh-links a:hover { background:#2980b9; }
.refresh-links a.refresh { background:#e74c3c; }
.refresh-links a.refresh:hover { background:#c0392b; }
.refresh-links a.export { background:#27ae60; }
.refresh-links a.export:hover { background:#229954; }
</style>
</head>
<body>
<div class="container">
  <div class="header">
    <h1>Claims Sync Report</h1>
    <p>Encircle System vs OneDrive Storage</p>
    {% if is_refresh %}
    <div class="cache-badge fresh">🔄 Fresh Data</div>
    {% else %}
    <div class="cache-badge cached">💾 Cached Data</div>
    {% endif %}
  </div>

  <div class="refresh-links">
    <a href="?">📊 View Report (Cached)</a>
    <a href="/sync/refresh/" class="refresh">🔄 Refresh Data</a>
    <a href="/sync/export/encircle/" class="export">📥 Export Encircle CSV</a>
    <a href="/sync/export/onedrive/" class="export">📥 Export OneDrive CSV</a>
  </div>

  <div class="status-bar">
    <div class="status-card encircle">
      <div class="label">Encircle Total</div>
      <div class="value">{{ summary.valid_encircle|default:0 }}</div>
      <div class="sub-value">{{ summary.encircle_test_data|default:0 }} test excluded</div>
    </div>
    <div class="status-card onedrive">
      <div class="label">OneDrive Total</div>
      <div class="value">{{ summary.valid_onedrive|default:0 }}</div>
      <div class="sub-value">{{ summary.onedrive_test_data|default:0 }} test excluded</div>
    </div>
    <div class="status-card matched">
      <div class="label">Matches</div>
      <div class="value">{{ summary.matches|default:0 }}</div>
      <div class="sub-value">real claims only</div>
    </div>
    <div class="status-card missing">
      <div class="label">Issues</div>
      <div class="value">{{ summary.issues_total|default:0 }}</div>
      <div class="sub-value">need attention</div>
    </div>
    <div class="status-card test-data">
      <div class="label">Test Data</div>
      <div class="value">{{ summary.test_total|default:0 }}</div>
      <div class="sub-value">auto-filtered</div>
    </div>
  </div>

  <div class="content">
    
    {# CLAIMS TO ADD TO ENCIRCLE - SHOW FIRST #}
    <div class="section">
      <div class="section-title">🚨 Claims to ADD to Encircle ({{ summary.onedrive_only }})</div>
      <div class="note" style="border-left-color:#f39c12; background:#fff9e6;">
        <strong>ACTION NEEDED:</strong> These OneDrive folders exist but have no matching Encircle claim. Create claims in Encircle for these.
      </div>
      
      {% if onedrive_extra %}
      <div class="claims-list">
        {% for folder in onedrive_extra %}
        <div class="claim-row extra">
          <div class="claim-name">➕ {{ folder.folder_name }}</div>
          <div class="claim-details">
            Contractor: {{ folder.contractor_identifier }} | 
            Files: {{ folder.file_count|default:0 }} | 
            Info Sheet: {% if folder.has_info_file %}✅{% else %}❌{% endif %}
          </div>
          <div class="claim-id">ID: {{ folder.folder_id|truncatechars:15 }}</div>
        </div>
        {% endfor %}
      </div>
      {% else %}
      <div class="note" style="background:#d4edda; border-left-color:#27ae60;">
        ✅ No missing claims found! All OneDrive folders have matching Encircle claims.
      </div>
      {% endif %}
    </div>

    {# CLAIMS MISSING FROM ONEDRIVE #}
    <div class="section">
      <div class="section-title">📁 Claims Missing from OneDrive ({{ summary.encircle_only }})</div>
      <div class="note" style="border-left-color:#e74c3c; background:#fff5f5;">
        <strong>ACTION REQUIRED:</strong> These Encircle claims have NO folders in OneDrive. Create folders immediately!
      </div>
      
      {% if encircle_missing_onedrive %}
      <div class="claims-list">
        {% for claim in encircle_missing_onedrive %}
        <div class="claim-row unmatched">
          <div class="claim-name">❌ {{ claim.policyholder_name }}</div>
          <div class="claim-details">
            Contractor: {{ claim.contractor_identifier }} | 
            Address: {{ claim.full_address|default:""|truncatewords:3 }}
          </div>
          <div class="claim-id">ID: {{ claim.id }}</div>
        </div>
        {% endfor %}
      </div>
      {% else %}
      <div class="note" style="background:#d4edda; border-left-color:#27ae60;">
        ✅ All Encircle claims have folders in OneDrive!
      </div>
      {% endif %}
    </div>

    {# MATCH BREAKDOWN #}
    {% if summary.match_breakdown %}
    <div class="section">
      <div class="section-title">📊 Match Breakdown</div>
      <div class="note"><strong>Enhanced Location Matching:</strong> Now comparing only first 4 characters of location codes (e.g., GA22 = GA22A = GA22P)</div>
      <div class="match-breakdown">
        <div class="breakdown-item">
          <div class="count">{{ summary.match_breakdown.high_confidence|default:0 }}</div>
          <div class="label">High Confidence (80%+)</div>
        </div>
        <div class="breakdown-item">
          <div class="count">{{ summary.match_breakdown.medium_confidence|default:0 }}</div>
          <div class="label">Medium Confidence (65-80%)</div>
        </div>
        <div class="breakdown-item">
          <div class="count">{{ summary.match_breakdown.low_confidence|default:0 }}</div>
          <div class="label">Low Confidence (Below 65%)</div>
        </div>
      </div>
    </div>
    {% endif %}

    {# SUCCESSFUL MATCHES - CONDENSED VIEW #}
    <div class="section">
      <div class="section-title">✅ Successfully Matched Claims ({{ summary.matches }})</div>
      <div class="note">
        <strong>Location Code Enhancement:</strong> GA22, GA22A, GA22P now all match as GA22
      </div>
      
      {% if matched_pairs %}
      <div class="claims-list">
        {% for pair in matched_pairs %}
        <div class="claim-row matched">
          <div class="claim-name">✓ {{ pair.encircle.policyholder_name }}</div>
          <div class="claim-details">
            → {{ pair.onedrive.folder_name }} | 
            Confidence: {{ pair.confidence }}
          </div>
          <div class="claim-id">ID: {{ pair.encircle.id }}</div>
        </div>
        {% endfor %}
      </div>
      {% else %}
      <div class="note" style="background:#fff5f5; border-left-color:#e74c3c;">
        ⚠️ No matches found. Check your data sources.
      </div>
      {% endif %}
    </div>

  </div>

  <div class="footer">
    <p><strong>Generated:</strong> {{ generated_date }} | 
       <strong>Status:</strong> {% if is_refresh %}Fresh Download{% else %}Using Cache{% endif %} |
       <strong>Enhanced Location Matching:</strong> 4-char codes only |
       <strong>Issues:</strong> {{ summary.issues_total|default:0 }} total</p>
  </div>
</div>
</body>
</html>
"""
    
    # Ensure all keys and precomputed fields exist
    comparison_results = _coerce_report_defaults(comparison_results)
    comparison_results['generated_date'] = dt.datetime.now().strftime("%Y-%m-%d %H:%M")
    comparison_results['is_refresh'] = is_refresh

    template = Template(template_content)
    context = Context(comparison_results)
    return template.render(context)

import csv
from django.http import HttpResponse
from django.views.decorators.http import require_GET
from django.core.cache import cache

# Add these new views to your existing code

@require_GET
def export_encircle_csv(request):
    """Export Encircle claims data as CSV"""
    try:
        # Get cached data (or fetch fresh if not available)
        encircle_claims = get_encircle_claims(use_cache=True)
        
        # Create HTTP response with CSV content type
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="encircle_claims_export.csv"'
        
        writer = csv.writer(response)
        
        # Write header row
        writer.writerow([
            'Claim ID',
            'Policyholder Name',
            'Contractor Identifier',
            'Full Address',
            'Insurance Company',
            'Status',
            'Created Date'
        ])
        
        # Write data rows
        for claim in encircle_claims:
            writer.writerow([
                claim.get('id', ''),
                claim.get('policyholder_name', ''),
                claim.get('contractor_identifier', ''),
                claim.get('full_address', ''),
                claim.get('insurance_company_name', ''),
                claim.get('status', ''),
                claim.get('created_date', '')
            ])
        
        return response
        
    except Exception as e:
        logger.error(f"Error exporting Encircle CSV: {str(e)}")
        return HttpResponse(f"Error exporting data: {str(e)}", status=500)


@require_GET
def export_onedrive_csv(request):
    """Export OneDrive claims data as CSV"""
    try:
        # Get cached data (or fetch fresh if not available)
        onedrive_claims = get_onedrive_claims(use_cache=True)
        
        # Create HTTP response with CSV content type
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="onedrive_folders_export.csv"'
        
        writer = csv.writer(response)
        
        # Write header row
        writer.writerow([
            'Folder Name',
            'Folder ID',
            'Contractor Identifier',
            'Normalized Name',
            'Has Info File',
            'File Count',
            'Error'
        ])
        
        # Write data rows
        for claim in onedrive_claims:
            writer.writerow([
                claim.get('folder_name', ''),
                claim.get('folder_id', ''),
                claim.get('contractor_identifier', ''),
                claim.get('normalized_name', ''),
                'Yes' if claim.get('has_info_file', False) else 'No',
                claim.get('file_count', 0),
                claim.get('error', '')
            ])
        
        return response
        
    except Exception as e:
        logger.error(f"Error exporting OneDrive CSV: {str(e)}")
        return HttpResponse(f"Error exporting data: {str(e)}", status=500)


# ==================== ENCIRCLE WEBHOOK ENDPOINTS ====================

@csrf_exempt
def encircle_webhook(request):
    """
    Webhook endpoint for Encircle API notifications.

    This endpoint receives webhook events from Encircle when:
    - Media is created/updated (including floorplans)
    - Claims are updated
    - Other events

    We specifically filter for floorplan creation events and send email notifications.

    Encircle Webhook Event Types:
    - media.created: New media uploaded
    - media.updated: Media updated
    - floor_plan.created: Floorplan created
    - floor_plan.updated: Floorplan updated
    - property_claim.updated: Claim data updated
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    try:
        # Parse the webhook payload
        try:
            payload = json.loads(request.body)
        except json.JSONDecodeError:
            logger.error("Invalid JSON in Encircle webhook payload")
            return JsonResponse({'error': 'Invalid JSON'}, status=400)

        # Log the webhook for debugging
        event_type = payload.get('event_type', payload.get('type', 'unknown'))
        logger.info(f"Encircle webhook received: event_type={event_type}")
        logger.debug(f"Encircle webhook payload: {json.dumps(payload, indent=2)}")

        # Verify webhook signature if configured
        webhook_secret = getattr(settings, 'ENCIRCLE_WEBHOOK_SECRET', '')
        if webhook_secret:
            signature = request.headers.get('X-Encircle-Signature', '')
            if not _verify_encircle_signature(request.body, signature, webhook_secret):
                logger.warning("Invalid Encircle webhook signature")
                return JsonResponse({'error': 'Invalid signature'}, status=401)

        # Process the webhook based on event type
        result = _process_encircle_webhook(payload)

        return JsonResponse({
            'success': True,
            'message': 'Webhook processed',
            'result': result
        })

    except Exception as e:
        logger.error(f"Error processing Encircle webhook: {str(e)}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)


def _verify_encircle_signature(payload_body, signature, secret):
    """Verify the Encircle webhook signature (if applicable)."""
    import hmac
    import hashlib

    if not signature or not secret:
        return True  # Skip verification if not configured

    expected = hmac.new(
        secret.encode('utf-8'),
        payload_body,
        hashlib.sha256
    ).hexdigest()

    return hmac.compare_digest(signature, expected)


def _process_encircle_webhook(payload):
    """
    Process the Encircle webhook payload and trigger appropriate actions.

    Encircle webhook structure (typical):
    {
        "event_type": "media.created" | "floor_plan.created" | etc,
        "data": {
            "property_claim_id": 12345,
            "media_type": "floor_plan" | "photo" | etc,
            "url": "https://...",
            ...
        },
        "claim": {
            "id": 12345,
            "name": "Claim Name",
            "address": "123 Main St",
            ...
        }
    }
    """
    from .tasks import send_floorplan_notification_task

    event_type = payload.get('event_type', payload.get('type', ''))
    data = payload.get('data', {})
    claim_data = payload.get('claim', payload.get('property_claim', {}))

    # Check if this is a floorplan event
    is_floorplan_event = False
    floorplan_url = None

    # Method 1: Direct floor_plan event type
    if 'floor_plan' in event_type.lower():
        is_floorplan_event = True
        floorplan_url = data.get('url') or data.get('download_url') or data.get('image_url')

    # Method 2: Media event with floor_plan type
    elif event_type in ['media.created', 'media.updated']:
        media_type = data.get('media_type', data.get('type', ''))
        if 'floor' in media_type.lower() or 'plan' in media_type.lower():
            is_floorplan_event = True
            floorplan_url = data.get('url') or data.get('download_url') or data.get('image_url')

    # Method 3: Check for floor_plan_dimensions or similar fields
    if not is_floorplan_event:
        if 'floor_plan_dimensions' in data or 'floor_plan' in data:
            is_floorplan_event = True
            fp_data = data.get('floor_plan_dimensions', data.get('floor_plan', {}))
            if isinstance(fp_data, dict):
                floorplan_url = fp_data.get('url') or fp_data.get('image_url')

    if is_floorplan_event:
        # Extract claim info
        claim_id = (
            claim_data.get('id') or
            data.get('property_claim_id') or
            data.get('claim_id') or
            payload.get('property_claim_id')
        )

        claim_info = {
            'encircle_id': claim_id,
            'name': claim_data.get('name', claim_data.get('contractor_identifier', f'Claim {claim_id}')),
            'address': claim_data.get('address', claim_data.get('location', {}).get('address', ''))
        }

        # Trigger the async task to send notification
        if claim_id:
            send_floorplan_notification_task.delay(
                claim_id=str(claim_id),
                floorplan_url=floorplan_url,
                claim_info=claim_info
            )
            logger.info(f"Floorplan notification task queued for claim {claim_id}")
            return {'action': 'floorplan_notification_queued', 'claim_id': claim_id}
        else:
            logger.warning("Floorplan event received but no claim_id found")
            return {'action': 'skipped', 'reason': 'no_claim_id'}

    # Not a floorplan event - log and acknowledge
    logger.info(f"Encircle webhook event '{event_type}' received but not a floorplan event - ignoring")
    return {'action': 'ignored', 'event_type': event_type}


@csrf_exempt
def encircle_webhook_test(request):
    """
    Test endpoint to manually trigger a floorplan notification.
    Useful for testing the webhook flow without actual Encircle events.

    POST /webhooks/encircle/test/
    {
        "claim_id": "12345",
        "claim_name": "Test Claim",
        "claim_address": "123 Test St",
        "floorplan_url": "https://example.com/floorplan.png"  (optional)
    }
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    try:
        data = json.loads(request.body)
        claim_id = data.get('claim_id', 'test_123')
        claim_name = data.get('claim_name', 'Test Claim')
        claim_address = data.get('claim_address', '123 Test Street')
        floorplan_url = data.get('floorplan_url', '')

        from .tasks import send_floorplan_notification_task

        claim_info = {
            'encircle_id': claim_id,
            'name': claim_name,
            'address': claim_address
        }

        # Queue the notification task
        send_floorplan_notification_task.delay(
            claim_id=str(claim_id),
            floorplan_url=floorplan_url,
            claim_info=claim_info
        )

        return JsonResponse({
            'success': True,
            'message': 'Test floorplan notification queued',
            'claim_id': claim_id
        })

    except Exception as e:
        logger.error(f"Error in encircle_webhook_test: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


# ==================== LEASE MANAGER DASHBOARD ====================

def lease_manager(request):
    """
    Main Lease Manager Dashboard
    Displays leases (not individual documents), activity feed, and pipeline status
    """
    from django.db.models import Count, Sum, Q
    from datetime import timedelta, date
    from .models import PipelineStageAssignment, LeaseStageCompletion

    # Get filter parameters
    status_filter = request.GET.get('status', '')
    client_filter = request.GET.get('client', '')
    date_filter = request.GET.get('date_range', '30')  # Default 30 days

    # Calculate date range
    try:
        days = int(date_filter)
    except ValueError:
        days = 30
    date_threshold = timezone.now() - timedelta(days=days)

    today = date.today()

    # Get all leases (not individual documents)
    leases_query = Lease.objects.select_related(
        'client', 'created_by', 'last_modified_by'
    ).prefetch_related('documents', 'stage_completions', 'stage_completions__assigned_user', 'stage_completions__completed_by')

    if status_filter:
        leases_query = leases_query.filter(status=status_filter)
    if client_filter:
        leases_query = leases_query.filter(client__id=client_filter)

    all_leases = leases_query.order_by('-created_at')[:100]

    # Get recent activity
    recent_activity = LeaseActivity.objects.select_related(
        'lease', 'lease__client', 'performed_by'
    ).filter(
        created_at__gte=date_threshold
    ).order_by('-created_at')[:50]

    # Pipeline statistics - CURRENT counts (how many are at each stage right now)
    pipeline_stats = Lease.objects.values('status').annotate(
        count=Count('id')
    ).order_by('status')

    # Convert to dict for easy template access
    status_counts = {item['status']: item['count'] for item in pipeline_stats}

    # CUMULATIVE pipeline counts - how many leases have reached or passed each stage
    # A lease at "signed" has passed draft, generated, review, sent_for_signature
    STATUS_ORDER = [
        'draft', 'generated', 'review', 'sent_for_signature',
        'signed', 'invoice_created', 'package_sent',
        'payment_pending', 'payment_received', 'completed'
    ]

    cumulative_counts = {}
    total_non_cancelled = Lease.objects.exclude(status='cancelled').count()

    for i, status in enumerate(STATUS_ORDER):
        # Count leases that are at this stage or have passed it
        statuses_at_or_past = STATUS_ORDER[i:]
        cumulative_counts[status] = Lease.objects.filter(
            status__in=statuses_at_or_past
        ).exclude(status='cancelled').count()

    # Get stage assignments for display (which team member handles each stage)
    stage_assignments = PipelineStageAssignment.objects.select_related('assigned_user').order_by('order')

    # Build pipeline steps with assignee info for template
    pipeline_steps = []
    for i, status_tuple in enumerate(Lease.LEASE_STATUS_CHOICES):
        status_value, status_label = status_tuple
        if status_value == 'cancelled':
            continue  # Skip cancelled in pipeline view

        assignment = stage_assignments.filter(stage=status_value).first()
        assignee_email = assignment.assigned_user.email if assignment and assignment.assigned_user else 'Unassigned'
        assignee_initials = ''.join([part[0].upper() for part in assignee_email.split('@')[0].split('.')[:2]]) if assignment and assignment.assigned_user else '?'

        pipeline_steps.append({
            'value': status_value,
            'label': status_label,
            'order': i,
            'assignee_email': assignee_email,
            'assignee_initials': assignee_initials,
            'current_count': status_counts.get(status_value, 0),
            'cumulative_count': cumulative_counts.get(status_value, 0),
        })

    # Calculate totals - ACTIVE means within date range and not cancelled/completed
    total_active = Lease.objects.filter(
        lease_start_date__lte=today,
        lease_end_date__gte=today
    ).exclude(
        status__in=['completed', 'cancelled']
    ).count()

    total_completed = Lease.objects.filter(status='completed').count()

    # Get expired leases (end date passed)
    total_expired = Lease.objects.filter(
        lease_end_date__lt=today
    ).exclude(
        status__in=['completed', 'cancelled']
    ).count()

    # Get leases organized by client/property owner
    clients_with_leases = Client.objects.filter(
        leases__isnull=False
    ).distinct().prefetch_related(
        'leases', 'leases__documents'
    ).annotate(
        lease_count=Count('leases', distinct=True),
        active_lease_count=Count(
            'leases',
            filter=Q(leases__lease_start_date__lte=today) & Q(leases__lease_end_date__gte=today) & ~Q(leases__status__in=['completed', 'cancelled']),
            distinct=True
        )
    ).order_by('-leases__created_at')

    # Get all clients for filter dropdown
    all_clients = Client.objects.all().order_by('pOwner')

    # Status choices for filter
    status_choices = Lease.LEASE_STATUS_CHOICES

    # Get total monthly rent for ACTIVE leases only (within date range)
    total_monthly_rent = Lease.objects.filter(
        lease_start_date__lte=today,
        lease_end_date__gte=today
    ).exclude(
        status__in=['completed', 'cancelled']
    ).aggregate(
        total=Sum('monthly_rent')
    )['total'] or 0

    context = {
        'leases': all_leases,
        'recent_activity': recent_activity,
        'status_counts': status_counts,
        'cumulative_counts': cumulative_counts,
        'pipeline_steps': pipeline_steps,
        'stage_assignments': stage_assignments,
        'total_active': total_active,
        'total_completed': total_completed,
        'total_expired': total_expired,
        'total_non_cancelled': total_non_cancelled,
        'clients_with_leases': clients_with_leases,
        'all_clients': all_clients,
        'status_choices': status_choices,
        'current_status_filter': status_filter,
        'current_client_filter': client_filter,
        'current_date_filter': date_filter,
        'total_monthly_rent': total_monthly_rent,
        'today': today,
    }

    return render(request, 'account/lease_manager.html', context)

def create_draft_lease(request):
    """
    Auto-create a draft lease when user starts inputting information for a client.
    Called via AJAX when user selects a client for lease generation.
    """
    from .models import PipelineStageAssignment, LeaseStageCompletion

    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    try:
        data = json.loads(request.body)
        client_id = data.get('client_id')
        client_name = data.get('client_name')

        if not client_id and not client_name:
            return JsonResponse({'error': 'client_id or client_name required'}, status=400)

        # Look up client by ID or name
        if client_id:
            client = Client.objects.get(id=client_id)
        else:
            client = Client.objects.get(pOwner=client_name)

        # Check if there's already a draft for this client
        existing_draft = Lease.objects.filter(
            client=client,
            status='draft'
        ).first()

        if existing_draft:
            return JsonResponse({
                'success': True,
                'lease_id': str(existing_draft.id),
                'message': 'Existing draft found',
                'is_new': False
            })

        # Create new draft lease with minimal data
        lease = Lease.objects.create(
            client=client,
            lessor_name='',  # Will be filled in during form entry
            property_address=client.pAddress or '',
            property_city=client.pCityStateZip.split(',')[0].strip() if client.pCityStateZip else '',
            status='draft',
            created_by=request.user if request.user.is_authenticated else None,
            last_modified_by=request.user if request.user.is_authenticated else None,
        )

        # Log activity
        LeaseActivity.objects.create(
            lease=lease,
            activity_type='draft',
            description=f'Draft lease created for {client.pOwner}',
            performed_by=request.user if request.user.is_authenticated else None
        )

        # Create stage completion records for all stages
        stage_assignments = PipelineStageAssignment.objects.all()
        for assignment in stage_assignments:
            LeaseStageCompletion.objects.create(
                lease=lease,
                stage=assignment.stage,
                assigned_user=assignment.assigned_user,
                is_completed=False
            )

        # Mark draft stage as completed since we just created it
        draft_completion = LeaseStageCompletion.objects.filter(
            lease=lease,
            stage='draft'
        ).first()
        if draft_completion:
            draft_completion.is_completed = True
            draft_completion.completed_by = request.user if request.user.is_authenticated else None
            draft_completion.completed_at = timezone.now()
            draft_completion.save()

        return JsonResponse({
            'success': True,
            'lease_id': str(lease.id),
            'message': 'Draft lease created',
            'is_new': True
        })

    except Client.DoesNotExist:
        return JsonResponse({'error': 'Client not found'}, status=404)
    except Exception as e:
        logger.error(f"Error creating draft lease: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


def update_lease_status(request):
    """API endpoint to update lease status"""
    from .models import LeaseStageCompletion

    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    try:
        data = json.loads(request.body)
        lease_id = data.get('lease_id')
        new_status = data.get('status')

        if not lease_id or not new_status:
            return JsonResponse({'error': 'Missing required fields'}, status=400)

        lease = Lease.objects.get(id=lease_id)
        old_status = lease.status

        # Update status
        lease.status = new_status
        lease.last_modified_by = request.user if request.user.is_authenticated else None

        # Update timestamp based on status
        now = timezone.now()
        status_timestamp_map = {
            'generated': 'generated_at',
            'review': 'reviewed_at',
            'sent_for_signature': 'sent_for_signature_at',
            'signed': 'signed_at',
            'invoice_created': 'invoice_created_at',
            'package_sent': 'package_sent_at',
            'payment_received': 'payment_received_at',
            'completed': 'completed_at',
        }

        if new_status in status_timestamp_map:
            setattr(lease, status_timestamp_map[new_status], now)

        lease.save()

        # Update stage completion record
        stage_completion = LeaseStageCompletion.objects.filter(
            lease=lease,
            stage=new_status
        ).first()

        if stage_completion and not stage_completion.is_completed:
            stage_completion.is_completed = True
            stage_completion.completed_by = request.user if request.user.is_authenticated else None
            stage_completion.completed_at = now
            stage_completion.save()

        # Log activity with status-aligned activity type
        LeaseActivity.objects.create(
            lease=lease,
            activity_type=new_status,  # Use new status as activity type
            description=f'Status changed from "{old_status}" to "{new_status}"',
            old_status=old_status,
            new_status=new_status,
            performed_by=request.user if request.user.is_authenticated else None
        )

        return JsonResponse({
            'success': True,
            'new_status': new_status,
            'status_display': lease.get_status_display()
        })

    except Lease.DoesNotExist:
        return JsonResponse({'error': 'Lease not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def get_leases_by_client(request, client_id):
    """API endpoint to get leases for a specific client"""
    from datetime import date
    today = date.today()

    try:
        client = Client.objects.get(id=client_id)
        leases = Lease.objects.filter(client=client).prefetch_related('documents').order_by('-created_at')

        leases_data = []
        for lease in leases:
            # Get documents for this lease
            docs = [{
                'id': str(doc.id),
                'document_type': doc.document_type,
                'document_type_display': doc.get_document_type_display(),
                'document_name': doc.document_name,
                'file_path': doc.file_path,
            } for doc in lease.documents.all()]

            leases_data.append({
                'id': str(lease.id),
                'lessor_name': lease.lessor_name,
                'property_address': lease.full_property_address,
                'monthly_rent': float(lease.monthly_rent) if lease.monthly_rent else None,
                'lease_start_date': lease.lease_start_date.isoformat() if lease.lease_start_date else None,
                'lease_end_date': lease.lease_end_date.isoformat() if lease.lease_end_date else None,
                'status': lease.status,
                'status_display': lease.get_status_display(),
                'status_color': lease.get_status_color(),
                'is_active': lease.is_active,
                'is_expired': lease.is_expired,
                'is_renewal': lease.is_renewal,
                'created_at': lease.created_at.isoformat(),
                'created_by': lease.created_by.email if lease.created_by else None,
                'documents': docs,
            })

        return JsonResponse({
            'success': True,
            'client': {
                'id': client.id,
                'name': client.pOwner,
                'address': client.pAddress,
            },
            'leases': leases_data
        })

    except Client.DoesNotExist:
        return JsonResponse({'error': 'Client not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def download_lease_document(request, document_id):
    """Download a specific lease document PDF"""
    import mimetypes

    try:
        lease_doc = LeaseDocument.objects.select_related('lease').get(id=document_id)

        if not lease_doc.file_path:
            return HttpResponse("Document file path not set", status=404)

        # Build full path from MEDIA_ROOT
        full_path = os.path.join(settings.MEDIA_ROOT, lease_doc.file_path)

        if not os.path.exists(full_path):
            return HttpResponse(f"Document file not found at {full_path}", status=404)

        # Log download activity
        LeaseActivity.objects.create(
            lease=lease_doc.lease,
            activity_type='downloaded',
            description=f'Downloaded {lease_doc.document_name}',
            performed_by=request.user if request.user.is_authenticated else None
        )

        # Serve the file
        with open(full_path, 'rb') as f:
            content = f.read()

        content_type, _ = mimetypes.guess_type(full_path)
        response = HttpResponse(content, content_type=content_type or 'application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{lease_doc.document_name}.pdf"'
        return response

    except LeaseDocument.DoesNotExist:
        return HttpResponse("Document not found", status=404)
    except Exception as e:
        return HttpResponse(f"Error downloading document: {str(e)}", status=500)


def view_lease_document(request, document_id):
    """View a specific lease document PDF in browser"""
    import mimetypes

    try:
        lease_doc = LeaseDocument.objects.select_related('lease').get(id=document_id)

        if not lease_doc.file_path:
            return HttpResponse("Document file path not set", status=404)

        # Build full path from MEDIA_ROOT
        full_path = os.path.join(settings.MEDIA_ROOT, lease_doc.file_path)

        if not os.path.exists(full_path):
            return HttpResponse(f"Document file not found", status=404)

        # Log view activity
        LeaseActivity.objects.create(
            lease=lease_doc.lease,
            activity_type='viewed',
            description=f'Viewed {lease_doc.document_name}',
            performed_by=request.user if request.user.is_authenticated else None
        )

        # Serve the file inline (opens in browser)
        with open(full_path, 'rb') as f:
            content = f.read()

        content_type, _ = mimetypes.guess_type(full_path)
        response = HttpResponse(content, content_type=content_type or 'application/pdf')
        # Use 'inline' to open in browser instead of download
        response['Content-Disposition'] = f'inline; filename="{lease_doc.document_name}.pdf"'
        return response

    except LeaseDocument.DoesNotExist:
        return HttpResponse("Document not found", status=404)
    except Exception as e:
        return HttpResponse(f"Error viewing document: {str(e)}", status=500)


def add_lease_note(request):
    """Add a note to a lease"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    try:
        data = json.loads(request.body)
        lease_id = data.get('lease_id')
        note = data.get('note', '').strip()

        if not lease_id or not note:
            return JsonResponse({'error': 'Missing required fields'}, status=400)

        lease = Lease.objects.get(id=lease_id)

        # Append note with timestamp
        timestamp = timezone.now().strftime('%Y-%m-%d %H:%M')
        user_name = request.user.email if request.user.is_authenticated else 'Anonymous'
        new_note = f"[{timestamp}] {user_name}: {note}"

        if lease.notes:
            lease.notes = f"{lease.notes}\n\n{new_note}"
        else:
            lease.notes = new_note

        lease.save()

        # Log activity
        LeaseActivity.objects.create(
            lease=lease,
            activity_type='note_added',
            description=f'Note added: {note[:100]}...' if len(note) > 100 else f'Note added: {note}',
            performed_by=request.user if request.user.is_authenticated else None
        )

        return JsonResponse({'success': True, 'note': new_note})

    except Lease.DoesNotExist:
        return JsonResponse({'error': 'Lease not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def lease_activity_feed(request):
    """API endpoint to get paginated activity feed"""
    page = int(request.GET.get('page', 1))
    per_page = int(request.GET.get('per_page', 20))
    client_filter = request.GET.get('client', '')

    offset = (page - 1) * per_page

    query = LeaseActivity.objects.select_related('lease', 'lease__client', 'performed_by')

    if client_filter:
        query = query.filter(lease__client__id=client_filter)

    total_count = query.count()
    activities = query.order_by('-created_at')[offset:offset + per_page]

    activity_data = []
    for activity in activities:
        activity_data.append({
            'id': str(activity.id),
            'activity_type': activity.activity_type,
            'activity_type_display': activity.get_activity_type_display(),
            'description': activity.description,
            'client_name': activity.lease.client.pOwner if activity.lease else 'Unknown',
            'client_id': activity.lease.client.id if activity.lease else None,
            'lease_id': str(activity.lease.id) if activity.lease else None,
            'performed_by': activity.performed_by.email if activity.performed_by else 'System',
            'created_at': activity.created_at.isoformat(),
            'time_ago': get_time_ago(activity.created_at),
        })

    return JsonResponse({
        'success': True,
        'activities': activity_data,
        'total_count': total_count,
        'page': page,
        'per_page': per_page,
        'has_more': offset + per_page < total_count
    })


def get_time_ago(dt):
    """Helper function to get human-readable time ago string"""
    now = timezone.now()
    diff = now - dt

    if diff.days > 30:
        return dt.strftime('%b %d, %Y')
    elif diff.days > 0:
        return f"{diff.days} day{'s' if diff.days > 1 else ''} ago"
    elif diff.seconds > 3600:
        hours = diff.seconds // 3600
        return f"{hours} hour{'s' if hours > 1 else ''} ago"
    elif diff.seconds > 60:
        minutes = diff.seconds // 60
        return f"{minutes} minute{'s' if minutes > 1 else ''} ago"
    else:
        return "Just now"


# ==================== SCOPE CHECKLIST VIEWS ====================

@login_required
def scope_checklist(request):
    """
    Main Scope Checklist page - Interior Inspection Work Scope
    Displays claim selection, room tabs, and detailed checklist with Xactimate codes
    """
    claims = Client.objects.all().order_by('-created_at')

    return render(request, 'account/scope_checklist.html', {
        'claims': claims
    })


@login_required
def scope_checklist_get_rooms(request, claim_id):
    """
    API endpoint to get rooms for a selected claim
    Returns room list with any saved scope checklist data
    """
    try:
        client = get_object_or_404(Client, id=claim_id)
        rooms = Room.objects.filter(client=client).order_by('sequence', 'room_name')

        rooms_data = []
        for room in rooms:
            room_info = {
                'id': str(room.id),
                'room_name': room.room_name,
                'sequence': room.sequence,
            }

            # Get saved scope checklist data if it exists
            try:
                from .models import RoomScopeChecklist
                scope_data = RoomScopeChecklist.objects.get(room=room, client=client)
                room_info['scope_data'] = scope_data.to_dict()
            except:
                room_info['scope_data'] = {}

            rooms_data.append(room_info)

        return JsonResponse({
            'success': True,
            'rooms': rooms_data,
            'client': {
                'id': client.id,
                'name': client.pOwner,
                'address': client.pAddress,
                'phone': client.cPhone,
                'insurance': client.insuranceCo_Name,
                'cause': client.causeOfLoss
            }
        })

    except Client.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Claim not found'}, status=404)
    except Exception as e:
        logger.error(f"Error getting rooms for scope checklist: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@csrf_exempt
@login_required
def scope_checklist_save(request):
    """
    API endpoint to save scope checklist data for all rooms
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    try:
        data = json.loads(request.body)
        claim_id = data.get('claim_id')
        rooms_data = data.get('rooms_data', {})

        if not claim_id:
            return JsonResponse({'error': 'Missing claim_id'}, status=400)

        client = get_object_or_404(Client, id=claim_id)
        from .models import RoomScopeChecklist

        saved_count = 0
        for room_id, room_fields in rooms_data.items():
            try:
                room = Room.objects.get(id=room_id)

                # Update or create scope checklist
                scope_data, created = RoomScopeChecklist.objects.update_or_create(
                    room=room,
                    client=client,
                    defaults={
                        'clg_material': room_fields.get('clg_material', ''),
                        'clg_construction': room_fields.get('clg_construction', ''),
                        'clg_finish': room_fields.get('clg_finish', ''),
                        'clg_activity': room_fields.get('clg_activity', ''),
                        'clg_sf': room_fields.get('clg_sf') or None,
                        'lit_type': room_fields.get('lit_type', ''),
                        'lit_activity': room_fields.get('lit_activity', ''),
                        'lit_qty': room_fields.get('lit_qty') or None,
                        'hvc_type': room_fields.get('hvc_type', ''),
                        'hvc_activity': room_fields.get('hvc_activity', ''),
                        'hvc_qty': room_fields.get('hvc_qty') or None,
                        'wal_material': room_fields.get('wal_material', ''),
                        'wal_finish': room_fields.get('wal_finish', ''),
                        'wal_activity': room_fields.get('wal_activity', ''),
                        'wal_sf': room_fields.get('wal_sf') or None,
                        'ele_outlets': room_fields.get('ele_outlets') or None,
                        'ele_switches': room_fields.get('ele_switches') or None,
                        'ele_sw3': room_fields.get('ele_sw3') or None,
                        'ele_activity': room_fields.get('ele_activity', ''),
                        'flr_type': room_fields.get('flr_type', ''),
                        'flr_activity': room_fields.get('flr_activity', ''),
                        'flr_sf': room_fields.get('flr_sf') or None,
                        'bb_height': room_fields.get('bb_height', ''),
                        'bb_activity': room_fields.get('bb_activity', ''),
                        'bb_lf': room_fields.get('bb_lf') or None,
                        'trim_type': room_fields.get('trim_type', ''),
                        'trim_crown': room_fields.get('trim_crown', ''),
                        'trim_chairrail': room_fields.get('trim_chairrail', ''),
                        'trim_activity': room_fields.get('trim_activity', ''),
                        'dor_type': room_fields.get('dor_type', ''),
                        'dor_activity': room_fields.get('dor_activity', ''),
                        'dor_qty': room_fields.get('dor_qty') or None,
                        'open_activity': room_fields.get('open_activity', ''),
                        'open_qty': room_fields.get('open_qty') or None,
                        'wdw_type': room_fields.get('wdw_type', ''),
                        'wdw_covers': room_fields.get('wdw_covers', ''),
                        'wdw_activity': room_fields.get('wdw_activity', ''),
                        'wdw_qty': room_fields.get('wdw_qty') or None,
                        'closet_type': room_fields.get('closet_type', ''),
                        'closet_rod': room_fields.get('closet_rod', ''),
                        'closet_lf': room_fields.get('closet_lf') or None,
                        'ins_type': room_fields.get('ins_type', ''),
                        'ins_rvalue': room_fields.get('ins_rvalue', ''),
                        'ins_sf': room_fields.get('ins_sf') or None,
                        'frm_type': room_fields.get('frm_type', ''),
                        'frm_activity': room_fields.get('frm_activity', ''),
                        'frm_lf': room_fields.get('frm_lf') or None,
                        'activity_notes': room_fields.get('activity_notes', ''),
                        'created_by': request.user,
                    }
                )
                saved_count += 1
            except Room.DoesNotExist:
                logger.warning(f"Room not found: {room_id}")
                continue
            except Exception as e:
                logger.error(f"Error saving room {room_id}: {str(e)}")
                continue

        return JsonResponse({
            'success': True,
            'saved_count': saved_count,
            'message': f'Saved data for {saved_count} rooms'
        })

    except Exception as e:
        logger.error(f"Error saving scope checklist: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@csrf_exempt
@login_required
def scope_checklist_generate_pdf(request):
    """
    Generate PDF for scope checklist
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    try:
        data = json.loads(request.body)
        claim_id = data.get('claim_id')
        rooms_data = data.get('rooms_data', {})

        if not claim_id:
            return JsonResponse({'error': 'Missing claim_id'}, status=400)

        client = get_object_or_404(Client, id=claim_id)

        # Generate PDF using ReportLab
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from io import BytesIO

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter),
                               rightMargin=30, leftMargin=30,
                               topMargin=30, bottomMargin=30)

        elements = []
        styles = getSampleStyleSheet()

        # Header style
        header_style = ParagraphStyle(
            'Header',
            parent=styles['Heading1'],
            fontSize=14,
            spaceAfter=12
        )

        # Title
        elements.append(Paragraph(f"Interior Inspection Work Scope", header_style))
        elements.append(Paragraph(f"<b>Client:</b> {client.pOwner}", styles['Normal']))
        elements.append(Paragraph(f"<b>Address:</b> {client.pAddress} {client.pCityStateZip}", styles['Normal']))
        elements.append(Paragraph(f"<b>Phone:</b> {client.cPhone}", styles['Normal']))
        elements.append(Paragraph(f"<b>Insurance:</b> {client.insuranceCo_Name}", styles['Normal']))
        elements.append(Paragraph(f"<b>Cause of Loss:</b> {client.causeOfLoss}", styles['Normal']))
        elements.append(Spacer(1, 20))

        # Get rooms
        rooms = Room.objects.filter(client=client).order_by('sequence', 'room_name')

        # Summary table headers
        headers = ['#', 'Room', 'CLG', 'LIT', 'HVC', 'WAL', 'ELE', 'FLR', 'BB', 'DOR', 'WDW', 'NOTES']

        # Build summary table data
        table_data = [headers]

        for idx, room in enumerate(rooms, 1):
            room_id = str(room.id)
            room_fields = rooms_data.get(room_id, {})

            row = [
                str(idx),
                room.room_name[:15],  # Truncate long names
                room_fields.get('clg_material', '') + '/' + room_fields.get('clg_activity', ''),
                room_fields.get('lit_type', ''),
                room_fields.get('hvc_type', ''),
                room_fields.get('wal_material', '') + '/' + room_fields.get('wal_activity', ''),
                f"OS:{room_fields.get('ele_outlets', '')} SW:{room_fields.get('ele_switches', '')}",
                room_fields.get('flr_type', '') + '/' + room_fields.get('flr_activity', ''),
                room_fields.get('bb_height', ''),
                room_fields.get('dor_type', ''),
                room_fields.get('wdw_type', ''),
                room_fields.get('activity_notes', '')[:30] if room_fields.get('activity_notes') else ''
            ]
            table_data.append(row)

        # Create summary table
        col_widths = [25, 80, 60, 40, 40, 60, 70, 60, 40, 40, 40, 120]
        summary_table = Table(table_data, colWidths=col_widths)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0066cc')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(summary_table)

        # Add detailed room pages
        elements.append(PageBreak())

        for idx, room in enumerate(rooms, 1):
            room_id = str(room.id)
            room_fields = rooms_data.get(room_id, {})

            elements.append(Paragraph(f"Room {idx}: {room.room_name}", header_style))
            elements.append(Spacer(1, 10))

            # Room details table
            detail_data = [
                ['SECTION', 'TYPE/MATERIAL', 'ACTIVITY', 'QTY/SF/LF', 'NOTES'],
                ['CEILING', room_fields.get('clg_material', ''), room_fields.get('clg_activity', ''), room_fields.get('clg_sf', ''), room_fields.get('clg_finish', '')],
                ['LIGHTS', room_fields.get('lit_type', ''), room_fields.get('lit_activity', ''), room_fields.get('lit_qty', ''), ''],
                ['HVAC', room_fields.get('hvc_type', ''), room_fields.get('hvc_activity', ''), room_fields.get('hvc_qty', ''), ''],
                ['WALLS', room_fields.get('wal_material', ''), room_fields.get('wal_activity', ''), room_fields.get('wal_sf', ''), room_fields.get('wal_finish', '')],
                ['ELECTRICAL', f"OS:{room_fields.get('ele_outlets', '')} SW:{room_fields.get('ele_switches', '')}", room_fields.get('ele_activity', ''), '', ''],
                ['FLOOR', room_fields.get('flr_type', ''), room_fields.get('flr_activity', ''), room_fields.get('flr_sf', ''), ''],
                ['BASEBOARD', room_fields.get('bb_height', ''), room_fields.get('bb_activity', ''), room_fields.get('bb_lf', ''), ''],
                ['TRIM', room_fields.get('trim_type', ''), room_fields.get('trim_activity', ''), '', f"CRN:{room_fields.get('trim_crown', '')} CHR:{room_fields.get('trim_chairrail', '')}"],
                ['DOORS', room_fields.get('dor_type', ''), room_fields.get('dor_activity', ''), room_fields.get('dor_qty', ''), ''],
                ['OPENINGS', '', room_fields.get('open_activity', ''), room_fields.get('open_qty', ''), ''],
                ['WINDOWS', room_fields.get('wdw_type', ''), room_fields.get('wdw_activity', ''), room_fields.get('wdw_qty', ''), room_fields.get('wdw_covers', '')],
                ['CLOSET', room_fields.get('closet_type', ''), '', room_fields.get('closet_lf', ''), f"ROD:{room_fields.get('closet_rod', '')}"],
                ['INSULATION', room_fields.get('ins_type', ''), '', room_fields.get('ins_sf', ''), room_fields.get('ins_rvalue', '')],
                ['FRAMING', room_fields.get('frm_type', ''), room_fields.get('frm_activity', ''), room_fields.get('frm_lf', ''), ''],
            ]

            detail_table = Table(detail_data, colWidths=[80, 120, 80, 80, 150])
            detail_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#333333')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#f0f0f0')),
            ]))
            elements.append(detail_table)

            # Activity notes
            if room_fields.get('activity_notes'):
                elements.append(Spacer(1, 10))
                elements.append(Paragraph(f"<b>Activity Notes:</b> {room_fields.get('activity_notes', '')}", styles['Normal']))

            elements.append(PageBreak())

        # Add scope codes reference page
        elements.append(Paragraph("XACTIMATE SCOPE CODES REFERENCE", header_style))
        elements.append(Spacer(1, 10))

        codes_data = [
            ['BUILDING MATERIALS', '', 'CONSTRUCTION TYPE', '', 'FINISH', ''],
            ['ACT', 'CLG TILES', 'FLT', 'FLAT', 'SMH', 'SMOOTH'],
            ['DRY', 'DRYWALL', 'VLT', 'VAULTED', 'TEX', 'TEXTURE'],
            ['PLA', 'PLASTER', 'TRY', 'TRAY', 'POP', 'POPCORN'],
            ['T&G', 'TONGUE&GROOVE', 'FRM', 'OPEN', '', ''],
            ['PNL', 'PANELING', '', '', '', ''],
            ['WPR', 'WALLPAPER', '', '', '', ''],
            ['CNC', 'CONCRETE', '', '', '', ''],
            ['MAS', 'BRICK/MASONRY', '', '', '', ''],
            ['', '', '', '', '', ''],
            ['FLOOR TYPES', '', 'DOOR TYPES', '', 'WINDOW TYPES', ''],
            ['FCC', 'CARPET', 'STD', 'STANDARD', 'WDW', 'WOOD'],
            ['FCS', 'STONE', 'BFD', 'BIFOLD', 'WDV', 'VINYL'],
            ['FCV', 'VINYL TILE', 'BYD', 'BYPASS/SLIDER', 'WDA', 'ALUMINUM'],
            ['FCW', 'HARDWOOD', 'BPM', 'MIRROR', 'BAY', 'BAY/BOW'],
            ['LAM', 'LAMINATE', '', '', '', ''],
            ['', '', '', '', '', ''],
            ['ACTIVITY CODES', '', '', '', '', ''],
            ['ALL', 'REPAIR ALL', 'CLN', 'CLEAN', 'R&R', 'REPLACE'],
            ['D&R', 'DETACH & RESET', 'MSK', 'MASK', 'S++', 'PRIME/SEAL'],
            ['PNT', 'PAINT', 'STN', 'STAIN & SEAL', 'SND', 'SAND'],
        ]

        codes_table = Table(codes_data, colWidths=[50, 100, 50, 100, 50, 100])
        codes_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0066cc')),
            ('BACKGROUND', (0, 10), (-1, 10), colors.HexColor('#0066cc')),
            ('BACKGROUND', (0, 17), (-1, 17), colors.HexColor('#0066cc')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('TEXTCOLOR', (0, 10), (-1, 10), colors.white),
            ('TEXTCOLOR', (0, 17), (-1, 17), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 10), (-1, 10), 'Helvetica-Bold'),
            ('FONTNAME', (0, 17), (-1, 17), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(codes_table)

        # Build PDF
        doc.build(elements)

        # Return PDF response
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="scope_checklist_{client.pOwner}.pdf"'
        return response

    except Exception as e:
        logger.error(f"Error generating scope checklist PDF: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@csrf_exempt
@login_required
def scope_checklist_send_email(request):
    """
    Send scope checklist PDF via email
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    try:
        data = json.loads(request.body)
        claim_id = data.get('claim_id')
        rooms_data = data.get('rooms_data', {})
        recipients = data.get('recipients', '')
        subject = data.get('subject', '')
        notes = data.get('notes', '')

        if not claim_id:
            return JsonResponse({'error': 'Missing claim_id'}, status=400)

        if not recipients:
            return JsonResponse({'error': 'Missing recipients'}, status=400)

        client = get_object_or_404(Client, id=claim_id)

        # Parse recipients
        recipient_list = [email.strip() for email in recipients.split(',') if email.strip()]

        if not recipient_list:
            return JsonResponse({'error': 'No valid recipients'}, status=400)

        # Generate PDF
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from io import BytesIO

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter),
                               rightMargin=30, leftMargin=30,
                               topMargin=30, bottomMargin=30)

        elements = []
        styles = getSampleStyleSheet()

        header_style = ParagraphStyle(
            'Header',
            parent=styles['Heading1'],
            fontSize=14,
            spaceAfter=12
        )

        # Build PDF content (same as generate_pdf)
        elements.append(Paragraph(f"Interior Inspection Work Scope", header_style))
        elements.append(Paragraph(f"<b>Client:</b> {client.pOwner}", styles['Normal']))
        elements.append(Paragraph(f"<b>Address:</b> {client.pAddress} {client.pCityStateZip}", styles['Normal']))
        elements.append(Paragraph(f"<b>Phone:</b> {client.cPhone}", styles['Normal']))
        elements.append(Paragraph(f"<b>Insurance:</b> {client.insuranceCo_Name}", styles['Normal']))
        elements.append(Paragraph(f"<b>Cause of Loss:</b> {client.causeOfLoss}", styles['Normal']))
        elements.append(Spacer(1, 20))

        rooms = Room.objects.filter(client=client).order_by('sequence', 'room_name')

        # Summary table
        headers = ['#', 'Room', 'CLG', 'LIT', 'HVC', 'WAL', 'ELE', 'FLR', 'BB', 'DOR', 'WDW', 'NOTES']
        table_data = [headers]

        for idx, room in enumerate(rooms, 1):
            room_id = str(room.id)
            room_fields = rooms_data.get(room_id, {})

            row = [
                str(idx),
                room.room_name[:15],
                room_fields.get('clg_material', '') + '/' + room_fields.get('clg_activity', ''),
                room_fields.get('lit_type', ''),
                room_fields.get('hvc_type', ''),
                room_fields.get('wal_material', '') + '/' + room_fields.get('wal_activity', ''),
                f"OS:{room_fields.get('ele_outlets', '')} SW:{room_fields.get('ele_switches', '')}",
                room_fields.get('flr_type', '') + '/' + room_fields.get('flr_activity', ''),
                room_fields.get('bb_height', ''),
                room_fields.get('dor_type', ''),
                room_fields.get('wdw_type', ''),
                room_fields.get('activity_notes', '')[:30] if room_fields.get('activity_notes') else ''
            ]
            table_data.append(row)

        col_widths = [25, 80, 60, 40, 40, 60, 70, 60, 40, 40, 40, 120]
        summary_table = Table(table_data, colWidths=col_widths)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0066cc')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(summary_table)

        doc.build(elements)
        buffer.seek(0)

        # Create and send email
        from django.core.mail import EmailMessage

        email_subject = subject or f"Interior Inspection Scope - {client.pOwner}"

        email_body = f"""
<html>
<body style="font-family: Arial, sans-serif;">
<h2>Interior Inspection Work Scope</h2>
<p><strong>Client:</strong> {client.pOwner}</p>
<p><strong>Address:</strong> {client.pAddress}</p>
<p><strong>Insurance:</strong> {client.insuranceCo_Name}</p>
<p><strong>Cause of Loss:</strong> {client.causeOfLoss}</p>
{f'<p><strong>Notes:</strong> {notes}</p>' if notes else ''}
<p>Please find the attached scope checklist PDF.</p>
<hr>
<p style="color: #666; font-size: 12px;">Generated by Claimet App</p>
</body>
</html>
"""

        email = EmailMessage(
            subject=email_subject,
            body=email_body,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=recipient_list
        )
        email.content_subtype = 'html'

        # Attach PDF
        pdf_filename = f"Scope_Checklist_{client.pOwner.replace(' ', '_')}.pdf"
        email.attach(pdf_filename, buffer.getvalue(), 'application/pdf')

        # Send email
        email.send()

        # Log sent email
        SentEmail.objects.create(
            subject=email_subject,
            recipients=recipient_list,
            body=email_body,
            sent_by=request.user,
            sent_at=timezone.now()
        )

        return JsonResponse({
            'success': True,
            'message': f'Email sent successfully to {len(recipient_list)} recipient(s)',
            'recipients_count': len(recipient_list)
        })

    except Exception as e:
        logger.error(f"Error sending scope checklist email: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
