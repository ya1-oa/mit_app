# docsAppR/template_generators.py

import io
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

logger = logging.getLogger('onedrive_sync')


class BaseTemplateGenerator:
    """Base class for all template generators"""

    def __init__(self, client):
        self.client = client

    def generate(self):
        """Generate template and return BytesIO object"""
        raise NotImplementedError("Subclasses must implement generate()")

    def _apply_thin_border(self, worksheet, min_row, max_row, min_col, max_col):
        """Apply thin borders to a range"""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in worksheet.iter_rows(
            min_row=min_row, max_row=max_row,
            min_col=min_col, max_col=max_col
        ):
            for cell in row:
                cell.border = thin_border


class InfoTemplateGenerator(BaseTemplateGenerator):
    """Generates 01-INFO.xlsx template"""

    FIELD_MAPPING = [
        ('Property Owner Name', 'pOwner'),
        ('Property Address: Street', 'pAddress'),
        ('Property City, State, Zip', 'pCityStateZip'),
        ('Customer Email', 'cEmail'),
        ('Customer Phone#', 'cPhone'),
        ('Co-Owner', 'coOwner2'),
        ('Co-Owner Phone', 'cPhone2'),
        ('Co-Owner Address', 'cAddress2'),
        ('Co-Owner City, State, Zip', 'cCityStateZip2'),
        ('Co-Owner Email', 'cEmail2'),
        ('Cause of Loss', 'causeOfLoss'),
        ('Date of Loss', 'dateOfLoss'),
        ('Year Built', 'yearBuilt'),
        ('Contract Date', 'contractDate'),
        ('Insurance Co. Name', 'insuranceCo_Name'),
        ('Claim #', 'claimNumber'),
        ('Policy #', 'policyNumber'),
        ('Desk Adjuster', 'deskAdjusterDA'),
        ('DA Phone', 'DAPhone'),
        ('DA Email', 'DAEmail'),
        ('Field Adjuster Name', 'fieldAdjusterName'),
        ('Field Adjuster Phone', 'phoneFieldAdj'),
        ('Field Adjuster Email', 'fieldAdjEmail'),
        ('Contractor Name', 'coName'),
        ('Contractor Website', 'coWebsite'),
        ('Contractor Address', 'coAddress'),
        ('Contractor Phone', 'coPhone'),
    ]

    def generate(self):
        """Generate the Excel file"""
        wb = Workbook()
        ws = wb.active
        ws.title = "jobinfo(2)"

        # Set column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 40

        # Header styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        # Write data
        for row_num, (label, field_name) in enumerate(self.FIELD_MAPPING, start=1):
            # Column A: Row number
            ws.cell(row=row_num, column=1, value=row_num)

            # Column B: Field label
            cell_b = ws.cell(row=row_num, column=2, value=label)
            cell_b.font = header_font
            cell_b.fill = header_fill
            cell_b.alignment = Alignment(horizontal='left', vertical='center')

            # Column C: Field value
            value = getattr(self.client, field_name, '')

            # Format value
            if value is None:
                value = ''
            elif isinstance(value, bool):
                value = 'Yes' if value else 'No'
            elif hasattr(value, 'strftime'):
                value = value.strftime('%Y-%m-%d')

            cell_c = ws.cell(row=row_num, column=3, value=str(value))
            cell_c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # Apply borders
        self._apply_thin_border(ws, 1, len(self.FIELD_MAPPING), 1, 3)

        # Freeze panes
        ws.freeze_panes = 'C2'

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        logger.info(f"Generated 01-INFO template for {self.client.pOwner}")
        return output


class RoomsTemplateGenerator(BaseTemplateGenerator):
    """Generates 01-ROOMS.xlsx template"""

    def generate(self):
        """Generate the Excel file"""
        wb = Workbook()
        ws = wb.active
        ws.title = "ROOMS#"

        # Header formatting
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        # Set column widths
        ws.column_dimensions['A'].width = 5   # Row number
        ws.column_dimensions['B'].width = 25  # Room name

        # Get work types
        from .models import WorkType
        work_types = WorkType.objects.filter(is_active=True).order_by('display_order')

        # Write headers
        headers = ['#', 'Room Name']

        # Add work type columns
        for idx, wt in enumerate(work_types, start=3):
            col_letter = chr(64 + idx)  # C, D, E, etc.
            ws.column_dimensions[col_letter].width = 12
            headers.append(f'WT{wt.work_type_id}')

        # Write header row
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Get rooms for this client
        rooms = self.client.rooms.all().order_by('sequence')

        # Write room data
        for row_num, room in enumerate(rooms, start=2):
            # Row number
            ws.cell(row=row_num, column=1, value=room.sequence)

            # Room name
            ws.cell(row=row_num, column=2, value=room.room_name)

            # Work type values
            wt_values = {}
            for wt_value in room.work_type_values.all():
                wt_values[wt_value.work_type.work_type_id] = wt_value.value_type

            # Fill in work type columns
            for col_num, wt in enumerate(work_types, start=3):
                value = wt_values.get(wt.work_type_id, 'NA')
                ws.cell(row=row_num, column=col_num, value=value)

        # Apply borders to all data
        if rooms:
            self._apply_thin_border(ws, 1, len(rooms) + 1, 1, len(headers))

        # Freeze panes (freeze headers and room name column)
        ws.freeze_panes = 'C2'

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        logger.info(f"Generated 01-ROOMS template for {self.client.pOwner} with {len(rooms)} rooms")
        return output
