"""
Excel import pipeline utilities.
Extracted from docsAppR/views.py to improve manageability.

Functions for importing client and room data from Excel files:
    import_client_from_info_file        - Parse 01-INFO.xlsx → client dict
    extract_client_data_from_jobinfo_openpyxl
    import_rooms_from_rooms_file        - Parse 01-ROOMS.xlsm → Room objects
    extract_room_data_from_rooms_sheet_openpyxl
    parse_excel_date_openpyxl
    determine_los_travel_value_enhanced
    ensure_work_types_exist
    create_rooms_for_client
    create_or_update_client
    import_from_master_insurer_file
    map_client_data_to_model
    parse_decimal
    parse_excel_date                    - Robust multi-format date parser
    clean_session_data                  - Recursively JSON-serialise arbitrary data
    extract_rooms_from_master_data
    normalize_header_for_mapping        - MASTER file exact-match header mapping
    normalize_header                    - Generic header normaliser
"""

import datetime as dt
import re
from decimal import Decimal, InvalidOperation
from io import BytesIO

import openpyxl
import pandas as pd
from django.utils import timezone
from openpyxl.utils import get_column_letter

from .models import Client, Room, WorkType, RoomWorkTypeValue

# ---------------------------------------------------------------------------
# INFO file import
# ---------------------------------------------------------------------------

def import_client_from_info_file(excel_file):
    """
    Extract client data from 01-INFO.xlsx file from jobinfo(2) tab.
    Uses data_only=True to resolve formula values.
    """
    try:
        excel_file.seek(0)
        wb = openpyxl.load_workbook(BytesIO(excel_file.read()), data_only=True)

        sheet_name = 'jobinfo(2)'
        if sheet_name not in wb.sheetnames:
            possible_sheets = [s for s in wb.sheetnames if 'jobinfo' in s.lower()]
            if possible_sheets:
                sheet_name = possible_sheets[0]
            else:
                raise ValueError(
                    f"No jobinfo sheet found in INFO file. Available sheets: {', '.join(wb.sheetnames)}"
                )

        ws = wb[sheet_name]
        client_data = extract_client_data_from_jobinfo_openpyxl(ws)
        wb.close()
        return client_data

    except Exception as e:
        raise Exception(f"Failed to process INFO file: {str(e)}")


def extract_client_data_from_jobinfo_openpyxl(worksheet):
    """
    Extract client data from jobinfo worksheet using openpyxl.
    Gets calculated values from formulas.
    """
    client_data = {}

    HEADER_COLUMN = 1
    DATA_COLUMN = 2

    for row in range(1, worksheet.max_row + 1):
        header_cell = worksheet.cell(row=row, column=HEADER_COLUMN)
        value_cell = worksheet.cell(row=row, column=DATA_COLUMN)

        header = str(header_cell.value).strip() if header_cell.value is not None else None
        value = value_cell.value

        if not header or header in ['None', 'nan', ''] or value is None:
            continue

        field_name = (header.lower()
                      .replace(' ', '_')
                      .replace('/', '_')
                      .replace('\\', '_')
                      .replace('.', '_')
                      .replace('-', '_')
                      .replace(':', '_')
                      .replace('__', '_')
                      .replace('#', 'num')
                      .strip('_'))

        if any(term in field_name for term in ['date', 'dol']):
            parsed_date = parse_excel_date_openpyxl(value)
            if parsed_date:
                if 'loss' in field_name:
                    client_data['date_of_loss'] = parsed_date
                else:
                    client_data[field_name] = parsed_date
            else:
                client_data[field_name] = value
            continue

        if isinstance(value, str) and value.lower() in ('yes', 'no', 'true', 'false', 'y', 'n'):
            value = value.lower() in ('yes', 'true', 'y')
        elif isinstance(value, (int, float)) and not isinstance(value, bool) and value in (0, 1):
            value = bool(value)

        client_data[field_name] = value

    return client_data


# ---------------------------------------------------------------------------
# ROOMS file import
# ---------------------------------------------------------------------------

def import_rooms_from_rooms_file(excel_file, client):
    """
    Extract room data from 01-ROOMS.xlsm file from ROOMS# tab.
    Uses data_only=True to resolve formula values.
    """
    try:
        excel_file.seek(0)
        wb = openpyxl.load_workbook(BytesIO(excel_file.read()), data_only=True)

        if 'ROOMS#' not in wb.sheetnames:
            raise ValueError(
                f"ROOMS# sheet not found in ROOMS file. Available sheets: {', '.join(wb.sheetnames)}"
            )

        ws = wb['ROOMS#']
        rooms_data = extract_room_data_from_rooms_sheet_openpyxl(ws)
        wb.close()

        if not rooms_data:
            raise ValueError("No room data found in ROOMS# sheet")

        rooms_created, wt_values_created = create_rooms_for_client(client, rooms_data)

        return {
            'rooms_processed': rooms_created,
            'work_type_values_created': wt_values_created,
            'total_rooms_found': len(rooms_data),
            'rooms_sample': rooms_data[:3] if rooms_data else []
        }

    except Exception as e:
        raise Exception(f"Failed to process ROOMS file: {str(e)}")


def extract_room_data_from_rooms_sheet_openpyxl(worksheet):
    """
    Extract room data using openpyxl with formula support.
    Columns U-BT, with each work type section being 5 columns apart.
    """
    rooms_data = []

    work_type_sections = {
        100: range(21, 26),
        200: range(26, 31),
        300: range(31, 36),
        400: range(36, 41),
        500: range(41, 46),
        800: range(46, 51),
        6100: range(51, 56),
        6200: range(56, 61),
        6300: range(61, 66),
        6400: range(66, 71),
    }

    max_rows_to_scan = min(worksheet.max_row, 200)

    for row in range(1, max_rows_to_scan + 1):
        first_section_cols = work_type_sections[100]
        room_name_parts = []

        for col in first_section_cols:
            if col <= worksheet.max_column:
                cell = worksheet.cell(row=row, column=col)
                if cell.value is not None and str(cell.value).strip():
                    room_name_parts.append(str(cell.value).strip())

        if room_name_parts:
            room_name = ' '.join(room_name_parts).strip()
            if (room_name and
                    room_name not in ['', 'Room Name', 'Room', 'ROOM NAME', 'ROOM'] and
                    not any(kw in room_name.upper() for kw in ['HEADER', 'TITLE', 'DESCRIPTION'])):

                room_data = {
                    'room_name': room_name,
                    'sequence': len(rooms_data) + 1,
                    'work_type_values': {}
                }

                for wt_id, col_range in work_type_sections.items():
                    value_col = list(col_range)[4]
                    if value_col <= worksheet.max_column:
                        cell = worksheet.cell(row=row, column=value_col)
                        if cell.value is not None and str(cell.value).strip():
                            value_str = str(cell.value).strip().upper()
                            value_type = determine_los_travel_value_enhanced(value_str)
                            if value_type != 'NA':
                                room_data['work_type_values'][wt_id] = value_type

                rooms_data.append(room_data)

    return rooms_data


# ---------------------------------------------------------------------------
# Date parsers
# ---------------------------------------------------------------------------

def parse_excel_date_openpyxl(value):
    """Enhanced date parser for openpyxl values."""
    if value is None:
        return None

    try:
        if isinstance(value, (dt.datetime, dt.date)):
            return value.date() if isinstance(value, dt.datetime) else value

        if isinstance(value, (int, float)):
            if value < 0 or value == 0:
                return None
            if value == 60:
                return dt.date(1900, 2, 28)
            if value < 60:
                value += 1
            return (dt.datetime(1899, 12, 30) + dt.timedelta(days=value)).date()

        if isinstance(value, str):
            value = value.strip()
            if not value or value.upper() in ('TBD', 'NA', 'N/A', 'UNKNOWN'):
                return None

            value = (value.replace('  ', ' ')
                     .replace('Sept', 'Sep')
                     .replace('Febr', 'Feb')
                     .replace('Dece', 'Dec')
                     .split(' ')[0])

            date_formats = [
                '%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%m/%d/%Y',
                '%d-%b-%y', '%d-%b-%Y', '%b %d, %Y', '%B %d, %Y',
                '%d-%m-%Y', '%d %b %Y', '%d %B %Y'
            ]
            for fmt in date_formats:
                try:
                    return dt.datetime.strptime(value, fmt).date()
                except ValueError:
                    continue

        return None

    except Exception:
        return None


def parse_excel_date(value):
    """Robust date parser that handles all cases."""
    try:
        if pd.isna(value) or value in ('', 'TBD', 'NA', 'N/A'):
            return None
    except (TypeError, ValueError):
        pass

    try:
        if isinstance(value, (int, float)):
            if value < 0 or value == 0:
                return None
            if value == 60:
                return dt.datetime(1900, 2, 28).date()
            if value < 60:
                value += 1
            return (dt.datetime(1899, 12, 30) + pd.Timedelta(days=value)).date()

        if not isinstance(value, str):
            value = str(value).strip()
        else:
            value = value.strip()

        if not value or value.upper() in ('TBD', 'NA', 'N/A', 'UNKNOWN'):
            return None

        value = (value.replace('\xa0', ' ')
                 .replace('Sept', 'Sep')
                 .replace('Febr', 'Feb')
                 .replace('Dece', 'Dec')
                 .split(' ')[0])

        date_formats = [
            '%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%m/%d/%Y',
            '%d-%b-%y', '%d-%b-%Y', '%b %d, %Y', '%B %d, %Y',
            '%d-%m-%Y', '%d %b %Y', '%d %B %Y'
        ]
        for fmt in date_formats:
            try:
                return dt.datetime.strptime(value, fmt).date()
            except ValueError:
                continue

        return None

    except Exception:
        return None


# ---------------------------------------------------------------------------
# Value determination
# ---------------------------------------------------------------------------

def determine_los_travel_value_enhanced(value_str):
    """Enhanced value determination for LOS/TRAVEL/NA."""
    if value_str is None:
        return 'NA'

    if isinstance(value_str, (int, float)):
        return 'LOS'

    value_clean = str(value_str).upper().strip()

    if not value_clean or value_clean in ['', 'N/A', 'NA', 'NULL', 'NONE', 'NAN']:
        return 'NA'

    if any(t in value_clean for t in ['TRAVEL', 'TRVL', 'TRV', 'TRAV', 'TRAVEL/AREA']):
        return 'TRAVEL'

    if any(t in value_clean for t in ['LOS', 'LINE OF SIGHT', 'SIGHT', 'L.O.S']):
        return 'LOS'

    try:
        numeric_value = ''.join(c for c in value_clean if c.isdigit() or c in '.-')
        if numeric_value and numeric_value not in ('-', '.'):
            float(numeric_value)
            return 'LOS'
    except ValueError:
        pass

    if any(char.isdigit() for char in value_clean):
        return 'LOS'

    return 'NA'


# ---------------------------------------------------------------------------
# Database helpers
# ---------------------------------------------------------------------------

def ensure_work_types_exist():
    """Make sure all work types are created in the database."""
    work_types = {}

    work_type_definitions = [
        (100, 'Overview'), (200, 'Source'), (300, 'CPS'),
        (400, 'PPR'), (500, 'Demo'), (600, 'Mitigation'), (700, 'HMR'),
        (6100, 'DAY 1 MC Readings'), (6200, 'DAY 2 MC Readings'),
        (6300, 'DAY 3 MC Readings'), (6400, 'DAY 4 MC Readings'),
    ]

    for wt_id, wt_name in work_type_definitions:
        work_type, _ = WorkType.objects.get_or_create(
            work_type_id=wt_id,
            defaults={'name': wt_name}
        )
        work_types[wt_id] = work_type

    return work_types


def create_rooms_for_client(client, rooms_data):
    """Create room and work type data for a client."""
    work_types = ensure_work_types_exist()

    existing_rooms_count = client.rooms.count()
    if existing_rooms_count > 0:
        client.rooms.all().delete()

    rooms_created = 0
    work_type_values_created = 0

    for room_info in rooms_data:
        room = Room.objects.create(
            client=client,
            room_name=room_info['room_name'],
            sequence=room_info['sequence']
        )
        rooms_created += 1

        for wt_id, value_type in room_info.get('work_type_values', {}).items():
            RoomWorkTypeValue.objects.create(
                room=room,
                work_type=work_types[wt_id],
                value_type=value_type
            )
            work_type_values_created += 1

    return rooms_created, work_type_values_created


def create_or_update_client(client_data):
    """Create or update a Client based on extracted data."""
    mapped_data = map_client_data_to_model(client_data)

    owner_name = mapped_data.get('pOwner', '')
    if not owner_name:
        raise ValueError("No property owner name found in INFO file")

    existing_client = Client.objects.filter(pOwner=owner_name).first()

    if existing_client:
        update_fields = {}
        for field, new_value in mapped_data.items():
            if hasattr(existing_client, field) and new_value is not None:
                if getattr(existing_client, field) != new_value:
                    update_fields[field] = new_value

        if update_fields:
            Client.objects.filter(pk=existing_client.pk).update(**update_fields)
            client_data['_action'] = 'updated'
        else:
            client_data['_action'] = 'unchanged'

        existing_client.refresh_from_db()
        return existing_client
    else:
        client = Client.objects.create(**mapped_data)
        client_data['_action'] = 'created'
        return client


# ---------------------------------------------------------------------------
# MASTER-Insurer file import
# ---------------------------------------------------------------------------

def import_from_master_insurer_file(excel_file):
    """
    Import multiple clients from MASTER-Insurer format.
    Column C = field names, Columns D onward = client data (one client per column).
    """
    try:
        excel_file.seek(0)
        wb = openpyxl.load_workbook(BytesIO(excel_file.read()), data_only=True)
        ws = wb.active

        HEADER_COLUMN = 3  # Column C
        field_mapping = {}

        for row in range(1, ws.max_row + 1):
            header_cell = ws.cell(row=row, column=HEADER_COLUMN)
            if header_cell.value and str(header_cell.value).strip():
                raw_header = str(header_cell.value).strip()
                field_name = normalize_header_for_mapping(raw_header)
                field_mapping[row] = field_name

        clients_data = []

        for col in range(4, ws.max_column + 1):
            client_data = {}
            has_data = False

            for row, field_name in field_mapping.items():
                value_cell = ws.cell(row=row, column=col)
                value = value_cell.value

                if value is not None and str(value).strip():
                    has_data = True

                    if any(date_term in field_name for date_term in ['date', 'dol']):
                        parsed_date = parse_excel_date_openpyxl(value)
                        client_data[field_name] = parsed_date if parsed_date else value
                    elif isinstance(value, str) and value.upper() in ('Y', 'N', 'YES', 'NO'):
                        client_data[field_name] = value.upper() in ('Y', 'YES')
                    else:
                        client_data[field_name] = value

            if has_data:
                rooms = []
                for i in range(1, 26):
                    room_field = f'room_area_{i}'
                    if room_field in client_data:
                        room_name = client_data.pop(room_field)
                        if room_name and str(room_name).strip():
                            rooms.append({'room_name': str(room_name).strip(), 'sequence': i})

                if rooms:
                    client_data['_rooms'] = rooms

                clients_data.append(client_data)

        wb.close()
        return clients_data

    except Exception as e:
        raise Exception(f"Failed to process MASTER-Insurer file: {str(e)}")


# ---------------------------------------------------------------------------
# Data mapping
# ---------------------------------------------------------------------------

def map_client_data_to_model(raw_data):
    """Map extracted data to Client model fields."""

    def get_val(key, default=''):
        value = raw_data.get(key)
        try:
            if value is None or pd.isna(value):
                return default
        except (TypeError, ValueError):
            pass
        if isinstance(value, str) and not value.strip():
            return default
        return value

    def get_bool(key, default=False):
        value = get_val(key)
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            return value.upper() in ['Y', 'YES', 'TRUE', '1']
        if isinstance(value, (int, float)):
            return bool(value)
        return default

    def get_date(key):
        value = get_val(key)
        if not value:
            return None
        if isinstance(value, dt.datetime):
            return value
        if isinstance(value, dt.date):
            return dt.datetime.combine(value, dt.time.min)
        date_val = parse_excel_date(value)
        if date_val:
            return dt.datetime.combine(date_val, dt.time.min)
        return None

    return {
        'pOwner': get_val('property_owner_name'),
        'pAddress': get_val('property_address_street'),
        'pCityStateZip': get_val('property_city_state_zip'),
        'cEmail': get_val('customer_email'),
        'cPhone': get_val('cst_owner_phonenum'),
        'coOwner2': get_val('co_owner_cst2'),
        'cPhone2': get_val('cst_ph_num_2'),
        'cAddress2': get_val('cst_address_num_2'),
        'cCityStateZip2': get_val('cst_city_state_zip_2'),
        'cEmail2': get_val('email_cst_num_2'),
        'causeOfLoss': get_val('cause_of_loss_2'),
        'dateOfLoss': get_date('date_of_loss_2') or timezone.now(),
        'rebuildType1': get_val('rebuild_type_1'),
        'rebuildType2': get_val('rebuild_type_2'),
        'rebuildType3': get_val('rebuild_type_3'),
        'demo': get_bool('demo'),
        'mitigation': get_bool('mitigation'),
        'otherStructures': get_bool('other_structures'),
        'replacement': get_bool('replacement'),
        'CPSCLNCONCGN': get_bool('cps_cln_con_cgn'),
        'yearBuilt': get_val('year_built'),
        'contractDate': get_date('contract_date') or timezone.now(),
        'breathingIssue': get_val('breathing_issue'),
        'hazardMaterialRemediation': get_val('hmr'),
        'insuranceCo_Name': get_val('insurance_co_name'),
        'claimNumber': get_val('claim_num'),
        'policyNumber': get_val('policy_num'),
        'emailInsCo': get_val('email_ins_co'),
        'deskAdjusterDA': get_val('desk_adjuster_da'),
        'DAPhone': get_val('da_phone'),
        'DAPhExt': get_val('da_ph_ext_num'),
        'DAEmail': get_val('da_email'),
        'fieldAdjusterName': get_val('field_adjuster_name'),
        'phoneFieldAdj': get_val('phone_num_field_adj'),
        'fieldAdjEmail': get_val('field_adj_email'),
        'adjContents': get_val('adj_contents'),
        'adjCpsPhone': get_val('adj_cps_phone_num'),
        'adjCpsEmail': get_val('adj_cps_email'),
        'emsAdj': get_val('tmp_adj'),
        'emsAdjPhone': get_val('tmp_adj_phone_num'),
        'emsTmpEmail': get_val('adj_tmp_email'),
        'attLossDraftDept': get_val('att_loss_draft_dept'),
        'insAddressOvernightMail': get_val('address_ins_overnight_mail'),
        'insCityStateZip': get_val('city_state_zip_ins'),
        'insuranceCoPhone': get_val('insurance_co_phone'),
        'insWebsite': get_val('website_ins_co'),
        'insMailingAddress': get_val('mailing_address_ins'),
        'insMailCityStateZip': get_val('mail_city_state_zip_ins'),
        'mortgageCoFax': get_val('fax_ins_co'),
        'newCustomerID': get_val('new_customer_num'),
        'roomID': get_val('room_id'),
        'mortgageCo': get_val('mortgage_co'),
        'mortgageAccountCo': get_val('account_num_mtge_co'),
        'mortgageContactPerson': get_val('contact_person_mtge'),
        'mortgagePhoneContact': get_val('phone_num_mtge_contact'),
        'mortgagePhoneExtContact': get_val('ph_ext_mtge_contact'),
        'mortgageAttnLossDraftDept': get_val('attn_loss_draft_dept'),
        'mortgageOverNightMail': get_val('mtge_ovn_mail'),
        'mortgageCityStZipOVN': get_val('city_st_zip_mtge_ovn'),
        'mortgageEmail': get_val('email_mtge'),
        'mortgageWebsite': get_val('mtge_website'),
        'mortgageInitialOfferPhase1ContractAmount': get_val('initial_offer_phase_1_contract_amount'),
        'drawRequest': get_val('draw_request'),
        'coName': get_val('co_name'),
        'coWebsite': get_val('co_website'),
        'coEmailstatus': get_val('co_emailstatus'),
        'coAddress': get_val('co_adress'),
        'coCityState': get_val('co_city_state'),
        'coAddress2': get_val('co_address_2'),
        'coCityState2': get_val('co_city_state_2'),
        'coCityState3': get_val('co_city_state_3'),
        'coLogo1': get_val('co_logo_1'),
        'coLogo2': get_val('co_logo_2'),
        'coLogo3': get_val('co_logo_3'),
        'coRepPH': get_val('co_rep_ph'),
        'coREPEmail': get_val('co_rep_email'),
        'coPhone2': get_val('co_ph_num_2'),
        'TinW9': get_val('tin_w9'),
        'fedExAccount': get_val('fedex_account_num'),
        'claimReportDate': get_date('claim_report_date') or timezone.now(),
        'insuranceCustomerServiceRep': get_val('co_represesntative'),
        'timeOfClaimReport': get_val('time_of_claim_report'),
        'phoneExt': get_val('phone_ext'),
        'tarpExtTMPOk': get_bool('tarp_ext_tmp_ok'),
        'IntTMPOk': get_bool('int_tmp_ok'),
        'DRYPLACUTOUTMOLDSPRAYOK': get_bool('drypla_cutout_mold_spray_ok'),
        'lossOfUseALE': get_val('ale_info'),
        'ale_lessee_name': get_val('tenant_lesee'),
        'ale_lessee_home_address': get_val('property_address_street_ale'),
        'ale_lessee_city_state_zip': get_val('property_city_state_zip_ale'),
        'ale_lessee_email': get_val('customer_email_ale'),
        'ale_lessee_phone': get_val('cst_owner_phonenum_ale'),
        'ale_rental_bedrooms': get_val('bedrooms'),
        'ale_rental_months': get_val('months'),
        'ale_rental_start_date': get_date('start_date') or timezone.now(),
        'ale_rental_end_date': get_date('end_date') or timezone.now(),
        'ale_rental_amount_per_month': parse_decimal(get_val('terms_amount')),
        'ale_lessor_name': get_val('lessor'),
        'ale_lessor_leased_address': get_val('leased_address'),
        'ale_lessor_city_zip': get_val('city_zip_lessor'),
        'ale_lessor_phone': get_val('phone_lessor'),
        'ale_lessor_email': get_val('email_lessor'),
        'ale_lessor_mailing_address': get_val('lessor_mailing_address'),
        'ale_lessor_mailing_city_zip': get_val('city_zip_lessor_mail'),
        'ale_lessor_contact_person': get_val('lessor_contact_person'),
        'ale_re_company_name': get_val('real_estate_company'),
        'ale_re_mailing_address': get_val('mailing_address_re'),
        'ale_re_city_zip': get_val('city_zip_re'),
        'ale_re_contact_person': get_val('contact_re'),
        'ale_re_phone': get_val('phone_re'),
        'ale_re_email': get_val('email_re'),
        'ale_re_owner_broker_name': get_val('owner_broker'),
        'ale_re_owner_broker_phone': get_val('phone_owner_broker'),
        'ale_re_owner_broker_email': get_val('email_owner_broker'),
    }


def parse_decimal(value):
    """Parse string to decimal for monetary values."""
    if not value:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass

    try:
        if isinstance(value, (int, float)):
            return Decimal(str(value))
        if isinstance(value, str):
            value = value.replace('$', '').replace(',', '').strip()
            if not value or value.upper() in ['NA', 'N/A', 'NULL', '']:
                return None
            return Decimal(value)
        return Decimal(str(value))
    except (ValueError, TypeError, InvalidOperation):
        return None


def clean_session_data(data):
    """Recursively convert non-serializable objects to JSON-serializable formats."""
    if isinstance(data, (dt.datetime, dt.date, dt.time)):
        return data.isoformat()
    elif isinstance(data, pd.Timestamp):
        return data.isoformat()
    elif isinstance(data, pd.Series):
        return data.to_dict()
    elif isinstance(data, pd.DataFrame):
        return data.to_dict(orient='records')
    elif isinstance(data, dict):
        return {k: clean_session_data(v) for k, v in data.items()}
    elif isinstance(data, (list, tuple, set)):
        return [clean_session_data(x) for x in data]
    elif hasattr(data, 'tolist'):
        return data.tolist()
    elif hasattr(data, 'isoformat'):
        return data.isoformat()
    elif isinstance(data, (int, float, str, bool)) or data is None:
        return data
    else:
        return str(data)


def extract_rooms_from_master_data(client_data):
    """Extract rooms from MASTER file client data for Room model."""
    rooms_data = []
    room_name_counts = {}

    for room_num in range(1, 26):
        room_field = f'room_area_{room_num}'
        room_value = client_data.get(room_field)

        if room_value:
            room_value = str(room_value).strip() if not isinstance(room_value, str) else room_value.strip()

        if (room_value and
                room_value not in ['', 'NA', 'N/A', 'None', 'nan'] and
                len(room_value) > 1):

            original_name = room_value
            if original_name in room_name_counts:
                room_name_counts[original_name] += 1
                room_value = f"{original_name} ({room_name_counts[original_name]})"
            else:
                room_name_counts[original_name] = 1

            rooms_data.append({
                'room_name': room_value,
                'sequence': room_num,
                'work_type_values': {}
            })

    return rooms_data


# ---------------------------------------------------------------------------
# Header normalisation
# ---------------------------------------------------------------------------

def normalize_header_for_mapping(header):
    """Normalize Excel header to field name for MASTER files (exact-match first)."""
    if not header:
        return ""

    try:
        if pd.isna(header):
            return ""
    except (TypeError, ValueError):
        pass

    header_str = str(header).strip()

    header_mapping = {
        # Property Owner Information
        'Property-Owner Name': 'property_owner_name',
        'Property address: street': 'property_address_street',
        'Property city, state, zip': 'property_city_state_zip',
        'Customer Email': 'customer_email',
        'Cst-owner Phone#': 'cst_owner_phonenum',
        # Co-Owner Information
        'Co-Owner.cst#2': 'co_owner_cst2',
        'cst ph # 2': 'cst_ph_num_2',
        'Cst address # 2': 'cst_address_num_2',
        'city, state-cst#2': 'cst_city_state_zip_2',
        'email-cst #2': 'email_cst_num_2',
        # Claim Information
        'Cause of Loss': 'cause_of_loss',
        'date of loss': 'date_of_loss',
        'rebuild  type 1': 'rebuild_type_1',
        'rebuild  type 2': 'rebuild_type_2',
        'rebuild  type 3': 'rebuild_type_3',
        'DEMO': 'demo',
        'Mitigation': 'mitigation',
        'Other Structures': 'other_structures',
        'Replacement': 'replacement',
        'CPS / CLN / CON/ CGN': 'cps_cln_con_cgn',
        'Year Built': 'year_built',
        'Contract Date': 'contract_date',
        'Loss of use/ ALE': 'loss_of_use_ale',
        'Breathing issue': 'breathing_issue',
        'HMR': 'hmr',
        # Insurance Information
        'Insurance Co. Name': 'insurance_co_name',
        'Claim #': 'claim_num',
        'policy #': 'policy_num',
        'Email INS. co.': 'email_ins_co',
        'DESK Adjuster DA': 'desk_adjuster_da',
        'DA Phone': 'da_phone',
        'DA Ph. Ext. #': 'da_ph_ext_num',
        'DA Email': 'da_email',
        'Field Adjuster Name': 'field_adjuster_name',
        'Phone # field adj': 'phone_num_field_adj',
        'Field adj email': 'field_adj_email',
        'adj contents': 'adj_contents',
        'adj CPS phone #': 'adj_cps_phone_num',
        'adj CPS email': 'adj_cps_email',
        'TMP adj': 'tmp_adj',
        'TMP adj phone #': 'tmp_adj_phone_num',
        'adj TMP email': 'adj_tmp_email',
        'ATT: Loss Draft Dept.': 'att_loss_draft_dept',
        'address ins overnight mail': 'address_ins_overnight_mail',
        'city, state-zip ins': 'city_state_zip_ins',
        'Insurance Co. Phone': 'insurance_co_phone',
        'Website Ins Co.': 'website_ins_co',
        'Mailing   address INS': 'mailing_address_ins',
        'Mail city, state, zip INS': 'mail_city_state_zip_ins',
        'FAX Ins. Co': 'fax_ins_co',
        # Rooms
        'NEW CUSTOMER #': 'new_customer_num',
        'ROOM ID': 'room_id',
        **{f'Room/Area {i}': f'room_area_{i}' for i in range(1, 26)},
        # Mortgage Information
        'Mortgage co': 'mortgage_co',
        'Account# Mtge Co.': 'account_num_mtge_co',
        'Loan status': 'loan_status',
        'contact person mtge': 'contact_person_mtge',
        'Phone # MTGE contact': 'phone_num_mtge_contact',
        'Ph. Ext. Mtge contact': 'ph_ext_mtge_contact',
        'Attn.: Loss Draft Dept': 'attn_loss_draft_dept',
        'Mtge OVN mail': 'mtge_ovn_mail',
        'city, St., zip ,mtge OVN': 'city_st_zip_mtge_ovn',
        'email mtge': 'email_mtge',
        'mtge website': 'mtge_website',
        'MTGE co. Fax #': 'mtge_co_fax_num',
        'Mailing   address mtge': 'mailing_address_mtge',
        'Initial Offer / phase 1 contract amount': 'initial_offer_phase_1_contract_amount',
        # Cash Flow
        'Draw Request': 'draw_request',
        'Cust id': 'cust_id',
        # Contractor Information
        'co name': 'co_name',
        'Co. website': 'co_website',
        'co. EMAIL/co. status': 'co_emailstatus',
        'co address': 'co_adress',
        'co. city state': 'co_city_state',
        'co. address 2': 'co_address_2',
        'co. city state 2': 'co_city_state_2',
        'co address 3': 'co_address_3',
        'co. city state 3': 'co_city_state_3',
        'Co. logo 1': 'co_logo_1',
        'Co. logo 2': 'co_logo_2',
        'Co. logo 3': 'co_logo_3',
        'Co. REP. / PH': 'co_rep_ph',
        'CO.REP. email': 'co_rep_email',
        'Co PH # 2': 'co_ph_num_2',
        'TIN W9': 'tin_w9',
        'FedEx     account #': 'fedex_account_num',
        # Claim Reporting
        'claim report date': 'claim_report_date',
        'Time OF CLAIM REPORT': 'time_of_claim_report',
        'co.represesntative': 'co_represesntative',
        'phone ext.': 'phone_ext',
        'Tarp ext. TMP ok': 'tarp_ext_tmp_ok',
        'Int TMP ok': 'int_tmp_ok',
        'DRY/PLA CUTOUT MOLD SPRAY  OK': 'drypla_cutout_mold_spray_ok',
        # ALE Information
        'ALE INFO  …            APC #': 'ale_info',
        'Lesse info / NAME': 'tenant_lesee',
        'HOME ADDRESS': 'property_address_street_ale',
        'Customer Phone#': 'cst_owner_phonenum_ale',
        'bedrooms': 'bedrooms',
        'months': 'months',
        'START DATE': 'start_date',
        'END DATE': 'end_date',
        'Amount / Month': 'terms_amount',
        # Lessor Information
        'LESSOR INFO / NAME': 'lessor',
        'Leased Address': 'leased_address',
        'Email lessor': 'email_lessor',
        'Lessor mailing Address': 'lessor_mailing_address',
        'LESSOR CONTACT PERSON': 'lessor_contact_person',
        # Real Estate Company
        'REAL ESTATE COMPANY': 'real_estate_company',
        'MAILING ADDRESS': 'mailing_address_re',
        'OWNER/BROKER': 'owner_broker',
    }

    if header_str in header_mapping:
        return header_mapping[header_str]

    # Fallback: generic normalisation
    field_name = header_str.lower()
    field_name = (field_name
                  .replace(' ', '_').replace('-', '_').replace('#', 'num')
                  .replace(':', '').replace('.', '').replace(',', '')
                  .replace('__', '_').strip('_'))
    return field_name


def normalize_header(header):
    """Normalize Excel header to field name (generic)."""
    if not header:
        return ""

    try:
        if pd.isna(header):
            return ""
    except (TypeError, ValueError):
        pass

    header_str = str(header).strip()
    field_name = (header_str.lower()
                  .replace(' ', '_').replace('/', '_').replace('\\', '_')
                  .replace('.', '_').replace('-', '_').replace(':', '_')
                  .replace('__', '_').replace('#', 'num').strip('_'))
    return field_name
